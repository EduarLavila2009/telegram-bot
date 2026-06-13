"""Aplicacion Telegram: texto/voz, Excel y restriccion por usuario."""

from __future__ import annotations

import asyncio
import logging
from logging.handlers import RotatingFileHandler
import re
import tempfile
import time
import unicodedata
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path

from telegram import InlineKeyboardButton, InlineKeyboardMarkup, KeyboardButton, ReplyKeyboardMarkup, Update
from telegram.error import Conflict, NetworkError, TelegramError, TimedOut
from telegram.ext import Application, CallbackQueryHandler, CommandHandler, ContextTypes, MessageHandler, filters

from . import config
from . import excel_store
from . import tributario_engine
from . import user_manager
from .factura_compra_parse import parse_factura_compra_text, _take_eol_label, _normalize_rif
from .transcription import transcribe_audio_file
from openpyxl import load_workbook

class CompanyContext:
    def __init__(self, user_id: int | str | None = None):
        self.user_id = str(user_id) if user_id is not None else None
        self.is_custom = False
        self.company_id = None
        self.company_name = "SUFEVICA"
        self.company_rif = config.EMITTER_RIF
        self.company_type = "Especial"
        self.company_email = config.DEFAULT_ACCOUNTANT_EMAIL or ""
        self.company_phone = ""
        self.company_address = ""
        self.color_primary = "#1A1B54" # Navy for SUFEVICA
        self.dir_path = Path(__file__).resolve().parent
        
        # Default SUFEVICA paths
        self.excel_path = config.EXCEL_PATH
        self.facturas_compra_path = config.FACTURAS_COMPRA_PATH
        self.facturas_recibidas_path = config.FACTURAS_RECIBIDAS_PATH
        self.facturas_emitidas_path = config.FACTURAS_EMITIDAS_PATH
        self.reportes_z_path = config.REPORTES_Z_PATH
        self.productos_path = config.PRODUCTOS_PATH
        self.retenciones_emitidas_dir = config.RETENCIONES_EMITIDAS_DIR
        self.retenciones_islr_dir = config.RETENCIONES_ISLR_DIR
        self.firma_sello_path = config.FIRMA_SELLO_PATH
        self.generados_dir = Path(__file__).resolve().parent / "modulo_cotizaciones" / "generados"
        self.historico_json_path = Path(__file__).resolve().parent / "modulo_cotizaciones" / "historico_documentos.json"
        
        if self.user_id:
            user = user_manager.get_user(self.user_id)
            if user and user.get("role") == "nueva_empresa":
                self.is_custom = True
                self.company_id = f"flashtax_{self.user_id}"
                self.company_name = user.get("company_name", "FlashTax")
                self.company_rif = user.get("company_rif", "J-00000000-0")
                self.company_type = user.get("company_type", "Especial")
                self.company_email = user.get("company_email", "")
                self.company_phone = user.get("company_phone", "")
                self.company_address = user.get("company_address", "")
                self.color_primary = "#4F46E5" # Indigo for FlashTax
                
                # Custom paths under empresas/flashtax_<user_id>/
                base_dir = Path(__file__).resolve().parent / "empresas" / self.company_id
                self.dir_path = base_dir
                
                self.excel_path = base_dir / "RETEN-REC.xlsx"
                self.facturas_compra_path = base_dir / "facturas_compra_recibidas.xlsx"
                self.facturas_recibidas_path = base_dir / "FACTURAS-RECIBIDAS-NUEVO.xlsx"
                self.facturas_emitidas_path = base_dir / "FACTURAS-EMITIDAS.xlsx"
                self.reportes_z_path = base_dir / "REPORTES-Z-NUEVO.xlsx"
                self.productos_path = base_dir / "inventario.xlsx"
                self.retenciones_emitidas_dir = base_dir / "RETENCIONES-EMITIDAS-NUEVO"
                self.retenciones_islr_dir = base_dir / "RETENCIONES-ISLR-EMITIDAS"
                self.generados_dir = base_dir / "generados"
                self.historico_json_path = base_dir / "historico_documentos.json"
                
                # Check for custom signature in company dir
                custom_firma = base_dir / "firma_sello_transparente.png"
                if custom_firma.exists():
                    self.firma_sello_path = custom_firma
                else:
                    self.firma_sello_path = config.FIRMA_SELLO_PATH

    def ensure_files(self) -> None:
        """Asegura que el directorio y los archivos Excel base existan."""
        if not self.is_custom:
            return
        
        self.dir_path.mkdir(parents=True, exist_ok=True)
        self.retenciones_emitidas_dir.mkdir(parents=True, exist_ok=True)
        self.retenciones_islr_dir.mkdir(parents=True, exist_ok=True)
        self.generados_dir.mkdir(parents=True, exist_ok=True)
        
        # Inicializar excels base usando excel_store
        excel_store.ensure_workbook(self.excel_path)
        excel_store.ensure_factura_compra_workbook(self.facturas_compra_path)
        excel_store.ensure_factura_compra_workbook(self.facturas_recibidas_path)
        excel_store.ensure_ventas_workbook(self.facturas_emitidas_path)
        excel_store.ensure_reporte_z_nuevo_workbook(self.reportes_z_path)
        
        # Inicializar inventario.xlsx si no existe
        if not self.productos_path.exists():
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Productos"
            ws.append(["Codigo", "Descripcion", "Precio", "Barras"])
            wb.save(self.productos_path)
            wb.close()

def _get_company_context(update: Update | None) -> CompanyContext:
    uid = None
    if update and update.effective_user:
        uid = update.effective_user.id
    ctx = CompanyContext(uid)
    if ctx.is_custom:
        ctx.ensure_files()
    return ctx

# Configuración inicial del logger raíz. Se sobreescribirá en _setup_logging() con el archivo de log.
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)


def _setup_logging() -> None:
    """Configura el logging para escribir tanto en la consola como en un archivo rotativo."""
    log_dir = Path(__file__).resolve().parent
    log_file = log_dir / "bot.log"
    
    formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
    
    # Handler para consola
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    console_handler.setLevel(logging.INFO)
    
    # Handler para archivo rotativo (5MB máximo por archivo, manteniendo hasta 5 respaldos)
    try:
        file_handler = RotatingFileHandler(
            log_file,
            maxBytes=5 * 1024 * 1024,
            backupCount=5,
            encoding="utf-8"
        )
        file_handler.setFormatter(formatter)
        file_handler.setLevel(logging.INFO)
    except Exception as e:
        logger.error("No se pudo configurar el RotatingFileHandler para el archivo de logs: %s", e)
        file_handler = None
        
    root_logger = logging.getLogger()
    root_logger.handlers.clear()
    root_logger.setLevel(logging.INFO)
    root_logger.addHandler(console_handler)
    if file_handler:
        root_logger.addHandler(file_handler)
        
    # Silenciar el ruido de logs de getUpdates constantes de la librería httpx
    logging.getLogger("httpx").setLevel(logging.WARNING)
    
    logger.info("Sistema de logs configurado. Archivo de logs: %s", log_file)

VOICE_BUTTON = "🎤 Activar comando de voz"
VOICE_CANCEL_BUTTON = "❌ Cancelar voz"
COTI_BUTTON = "📋 Nueva Cotización"
NOTA_BUTTON = "📦 Nueva Nota de Entrega"
HISTORIAL_BUTTON = "📂 Historial de Documentos"
RETENTION_RATE = Decimal("0.75")

# Nuevos botones de menú y submenú
TRIBUTOS_BUTTON = "🏛️ Tributos"
SUBMENU_CARGAR_FACTURA = "📥 Cargar Facturas"
SUBMENU_RETENCION_RECIBIDA = "🧾 Retenciones Recibidas"
SUBMENU_REPORTE_Z = "📊 Reportes Z"
SUBMENU_FACTURA_EMITIDA = "📈 Facturas Emitidas"
SUBMENU_GENERAR_RETENCION = "✍️ Generar Retención"
SUBMENU_ELIMINAR_RETENCION = "❌ Eliminar Retención"
SUBMENU_GENERAR_REPORTES = "📋 Generar Reportes"
SUBMENU_VOLVER = "🔙 Volver al Menú Principal"

# Botones del submenú de reportes
REPORT_IVA_BUTTON = "📉 Reporte IVA (Quincenal)"
REPORT_RETENCIONES_BUTTON = "🧾 Reporte Retenciones Recibidas"
REPORT_FACTURAS_BUTTON = "📥 Reporte Facturas Cargadas"
REPORT_PENDIENTES_BUTTON = "⚠️ Facturas sin Retención"
REPORT_VOLVER_TRIBUTOS = "🔙 Volver al Menú de Tributos"

# Botón administrativo
ADMIN_PANEL_BUTTON = "⚙️ Panel Admin"


def _main_keyboard(user_id: int | str = "") -> ReplyKeyboardMarkup:
    role = None
    if user_id:
        user = user_manager.get_user(user_id)
        if user:
            role = user.get("role")
            
    buttons = []
    
    # Si no hay ID, o es el admin, o tiene rol completo: mostrar todo
    if not user_id or str(user_id) == str(config.ALLOWED_USER_ID) or role in ("admin", "full_access", "nueva_empresa", "tributos_and_cotizaciones"):
        buttons.append([KeyboardButton(TRIBUTOS_BUTTON)])
        buttons.append([KeyboardButton(COTI_BUTTON), KeyboardButton(NOTA_BUTTON)])
        buttons.append([KeyboardButton(HISTORIAL_BUTTON)])
    else:
        if role == "tributos_only":
            buttons.append([KeyboardButton(TRIBUTOS_BUTTON)])
        elif role == "cotizaciones_only":
            buttons.append([KeyboardButton(COTI_BUTTON), KeyboardButton(NOTA_BUTTON)])
            buttons.append([KeyboardButton(HISTORIAL_BUTTON)])
            
    buttons.append([KeyboardButton(VOICE_BUTTON), KeyboardButton(VOICE_CANCEL_BUTTON)])
    
    if user_id:
        if role in ("admin", "nueva_empresa") or str(user_id) == str(config.ALLOWED_USER_ID):
            buttons.append([KeyboardButton("🚀 Menú de Inicio")])
            
    return ReplyKeyboardMarkup(
        keyboard=buttons,
        resize_keyboard=True,
        one_time_keyboard=False,
    )


def _tributos_submenu_keyboard(user_id: int | str = "") -> ReplyKeyboardMarkup:
    # Si es un Contribuyente Ordinario de nueva_empresa, se quita SUBMENU_GENERAR_RETENCION ("✍️ Generar Retención")
    company_is_ordinario = False
    user_is_tributos_only = False
    if user_id:
        user = user_manager.get_user(user_id)
        if user:
            role = user.get("role")
            if role == "nueva_empresa" and user.get("company_type") == "Ordinario":
                company_is_ordinario = True
            elif role == "tributos_only":
                user_is_tributos_only = True
                
    if user_is_tributos_only:
        kb_layout = [
            [KeyboardButton(SUBMENU_GENERAR_REPORTES)],
            [KeyboardButton(SUBMENU_VOLVER)],
        ]
    elif company_is_ordinario:
        kb_layout = [
            [KeyboardButton(SUBMENU_CARGAR_FACTURA), KeyboardButton(SUBMENU_RETENCION_RECIBIDA)],
            [KeyboardButton(SUBMENU_REPORTE_Z), KeyboardButton(SUBMENU_FACTURA_EMITIDA)],
            [KeyboardButton(SUBMENU_GENERAR_REPORTES)],
            [KeyboardButton(SUBMENU_VOLVER)],
        ]
    else:
        kb_layout = [
            [KeyboardButton(SUBMENU_CARGAR_FACTURA), KeyboardButton(SUBMENU_RETENCION_RECIBIDA)],
            [KeyboardButton(SUBMENU_REPORTE_Z), KeyboardButton(SUBMENU_FACTURA_EMITIDA)],
            [KeyboardButton(SUBMENU_GENERAR_RETENCION), KeyboardButton(SUBMENU_ELIMINAR_RETENCION)],
            [KeyboardButton(SUBMENU_GENERAR_REPORTES)],
            [KeyboardButton(SUBMENU_VOLVER)],
        ]
    return ReplyKeyboardMarkup(
        keyboard=kb_layout,
        resize_keyboard=True,
        one_time_keyboard=False,
    )


def _reportes_submenu_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(REPORT_IVA_BUTTON)],
            [KeyboardButton(REPORT_RETENCIONES_BUTTON)],
            [KeyboardButton(REPORT_FACTURAS_BUTTON)],
            [KeyboardButton(REPORT_PENDIENTES_BUTTON)],
            [KeyboardButton(REPORT_VOLVER_TRIBUTOS)],
        ],
        resize_keyboard=True,
        one_time_keyboard=False,
    )


def _allowed(update: Update) -> bool:
    u = update.effective_user
    if u is None:
        return False
    # El administrador principal (config.ALLOWED_USER_ID) siempre está permitido y tiene rol admin
    if u.id == config.ALLOWED_USER_ID:
        # Asegurar que esté registrado en el JSON como admin
        user_manager.register_user(
            user_id=u.id,
            name="Administrador Principal",
            role="admin",
            expiration_date="never",
            limit_ops=-1
        )
        return True
    
    # Comprobar si tiene suscripción activa
    return user_manager.is_subscription_active(u.id)


async def _deny(update: Update, context: ContextTypes.DEFAULT_TYPE = None) -> None:
    msg = update.effective_message
    u = update.effective_user
    if msg:
        if u is not None:
            kb = None
            if context:
                kb = InlineKeyboardMarkup([
                    [
                        InlineKeyboardButton("📥 Solicitar como Cliente", callback_data="user_request_access:cliente"),
                        InlineKeyboardButton("📥 Solicitar como SUFEVICA", callback_data="user_request_access:sufevica")
                    ]
                ])
            await msg.reply_text(
                "❌ *Acceso no autorizado / Suscripción Expirada*\n\n"
                f"Tu ID de Telegram es: `{u.id}`\n\n"
                "Por favor, ponte en contacto con el administrador para solicitar acceso o renovar tu plan.",
                parse_mode="Markdown",
                reply_markup=kb
            )
        else:
            await msg.reply_text("Acceso no autorizado.")


async def _handle_solicitar_access_flow(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    req_type: str = "cliente",
    callback_query = None
) -> None:
    import html
    msg = update.effective_message
    if not msg:
        return
    u = update.effective_user
    if not u:
        return
    
    if _allowed(update):
        text_already = "✅ Ya tienes acceso activo y autorizado al bot. ¡Usa el menú inferior para navegar!"
        if callback_query:
            await callback_query.edit_message_text(text_already)
        else:
            await msg.reply_text(text_already, reply_markup=_main_keyboard(u.id))
        return
        
    admin_id = config.ALLOWED_USER_ID
    if req_type == "sufevica":
        kb = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("✅ Autorizar SUFEVICA", callback_data=f"admin_req_sufevica_start:{u.id}"),
                InlineKeyboardButton("❌ Rechazar", callback_data=f"admin_req_reject:{u.id}")
            ]
        ])
        type_lbl = "SUFEVICA (Interno)"
    else:
        kb = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("✅ Autorizar Cliente", callback_data=f"admin_req_cliente_start:{u.id}"),
                InlineKeyboardButton("❌ Rechazar", callback_data=f"admin_req_reject:{u.id}")
            ]
        ])
        type_lbl = "Cliente (FlashTax)"

    name_str = f"{u.first_name} {u.last_name or ''}".strip()
    name_escaped = html.escape(name_str)
    username_escaped = html.escape(u.username) if u.username else "sin_username"
    user_mention = f'<a href="tg://user?id={u.id}">{name_escaped}</a>'
    try:
        await context.bot.send_message(
            chat_id=admin_id,
            text=f"🔔 <b>Nueva solicitud de acceso al bot:</b>\n\n"
                 f"• <b>Tipo:</b> {type_lbl}\n"
                 f"• <b>Usuario:</b> {user_mention}\n"
                 f"• <b>Nombre:</b> {name_escaped}\n"
                 f"• <b>ID de Telegram:</b> <code>{u.id}</code>\n"
                 f"• <b>Username:</b> @{username_escaped}\n\n"
                 f"¿Deseas autorizar a este usuario?",
            parse_mode="HTML",
            reply_markup=kb
        )
        success_text = (
            "📨 *Solicitud de Acceso Enviada*\n\n"
            "Tu solicitud ha sido recibida por el administrador. "
            "Te enviaremos una notificación automática por este chat cuando sea aprobada."
        )
        if callback_query:
            await callback_query.edit_message_text(success_text, parse_mode="Markdown")
        else:
            await msg.reply_text(success_text, parse_mode="Markdown")
    except Exception as e:
        logger.error(f"Error al enviar solicitud de acceso al admin: {e}")
        err_text = (
            "❌ *Error al enviar solicitud*\n\n"
            "No se pudo enviar la solicitud al administrador en este momento. "
            "Por favor, intenta de nuevo más tarde o contacta al administrador directamente."
        )
        if callback_query:
            await callback_query.edit_message_text("❌ Ocurrió un error al enviar la solicitud. Intenta de nuevo más tarde.")
        else:
            await msg.reply_text(err_text, parse_mode="Markdown")


async def handle_user_request_access_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    q = update.callback_query
    if not q:
        return
    await q.answer()
    u = update.effective_user
    if not u:
        return
        
    data = (q.data or "").strip()
    req_type = "cliente"
    if ":" in data:
        req_type = data.split(":")[1]
        
    await _handle_solicitar_access_flow(update, context, req_type=req_type, callback_query=q)



def _is_public_or_sufevica(update: Update) -> bool:
    is_channel = update.channel_post is not None or update.edited_channel_post is not None
    return is_channel or _is_sufevica_chat(update)


def _check_permission(update: Update, module: str) -> bool:
    if _is_public_or_sufevica(update):
        return True
    u = update.effective_user
    if u is None:
        return False
    if u.id == config.ALLOWED_USER_ID:
        return True
        
    user = user_manager.get_user(u.id)
    if not user:
        return False
    role = user.get("role")
    if role == "admin":
        return True
    if module == "tributos":
        return role in ("full_access", "tributos_only", "nueva_empresa", "tributos_and_cotizaciones")
    if module == "cotizaciones":
        return role in ("full_access", "cotizaciones_only", "nueva_empresa", "tributos_and_cotizaciones")
    return False


def _can_modify_tributos(update: Update) -> bool:
    if _is_public_or_sufevica(update):
        return True
    u = update.effective_user
    if u is None:
        return False
    if u.id == config.ALLOWED_USER_ID:
        return True
    user = user_manager.get_user(u.id)
    if not user:
        return False
    role = user.get("role")
    if role in ("admin", "full_access", "nueva_empresa"):
        return True
    return False


def _get_next_retencion_emitida_number_for_user(user_id: int | str, monthly_path: Path, emission_date: date) -> str:
    user = user_manager.get_user(user_id)
    config_correlative = user.get("last_correlative") if user else None
    prefix = emission_date.strftime("%Y%m")
    
    max_seq_excel = excel_store.max_seq_retencion_emitida(monthly_path.parent, emission_date=emission_date)
    max_seq_config = 0
    if config_correlative and str(config_correlative).startswith(prefix):
        seq_part = str(config_correlative)[len(prefix):]
        if seq_part.isdigit():
            max_seq_config = int(seq_part)
            
    max_seq = max(max_seq_excel, max_seq_config)
    return f"{prefix}{max_seq + 1:08d}"



async def _check_and_consume_quota(update: Update) -> bool:
    if _is_public_or_sufevica(update):
        return True
    u = update.effective_user
    if u is None:
        return False
    if u.id == config.ALLOWED_USER_ID:
        return True
    user = user_manager.get_user(u.id)
    if user and user.get("role") == "admin":
        return True
        
    if not user_manager.has_quota(u.id):
        msg = update.effective_message
        if msg:
            await msg.reply_text(
                "⚠️ *Límite de operaciones alcanzado*\n\n"
                "Has consumido tu límite mensual de operaciones permitidas en tu plan.\n"
                "Ponte en contacto con el administrador para solicitar una ampliación o renovación de cuota.",
                parse_mode="Markdown"
            )
        return False
    user_manager.increment_usage(u.id)
    return True

async def _notify_same_source_channel(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    text: str,
) -> None:
    """Notifica en SUFEVICA cuando el mensaje viene de ese chat."""
    msg = update.effective_message
    if not msg:
        return
    is_channel = update.channel_post is not None or update.edited_channel_post is not None
    if _is_sufevica_chat(update) or is_channel:
        try:
            await context.bot.send_message(chat_id=msg.chat_id, text=text)
        except Exception:  # noqa: BLE001
            # Si el bot no tiene permiso para publicar en el canal,
            # al menos intentamos responder como fallback.
            await msg.reply_text(text)
        return
    await msg.reply_text(text)


def _extract_labeled_value(text: str, labels: tuple[str, ...]) -> str:
    escaped = [re.escape(lbl) for lbl in labels]
    pattern = (
        r"(?:^|\n|[,;])\s*(?:" + "|".join(escaped) + r")\s*[:=]\s*([^\n,;]+)"
    )
    m = re.search(pattern, text, flags=re.IGNORECASE)
    return m.group(1).strip() if m else ""


def _extract_labeled_value_eol(text: str, labels: tuple[str, ...]) -> str:
    """Busca etiqueta: valor hasta fin de línea (permite comas en montos)."""
    labels_norm = tuple(_normalize_text(lbl).strip() for lbl in labels)
    for raw in text.splitlines():
        line = raw.strip()
        if ":" not in line:
            continue
        key, _, rest = line.partition(":")
        key_norm = _normalize_text(re.sub(r"^[\s*•\-\d.\)\(]+", "", key)).strip()
        if not key_norm:
            continue
        if any(
            key_norm == lbl
            or key_norm.startswith(lbl + " ")
            or lbl in key_norm
            for lbl in labels_norm
        ):
            val = re.sub(r"\*+", "", rest).strip()
            if val:
                return val
    return ""


def _parse_retencion_from_any_text(text: str) -> dict[str, str] | None:
    """
    Parseo flexible para textos del canal SUFEVICA sin exigir verbo de intención.
    """
    t = text.strip()
    data = {
        "fecha_emision": _extract_labeled_value_eol(
            t,
            ("fecha_emision", "fecha emision", "fecha de emision", "fecha de emisión", "fecha"),
        ),
        "numero_comprobante": _extract_labeled_value_eol(
            t,
            (
                "numero_comprobante",
                "nro_comprobante",
                "nro comprobante",
                "numero de comprobante",
                "número de comprobante",
                "comprobante",
            ),
        ),
        "rif": _extract_labeled_value_eol(t, ("rif",)),
        "fechas_facturas": _extract_labeled_value_eol(
            t,
            ("fechas_facturas", "fecha_factura", "fecha factura", "fechas facturas"),
        ),
        "numeros_facturas": _extract_labeled_value_eol(
            t,
            (
                "numeros_facturas",
                "numero_factura",
                "nro_factura",
                "nro factura",
                "número factura",
                "número de factura",
                "numero de factura",
                "factura afectada",
                "factura",
            ),
        ),
        "controles_facturas": _extract_labeled_value_eol(
            t,
            (
                "controles_facturas",
                "control_factura",
                "control",
                "nro control",
                "numero de control",
                "número de control",
                "numero de control",
            ),
        ),
        "total_compra_con_iva": _extract_labeled_value_eol(
            t,
            (
                "total_compra_con_iva",
                "total compra",
                "total compra con iva",
                "total",
                "total factura",
                "total general",
            ),
        ),
        "base_imponible": _extract_labeled_value_eol(
            t,
            ("base_imponible", "base imponible", "base gravable", "base"),
        ),
        "iva_retenido": _extract_labeled_value_eol(
            t,
            (
                "iva_retenido",
                "iva retenido",
                "monto iva retenido",
                "iva retenido total",
                "impuesto retenido",
                "monto retenido",
                "retenido",
            ),
        ),
    }
    required = ("fecha_emision", "numero_comprobante", "rif", "iva_retenido")
    if any(not data[k] for k in required):
        return None
    return data


def _parse_retencion_entry_request(text: str) -> dict[str, str] | None:
    t = text.strip()
    t_norm = _normalize_text(t)
    # Acepta "registrar una retencion", "quiero guardar retencion", etc.
    has_ret = "retencion" in t_norm or "retenciones" in t_norm
    intent_cmd = any(
        k in t_norm
        for k in (
            "registrar retencion",
            "guardar retencion",
            "nueva retencion",
            "registro retencion",
        )
    )
    intent_natural = has_ret and any(
        w in t_norm
        for w in (
            "registrar",
            "guardar",
            "nueva",
            "agregar",
            "anadir",
            "ingresar",
            "cargar",
            "reportar",
        )
    )
    if not (intent_cmd or intent_natural):
        return None

    data = {
        "fecha_emision": _extract_labeled_value(
            t,
            ("fecha_emision", "fecha emision", "fecha"),
        ),
        "numero_comprobante": _extract_labeled_value(
            t,
            ("numero_comprobante", "nro_comprobante", "nro comprobante", "comprobante"),
        ),
        "rif": _extract_labeled_value(t, ("rif",)),
        "fechas_facturas": _extract_labeled_value(
            t,
            ("fechas_facturas", "fecha_factura", "fecha factura"),
        ),
        "numeros_facturas": _extract_labeled_value(
            t,
            (
                "numeros_facturas",
                "numero_factura",
                "nro_factura",
                "nro factura",
                "número de factura",
                "numero de factura",
                "factura afectada",
                "factura",
            ),
        ),
        "controles_facturas": _extract_labeled_value(
            t,
            (
                "controles_facturas",
                "control_factura",
                "control",
                "nro control",
                "número de control",
                "numero de control",
            ),
        ),
        "total_compra_con_iva": _extract_labeled_value(
            t,
            (
                "total_compra_con_iva",
                "total compra",
                "total",
                "total factura",
                "total general",
            ),
        ),
        "base_imponible": _extract_labeled_value(
            t,
            ("base_imponible", "base imponible", "base"),
        ),
        "iva_retenido": _extract_labeled_value(
            t,
            (
                "iva_retenido",
                "iva retenido",
                "monto iva retenido",
                "impuesto retenido",
                "monto retenido",
                "retenido",
            ),
        ),
    }
    required = ("fecha_emision", "numero_comprobante", "rif", "iva_retenido")
    if any(not data[k] for k in required):
        return None
    return data


def _parse_retencion_islr_entry_request(text: str) -> dict[str, str] | None:
    t = text.strip()
    t_norm = _normalize_text(t)
    if "retencion islr" not in t_norm and "retenciones islr" not in t_norm and "retencion de islr" not in t_norm:
        return None
    data = {
        "fecha_emision": _extract_labeled_value(t, ("fecha_emision", "fecha emision", "fecha")),
        "numero_comprobante": _extract_labeled_value(t, ("numero_comprobante", "nro_comprobante", "nro comprobante", "comprobante")),
        "proveedor": _extract_labeled_value(t, ("proveedor", "razon_social", "razon social")),
        "proveedor_rif": _extract_labeled_value(t, ("rif", "proveedor_rif")),
        "concepto_retencion": _extract_labeled_value(t, ("concepto_retencion", "concepto retencion", "concepto")),
        "numero_documento": _extract_labeled_value(t, ("numero_documento", "numero_factura", "numero factura", "nro factura", "factura")),
        "numero_control": _extract_labeled_value(t, ("numero_control", "control")),
        "base_imponible": _extract_labeled_value(t, ("base_imponible", "base imponible", "base")),
        "porcentaje_retencion": _extract_labeled_value(t, ("porcentaje_retencion", "porcentaje retencion", "tasa", "porcentaje", "alicuota")),
        "islr_retenido": _extract_labeled_value(t, ("islr_retenido", "islr retenido", "retenido", "monto retenido")),
        "total_factura": _extract_labeled_value(t, ("total_factura", "total factura", "total")),
    }
    required = ("fecha_emision", "numero_comprobante", "proveedor_rif", "islr_retenido")
    if any(not data[k] for k in required):
        return None
    return data


def _parse_retencion_islr_from_any_text(text: str) -> dict[str, str] | None:
    t = text.strip()
    t_norm = _normalize_text(t)
    if "islr" not in t_norm:
        return None
    data = {
        "fecha_emision": _extract_labeled_value_eol(t, ("fecha_emision", "fecha emision", "fecha de emision", "fecha")),
        "numero_comprobante": _extract_labeled_value_eol(t, ("numero_comprobante", "nro comprobante", "comprobante")),
        "proveedor": _extract_labeled_value_eol(t, ("proveedor", "razon social")),
        "proveedor_rif": _extract_labeled_value_eol(t, ("rif", "proveedor_rif")),
        "concepto_retencion": _extract_labeled_value_eol(t, ("concepto retencion", "concepto")),
        "numero_documento": _extract_labeled_value_eol(t, ("numero_documento", "nro factura", "factura")),
        "numero_control": _extract_labeled_value_eol(t, ("numero_control", "control")),
        "base_imponible": _extract_labeled_value_eol(t, ("base_imponible", "base")),
        "porcentaje_retencion": _extract_labeled_value_eol(t, ("porcentaje retencion", "porcentaje", "tasa")),
        "islr_retenido": _extract_labeled_value_eol(t, ("islr_retenido", "islr retenido", "retenido")),
        "total_factura": _extract_labeled_value_eol(t, ("total factura", "total")),
    }
    required = ("fecha_emision", "numero_comprobante", "proveedor_rif", "islr_retenido")
    if any(not data[k] for k in required):
        return None
    return data


def _parse_retenciones_report_request(
    text: str,
) -> tuple[date, date, str] | None:
    t = text.lower().strip()
    t_norm = _normalize_text(t)
    if "retencion" not in t_norm and "retenciones" not in t_norm:
        return None
    fmt = "excel"
    if " pdf" in f" {t_norm} ":
        fmt = "pdf"
    elif " excel" in f" {t_norm} ":
        fmt = "excel"
    m = re.search(
        r"(\d{1,2}[\/\-\s]\d{1,2}[\/\-\s]\d{2,4}).*?(\d{1,2}[\/\-\s]\d{1,2}[\/\-\s]\d{2,4})",
        t,
    )
    if not m:
        return None
    d1 = _parse_user_date(m.group(1))
    d2 = _parse_user_date(m.group(2))
    if d1 is None or d2 is None:
        return None
    date_from, date_to = (d1, d2) if d1 <= d2 else (d2, d1)
    return date_from, date_to, fmt


def _normalize_text(text: str) -> str:
    base = unicodedata.normalize("NFKD", text)
    plain = "".join(ch for ch in base if not unicodedata.combining(ch))
    return plain.lower().strip()


def _parse_user_date(s: str) -> date | None:
    s = re.sub(r"\s+", "/", s.strip()).replace("-", "/")
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_emitir_retencion_request(text: str) -> list[str] | None:
    """
    Ejemplo: Emitir retencion de facturas:06949950|06949951
    """
    m = re.search(
        r"(?is)\bemitir\s+retencion\s+de\s+facturas?\s*:\s*(.+)$",
        text,
    )
    if not m:
        return None
    raw = m.group(1)
    docs = [x.strip() for x in re.split(r"[|,]", raw) if x.strip()]
    # Acepta números o alfanuméricos (ej. B00093101). Normaliza removiendo espacios.
    docs = [re.sub(r"\s+", "", x).upper() for x in docs]
    docs = [x for x in docs if re.search(r"[A-Z0-9]", x)]
    if not docs:
        return None
    # preservar orden sin duplicados
    seen: set[str] = set()
    out: list[str] = []
    for d in docs:
        if d in seen:
            continue
        seen.add(d)
        out.append(d)
    return out


def _periodo_fiscal(d: date) -> str:
    return d.strftime("%Y-%m")


def _reten_emit_monthly_path(emission_date: date, user_id: int | str | None = None) -> Path:
    ctx = CompanyContext(user_id)
    base_dir = ctx.retenciones_emitidas_dir
    return excel_store.monthly_retencion_emitida_path(base_dir, emission_date)


def _totals_for_items(items: list[excel_store.FacturaCompraRow]) -> tuple[Decimal, Decimal, Decimal]:
    base_total = Decimal("0")
    iva_total = Decimal("0")
    for it in items:
        base_total += it.base_imponible or Decimal("0")
        iva_total += it.monto_iva or Decimal("0")
    retenido = (iva_total * RETENTION_RATE).quantize(Decimal("0.01"))
    return base_total, iva_total, retenido
async def _start_emitir_retencion_flow(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    doc_numbers: list[str],
) -> None:
    msg = update.effective_message
    if not msg:
        return
    ctx = _get_company_context(update)
    if ctx.is_custom and ctx.company_type == "Ordinario":
        await msg.reply_text("⚠️ Opción Bloqueada: Los Contribuyentes Ordinarios no emiten comprobantes de retención de acuerdo con las normativas del SENIAT.")
        return
    items = excel_store.load_facturas_by_document_numbers(
        ctx.facturas_recibidas_path,
        doc_numbers,
    )
    if not items:
        await msg.reply_text(
            "No encontré facturas en FACTURAS-RECIBIDAS para esos números de documento."
        )
        return
    found_docs = {it.numero_documento for it in items}
    missing = [d for d in doc_numbers if d not in found_docs]
    if missing:
        await msg.reply_text(
            "Faltan estas facturas en el Excel de recibidas: " + ", ".join(missing)
        )
        return
    rifs = {it.proveedor_rif for it in items if it.proveedor_rif}
    if len(rifs) > 1:
        await msg.reply_text(
            "Las facturas seleccionadas tienen proveedores distintos. Emite por proveedor."
        )
        return
    provider = (items[0].proveedor or "").strip()
    provider_rif = (items[0].proveedor_rif or "").strip()
    provider_phone = (items[0].proveedor_telefono or "").strip()
    provider_address = (items[0].direccion_fiscal_proveedor or "").strip()
    base_total, iva_total, retenido = _totals_for_items(items)
    emission_date = date.today()
    monthly_path = _reten_emit_monthly_path(emission_date, update.effective_user.id)
    next_num = _get_next_retencion_emitida_number_for_user(
        update.effective_user.id,
        monthly_path,
        emission_date=emission_date,
    )
    context.user_data["pending_emit_ret"] = {
        "docs": doc_numbers,
        "provider": provider,
        "provider_rif": provider_rif,
        "provider_phone": provider_phone,
        "provider_address": provider_address,
        "base_total": str(base_total),
        "iva_total": str(iva_total),
        "retenido": str(retenido),
        "seq_mode": "auto",
    }
    kb = InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton("SI, seguir secuencia", callback_data="emit_seq_yes"),
                InlineKeyboardButton("NO, colocar manual", callback_data="emit_seq_no"),
            ]
        ]
    )
    await msg.reply_text(
        "Comprobante de retención listo para generar.\n"
        f"Proveedor: {provider or '—'} ({provider_rif or '—'})\n"
        f"Facturas: {' | '.join(doc_numbers)}\n"
        f"IVA total: {iva_total} | IVA retenido (75%): {retenido}\n"
        f"Siguiente correlativo según Excel ({monthly_path.parent.name}): {next_num}\n\n"
        "¿Deseas seguir la secuencia automática del Excel?",
        reply_markup=kb,
    )

def _parse_reporte_z_nuevo(text: str) -> dict[str, str] | None:
    t = text.strip()
    # Verificar si tiene el encabezado de Reporte Z y campos característicos
    if "reporte z" not in t.lower():
        return None
    if "desglose financiero" not in t.lower() and "sub-total" not in t.lower():
        return None

    # Extraer el número del reporte Z
    m_num = re.search(r"reporte z\s*:\s*(\d+)", t, flags=re.IGNORECASE)
    if not m_num:
        m_num = re.search(r"reporte z\s+(\d+)", t, flags=re.IGNORECASE)
        if not m_num:
            return None
    num_rep = m_num.group(1)

    # Extraer fecha de emisión
    m_fecha = re.search(r"fecha de emisi[oó]n\s*:\s*([^\n]+)", t, flags=re.IGNORECASE)
    if not m_fecha:
        m_fecha = re.search(r"fecha\s*:\s*([^\n]+)", t, flags=re.IGNORECASE)
        if not m_fecha:
            return None
    fecha_val = m_fecha.group(1).strip()

    # Extraer montos
    m_sub = re.search(r"sub-total\s*:\s*(?:bs\.?\s*)?([^\n]+)", t, flags=re.IGNORECASE)
    m_base = re.search(r"base imponible\s*:\s*(?:bs\.?\s*)?([^\n]+)", t, flags=re.IGNORECASE)
    m_exento = re.search(r"monto exento\s*:\s*(?:bs\.?\s*)?([^\n]+)", t, flags=re.IGNORECASE)
    m_iva = re.search(r"iva\s*(?:\(16%\))?\s*:\s*(?:bs\.?\s*)?([^\n]+)", t, flags=re.IGNORECASE)
    m_total = re.search(r"total general\s*(?:bs\.?\s*)?\s*:\s*([^\n]+)", t, flags=re.IGNORECASE)

    sub_str = m_sub.group(1).strip() if m_sub else ""
    base_str = m_base.group(1).strip() if m_base else ""
    exento_str = m_exento.group(1).strip() if m_exento else "0,00"
    iva_str = m_iva.group(1).strip() if m_iva else ""
    tot_str = m_total.group(1).strip() if m_total else ""

    # Validar que al menos tengamos base o total
    if not (base_str or tot_str):
        return None

    # Parsear y limpiar montos a strings decimales estándar
    try:
        sub_dec = excel_store.parse_amount_ves_string(sub_str)
        base_dec = excel_store.parse_amount_ves_string(base_str)
        exento_dec = excel_store.parse_amount_ves_string(exento_str) or Decimal("0")
        iva_dec = excel_store.parse_amount_ves_string(iva_str)
        tot_dec = excel_store.parse_amount_ves_string(tot_str)

        # Autocompletado si faltan campos
        if base_dec is not None and iva_dec is None:
            iva_dec = (base_dec * Decimal("0.16")).quantize(Decimal("0.01"))
        if base_dec is not None and tot_dec is None:
            tot_dec = base_dec + exento_dec + (iva_dec or Decimal("0"))
        if tot_dec is not None and base_dec is None:
            base_dec = ((tot_dec - exento_dec) / Decimal("1.16")).quantize(Decimal("0.01"))
            iva_dec = tot_dec - exento_dec - base_dec

        return {
            "numero_reporte": num_rep,
            "fecha_emision": fecha_val,
            "sub_total": str(sub_dec) if sub_dec is not None else (str(base_dec) if base_dec is not None else ""),
            "base_imponible": str(base_dec) if base_dec is not None else "",
            "monto_exento": str(exento_dec),
            "iva": str(iva_dec) if iva_dec is not None else "",
            "total": str(tot_dec) if tot_dec is not None else "",
        }
    except Exception:
        return None


def _parse_venta_o_reportez(text: str) -> dict[str, str] | None:
    t = text.strip()
    t_norm = _normalize_text(t)
    
    is_venta = (
        "venta" in t_norm
        or "factura emitida" in t_norm
        or "facturas emitidas" in t_norm
        or "factura manual" in t_norm
        or "facturas manuales" in t_norm
        or "factura de venta" in t_norm
    )
    is_z = (
        "reporte z" in t_norm
        or "reportez" in t_norm
        or "impresora fiscal" in t_norm
        or "resumen diario" in t_norm
        or "resumen diario z" in t_norm
        or "cierre z" in t_norm
        or "cierre diario" in t_norm
    )
    
    if not (is_venta or is_z):
        return None
        
    clasif = "Reporte Z" if is_z else "Factura Emitida"
    
    # 1. Intentar extracción con etiquetas estándar (con dos puntos o igual)
    fecha = _extract_labeled_value(t, ("fecha", "fecha_emision", "fecha emision"))
    num_doc = _extract_labeled_value(t, ("factura", "numero", "nro", "numero_documento", "documento"))
    cliente = _extract_labeled_value(t, ("cliente", "razon_social", "razon social", "comprador"))
    rif = _extract_labeled_value(t, ("rif", "rif_cliente", "rif cliente"))
    base = _extract_labeled_value(t, ("base", "base_imponible", "base imponible"))
    iva = _extract_labeled_value(t, ("iva", "impuesto"))
    total = _extract_labeled_value(t, ("total", "total_venta", "monto_total"))
    
    if not fecha:
        fecha = _extract_labeled_value_eol(t, ("fecha", "fecha emision"))
    if not num_doc:
        num_doc = _extract_labeled_value_eol(t, ("factura", "numero", "nro", "documento"))
    if not cliente:
        cliente = _extract_labeled_value_eol(t, ("cliente", "razon social"))
    if not rif:
        rif = _extract_labeled_value_eol(t, ("rif",))
    if not base:
        base = _extract_labeled_value_eol(t, ("base", "base imponible"))
    if not iva:
        iva = _extract_labeled_value_eol(t, ("iva",))
    if not total:
        total = _extract_labeled_value_eol(t, ("total",))

    # 2. Robustez: Extracción de fecha sin dos puntos (patrones DD/MM/YYYY o DD-MM-YYYY o YYYY-MM-DD)
    if not fecha:
        m_date = re.search(r"\b(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})\b", t)
        if m_date:
            day_str, month_str, year_str = m_date.groups()
            if len(year_str) == 2:
                year_str = "20" + year_str
            fecha = f"{day_str.zfill(2)}/{month_str.zfill(2)}/{year_str}"
        else:
            fecha = date.today().strftime("%d/%m/%Y")
            
    # 3. Robustez: Extracción de número de documento / reporte
    if not num_doc:
        if is_z:
            m_num = re.search(r"(?:z|reporte|nro|numero|#)\s*(\d+)", t, flags=re.IGNORECASE)
            if m_num:
                num_doc = m_num.group(1)
            else:
                parsed_date = tributario_engine._parse_row_date(fecha) or date.today()
                num_doc = f"Z-{parsed_date.strftime('%Y%m%d')}"
        else:
            num_doc = "S/N"
            
    # 4. Robustez: Extracción de montos sin dos puntos
    if not base:
        m_base = re.search(r"\b(?:base|imponible)\s+(\d+(?:[\.,]\d+)?)", t, flags=re.IGNORECASE)
        if m_base:
            base = m_base.group(1)
    if not iva:
        m_iva = re.search(r"\b(?:iva|impuesto)\s+(\d+(?:[\.,]\d+)?)", t, flags=re.IGNORECASE)
        if m_iva:
            iva = m_iva.group(1)
    if not total:
        m_total = re.search(r"\b(?:total)\s+(\d+(?:[\.,]\d+)?)", t, flags=re.IGNORECASE)
        if m_total:
            total = m_total.group(1)

    # 5. Valores predeterminados para Reporte Z
    if is_z:
        if not cliente:
            cliente = "CONSUMIDOR FINAL (VENTAS DIARIAS)"
        if not rif:
            rif = "V-00000000-0"

    # Verificar que al menos tengamos algún monto
    if not (base or iva or total):
        return None
        
    return {
        "clasificacion": clasif,
        "estado": "REGISTRADO",
        "fecha": fecha,
        "numero_documento": num_doc,
        "razon_social": cliente,
        "rif": rif,
        "base_imponible": base or "",
        "iva": iva or "",
        "total": total or "",
    }


def _match_intent(text: str) -> str | None:
    t = text.lower().strip()
    t_norm = _normalize_text(t)
    if any(
        p in t_norm
        for p in (
            "excel facturas compra",
            "excel de facturas compra",
            "enviar excel facturas compra",
            "mandar excel facturas compra",
            "descargar excel facturas compra",
            "facturas compra xlsx",
            "excel facturas recibidas",
        )
    ):
        return "send_facturas_compra_excel"
    if any(
        p in t_norm
        for p in (
            "enviar excel",
            "manda el excel",
            "mandar excel",
            "excel consolidado",
            "descargar excel",
            "envíame el excel",
            "enviame el excel",
        )
    ):
        return "send_excel"
    if any(
        p in t_norm
        for p in (
            "resumen de hoy",
            "resumen hoy",
            "dame el resumen",
            "el resumen de hoy",
        )
    ):
        return "summary_today"
    if any(
        p in t_norm
        for p in (
            "tributos",
            "compromisos tributarios",
            "iva por pagar",
            "reporte quincenal",
            "quincena actual",
            "corte de quincena",
        )
    ):
        return "tributos_report"
    if _parse_retenciones_report_request(text) is not None:
        return "retenciones_report"
    if any(
        p in t_norm
        for p in (
            "generar txt seniat",
            "txt seniat",
            "descargar txt seniat",
            "archivos txt seniat",
            "txt de retenciones",
            "txt iva",
        )
    ):
        return "seniat_txt"
    return None


def _is_sufevica_chat(update: Update) -> bool:
    msg = update.effective_message
    if not msg or not msg.chat:
        return False
    if config.SOURCE_CHANNEL_ID is not None and msg.chat_id == config.SOURCE_CHANNEL_ID:
        return True
    # Fallback por nombre para evitar caída si cambió el chat_id del canal/grupo.
    title = _normalize_text(getattr(msg.chat, "title", "") or "")
    username = _normalize_text(getattr(msg.chat, "username", "") or "")
    if "sufevica" in title or "sufevica" in username:
        return True
    return False


def _match_items_against_inventory(productos_path, items_list) -> list:
    matched_items = []
    for it in items_list:
        code = (it.get("code") or "").strip()
        desc = (it.get("desc") or "").strip()
        try:
            qty = float(it.get("qty") or 1.0)
        except Exception:
            qty = 1.0
        try:
            price = float(it.get("priceUsd") or 0.0)
        except Exception:
            price = 0.0

        found_products = []
        if code:
            found_products = excel_store.search_products_in_excel(productos_path, code, search_by="code")
        if not found_products and desc:
            found_products = excel_store.search_products_in_excel(productos_path, desc, search_by="desc")

        if found_products:
            matched_p = found_products[0]
            matched_items.append({
                "code": matched_p["code"],
                "desc": matched_p["description"],
                "qty": qty,
                "priceUsd": matched_p["price"],
                "totalUsd": qty * matched_p["price"]
            })
        else:
            matched_items.append({
                "code": code,
                "desc": desc,
                "qty": qty,
                "priceUsd": price,
                "totalUsd": qty * price
            })
    return matched_items


def _register_document_in_history(update: Update, doc_type: str, doc_number: str, client_name: str, client_rif: str, total_amount: str, temp_pdf_path: Path | str) -> Path:
    """
    Registra un documento generado (cotización, nota, retención) en el archivo JSON
    de historial del tenant y guarda el PDF de forma permanente en su directorio.
    Retorna la ruta permanente del PDF.
    """
    ctx = _get_company_context(update)
    ctx.generados_dir.mkdir(parents=True, exist_ok=True)
    
    src_path = Path(temp_pdf_path)
    dest_path = ctx.generados_dir / src_path.name
    
    if src_path.resolve() != dest_path.resolve():
        try:
            import shutil
            shutil.copy2(src_path, dest_path)
        except Exception as e:
            logger.error(f"Error al copiar archivo PDF al historial: {e}")
        
    history = []
    if ctx.historico_json_path.exists():
        try:
            import json
            with open(ctx.historico_json_path, "r", encoding="utf-8") as f:
                history = json.load(f)
        except Exception as e:
            logger.error(f"Error al leer historial JSON: {e}")
            
    from datetime import date
    import time
    entry = {
        "doc_type": doc_type,
        "doc_number": doc_number,
        "date": date.today().strftime("%Y-%m-%d"),
        "client_name": client_name,
        "client_rif": client_rif,
        "total_amount": total_amount,
        "pdf_filename": src_path.name,
        "timestamp": time.time()
    }
    history.append(entry)
    
    try:
        import json
        with open(ctx.historico_json_path, "w", encoding="utf-8") as f:
            json.dump(history, f, indent=4, ensure_ascii=False)
    except Exception as e:
        logger.error(f"Error al guardar historial JSON: {e}")
        
    return dest_path


async def _send_history_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    msg = update.effective_message
    if not msg:
        return
        
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("📋 Cotizaciones", callback_data="history_cat:cotizacion")],
        [InlineKeyboardButton("📦 Notas de Entrega", callback_data="history_cat:nota")],
        [InlineKeyboardButton("✍️ Retenciones de IVA", callback_data="history_cat:retencion_iva")],
        [InlineKeyboardButton("🏛️ Retenciones de ISLR", callback_data="history_cat:retencion_islr")],
        [InlineKeyboardButton("🔍 Reimprimir por Número", callback_data="history_reprint_manual")],
        [InlineKeyboardButton("❌ Cerrar", callback_data="history_close")]
    ])
    
    await msg.reply_text(
        "📂 *HISTORIAL DE DOCUMENTOS GENERADOS* 📂\n\n"
        "Selecciona una categoría de documento para consultar o reimprimir:",
        reply_markup=kb,
        parse_mode="Markdown"
    )


async def handle_history_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    q = update.callback_query
    if not q:
        return
    await q.answer()
    data = q.data
    msg = q.message
    if not msg:
        return
        
    ctx = _get_company_context(update)
    
    if data == "history_menu":
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("📋 Cotizaciones", callback_data="history_cat:cotizacion")],
            [InlineKeyboardButton("📦 Notas de Entrega", callback_data="history_cat:nota")],
            [InlineKeyboardButton("✍️ Retenciones de IVA", callback_data="history_cat:retencion_iva")],
            [InlineKeyboardButton("🏛️ Retenciones de ISLR", callback_data="history_cat:retencion_islr")],
            [InlineKeyboardButton("🔍 Reimprimir por Número", callback_data="history_reprint_manual")],
            [InlineKeyboardButton("❌ Cerrar", callback_data="history_close")]
        ])
        await q.edit_message_text(
            "📂 *HISTORIAL DE DOCUMENTOS GENERADOS* 📂\n\n"
            "Selecciona una categoría de documento para consultar o reimprimir:",
            reply_markup=kb,
            parse_mode="Markdown"
        )
        
    elif data.startswith("history_cat:"):
        cat = data.split(":", 1)[1]
        
        history = []
        if ctx.historico_json_path.exists():
            try:
                import json
                with open(ctx.historico_json_path, "r", encoding="utf-8") as f:
                    history = json.load(f)
            except Exception as e:
                logger.error(f"Error al leer historial: {e}")
                
        filtered = [(idx, entry) for idx, entry in enumerate(history) if entry.get("doc_type") == cat]
        
        cat_names = {
            "cotizacion": "Cotizaciones",
            "nota": "Notas de Entrega",
            "retencion_iva": "Retenciones de IVA",
            "retencion_islr": "Retenciones de ISLR"
        }
        cat_name = cat_names.get(cat, cat.capitalize())
        
        if not filtered:
            kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Volver", callback_data="history_menu")]])
            await q.edit_message_text(
                f"📂 *HISTORIAL > {cat_name.upper()}*\n\n"
                f"⚠️ No se encontraron documentos registrados en esta categoría.",
                reply_markup=kb,
                parse_mode="Markdown"
            )
            return
            
        filtered.sort(key=lambda x: x[1].get("timestamp", 0), reverse=True)
        recent = filtered[:8]
        
        kb_list = []
        for idx, entry in recent:
            doc_num = entry.get("doc_number", "—")
            client = entry.get("client_name", "Cliente")
            amount = entry.get("total_amount", "—")
            label = f"Nro {doc_num} - {client[:18]} ({amount})"
            kb_list.append([InlineKeyboardButton(label, callback_data=f"history_view:{idx}")])
            
        kb_list.append([InlineKeyboardButton("🔙 Volver", callback_data="history_menu")])
        kb = InlineKeyboardMarkup(kb_list)
        
        await q.edit_message_text(
            f"📂 *HISTORIAL > {cat_name.upper()}*\n\n"
            f"Selecciona un documento para ver sus detalles o reimprimirlo:",
            reply_markup=kb,
            parse_mode="Markdown"
        )
        
    elif data.startswith("history_view:"):
        idx_str = data.split(":", 1)[1]
        try:
            idx = int(idx_str)
        except ValueError:
            return
            
        history = []
        if ctx.historico_json_path.exists():
            try:
                import json
                with open(ctx.historico_json_path, "r", encoding="utf-8") as f:
                    history = json.load(f)
            except Exception as e:
                logger.error(f"Error: {e}")
                
        if idx < 0 or idx >= len(history):
            await q.edit_message_text("❌ Documento no encontrado.")
            return
            
        entry = history[idx]
        doc_type = entry.get("doc_type")
        doc_num = entry.get("doc_number", "—")
        client = entry.get("client_name", "—")
        rif = entry.get("client_rif", "—")
        date_str = entry.get("date", "—")
        amount = entry.get("total_amount", "—")
        
        doc_names = {
            "cotizacion": "Cotización",
            "nota": "Nota de Entrega",
            "retencion_iva": "Retención de IVA",
            "retencion_islr": "Retención de ISLR"
        }
        doc_name = doc_names.get(doc_type, "Documento")
        
        details = (
            f"📄 *DETALLE DEL DOCUMENTO*\n\n"
            f"📌 *Tipo:* {doc_name}\n"
            f"🔢 *Número:* {doc_num}\n"
            f"📅 *Fecha:* {date_str}\n"
            f"👤 *Cliente / Proveedor:* {client}\n"
            f"🆔 *RIF/CI:* {rif}\n"
            f"💰 *Monto Total:* {amount}\n"
        )
        
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("📥 Reimprimir / Descargar (PDF)", callback_data=f"history_print:{idx}")],
            [InlineKeyboardButton("🔙 Volver a la Lista", callback_data=f"history_cat:{doc_type}")],
            [InlineKeyboardButton("❌ Cerrar", callback_data="history_close")]
        ])
        
        await q.edit_message_text(details, reply_markup=kb, parse_mode="Markdown")
        
    elif data.startswith("history_print:"):
        idx_str = data.split(":", 1)[1]
        try:
            idx = int(idx_str)
        except ValueError:
            return
            
        history = []
        if ctx.historico_json_path.exists():
            try:
                import json
                with open(ctx.historico_json_path, "r", encoding="utf-8") as f:
                    history = json.load(f)
            except Exception as e:
                logger.error(f"Error: {e}")
                
        if idx < 0 or idx >= len(history):
            await q.answer("❌ Documento no encontrado.", show_alert=True)
            return
            
        entry = history[idx]
        pdf_filename = entry.get("pdf_filename")
        if not pdf_filename:
            await q.answer("❌ Nombre de archivo no registrado.", show_alert=True)
            return
            
        pdf_path = ctx.generados_dir / pdf_filename
        if not pdf_path.exists():
            await q.answer("⚠️ El archivo PDF ya no se encuentra en el servidor.", show_alert=True)
            return
            
        doc_names = {
            "cotizacion": "Cotizacion",
            "nota": "Nota_de_Entrega",
            "retencion_iva": "Retencion_IVA",
            "retencion_islr": "Retencion_ISLR"
        }
        doc_name = doc_names.get(entry.get("doc_type"), "Documento")
        
        await msg.reply_document(
            document=str(pdf_path),
            filename=f"{doc_name}_{entry.get('doc_number')}.pdf",
            caption=f"📄 *Reimpresión:* {doc_name} Nro {entry.get('doc_number')}",
            parse_mode="Markdown"
        )
        
    elif data == "history_reprint_manual":
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("📋 Cotización", callback_data="history_repman:cotizacion")],
            [InlineKeyboardButton("📦 Nota de Entrega", callback_data="history_repman:nota")],
            [InlineKeyboardButton("✍️ Retención de IVA", callback_data="history_repman:retencion_iva")],
            [InlineKeyboardButton("🏛️ Retenciones de ISLR", callback_data="history_repman:retencion_islr")],
            [InlineKeyboardButton("🔙 Volver", callback_data="history_menu")]
        ])
        await q.edit_message_text(
            "🔍 *Reimprimir por Número*\n\n"
            "Selecciona el tipo de documento que deseas buscar y reimprimir:",
            reply_markup=kb,
            parse_mode="Markdown"
        )
        
    elif data.startswith("history_repman:"):
        doc_type = data.split(":", 1)[1]
        context.user_data["awaiting_reprint_num"] = doc_type
        
        doc_names = {
            "cotizacion": "Cotización",
            "nota": "Nota de Entrega",
            "retencion_iva": "Retención de IVA",
            "retencion_islr": "Retención de ISLR"
        }
        doc_name = doc_names.get(doc_type, "Documento")
        
        sample_nums = {
            "cotizacion": "000018",
            "nota": "000018",
            "retencion_iva": "20260600000381",
            "retencion_islr": "20260600000002"
        }
        sample_num = sample_nums.get(doc_type, "18")
        
        try:
            await q.delete_message()
        except Exception:
            pass
            
        await msg.reply_text(
            f"✏️ *Reimpresión de {doc_name}*\n\n"
            f"Envía por mensaje el número del documento que deseas reimprimir (ejemplo: `{sample_num}`):",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancelar", callback_data="history_menu")]])
        )

    elif data == "history_close":
        context.user_data.pop("awaiting_reprint_num", None)
        try:
            await msg.delete()
        except Exception:
            pass


def _parse_document_text_explicit(text: str, forced_type: str) -> dict | None:
    doc_data = _parse_document_text(text)
    if doc_data is None:
        doc_data = _parse_document_text_relaxed(text, forced_type)
    if doc_data is not None:
        doc_data["docType"] = forced_type
    return doc_data


def _parse_document_items_master(text: str) -> list:
    lines = text.split('\n')
    if not lines:
        return []
    
    stitched_lines = [lines[0].strip()]
    for i in range(1, len(lines)):
        prev_line = stitched_lines[-1].strip()
        curr_line = lines[i].strip()
        
        if not prev_line or not curr_line:
            if curr_line:
                stitched_lines.append(curr_line)
            continue
            
        prev_words = prev_line.split()
        curr_words = curr_line.split()
        
        if not prev_words or not curr_words:
            stitched_lines.append(curr_line)
            continue
            
        last_word = prev_words[-1]
        first_word = curr_words[0]
        
        should_stitch = False
        
        # Case 1: Purely alphabetic word split (e.g. "TUBER", "IA" or "CAN", "T")
        if last_word.isalpha() and first_word.isalpha():
            if len(last_word) <= 2 or len(first_word) <= 2:
                should_stitch = True
                
        # Case 2: Split inside fractions or decimals (e.g. "1", "-1/2" or "33,", "05")
        elif re.match(r'.*\d$', last_word) and re.match(r'^[-/]', first_word):
            should_stitch = True
        elif re.match(r'.*[-/]$', last_word) and re.match(r'^\d', first_word):
            should_stitch = True
        elif re.match(r'.*[,.]$', last_word) and re.match(r'^\d', first_word):
            should_stitch = True
        elif re.match(r'.*\d$', last_word) and re.match(r'^[,.]', first_word):
            should_stitch = True
            
        if should_stitch:
            prev_line_rest = " ".join(prev_words[:-1])
            stitched_word = last_word + first_word
            
            if prev_line_rest:
                stitched_lines[-1] = prev_line_rest + " " + stitched_word
            else:
                stitched_lines[-1] = stitched_word
                
            curr_line_rest = " ".join(curr_words[1:])
            if curr_line_rest:
                stitched_lines.append(curr_line_rest)
        else:
            stitched_lines.append(curr_line)
            
    cleaned_unified = " ".join([l for l in stitched_lines if l.strip()])
    sp_parts = cleaned_unified.split()
    
    def parse_token_val(tok):
        clean = re.sub(r"[^\d.,\-]", "", tok)
        if "," in clean and "." in clean:
            clean = clean.replace(".", "").replace(",", ".")
        elif "," in clean:
            clean = clean.replace(",", ".")
        return float(clean) if clean else 0.0

    matches = []
    j = 2
    while j < len(sp_parts):
        try:
            qty_val = parse_token_val(sp_parts[j-2])
            price_val = parse_token_val(sp_parts[j-1])
            total_val = parse_token_val(sp_parts[j])
            
            if qty_val > 0 and price_val > 0 and total_val > 0:
                calc = qty_val * price_val
                if abs(calc - total_val) < 0.05 * total_val:
                    matches.append(j)
                    j += 3 # skip to avoid overlapping matches
                    continue
        except Exception:
            pass
        j += 1

    items = []
    headers_to_strip = {"codigo", "producto", "cant", "p.u.", "total", "pu", "cant.", "total."}
    common_desc_words = {"de", "para", "con", "en", "tubo", "codo", "anillo", "brida", "niple", "tee", "af", "pvc", "hg", "galvanizado", "soldadura"}
    
    for k in range(len(matches)):
        idx = matches[k]
        start = matches[k-1] + 1 if k > 0 else 0
        
        qty_val = parse_token_val(sp_parts[idx-2])
        price_val = parse_token_val(sp_parts[idx-1])
        
        text_tokens = sp_parts[start : idx-2]
        
        if k == 0:
            text_tokens = [t for t in text_tokens if t.lower() not in headers_to_strip]
            
        if not text_tokens:
            continue
            
        first_tok = text_tokens[0]
        is_code = len(first_tok) <= 15 and (
            any(c.isdigit() for c in first_tok) or 
            first_tok.lower() not in common_desc_words
        )
        
        if is_code and len(text_tokens) > 1:
            code = first_tok.upper()
            desc = " ".join(text_tokens[1:])
        else:
            code = ""
            desc = " ".join(text_tokens)
            
        items.append({
            "code": code,
            "desc": desc.strip(),
            "qty": qty_val,
            "priceUsd": price_val
        })
        
    return items


_bcv_rate_cache = {
    "rate": 550.00,
    "last_fetched": 0.0
}

def _fetch_bcv_usd_rate() -> float:
    """Obtiene la tasa oficial de cambio USD/VES del BCV desde DolarAPI con fallback raspando la web del BCV."""
    url_api = "https://ve.dolarapi.com/v1/dolares/oficial"
    headers = {'User-Agent': 'Mozilla/5.0'}
    
    import urllib.request
    import json
    import ssl
    import re
    
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    
    try:
        req = urllib.request.Request(url_api, headers=headers)
        with urllib.request.urlopen(req, context=ctx, timeout=5) as response:
            data = json.loads(response.read().decode('utf-8'))
            if data and "promedio" in data:
                return float(data["promedio"])
    except Exception as e:
        logger.warning(f"Error al obtener tasa BCV de DolarAPI: {e}. Intentando raspado directo...")
        
    url_bcv = "https://www.bcv.org.ve"
    try:
        req = urllib.request.Request(url_bcv, headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'})
        with urllib.request.urlopen(req, context=ctx, timeout=8) as response:
            html = response.read().decode('utf-8')
            match = re.search(r'id="dolar"[^>]*>.*?<strong[^>]*>\s*([\d,.]+)\s*</strong>', html, re.DOTALL)
            if match:
                val = match.group(1).strip()
                val_float = float(val.replace('.', '').replace(',', '.'))
                return val_float
    except Exception as ex:
        logger.error(f"Error al raspar página oficial del BCV: {ex}")
        
    return 550.00

def get_current_bcv_rate() -> float:
    import time
    now = time.time()
    if now - _bcv_rate_cache["last_fetched"] > 21600:
        try:
            rate = _fetch_bcv_usd_rate()
            _bcv_rate_cache["rate"] = rate
            _bcv_rate_cache["last_fetched"] = now
        except Exception as e:
            logger.error(f"Error en get_current_bcv_rate: {e}")
    return _bcv_rate_cache["rate"]


SYNC_FILES = {
    "reten_rec": "RETEN-REC.xlsx",
    "facturas_recibidas": "FACTURAS-RECIBIDAS-NUEVO.xlsx",
    "facturas_emitidas": "FACTURAS-EMITIDAS.xlsx",
    "reportes_z": "REPORTES-Z-NUEVO.xlsx",
    "productos": config.PRODUCTOS_PATH.name,
    "usuarios": "usuarios.json"
}

def get_sync_file_path(key: str, user_id: int | str | None = None) -> Path | None:
    ctx = CompanyContext(user_id)
    if key == "reten_rec":
        return ctx.excel_path
    elif key == "facturas_recibidas":
        return ctx.facturas_recibidas_path
    elif key == "facturas_emitidas":
        return ctx.facturas_emitidas_path
    elif key == "reportes_z":
        return ctx.reportes_z_path
    elif key == "productos":
        return ctx.productos_path
    elif key == "usuarios":
        return Path(__file__).resolve().parent / "usuarios.json"
    return None

_last_mtime_cache = {}
PIN_PREFIX = "[SUFEVICA_BACKUP_STATE] "

async def _get_pinned_state(bot) -> dict:
    import json
    try:
        chat = await bot.get_chat(chat_id=config.ALLOWED_USER_ID)
        if chat.pinned_message and chat.pinned_message.text and chat.pinned_message.text.startswith(PIN_PREFIX):
            json_str = chat.pinned_message.text[len(PIN_PREFIX):].strip()
            return json.loads(json_str)
    except Exception as e:
        logger.error(f"Error al obtener el estado fijado de respaldos: {e}")
    return {}

async def _save_pinned_state(bot, state: dict) -> None:
    import json
    text = f"{PIN_PREFIX}{json.dumps(state)}"
    try:
        chat = await bot.get_chat(chat_id=config.ALLOWED_USER_ID)
        if chat.pinned_message and chat.pinned_message.text and chat.pinned_message.text.startswith(PIN_PREFIX):
            try:
                await bot.edit_message_text(
                    chat_id=config.ALLOWED_USER_ID,
                    message_id=chat.pinned_message.message_id,
                    text=text
                )
                return
            except Exception:
                pass
        msg = await bot.send_message(chat_id=config.ALLOWED_USER_ID, text=text)
        await bot.pin_chat_message(chat_id=config.ALLOWED_USER_ID, message_id=msg.message_id, disable_notification=True)
    except Exception as e:
        logger.error(f"Error al guardar el estado fijado de respaldos: {e}")

async def check_and_sync_files(context: ContextTypes.DEFAULT_TYPE) -> None:
    import os
    changed = False
    current_state = await _get_pinned_state(context.bot)
    
    files_to_sync = []
    # Archivos estáticos
    for key, filename in SYNC_FILES.items():
        path = get_sync_file_path(key)
        if path:
            files_to_sync.append((key, path, filename))
            
    # Archivos dinámicos de retenciones mensuales (contexto SUFEVICA por defecto)
    ctx = CompanyContext(None)
    if ctx.retenciones_emitidas_dir.is_dir():
        for path in ctx.retenciones_emitidas_dir.glob("RETEN-EMIT-*.xlsx"):
            key = f"dynamic_emit_{path.name.lower()}"
            files_to_sync.append((key, path, path.name))
            
    if ctx.retenciones_islr_dir.is_dir():
        for path in ctx.retenciones_islr_dir.glob("RETEN-ISLR-*.xlsx"):
            key = f"dynamic_islr_{path.name.lower()}"
            files_to_sync.append((key, path, path.name))
            
    # Archivos de empresas custom (Premium)
    try:
        users_data = user_manager.load_users().get("users", {})
        for uid, u_info in users_data.items():
            if u_info.get("role") == "nueva_empresa":
                c_ctx = CompanyContext(uid)
                # Archivos base de la empresa custom
                for base_key, filename in SYNC_FILES.items():
                    if base_key == "usuarios":
                        continue
                    path = get_sync_file_path(base_key, uid)
                    if path and path.exists():
                        # Usar prefijo único para evitar colisiones
                        key = f"company_{uid}_{base_key}"
                        files_to_sync.append((key, path, f"company_{uid}_{filename}"))
                
                # Archivos dinámicos de la empresa custom
                if c_ctx.retenciones_emitidas_dir.is_dir():
                    for path in c_ctx.retenciones_emitidas_dir.glob("RETEN-EMIT-*.xlsx"):
                        key = f"company_{uid}_dynamic_emit_{path.name.lower()}"
                        files_to_sync.append((key, path, f"company_{uid}_{path.name}"))
                        
                if c_ctx.retenciones_islr_dir.is_dir():
                    for path in c_ctx.retenciones_islr_dir.glob("RETEN-ISLR-*.xlsx"):
                        key = f"company_{uid}_dynamic_islr_{path.name.lower()}"
                        files_to_sync.append((key, path, f"company_{uid}_{path.name}"))
    except Exception as e:
        logger.error(f"Error al listar archivos de empresas para sincronizar: {e}")
            
    for key, path, filename in files_to_sync:
        if not path or not path.exists():
            continue
        mtime = os.path.getmtime(path)
        last_mtime = _last_mtime_cache.get(key)
        if last_mtime is None:
            _last_mtime_cache[key] = mtime
            continue
        if mtime > last_mtime:
            logger.info(f"Detectado cambio en {filename}. Subiendo respaldo...")
            try:
                with open(path, "rb") as f:
                    sent_msg = await context.bot.send_document(
                        chat_id=config.ALLOWED_USER_ID,
                        document=f,
                        filename=filename,
                        caption=f"📦 Respaldo automático de {filename}",
                        disable_notification=True
                    )
                new_file_id = sent_msg.document.file_id
                current_state[key] = new_file_id
                _last_mtime_cache[key] = mtime
                changed = True
                
                # Borrar el mensaje de respaldo del chat para no saturar la conversación
                try:
                    await context.bot.delete_message(
                        chat_id=config.ALLOWED_USER_ID,
                        message_id=sent_msg.message_id
                    )
                except Exception as del_err:
                    logger.warning(f"No se pudo borrar el mensaje de respaldo: {del_err}")
            except Exception as e:
                logger.error(f"Error subiendo respaldo de {filename}: {e}")
    if changed:
        await _save_pinned_state(context.bot, current_state)

async def restore_files_from_backup(bot) -> None:
    import os
    current_state = await _get_pinned_state(bot)
    if not current_state:
        logger.info("No se encontró ningún respaldo fijado. Iniciando con archivos locales actuales.")
        return
    logger.info("Restaurando archivos de Excel desde el respaldo de Telegram...")
    
    # 1. Restaurar usuarios.json primero
    usuarios_path = get_sync_file_path("usuarios")
    if "usuarios" in current_state and usuarios_path:
        try:
            logger.info("Descargando usuarios.json desde Telegram...")
            tg_file = await bot.get_file(current_state["usuarios"])
            usuarios_path.parent.mkdir(parents=True, exist_ok=True)
            await tg_file.download_to_drive(custom_path=str(usuarios_path))
            logger.info("usuarios.json restaurado con éxito.")
            _last_mtime_cache["usuarios"] = os.path.getmtime(usuarios_path)
        except Exception as e:
            logger.error(f"Error descargando respaldo de usuarios.json: {e}")

    # 2. Restaurar el resto de archivos
    ctx = CompanyContext(None)
    for key, file_id in current_state.items():
        if key == "usuarios":
            continue
            
        path = None
        if key.startswith("company_"):
            # Formato: company_{uid}_{base_key} o company_{uid}_dynamic_emit_{filename}
            parts = key.split("_")
            if len(parts) >= 3:
                uid = parts[1]
                sub_key = "_".join(parts[2:])
                c_ctx = CompanyContext(uid)
                if sub_key == "reten_rec":
                    path = c_ctx.excel_path
                elif sub_key == "facturas_recibidas":
                    path = c_ctx.facturas_recibidas_path
                elif sub_key == "facturas_emitidas":
                    path = c_ctx.facturas_emitidas_path
                elif sub_key == "reportes_z":
                    path = c_ctx.reportes_z_path
                elif sub_key == "productos":
                    path = c_ctx.productos_path
                elif sub_key.startswith("dynamic_emit_"):
                    filename = sub_key[len("dynamic_emit_"):]
                    path = c_ctx.retenciones_emitidas_dir / filename
                elif sub_key.startswith("dynamic_islr_"):
                    filename = sub_key[len("dynamic_islr_"):]
                    path = c_ctx.retenciones_islr_dir / filename
                else:
                    continue
                filename = path.name
            else:
                continue
        elif key.startswith("dynamic_emit_"):
            filename = key[len("dynamic_emit_"):]
            path = ctx.retenciones_emitidas_dir / filename
        elif key.startswith("dynamic_islr_"):
            filename = key[len("dynamic_islr_"):]
            path = ctx.retenciones_islr_dir / filename
        else:
            path = get_sync_file_path(key)
            if not path:
                continue
            filename = SYNC_FILES.get(key, path.name)
            
        if not path:
            continue
            
        try:
            logger.info(f"Descargando {filename} desde Telegram...")
            tg_file = await bot.get_file(file_id)
            path.parent.mkdir(parents=True, exist_ok=True)
            await tg_file.download_to_drive(custom_path=str(path))
            logger.info(f"Archivo {filename} restaurado con éxito.")
            _last_mtime_cache[key] = os.path.getmtime(path)
        except Exception as e:
            logger.error(f"Error descargando respaldo de {filename} (clave: {key}): {e}")

async def post_init(application: Application) -> None:
    # 1. Restaurar archivos desde el respaldo de Telegram al iniciar
    await restore_files_from_backup(application.bot)
    # 2. Iniciar job periódico de monitoreo cada 15 segundos
    if application.job_queue:
        application.job_queue.run_repeating(check_and_sync_files, interval=15, first=10)

async def descargar_excel_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not _allowed(update):
        await _deny(update)
        return
    msg = update.effective_message
    if not msg:
        return
    sent_any = False
    for key, filename in SYNC_FILES.items():
        path = get_sync_file_path(key, update.effective_user.id)
        if path and path.exists():
            try:
                with open(path, "rb") as f:
                    await msg.reply_document(
                        document=f,
                        filename=filename,
                        caption=f"📊 Archivo: {filename}"
                    )
                sent_any = True
            except Exception as e:
                logger.error(f"Error al enviar {filename}: {e}")
    if not sent_any:
        await msg.reply_text("⚠️ No se encontraron archivos de Excel locales en el servidor.")
async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    msg = update.effective_message
    if not msg or not msg.document:
        return
    filename = msg.document.file_name
    logger.info(f"Recibido documento: {filename} (chat_id={msg.chat_id})")
    
    if not _allowed(update) and not _is_sufevica_chat(update):
        logger.warning(f"Documento rechazado por falta de permisos (chat_id={msg.chat_id})")
        await _deny(update)
        return

    user = user_manager.get_user(update.effective_user.id)
    if user and user.get("role") == "tributos_only":
        await msg.reply_text("❌ Tu nivel de autorización (\"Tributos Only\") solo te permite consultar y generar reportes, no subir o reemplazar archivos en el bot.")
        return

    # Interceptar firma y sello para admin (cargar firma y sello)
    if context.user_data.get("awaiting_admin_company_signature"):
        target_uid = context.user_data.get("admin_edit_target_uid")
        context.user_data.pop("awaiting_admin_company_signature", None)
        context.user_data.pop("admin_edit_target_uid", None)
        
        ctx = CompanyContext(target_uid)
        if not ctx.is_custom:
            await msg.reply_text("❌ Solo las empresas personalizadas pueden subir firma y sello.")
            return
            
        status_msg = await msg.reply_text("📥 *Guardando imagen de firma y sello del cliente...*", parse_mode="Markdown")
        try:
            tg_file = await context.bot.get_file(msg.document.file_id)
            target_path = ctx.dir_path / "firma_sello_transparente.png"
            target_path.parent.mkdir(parents=True, exist_ok=True)
            await tg_file.download_to_drive(custom_path=str(target_path))
            await status_msg.edit_text(f"✅ *Firma y Sello del cliente actualizados con éxito!*", parse_mode="Markdown")
            await _show_admin_user_detail(update, context, target_uid)
        except Exception as e:
            logger.error(f"Error al guardar firma y sello de cliente desde documento: {e}")
            await status_msg.edit_text(f"❌ *Error al guardar la firma y sello de cliente:* `{e}`", parse_mode="Markdown")
            await _show_admin_user_detail(update, context, target_uid)
        return

    # Interceptar firma y sello
    if context.user_data.get("awaiting_company_signature"):
        context.user_data.pop("awaiting_company_signature", None)
        ctx = _get_company_context(update)
        if not ctx.is_custom:
            await msg.reply_text("❌ Solo las empresas personalizadas pueden subir firma y sello.")
            return
        
        status_msg = await msg.reply_text("📥 *Guardando imagen de firma y sello...*", parse_mode="Markdown")
        try:
            tg_file = await context.bot.get_file(msg.document.file_id)
            target_path = ctx.dir_path / "firma_sello_transparente.png"
            target_path.parent.mkdir(parents=True, exist_ok=True)
            await tg_file.download_to_drive(custom_path=str(target_path))
            await status_msg.edit_text(f"✅ *Firma y Sello actualizados con éxito!*\nGuardado en `{target_path.name}`", parse_mode="Markdown")
            await _show_company_config_menu(update, context)
        except Exception as e:
            logger.error(f"Error al guardar firma y sello desde documento: {e}")
            await status_msg.edit_text(f"❌ *Error al guardar la firma y sello:* `{e}`", parse_mode="Markdown")
        return

    matched_key = None
    for key, name in SYNC_FILES.items():
        if filename.lower() == name.lower():
            matched_key = key
            break
            
    if not matched_key and filename.lower().endswith(".xlsx"):
        fn_lower = filename.lower()
        if fn_lower.startswith("productos") or fn_lower.startswith("inventario"):
            matched_key = "productos"
            
    if matched_key:
        path = get_sync_file_path(matched_key, update.effective_user.id)
        status_msg = await msg.reply_text(f"📥 *Recibido {filename}. Procesando y reemplazando archivo local...*", parse_mode="Markdown")
        try:
            tg_file = await context.bot.get_file(msg.document.file_id)
            path.parent.mkdir(parents=True, exist_ok=True)
            await tg_file.download_to_drive(custom_path=str(path))
            import os
            _last_mtime_cache[matched_key] = os.path.getmtime(path)
            
            ctx = _get_company_context(update)
            if not ctx.is_custom:
                current_state = await _get_pinned_state(context.bot)
                current_state[matched_key] = msg.document.file_id
                await _save_pinned_state(context.bot, current_state)
            await status_msg.edit_text(f"✅ *¡Archivo `{filename}` actualizado con éxito!* El bot trabajará con esta nueva versión.", parse_mode="Markdown")
        except Exception as e:
            logger.error(f"Error al descargar archivo subido {filename}: {e}")
            await status_msg.edit_text(f"❌ *Error al procesar el archivo:* `{e}`", parse_mode="Markdown")

def _parse_document_text_relaxed(text: str, doc_type: str) -> dict | None:
    client_name = _take_eol_label(text, ("cliente", "razon social", "razón social", "nombre", "comprador", "dirigido a"))
    rif = _take_eol_label(text, ("rif", "c.i.", "ci", "cedula", "cédula", "identificacion", "rif cliente"))
    address = _take_eol_label(text, ("direccion", "dirección", "direccion fiscal", "dirección fiscal"))
    phone = _take_eol_label(text, ("telefono", "teléfono", "telefonos", "teléfonos", "telf", "celular"))
    doc_number = _take_eol_label(text, ("numero", "nro", "número", "documento", "cotizacion nro", "cotización nro", "nota nro"))
    salesman = _take_eol_label(text, ("vendedor", "ejecutivo", "atendido por")) or "FREDDY LOPEZ"
    sale_type = _take_eol_label(text, ("tipo", "tipo de venta", "condición de pago", "condicion de pago")) or "Contado"
    
    if not client_name:
        m_cli = re.search(r"(?im)(?:cliente|razon social|sres|señores)\s*[:\-]\s*(.+)$", text)
        if m_cli:
            client_name = m_cli.group(1).strip()
    if not rif:
        m_rif = re.search(r"\b([VEJPGvejpg][\s.\-]*\d{6,12}(?:[\s.\-]\d)?)\b", text)
        if m_rif:
            rif = re.sub(r"\s+", "", m_rif.group(1).strip())
            
    if rif:
        rif = _normalize_rif(rif)
        
    if not doc_number:
        m_num = re.search(r"(?im)(?:nro|numero|#)\s*[:\-]?\s*(\d+)\b", text)
        if m_num:
            doc_number = m_num.group(1).strip()
        else:
            doc_number = ""
            
    # Try the master stream parser first
    items = _parse_document_items_master(text)
    if items:
        return {
            "docType": doc_type,
            "currency": "usd",
            "exchangeRate": get_current_bcv_rate(),
            "docNumber": doc_number,
            "docDate": date.today().strftime("%Y-%m-%d"),
            "client": {
                "name": client_name or "CLIENTE REGISTRADO POR TEXTO",
                "address": address or "DIRECCIÓN DE CLIENTE",
                "rif": rif or "V-00000000-0",
                "phone": phone or "—",
                "salesman": salesman,
                "saleType": sale_type
            },
            "items": items
        }
            
    items = []
    lines = text.splitlines()
    
    for line in lines:
        line_clean = line.strip()
        if not line_clean:
            continue
            
        t_line = _normalize_text(line_clean)
        if any(w in t_line for w in ("total", "subtotal", "sub-total", "iva", "cliente", "rif", "direccion", "vendedor", "fecha", "forma libre")):
            continue
            
        parts = [p.strip() for p in re.split(r"\s{2,}|\t", line_clean) if p.strip()]
        if len(parts) >= 3:
            nums = []
            for i, p in enumerate(parts):
                clean_num = re.sub(r"[^\d.,\-]", "", p)
                try:
                    val = float(clean_num.replace(".", "").replace(",", "."))
                    nums.append((i, val))
                except ValueError:
                    try:
                        val = float(clean_num.replace(",", "."))
                        nums.append((i, val))
                    except ValueError:
                        pass
            
            qty = None
            price = None
            code = ""
            desc = ""
            matched_math = False
            
            if len(nums) >= 2:
                for a in range(len(nums)):
                    for b in range(len(nums)):
                        if a == b: continue
                        for c in range(len(nums)):
                            if c == a or c == b: continue
                            val_a = nums[a][1]
                            val_b = nums[b][1]
                            val_c = nums[c][1]
                            if val_a > 0 and val_b > 0 and abs(val_a * val_b - val_c) < 0.05 * val_c:
                                qty = val_a
                                price = val_b
                                matched_math = True
                                break
                        if matched_math: break
                    if matched_math: break
                    
                if matched_math:
                    num_indices = {nums[a][0], nums[b][0], nums[c][0]}
                    text_parts = [parts[i] for i in range(len(parts)) if i not in num_indices]
                    if len(text_parts) >= 2:
                        code = text_parts[0].upper()
                        desc = " ".join(text_parts[1:])
                    elif len(text_parts) == 1:
                        desc = text_parts[0]
                else:
                    try:
                        first_clean = re.sub(r"[^\d.,\-]", "", parts[0])
                        qty = float(first_clean.replace(".", "").replace(",", "."))
                        
                        last_clean = re.sub(r"[^\d.,\-]", "", parts[-1])
                        price = float(last_clean.replace(".", "").replace(",", "."))
                        
                        if len(parts) >= 4:
                            penult_clean = re.sub(r"[^\d.,\-]", "", parts[-2])
                            try:
                                price_p = float(penult_clean.replace(".", "").replace(",", "."))
                                if abs(qty * price_p - price) < 0.1 * price:
                                    price = price_p
                            except ValueError:
                                pass
                                
                        text_parts = parts[1:-1] if len(parts) >= 3 else [parts[1]]
                        if len(text_parts) >= 2:
                            code = text_parts[0].upper()
                            desc = " ".join(text_parts[1:])
                        else:
                            desc = text_parts[0]
                    except Exception:
                        pass
                        
                if qty is not None and price is not None and qty > 0 and price >= 0:
                    items.append({
                        "code": code,
                        "desc": desc.strip(),
                        "qty": qty,
                        "priceUsd": price
                    })
                    continue

        # 1.5 Estrategia de Columnas al Final (para tablas copiadas de Excel con formato: CODIGO PRODUCTO CANT P.U. TOTAL)
        sp_parts = [p.strip() for p in line_clean.split() if p.strip()]
        if len(sp_parts) >= 4:
            try:
                def parse_token_val(tok):
                    clean = re.sub(r"[^\d.,\-]", "", tok)
                    if "," in clean and "." in clean:
                        clean = clean.replace(".", "").replace(",", ".")
                    elif "," in clean:
                        clean = clean.replace(",", ".")
                    return float(clean)

                qty_val = parse_token_val(sp_parts[-3])
                price_val = parse_token_val(sp_parts[-2])
                total_val = parse_token_val(sp_parts[-1])
                
                if qty_val > 0 and price_val >= 0 and total_val >= 0 and abs(qty_val * price_val - total_val) < 0.05 * total_val:
                    remaining_tokens = sp_parts[:-3]
                    first_tok = remaining_tokens[0]
                    
                    is_code = len(first_tok) <= 15 and (
                        any(c.isdigit() for c in first_tok) or 
                        not any(v in first_tok.lower() for v in ("de", "para", "con", "en", "tubo", "codo", "anillo", "brida", "niple", "tee"))
                    )
                    
                    if is_code and len(remaining_tokens) > 1:
                        code = first_tok.upper()
                        desc = " ".join(remaining_tokens[1:])
                    else:
                        code = ""
                        desc = " ".join(remaining_tokens)
                        
                    items.append({
                        "code": code,
                        "desc": desc.strip(),
                        "qty": qty_val,
                        "priceUsd": price_val
                    })
                    continue
            except Exception:
                pass

        m = re.match(
            r"^\s*[-*•]?\s*(\d+(?:[\.,]\d+)?)\s*(?:x|unidades|uds|unds|pzs|unid)?\s+([A-Z0-9\-]{3,15})?\s+(.+?)\s*(?:a|precio|costo)?\s+(\d+(?:[\.,]\d+)?)\s*$",
            line_clean,
            flags=re.IGNORECASE
        )
        if m:
            qty_str, code_str, desc_str, price_str = m.groups()
            try:
                qty = float(qty_str.replace(",", "."))
                price = float(price_str.replace(",", "."))
                items.append({
                    "code": code_str.upper() if code_str else "",
                    "desc": desc_str.strip(),
                    "qty": qty,
                    "priceUsd": price
                })
                continue
            except Exception:
                pass
                
        m_simple = re.match(
            r"^\s*[-*•]?\s*(\d+(?:[\.,]\d+)?)\s*(?:x|unidades|uds|unds|pzs|unid)?\s+(.+?)\s*(?:a|precio|costo)?\s+(\d+(?:[\.,]\d+)?)\s*$",
            line_clean,
            flags=re.IGNORECASE
        )
        if m_simple:
            qty_str, desc_str, price_str = m_simple.groups()
            try:
                qty = float(qty_str.replace(",", "."))
                price = float(price_str.replace(",", "."))
                items.append({
                    "code": "",
                    "desc": desc_str.strip(),
                    "qty": qty,
                    "priceUsd": price
                })
                continue
            except Exception:
                pass

    if not items:
        bracket_items = re.findall(r"\[\s*([^\]]+)\s*\]", text)
        for bit in bracket_items:
            parts = [p.strip() for p in bit.split("|")]
            if len(parts) >= 2:
                try:
                    qty = float(parts[0].replace(",", "."))
                    code = parts[1].upper() if len(parts) == 4 else ""
                    desc = parts[2] if len(parts) == 4 else (parts[1] if len(parts) == 3 else parts[0])
                    price = float(parts[-1].replace(",", "."))
                    items.append({
                        "code": code,
                        "desc": desc,
                        "qty": qty,
                        "priceUsd": price
                    })
                except Exception:
                    pass
                    
    if not items:
        return None
        
    return {
        "docType": doc_type,
        "currency": "usd",
        "exchangeRate": get_current_bcv_rate(),
        "docNumber": doc_number,
        "docDate": date.today().strftime("%Y-%m-%d"),
        "client": {
            "name": client_name or "CLIENTE REGISTRADO POR TEXTO",
            "address": address or "DIRECCIÓN DE CLIENTE",
            "rif": rif or "V-00000000-0",
            "phone": phone or "—",
            "salesman": salesman,
            "saleType": sale_type
        },
        "items": items
    }


def _parse_document_text(text: str) -> dict | None:
    t_norm = _normalize_text(text)
    
    # 1. Identificar tipo de documento
    is_cotizacion = "cotizacion" in t_norm or "presupuesto" in t_norm
    is_nota = "nota" in t_norm or "entrega" in t_norm
    
    if not (is_cotizacion or is_nota):
        return None
        
    # 2. Evitar procesar como cotización si parece un comprobante de retención o reporte tributario
    if any(k in t_norm for k in ("comprobante de retencion", "numero de comprobante", "iva retenido", "tributos", "anticipo islr", "debito fiscal")):
        return None
        
    # 3. Si parece una factura de compra recibida de un proveedor externo, no lo procesamos aquí (irá al flujo de facturas)
    is_compra = "proveedor" in t_norm or "emisor" in t_norm or "compra" in t_norm
    has_sufevica_emisor = "sufevica" in t_norm or "vittoria" in t_norm
    if is_compra and not has_sufevica_emisor and ("factura" in t_norm or "compras" in t_norm):
        return None

    # Debe ser una intención de creación explícita o implícita
    has_explicit = any(w in t_norm for w in ("crear", "nueva", "nuevo", "generar", "hacer", "registrar"))
    
    if not has_explicit:
        keywords_doc = ("cotizacion", "cotización", "presupuesto", "nota de entrega", "nota de despacho")
        if not any(k in t_norm for k in keywords_doc):
            return None
            
    doc_type = "cotizacion" if is_cotizacion else "nota"
    
    # Extraer campos cliente
    client_name = _take_eol_label(text, ("cliente", "razon social", "razón social", "nombre", "comprador", "dirigido a"))
    rif = _take_eol_label(text, ("rif", "c.i.", "ci", "cedula", "cédula", "identificacion", "rif cliente"))
    address = _take_eol_label(text, ("direccion", "dirección", "direccion fiscal", "dirección fiscal"))
    phone = _take_eol_label(text, ("telefono", "teléfono", "telefonos", "teléfonos", "telf", "celular"))
    doc_number = _take_eol_label(text, ("numero", "nro", "número", "documento", "cotizacion nro", "cotización nro", "nota nro"))
    salesman = _take_eol_label(text, ("vendedor", "ejecutivo", "atendido por")) or "FREDDY LOPEZ"
    sale_type = _take_eol_label(text, ("tipo", "tipo de venta", "condición de pago", "condicion de pago")) or "Contado"
    
    # Fallbacks de extracción de cliente en caso de texto desestructurado o pegado
    if not client_name:
        m_cli = re.search(r"(?im)(?:cliente|razon social|sres|señores)\s*[:\-]\s*(.+)$", text)
        if m_cli:
            client_name = m_cli.group(1).strip()
    if not rif:
        m_rif = re.search(r"\b([VEJPGvejpg][\s.\-]*\d{6,12}(?:[\s.\-]\d)?)\b", text)
        if m_rif:
            rif = re.sub(r"\s+", "", m_rif.group(1).strip())
            
    if rif:
        rif = _normalize_rif(rif)
        
    if not doc_number:
        m_num = re.search(r"(?im)(?:nro|numero|#|cotizacion|nota)\s*[:\-]?\s*(\d+)\b", text)
        if m_num:
            doc_number = m_num.group(1).strip()
        else:
            import random
            doc_number = f"{random.randint(1, 999):03d}{random.randint(1, 999):03d}"
            
    # Try the master stream parser first
    items = _parse_document_items_master(text)
    if items:
        return {
            "docType": doc_type,
            "currency": "usd",
            "exchangeRate": get_current_bcv_rate(),
            "docNumber": doc_number,
            "docDate": date.today().strftime("%Y-%m-%d"),
            "client": {
                "name": client_name or "CLIENTE REGISTRADO POR TEXTO",
                "address": address or "DIRECCIÓN DE CLIENTE",
                "rif": rif or "V-00000000-0",
                "phone": phone or "—",
                "salesman": salesman,
                "saleType": sale_type
            },
            "items": items
        }
            
    # Extraer ítems
    items = []
    lines = text.splitlines()
    
    for line in lines:
        line_clean = line.strip()
        if not line_clean:
            continue
            
        t_line = _normalize_text(line_clean)
        if any(w in t_line for w in ("total", "subtotal", "sub-total", "iva", "cliente", "rif", "direccion", "vendedor", "fecha", "cotizacion", "nota de entrega", "forma libre")):
            continue
            
        # 1. Estrategia Tabular (Múltiples espacios o tabulaciones)
        parts = [p.strip() for p in re.split(r"\s{2,}|\t", line_clean) if p.strip()]
        if len(parts) >= 3:
            nums = []
            for i, p in enumerate(parts):
                clean_num = re.sub(r"[^\d.,\-]", "", p)
                try:
                    val = float(clean_num.replace(".", "").replace(",", "."))
                    nums.append((i, val))
                except ValueError:
                    try:
                        val = float(clean_num.replace(",", "."))
                        nums.append((i, val))
                    except ValueError:
                        pass
            
            qty = None
            price = None
            code = ""
            desc = ""
            matched_math = False
            
            if len(nums) >= 2:
                for a in range(len(nums)):
                    for b in range(len(nums)):
                        if a == b: continue
                        for c in range(len(nums)):
                            if c == a or c == b: continue
                            val_a = nums[a][1]
                            val_b = nums[b][1]
                            val_c = nums[c][1]
                            if val_a > 0 and val_b > 0 and abs(val_a * val_b - val_c) < 0.05 * val_c:
                                qty = val_a
                                price = val_b
                                matched_math = True
                                break
                        if matched_math: break
                    if matched_math: break
                    
                if matched_math:
                    num_indices = {nums[a][0], nums[b][0], nums[c][0]}
                    text_parts = [parts[i] for i in range(len(parts)) if i not in num_indices]
                    if len(text_parts) >= 2:
                        code = text_parts[0].upper()
                        desc = " ".join(text_parts[1:])
                    elif len(text_parts) == 1:
                        desc = text_parts[0]
                else:
                    try:
                        first_clean = re.sub(r"[^\d.,\-]", "", parts[0])
                        qty = float(first_clean.replace(".", "").replace(",", "."))
                        
                        last_clean = re.sub(r"[^\d.,\-]", "", parts[-1])
                        price = float(last_clean.replace(".", "").replace(",", "."))
                        
                        if len(parts) >= 4:
                            penult_clean = re.sub(r"[^\d.,\-]", "", parts[-2])
                            try:
                                price_p = float(penult_clean.replace(".", "").replace(",", "."))
                                if abs(qty * price_p - price) < 0.1 * price:
                                    price = price_p
                            except ValueError:
                                pass
                                
                        text_parts = parts[1:-1] if len(parts) >= 3 else [parts[1]]
                        if len(text_parts) >= 2:
                            code = text_parts[0].upper()
                            desc = " ".join(text_parts[1:])
                        else:
                            desc = text_parts[0]
                    except Exception:
                        pass
                        
                if qty is not None and price is not None and qty > 0 and price >= 0:
                    items.append({
                        "code": code,
                        "desc": desc.strip(),
                        "qty": qty,
                        "priceUsd": price
                    })
                    continue

        # 1.5 Estrategia de Columnas al Final (para tablas copiadas de Excel con formato: CODIGO PRODUCTO CANT P.U. TOTAL)
        sp_parts = [p.strip() for p in line_clean.split() if p.strip()]
        if len(sp_parts) >= 4:
            try:
                def parse_token_val(tok):
                    clean = re.sub(r"[^\d.,\-]", "", tok)
                    if "," in clean and "." in clean:
                        clean = clean.replace(".", "").replace(",", ".")
                    elif "," in clean:
                        clean = clean.replace(",", ".")
                    return float(clean)

                qty_val = parse_token_val(sp_parts[-3])
                price_val = parse_token_val(sp_parts[-2])
                total_val = parse_token_val(sp_parts[-1])
                
                if qty_val > 0 and price_val >= 0 and total_val >= 0 and abs(qty_val * price_val - total_val) < 0.05 * total_val:
                    remaining_tokens = sp_parts[:-3]
                    first_tok = remaining_tokens[0]
                    
                    is_code = len(first_tok) <= 15 and (
                        any(c.isdigit() for c in first_tok) or 
                        not any(v in first_tok.lower() for v in ("de", "para", "con", "en", "tubo", "codo", "anillo", "brida", "niple", "tee"))
                    )
                    
                    if is_code and len(remaining_tokens) > 1:
                        code = first_tok.upper()
                        desc = " ".join(remaining_tokens[1:])
                    else:
                        code = ""
                        desc = " ".join(remaining_tokens)
                        
                    items.append({
                        "code": code,
                        "desc": desc.strip(),
                        "qty": qty_val,
                        "priceUsd": price_val
                    })
                    continue
            except Exception:
                pass

        # 2. Estrategia de Expresiones Regulares para líneas de texto simples
        m = re.match(
            r"^\s*[-*•]?\s*(\d+(?:[\.,]\d+)?)\s*(?:x|unidades|uds|unds|pzs|unid)?\s+([A-Z0-9\-]{3,15})?\s+(.+?)\s*(?:a|precio|costo)?\s+(\d+(?:[\.,]\d+)?)\s*$",
            line_clean,
            flags=re.IGNORECASE
        )
        if m:
            qty_str, code_str, desc_str, price_str = m.groups()
            try:
                qty = float(qty_str.replace(",", "."))
                price = float(price_str.replace(",", "."))
                items.append({
                    "code": code_str.upper() if code_str else "",
                    "desc": desc_str.strip(),
                    "qty": qty,
                    "priceUsd": price
                })
                continue
            except Exception:
                pass
                
        m_simple = re.match(
            r"^\s*[-*•]?\s*(\d+(?:[\.,]\d+)?)\s*(?:x|unidades|uds|unds|pzs|unid)?\s+(.+?)\s*(?:a|precio|costo)?\s+(\d+(?:[\.,]\d+)?)\s*$",
            line_clean,
            flags=re.IGNORECASE
        )
        if m_simple:
            qty_str, desc_str, price_str = m_simple.groups()
            try:
                qty = float(qty_str.replace(",", "."))
                price = float(price_str.replace(",", "."))
                items.append({
                    "code": "",
                    "desc": desc_str.strip(),
                    "qty": qty,
                    "priceUsd": price
                })
                continue
            except Exception:
                pass

    if not items:
        bracket_items = re.findall(r"\[\s*([^\]]+)\s*\]", text)
        for bit in bracket_items:
            parts = [p.strip() for p in bit.split("|")]
            if len(parts) >= 2:
                try:
                    qty = float(parts[0].replace(",", "."))
                    code = parts[1].upper() if len(parts) == 4 else ""
                    desc = parts[2] if len(parts) == 4 else (parts[1] if len(parts) == 3 else parts[0])
                    price = float(parts[-1].replace(",", "."))
                    items.append({
                        "code": code,
                        "desc": desc,
                        "qty": qty,
                        "priceUsd": price
                    })
                except Exception:
                    pass
                    
    if not items:
        return None
        
    return {
        "docType": doc_type,
        "currency": "usd",
        "exchangeRate": get_current_bcv_rate(),
        "docNumber": doc_number,
        "docDate": date.today().strftime("%Y-%m-%d"),
        "client": {
            "name": client_name or "CLIENTE REGISTRADO POR TEXTO",
            "address": address or "DIRECCIÓN DE CLIENTE",
            "rif": rif or "V-00000000-0",
            "phone": phone or "—",
            "salesman": salesman,
            "saleType": sale_type
        },
        "items": items
    }


def _normalize_phone_for_whatsapp(phone: str) -> str:
    # Remover todo lo que no sea dígito
    digits = re.sub(r"\D", "", phone)
    if not digits:
        return ""
    if len(digits) == 11 and digits.startswith("0"):
        digits = "58" + digits[1:]
    elif len(digits) == 10 and not digits.startswith("58"):
        digits = "58" + digits
    return digits

async def _generate_document_from_parsed_data(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    doc_data: dict,
) -> None:
    msg = update.effective_message
    if not msg:
        return
        
    if not await _check_and_consume_quota(update):
        return
        
    doc_type = doc_data["docType"]
    doc_num = doc_data["docNumber"]
    title_up = "COTIZACIÓN" if doc_type == "cotizacion" else "NOTA DE ENTREGA"
    emoji = "📋" if doc_type == "cotizacion" else "📦"
    
    try:
        ctx = _get_company_context(update)
        generados_dir = ctx.generados_dir
        generados_dir.mkdir(parents=True, exist_ok=True)
        # Enriquecer doc_data con los datos dinámicos de la empresa
        doc_data["company_brand"] = "FlashTax" if ctx.is_custom else "SUFEVICA"
        doc_data["company_name"] = ctx.company_name
        doc_data["company_rif"] = ctx.company_rif
        doc_data["primary_color"] = ctx.color_primary
        if ctx.company_address:
            doc_data["company_address"] = f"{ctx.company_address}\nTeléfono: {ctx.company_phone}\nE-mail: {ctx.company_email}"
        doc_data["signature_path"] = str(ctx.firma_sello_path)

        # Generar el PDF oficial y estético en segundo plano automáticamente
        from .pdf_generator import generate_document_pdf
        pdf_filename = f"{title_up}_{doc_num}.pdf"
        pdf_output_path = generados_dir / pdf_filename
        generate_document_pdf(doc_data, pdf_output_path)
        
        # Calcular montos para el mensaje pre-completado de WhatsApp
        client_name = doc_data['client'].get('name', 'Cliente')
        total_amount = 0.0
        for it in doc_data['items']:
            qty = float(it.get('qty', 1.0))
            price = float(it.get('priceUsd', 0.0))
            total_amount += (qty * price)
            
        currency = doc_data.get("currency", "usd")
        rate = float(doc_data.get("exchangeRate", get_current_bcv_rate())) if currency == "ves" else 1.0
        total_conv = total_amount * rate
        
        formatted_total = f"{total_conv:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        symbol = "$" if currency == "usd" else "Bs."
        
        # Construir mensaje de WhatsApp
        import urllib.parse
        brand_name = "FlashTax" if ctx.is_custom else "SUFEVICA"
        user_name = update.effective_user.first_name if update.effective_user else "Administrador"
        wa_msg = (
            f"Estimado/a *{client_name}*,\n\n"
            f"Le adjunto su *{title_up} Nro {doc_num}* por un monto total de *{symbol} {formatted_total}*.\n\n"
            f"Quedo a su entera disposición.\n\n"
            f"Atentamente,\n"
            f"*{user_name.upper()}* ({brand_name})"
        )
        encoded_text = urllib.parse.quote(wa_msg)
        normalized_phone = _normalize_phone_for_whatsapp(doc_data['client'].get('phone', ''))
        
        if normalized_phone:
            wa_url = f"https://api.whatsapp.com/send?phone={normalized_phone}&text={encoded_text}"
        else:
            wa_url = f"https://api.whatsapp.com/send?text={encoded_text}"
            
        # Registrar el documento generado en el historial
        _register_document_in_history(
            update=update,
            doc_type=doc_type,
            doc_number=doc_num,
            client_name=client_name,
            client_rif=doc_data['client'].get('rif', ''),
            total_amount=f"{symbol} {formatted_total}",
            temp_pdf_path=pdf_output_path
        )
            
        # Almacenar en user_data para el flujo de envío de correo
        context.user_data["share_doc"] = {
            "pdf_path": str(pdf_output_path),
            "pdf_filename": pdf_filename,
            "title": title_up,
            "doc_number": doc_num,
            "client_name": client_name,
            "total_amount": f"{symbol} {formatted_total}",
            "awaiting": None
        }
        
        # Crear los botones de compartir
        kb = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("🟢 Enviar por WhatsApp", url=wa_url),
                InlineKeyboardButton("📧 Enviar por Correo", callback_data="share_email")
            ]
        ])
        
        text = (
            f"✅ *{title_up} GENERADA CON ÉXITO* {emoji}\n\n"
            f"👤 *Cliente:* {doc_data['client']['name']}\n"
            f"🆔 *RIF:* {doc_data['client']['rif']}\n"
            f"🔢 *Número:* {doc_num}\n"
            f"🛒 *Productos:* {len(doc_data['items'])} ítems cargados\n\n"
            f"🔗 *Enlace en tu Servidor PC (Local):*\n"
            f"• *PDF:* `file:///c:/Users/Freddy%20Lopez/Documents/Telegram%20bot/bot_financiero_telegram/modulo_cotizaciones/generados/{pdf_filename}`\n\n"
            f"💡 *¿Cómo compartir el archivo PDF por WhatsApp?*\n"
            f"Debido a limitaciones de WhatsApp, los enlaces web no pueden adjuntar archivos locales. Para que el PDF le llegue al cliente por WhatsApp:\n"
            f"1. Mantén presionado el archivo PDF que el bot envió arriba (o haz clic derecho en PC).\n"
            f"2. Selecciona *Reenviar* o *Compartir* y elije *WhatsApp* para enviárselo directamente.\n"
            f"_(El botón verde de abajo te ayuda abriendo el chat de WhatsApp con el saludo de texto pre-escrito)._\n\n"
            f"📧 También puedes enviarlo automáticamente por Correo SMTP presionando el botón de abajo.\n\n"
            f"👇 *Elige una opción:*"
        )
        
        await msg.reply_text(text, parse_mode="Markdown")
        
        # Enviar PDF oficial (diseño corporativo premium) con los botones de compartir adjuntos
        await msg.reply_document(
            document=str(pdf_output_path),
            filename=pdf_filename,
            caption=f"📄 PDF Oficial {title_up} Nro {doc_num} - SUFEVICA",
            reply_markup=kb
        )
        
    except Exception as e:
        logger.exception("Error al generar el documento desde texto")
        await msg.reply_text(f"❌ Ocurrió un error al generar el documento: {e!s}")


async def _process_intent(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    text: str,
) -> None:
    msg = update.effective_message
    if not msg:
        return
        
    if not _check_permission(update, "tributos"):
        await msg.reply_text("❌ No tienes privilegios para acceder al módulo de Tributos.")
        return

    if not _can_modify_tributos(update):
        await msg.reply_text("❌ Tu nivel de autorización (\"Tributos Only\") solo te permite consultar y generar reportes, no registrar nueva información en el bot.")
        return

    ctx = _get_company_context(update)
    is_channel = update.channel_post is not None or update.edited_channel_post is not None
    emit_docs = _parse_emitir_retencion_request(text)
    if emit_docs is not None:
        if ctx.is_custom and ctx.company_type == "Ordinario":
            await msg.reply_text("⚠️ Opción Bloqueada: Los Contribuyentes Ordinarios no emiten comprobantes de retención de acuerdo con las normativas del SENIAT.")
            return
        await _start_emitir_retencion_flow(update, context, emit_docs)
        return
    # Registrar automáticamente si el texto tiene formato de retención de ISLR
    islr_data = _parse_retencion_islr_entry_request(text)
    if islr_data is None:
        islr_data = _parse_retencion_islr_from_any_text(text)
    if islr_data is not None:
        if not await _check_and_consume_quota(update):
            return
        emission_date = _parse_user_date(islr_data["fecha_emision"]) or date.today()
        monthly_path = excel_store.monthly_retencion_islr_path(ctx.retenciones_islr_dir, emission_date)
        
        # Validar duplicados de comprobantes de ISLR
        dup_info = excel_store.check_retencion_islr_exists(ctx.retenciones_islr_dir, islr_data["numero_comprobante"])
        if dup_info:
            context.user_data["pending_replace_islr"] = {
                "type": "text",
                "data": islr_data,
                "monthly_path": str(monthly_path),
            }
            kb = InlineKeyboardMarkup([
                [
                    InlineKeyboardButton("✅ Sí, sustituir", callback_data=f"rep_ret_conf:islr:{islr_data['numero_comprobante']}"),
                    InlineKeyboardButton("❌ Cancelar", callback_data="rep_ret_cancel")
                ]
            ])
            await msg.reply_text(
                f"⚠️ La retención de ISLR Nro {islr_data['numero_comprobante']} ya existe.\n"
                f"¿Deseas sustituir el comprobante de ISLR existente?",
                reply_markup=kb
            )
            return

        await _save_retencion_islr_text(update, context, islr_data, ctx, monthly_path)
        return

    ret_data = _parse_retencion_entry_request(text)
    # Registrar automáticamente si el texto ya contiene los campos en cualquier chat (privado o grupal).
    if ret_data is None:
        ret_data = _parse_retencion_from_any_text(text)
    if ret_data is not None:
        if not await _check_and_consume_quota(update):
            return
        try:
            inserted = excel_store.append_record(
                ctx.excel_path,
                fecha_emision=ret_data["fecha_emision"],
                numero_comprobante=ret_data["numero_comprobante"],
                rif=ret_data["rif"],
                fechas_facturas=ret_data["fechas_facturas"],
                numeros_facturas=ret_data["numeros_facturas"],
                controles_facturas=ret_data["controles_facturas"],
                total_compra_con_iva=ret_data["total_compra_con_iva"],
                base_imponible=ret_data["base_imponible"],
                iva_retenido=ret_data["iva_retenido"],
                ocr_snippet=(
                    "Comprobante de retencion (registro manual texto/voz). "
                    f"{text[:420]}"
                ),
            )
        except ValueError as e:
            await msg.reply_text(
                "No pude guardar en Excel por estructura incompatible.\n"
                f"Detalle: {e!s}"
            )
            return
        if not inserted:
            await _notify_same_source_channel(
                update,
                context,
                "Esa retencion parece duplicada; no se agregó de nuevo.",
            )
            return
        await _notify_same_source_channel(
            update,
            context,
            f"✅ Datos registrados correctamente en {ctx.excel_path.name}.",
        )
        return

    # Intentar primero el nuevo formato de Reporte Z
    z_nuevo = _parse_reporte_z_nuevo(text)
    if z_nuevo is not None:
        if not await _check_and_consume_quota(update):
            return
        try:
            inserted = excel_store.append_reporte_z_nuevo(
                ctx.reportes_z_path,
                numero_reporte=z_nuevo["numero_reporte"],
                fecha_emision=z_nuevo["fecha_emision"],
                sub_total=z_nuevo["sub_total"],
                base_imponible=z_nuevo["base_imponible"],
                monto_exento=z_nuevo["monto_exento"],
                iva=z_nuevo["iva"],
                total=z_nuevo["total"],
                texto_origen=text,
            )
        except Exception as e:
            await msg.reply_text(
                "No pude registrar el nuevo reporte Z en Excel.\n"
                f"Detalle: {e!s}"
            )
            return
            
        if not inserted:
            await _notify_same_source_channel(
                update,
                context,
                f"⚠️ El Reporte Z Nro {z_nuevo['numero_reporte']} ya se encuentra registrado en el sistema; no se agregó de nuevo.",
            )
            return

        await _notify_same_source_channel(
            update,
            context,
            f"✅ Reporte Z Nro {z_nuevo['numero_reporte']} registrado correctamente en {ctx.reportes_z_path.name}.",
        )
        return

    v_data = _parse_venta_o_reportez(text)
    if v_data is not None:
        if not await _check_and_consume_quota(update):
            return
        try:
            b_dec = excel_store.parse_amount_ves_string(v_data["base_imponible"])
            i_dec = excel_store.parse_amount_ves_string(v_data["iva"])
            t_dec = excel_store.parse_amount_ves_string(v_data["total"])
            
            from decimal import Decimal
            if b_dec is not None and i_dec is None:
                i_dec = (b_dec * Decimal("0.16")).quantize(Decimal("0.01"))
            if b_dec is not None and t_dec is None:
                t_dec = b_dec + (i_dec or Decimal("0"))
            if t_dec is not None and b_dec is None:
                b_dec = (t_dec / Decimal("1.16")).quantize(Decimal("0.01"))
                i_dec = t_dec - b_dec
                
            base_str = str(b_dec) if b_dec is not None else v_data["base_imponible"]
            iva_str = str(i_dec) if i_dec is not None else v_data["iva"]
            tot_str = str(t_dec) if t_dec is not None else v_data["total"]
            
            if v_data["clasificacion"] == "Reporte Z":
                inserted = excel_store.append_reporte_z_nuevo(
                    ctx.reportes_z_path,
                    numero_reporte=v_data["numero_documento"],
                    fecha_emision=v_data["fecha"],
                    sub_total=base_str,
                    base_imponible=base_str,
                    monto_exento="0,00",
                    iva=iva_str,
                    total=tot_str,
                    texto_origen=text,
                )
            else:
                inserted = excel_store.append_venta_record(
                    ctx.facturas_emitidas_path,
                    clasificacion=v_data["clasificacion"],
                    estado=v_data["estado"],
                    fecha=v_data["fecha"],
                    numero_documento=v_data["numero_documento"],
                    razon_social=v_data["razon_social"],
                    rif=v_data["rif"],
                    base_imponible=base_str,
                    iva=iva_str,
                    total=tot_str,
                    texto_origen=text,
                )
        except Exception as e:
            await msg.reply_text(
                "No pude registrar la venta en Excel.\n"
                f"Detalle: {e!s}"
            )
            return
            
        if not inserted:
            await _notify_same_source_channel(
                update,
                context,
                f"⚠️ Esa {v_data['clasificacion']} Nro {v_data['numero_documento']} ya se encuentra registrada; no se agregó de nuevo.",
            )
            return
            
        path_name = ctx.reportes_z_path.name if v_data["clasificacion"] == "Reporte Z" else ctx.facturas_emitidas_path.name
        await _notify_same_source_channel(
            update,
            context,
            f"✅ {v_data['clasificacion']} Nro {v_data['numero_documento']} registrado correctamente en {path_name}.",
        )
        return

    fc = parse_factura_compra_text(text)
    if fc is not None:
        if not await _check_and_consume_quota(update):
            return
        is_sale = False
        if fc.proveedor_rif:
            clean_rif = re.sub(r"\D", "", str(fc.proveedor_rif))
            clean_emitter_rif = re.sub(r"\D", "", str(ctx.company_rif))
            if clean_emitter_rif and clean_emitter_rif in clean_rif:
                is_sale = True
        if not is_sale and fc.proveedor and ctx.company_name.upper() in str(fc.proveedor).upper():
            is_sale = True

        if is_sale:
            # Factura emitida por SUFEVICA -> Venta
            try:
                inserted = excel_store.append_venta_record(
                    ctx.facturas_emitidas_path,
                    clasificacion="Factura Emitida",
                    estado="REGISTRADO",
                    fecha=fc.fecha_emision,
                    numero_documento=fc.numero_documento,
                    razon_social=fc.receptor,
                    rif=fc.receptor_rif,
                    base_imponible=fc.base_imponible or fc.subtotal,
                    iva=fc.monto_iva,
                    total=fc.total,
                    texto_origen=text,
                )
            except Exception as e:
                await msg.reply_text(
                    f"No pude guardar la factura de venta en Excel: {e!s}"
                )
                return
                
            if not inserted:
                await _notify_same_source_channel(
                    update,
                    context,
                    f"⚠️ La factura de venta Nro {fc.numero_documento} ya se encuentra registrada; no se agregó de nuevo.",
                )
                return
            await _notify_same_source_channel(
                update,
                context,
                f"✅ Factura de venta Nro {fc.numero_documento or '—'} registrada correctamente en {ctx.facturas_emitidas_path.name}.",
            )
            return
        else:
            # Factura recibida de proveedor -> Compra
            warn = ""
            try:
                base = excel_store.parse_amount_ves_string(fc.base_imponible or fc.subtotal)
                exento = excel_store.parse_amount_ves_string(fc.monto_exento) or Decimal("0")
                iva = excel_store.parse_amount_ves_string(fc.monto_iva)
                total = excel_store.parse_amount_ves_string(fc.total)
                if base is not None and iva is not None and total is not None:
                    diff = (base + exento + iva - total).copy_abs()
                    if diff > Decimal("0.02"):
                        warn = (
                            f" [ADVERTENCIA: inconsistencia matemática "
                            f"base({base})+exento({exento})+iva({iva})!=total({total})]"
                        )
            except Exception:
                warn = ""
            try:
                inserted = excel_store.append_factura_compra(
                    ctx.facturas_recibidas_path,
                    tipo_documento=fc.tipo_documento,
                    fecha_emision=fc.fecha_emision,
                    fecha_vencimiento=fc.fecha_vencimiento,
                    numero_documento=fc.numero_documento,
                    numero_control=fc.numero_control,
                    proveedor=fc.proveedor,
                    proveedor_rif=fc.proveedor_rif,
                    proveedor_telefono=fc.proveedor_telefono,
                    direccion_fiscal_proveedor=fc.direccion_fiscal_proveedor,
                    receptor=fc.receptor,
                    receptor_rif=fc.receptor_rif,
                    subtotal=fc.subtotal,
                    monto_exento=fc.monto_exento,
                    base_imponible=fc.base_imponible,
                    monto_iva=fc.monto_iva,
                    total=fc.total,
                    texto_resumen=(text[:420] + warn)[:500],
                )
            except ValueError as e:
                await msg.reply_text(
                    f"No pude guardar la factura de compra: {e!s}"
                )
                return
                
            if not inserted:
                await _notify_same_source_channel(
                    update,
                    context,
                    f"⚠️ La factura recibida Nro {fc.numero_documento} del proveedor {fc.proveedor or ''} ({fc.proveedor_rif or ''}) ya se encuentra registrada; no se agregó de nuevo.",
                )
                return
                
            await _notify_same_source_channel(
                update,
                context,
                "✅ Datos registrados correctamente en "
                f"{ctx.facturas_recibidas_path.name} (Doc {fc.numero_documento or '—'}).",
            )
            return

    intent = _match_intent(text)
    if intent == "send_facturas_compra_excel":
        excel_store.ensure_factura_compra_workbook(ctx.facturas_recibidas_path)
        await msg.reply_document(
            document=str(ctx.facturas_recibidas_path),
            filename=ctx.facturas_recibidas_path.name,
            caption="Facturas de compra / recibidas (Subtotal, IVA, Total, etc.).",
        )
        return
    if intent == "send_excel":
        if not ctx.excel_path.exists():
            await msg.reply_text(
                "Todavía no existe el archivo Excel. Registra una retención primero "
                "o revisa EXCEL_PATH en .env."
            )
            return
        await msg.reply_document(
            document=str(ctx.excel_path),
            filename=ctx.excel_path.name,
            caption="consolidado_financiero.xlsx (retenciones; no es el de facturas compra).",
        )
        return
    if intent == "summary_today":
        today = date.today()
        n, total = excel_store.summary_for_date(ctx.excel_path, today)
        await msg.reply_text(
            f"Resumen del {today.strftime('%d/%m/%Y')}: {n} registro(s). "
            f"Suma de montos: {total}"
        )
        return
    if intent == "tributos_report":
        today = date.today()
        fortnight = 1 if today.day <= 15 else 2
        report = tributario_engine.get_compromiso_tributario_report(
            today.year, today.month, fortnight,
            facturas_emitidas_path=ctx.facturas_emitidas_path,
            reportes_z_path=ctx.reportes_z_path,
            retenciones_emitidas_dir=ctx.retenciones_emitidas_dir,
            excel_path=ctx.excel_path,
            retenciones_islr_dir=ctx.retenciones_islr_dir
        )
        text = format_tributos_report(report)
        kb = _tributos_keyboard(today.year, today.month, fortnight, _generate_short_summary(report))
        await msg.reply_text(text, reply_markup=kb, parse_mode="Markdown")
        return
    if intent == "seniat_txt":
        today = date.today()
        fortnight = 1 if today.day <= 15 else 2
        status_msg = await msg.reply_text(
            f"⏳ *Generando archivos TXT según la normativa del SENIAT para la quincena actual...*",
            parse_mode="Markdown"
        )
        asyncio.create_task(_send_seniat_txt_telegram_async(update, context, today.year, today.month, fortnight, status_msg))
        return
    if intent == "retenciones_report":
        parsed = _parse_retenciones_report_request(text)
        if parsed is None:
            await msg.reply_text(
                "Formato no reconocido. Usa por ejemplo: "
                "«retenciones recibidas del 01/05/2026 al 31/05/2026 en excel» "
                "o «... en pdf»."
            )
            return
        date_from, date_to, out_fmt = parsed
        records = excel_store.retenciones_by_document_date(
            ctx.excel_path,
            date_from=date_from,
            date_to=date_to,
        )
        if not records:
            await msg.reply_text(
                "No encontré retenciones en ese rango por fecha del documento."
            )
            return
        suffix = ".xlsx" if out_fmt == "excel" else ".pdf"
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            out_path = Path(tmp.name)
        try:
            if out_fmt == "excel":
                excel_store.export_retenciones_excel(records, out_path)
            else:
                excel_store.export_retenciones_pdf(
                    records,
                    out_path,
                    date_from=date_from,
                    date_to=date_to,
                )
            if out_fmt == "excel":
                filename = "RETEN-REC.xlsx"
            else:
                report_name = (
                    "detallado_de_retenciones_recibidas_de_clientes_del_"
                    f"{date_from.strftime('%d-%m-%Y')}_al_{date_to.strftime('%d-%m-%Y')}"
                )
                filename = f"{report_name}{suffix}"
            await msg.reply_document(
                document=str(out_path),
                filename=filename,
                caption=(
                    "Reporte generado: "
                    "detallado de retenciones recibidas de clientes del "
                    f"{date_from.strftime('%d/%m/%Y')} al {date_to.strftime('%d/%m/%Y')} "
                    "(filtrado por fecha del documento)."
                ),
            )
        finally:
            out_path.unlink(missing_ok=True)
        return
    await msg.reply_text(
        "No reconoci la orden.\n"
        "Ejemplos:\n"
        "1) «resumen de hoy»\n"
        "2) «enviar excel» (retenciones consolidado) o «enviar excel facturas compra»\n"
        "3) «retenciones recibidas del 01/05/2026 al 31/05/2026 en excel»\n"
        "4) «registrar retencion, fecha:24/04/2026, comprobante:20260400000381, "
        "rif:J-41278020-4, fecha factura:23/04/2026, nro factura:00007553, "
        "control:Z7C7018762, total:1.400,00, base imponible:1.206,90, iva retenido:144,83»\n"
        "5) Pega texto de factura de compra / factura recibida (Febeca, totales en Bs, etc.).\n"
        "6) Emitir retencion de facturas:06949950|06949951"
    )


async def _save_retencion_islr_text(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    islr_data: dict,
    ctx: CompanyContext,
    monthly_path: Path
) -> None:
    msg = update.effective_message
    if not msg:
        return
    try:
        base_val = excel_store.parse_amount_ves_string(islr_data["base_imponible"]) or Decimal("0")
        rate_val = excel_store.parse_amount_ves_string(islr_data["porcentaje_retencion"]) or Decimal("0")
        if rate_val > 1:
            rate_val = rate_val / 100
        islr_val = excel_store.parse_amount_ves_string(islr_data["islr_retenido"]) or Decimal("0")
        total_val = excel_store.parse_amount_ves_string(islr_data["total_factura"]) or Decimal("0")
        
        emission_date = _parse_user_date(islr_data["fecha_emision"]) or date.today()

        excel_store.append_retencion_islr(
            monthly_path,
            numero_comprobante=islr_data["numero_comprobante"],
            fecha_emision=islr_data["fecha_emision"],
            periodo_fiscal=_periodo_fiscal(emission_date),
            proveedor=islr_data["proveedor"],
            proveedor_rif=islr_data["proveedor_rif"],
            concepto_retencion=islr_data["concepto_retencion"],
            numero_documento=islr_data["numero_documento"],
            numero_control=islr_data["numero_control"],
            base_imponible=base_val,
            porcentaje_retencion=rate_val,
            islr_retenido=islr_val,
            total_factura=total_val,
        )
        await _notify_same_source_channel(
            update,
            context,
            f"✅ Retención de ISLR Nro {islr_data['numero_comprobante']} registrada correctamente en {monthly_path.name}.",
        )
    except Exception as e:
        await msg.reply_text(
            "No pude registrar la retención de ISLR en Excel.\n"
            f"Detalle: {e!s}"
        )


async def _save_retencion_islr_ocr(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    pending: dict,
    ctx: CompanyContext,
    monthly_path: Path
) -> None:
    msg = update.effective_message
    if not msg:
        return
    try:
        base_val = excel_store.parse_amount_ves_string(pending["base_imponible"]) or Decimal("0")
        rate_val = excel_store.parse_amount_ves_string(pending["porcentaje_retencion"]) or Decimal("0")
        if rate_val > 1:
            rate_val = rate_val / 100
        islr_val = excel_store.parse_amount_ves_string(pending["islr_retenido"]) or Decimal("0")
        total_val = excel_store.parse_amount_ves_string(pending["total_factura"]) or Decimal("0")
        
        emission_date = tributario_engine._parse_row_date(pending["fecha_emision"]) or date.today()

        excel_store.append_retencion_islr(
            monthly_path,
            numero_comprobante=pending["numero_comprobante"],
            fecha_emision=pending["fecha_emision"],
            periodo_fiscal=_periodo_fiscal(emission_date),
            proveedor=pending["proveedor"],
            proveedor_rif=pending["proveedor_rif"],
            concepto_retencion=pending["concepto_retencion"],
            numero_documento=pending["numero_documento"],
            numero_control=pending["numero_control"],
            base_imponible=base_val,
            porcentaje_retencion=rate_val,
            islr_retenido=islr_val,
            total_factura=total_val,
        )
        await msg.reply_text(f"✅ Retención de ISLR Nro {pending['numero_comprobante']} registrada con éxito en {monthly_path.name}.")
    except Exception as e:
        logger.exception("Error al guardar retención ISLR desde OCR")
        await msg.reply_text(f"❌ Error al registrar retención ISLR: {e!s}")


async def _emitir_retencion_generate(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
) -> None:
    msg = update.effective_message
    pending = context.user_data.get("pending_emit_ret")
    if not msg or not pending:
        return
        
    if not await _check_and_consume_quota(update):
        return
    docs = list(pending.get("docs", []))
    ctx = _get_company_context(update)
    items = excel_store.load_facturas_by_document_numbers(ctx.facturas_recibidas_path, docs)
    if not items:
        await msg.reply_text("No encontré las facturas solicitadas al momento de emitir.")
        context.user_data.pop("pending_emit_ret", None)
        return
    emission_date_str = str(pending.get("emission_date") or "").strip()
    emission_date = _parse_user_date(emission_date_str) if emission_date_str else None
    if emission_date is None:
        emission_date = date.today()
        emission_date_str = emission_date.strftime("%d/%m/%Y")
    monthly_path = _reten_emit_monthly_path(emission_date, update.effective_user.id)
    seq_mode = str(pending.get("seq_mode") or "auto")
    if seq_mode == "manual":
        num_comp = str(pending.get("manual_num") or "").strip()
        if not (len(num_comp) == 14 and num_comp.isdigit()):
            await msg.reply_text(
                "Número manual inválido. Debe tener 14 dígitos (YYYYMM + 8 secuencial)."
            )
            return
    else:
        num_comp = _get_next_retencion_emitida_number_for_user(
            update.effective_user.id,
            monthly_path,
            emission_date=emission_date,
        )
        logger.info(
            "Correlativo asignado desde Excel %s: %s",
            monthly_path.parent,
            num_comp,
        )
    out_fmt = str(pending.get("format") or "pdf").strip().lower()
    provider = str(pending.get("provider") or "").strip()
    provider_rif = str(pending.get("provider_rif") or "").strip()
    provider_phone = str(pending.get("provider_phone") or "").strip()
    provider_address = str(pending.get("provider_address") or "").strip()
    base_total, iva_total, retenido = _totals_for_items(items)

    # Validar duplicados de comprobantes de IVA
    dup_info = excel_store.check_retencion_emitida_exists(ctx.retenciones_emitidas_dir, num_comp)
    if dup_info and not pending.get("replace_confirmed"):
        kb = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("✅ Sí, sustituir", callback_data=f"rep_ret_conf:iva:{num_comp}"),
                InlineKeyboardButton("❌ Cancelar", callback_data="rep_ret_cancel")
            ]
        ])
        await msg.reply_text(
            f"⚠️ El comprobante de IVA Nro {num_comp} ya existe.\n"
            f"¿Deseas sustituir el correlativo existente y reactivar las facturas previas?",
            reply_markup=kb
        )
        return

    if dup_info and pending.get("replace_confirmed"):
        excel_store.delete_retencion_emitida_row(dup_info[0], dup_info[1])
        logger.info("Deleted duplicate IVA retention %s at row %d", dup_info[0], dup_info[1])

    excel_store.append_retencion_emitida(
        monthly_path,
        numero_comprobante=num_comp,
        fecha_emision=emission_date_str,
        periodo_fiscal=_periodo_fiscal(emission_date),
        proveedor=provider,
        proveedor_rif=provider_rif,
        direccion_fiscal_prov=provider_address,
        documentos="|".join(docs),
        controles="|".join(str(it.numero_control or "") for it in items),
        base_imponible_total=base_total,
        iva_total=iva_total,
        porcentaje_retencion=RETENTION_RATE,
        iva_retenido_total=retenido,
        formato_salida=out_fmt,
    )
    suffix = ".xlsx" if out_fmt == "excel" else ".pdf"
    ctx.generados_dir.mkdir(parents=True, exist_ok=True)
    out_path = ctx.generados_dir / f"COMPROBANTE-RET-{num_comp}{suffix}"
    try:
        if out_fmt == "excel":
            excel_store.export_comprobante_emitido_excel(
                out_path=out_path,
                numero_comprobante=num_comp,
                fecha_emision=emission_date_str,
                periodo_fiscal=_periodo_fiscal(emission_date),
                proveedor=provider,
                proveedor_rif=provider_rif,
                proveedor_telefono=provider_phone,
                direccion_fiscal_prov=provider_address,
                items=items,
                porcentaje_retencion=RETENTION_RATE,
                emisor_nombre=ctx.company_name,
                emisor_rif=ctx.company_rif,
            )
        else:
            excel_store.export_comprobante_emitido_pdf(
                out_path=out_path,
                numero_comprobante=num_comp,
                fecha_emision=emission_date_str,
                periodo_fiscal=_periodo_fiscal(emission_date),
                proveedor=provider,
                proveedor_rif=provider_rif,
                proveedor_telefono=provider_phone,
                direccion_fiscal_prov=provider_address,
                items=items,
                porcentaje_retencion=RETENTION_RATE,
                firma_sello_path=ctx.firma_sello_path,
                emisor_nombre=ctx.company_name,
                emisor_rif=ctx.company_rif,
            )
        
        # Registrar en el historial
        _register_document_in_history(
            update=update,
            doc_type="retencion_iva",
            doc_number=num_comp,
            client_name=provider,
            client_rif=provider_rif,
            total_amount=f"Bs. {retenido:,.2f}",
            temp_pdf_path=out_path
        )
        
        await msg.reply_document(
            document=str(out_path),
            filename=f"COMPROBANTE-RET-{num_comp}{suffix}",
            caption=(
                f"✅ Comprobante emitido ({num_comp}) y guardado en {monthly_path.name}."
            ),
        )
    finally:
        pass
    context.user_data.pop("pending_emit_ret", None)


async def handle_emit_retention_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
) -> None:
    q = update.callback_query
    if not q:
        return
    await q.answer()
    data = (q.data or "").strip()
    msg = q.message
    if not msg:
        return
    pending = context.user_data.get("pending_emit_ret")
    if not pending:
        await msg.reply_text("No hay una emisión pendiente. Envía el comando nuevamente.")
        return
    if data == "emit_seq_yes":
        pending["seq_mode"] = "auto"
        emission_date = _parse_user_date(str(pending.get("emission_date") or ""))
        if emission_date is None:
            emission_date = date.today()
        monthly_path = _reten_emit_monthly_path(emission_date, update.effective_user.id)
        next_num = _get_next_retencion_emitida_number_for_user(
            update.effective_user.id,
            monthly_path,
            emission_date=emission_date,
        )
        kb = InlineKeyboardMarkup(
            [[
                InlineKeyboardButton("Fecha de hoy", callback_data="emit_date_today"),
                InlineKeyboardButton("Fecha manual", callback_data="emit_date_manual"),
            ]]
        )
        await msg.reply_text(
            f"Secuencia automática desde Excel. Correlativo a asignar: {next_num}\n"
            "¿Qué fecha de emisión deseas?",
            reply_markup=kb,
        )
        return
    if data == "emit_seq_no":
        pending["seq_mode"] = "manual"
        pending["awaiting"] = "manual_number"
        await msg.reply_text(
            "Envía el número de comprobante manual (14 dígitos, formato YYYYMM########)."
        )
        return
    if data == "emit_date_today":
        pending["emission_date"] = date.today().strftime("%d/%m/%Y")
        kb = InlineKeyboardMarkup(
            [[
                InlineKeyboardButton("PDF", callback_data="emit_fmt_pdf"),
                InlineKeyboardButton("Excel", callback_data="emit_fmt_excel"),
            ]]
        )
        await msg.reply_text("Formato del comprobante a generar:", reply_markup=kb)
        return
    if data == "emit_date_manual":
        pending["awaiting"] = "manual_date"
        await msg.reply_text("Envía la fecha de emisión en formato DD/MM/AAAA.")
        return
    if data in ("emit_fmt_pdf", "emit_fmt_excel"):
        pending["format"] = "pdf" if data.endswith("pdf") else "excel"
        await _emitir_retencion_generate(update, context)
        return
async def handle_replace_retention_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
) -> None:
    q = update.callback_query
    if not q:
        return
    await q.answer()
    data = (q.data or "").strip()
    msg = q.message
    if not msg:
        return
    
    if data.startswith("rep_ret_conf:"):
        parts = data.split(":")
        if len(parts) < 3:
            return
        ret_type = parts[1]
        num_comp = parts[2]
        
        ctx = _get_company_context(update)
        
        if ret_type == "iva":
            pending = context.user_data.get("pending_emit_ret")
            if not pending:
                await msg.reply_text("No hay una emisión de IVA pendiente en tu sesión.")
                return
            
            # 1. Delete the existing IVA retention
            dup_info = excel_store.check_retencion_emitida_exists(ctx.retenciones_emitidas_dir, num_comp)
            if dup_info:
                path, row_idx, _ = dup_info
                excel_store.delete_retencion_emitida_row(path, row_idx)
                logger.info("Deleted existing IVA retention %s at row %d", path, row_idx)
            
            # 2. Re-emit the retention with replace_confirmed=True
            pending["replace_confirmed"] = True
            
            try:
                await q.delete_message()
            except Exception:
                pass
                
            await _emitir_retencion_generate(update, context)
            
        elif ret_type == "islr":
            pending_replace = context.user_data.get("pending_replace_islr")
            if not pending_replace:
                await msg.reply_text("No hay una retención de ISLR pendiente para sustituir.")
                return
            
            # 1. Delete the existing ISLR retention
            dup_info = excel_store.check_retencion_islr_exists(ctx.retenciones_islr_dir, num_comp)
            if dup_info:
                path, row_idx, _ = dup_info
                excel_store.delete_retencion_islr_row(path, row_idx)
                logger.info("Deleted existing ISLR retention %s at row %d", path, row_idx)
                
            # 2. Save the new one
            is_type = pending_replace.get("type")
            islr_data = pending_replace.get("data")
            
            emission_date_str = islr_data.get("fecha_emision")
            emission_date = _parse_user_date(emission_date_str) or tributario_engine._parse_row_date(emission_date_str) or date.today()
            monthly_path = excel_store.monthly_retencion_islr_path(ctx.retenciones_islr_dir, emission_date)
            
            try:
                await q.delete_message()
            except Exception:
                pass
                
            if is_type == "text":
                await _save_retencion_islr_text(update, context, islr_data, ctx, monthly_path)
            elif is_type == "ocr":
                await _save_retencion_islr_ocr(update, context, islr_data, ctx, monthly_path)
                
            context.user_data.pop("pending_replace_islr", None)
            
    elif data == "rep_ret_cancel":
        context.user_data.pop("pending_emit_ret", None)
        context.user_data.pop("pending_replace_islr", None)
        
        try:
            await q.delete_message()
        except Exception:
            pass
            
        await msg.reply_text("❌ Operación cancelada. El comprobante existente no ha sido modificado.")

async def mi_id_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Para comprobar el id que Telegram usa (cualquier usuario)."""
    del context
    msg = update.effective_message
    u = update.effective_user
    if not msg or not u:
        return
    if _allowed(update):
        await msg.reply_text(
            f"Estas autorizado. Tu id: {u.id}. TELEGRAM_ALLOWED_USER_ID en .env debe ser ese número."
        )
    else:
        await msg.reply_text(
            f"Tu id de Telegram es: {u.id}\n"
            "Ponlo en TELEGRAM_ALLOWED_USER_ID del archivo .env del bot y reinicia."
        )


async def _show_startup_menu(update: Update, context: ContextTypes.DEFAULT_TYPE, msg_to_edit=None) -> None:
    u = update.effective_user
    if u is None:
        return
    user = user_manager.get_user(u.id)
    if not user:
        return
    
    role = user.get("role")
    kb_list = []
    
    if role == "admin":
        kb_list.append([InlineKeyboardButton("🏢 Operar SUFEVICA", callback_data="work_panel:sufevica")])
        kb_list.append([InlineKeyboardButton("👥 Administrar Clientes", callback_data="work_panel:admin_clients")])
        kb_list.append([InlineKeyboardButton("✉️ Enviar Invitación", callback_data="work_panel:send_invite")])
        welcome_text = "🚀 *Menú de Inicio - Panel de Control* 🚀\n\nBienvenido, Administrador. Selecciona tu panel de trabajo:"
    elif role == "nueva_empresa":
        kb_list.append([InlineKeyboardButton("🏢 Operar Mi Empresa", callback_data="work_panel:client_operate")])
        kb_list.append([InlineKeyboardButton("⚙️ Configurar Empresa", callback_data="work_panel:client_config")])
        welcome_text = "🚀 *Menú de Inicio - Cliente* 🚀\n\nBienvenido. Selecciona tu panel de trabajo:"
    else:
        kb_list.append([InlineKeyboardButton("🏢 Operar SUFEVICA", callback_data="work_panel:sufevica")])
        welcome_text = "🚀 *Menú de Inicio* 🚀\n\nBienvenido. Presiona el botón de abajo para operar:"
        
    kb = InlineKeyboardMarkup(kb_list)
    
    if msg_to_edit:
        await msg_to_edit.edit_text(welcome_text, reply_markup=kb, parse_mode="Markdown")
    else:
        msg = update.effective_message
        if msg:
            await msg.reply_text(welcome_text, reply_markup=kb, parse_mode="Markdown")


async def handle_work_panel_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    q = update.callback_query
    if not q:
        return
    await q.answer()
    data = q.data
    msg = q.message
    if not msg:
        return
        
    parts = data.split(":")
    action = parts[1]
    
    uid = update.effective_user.id
    user = user_manager.get_user(uid)
    if not user:
        return
        
    if action == "sufevica":
        await q.delete_message()
        await context.bot.send_message(
            chat_id=msg.chat_id,
            text="🏢 *Módulo SUFEVICA Activo*\n\nOperando bajo el contexto de SUFEVICA. Utiliza el teclado inferior para navegar:",
            reply_markup=_main_keyboard(uid),
            parse_mode="Markdown"
        )
    elif action == "admin_clients":
        await _show_admin_panel(update, context, msg_to_edit=msg)
    elif action == "client_operate":
        await q.delete_message()
        await context.bot.send_message(
            chat_id=msg.chat_id,
            text="🏢 *Módulo de Operaciones Activo*\n\nOperando bajo el contexto de tu empresa. Utiliza el teclado inferior para navegar:",
            reply_markup=_main_keyboard(uid),
            parse_mode="Markdown"
        )
    elif action == "client_config":
        await _show_company_config_menu(update, context, msg_to_edit=msg)
    elif action == "send_invite":
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("👤 Cliente (FlashTax)", callback_data="work_panel:invite_select:cliente")],
            [InlineKeyboardButton("🛡️ Usuario SUFEVICA", callback_data="work_panel:invite_select:sufevica")],
            [InlineKeyboardButton("🔙 Volver al Inicio", callback_data="work_panel:start")]
        ])
        await msg.edit_text(
            "✉️ *Enviar Enlace de Invitación*\n\n"
            "Selecciona el destinatario de la invitación:",
            reply_markup=kb,
            parse_mode="Markdown"
        )
    elif action == "invite_select":
        import urllib.parse
        target_type = parts[2]
        bot_username = context.bot.username
        
        if target_type == "sufevica":
            link = f"https://t.me/{bot_username}?start=solicitar_sufevica"
            text_msg = (
                "Hola, te invito a unirte al bot financiero como Usuario SUFEVICA.\n"
                "Por favor, presiona el siguiente enlace para solicitar tu nivel de acceso:\n"
            )
            title_lbl = "Usuario SUFEVICA"
        else:
            link = f"https://t.me/{bot_username}?start=solicitar_cliente"
            text_msg = (
                "Hola, te invito a registrarte como Cliente (FlashTax) en el bot financiero.\n"
                "Por favor, presiona el siguiente enlace para iniciar el registro de tu empresa:\n"
            )
            title_lbl = "Cliente (FlashTax)"
            
        share_url = f"https://t.me/share/url?url={urllib.parse.quote(link)}&text={urllib.parse.quote(text_msg)}"
        
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("🔗 Compartir con mis Contactos", url=share_url)],
            [InlineKeyboardButton("🔙 Atrás", callback_data="work_panel:send_invite")]
        ])
        
        await msg.edit_text(
            f"✉️ *Compartir Invitación - {title_lbl}*\n\n"
            f"Se ha generado el enlace de invitación:\n"
            f"`{link}`\n\n"
            f"Pulsa el botón de abajo para abrir tu lista de contactos en Telegram y enviarles el enlace automáticamente.",
            reply_markup=kb,
            parse_mode="Markdown"
        )
    elif action == "start":
        await _show_startup_menu(update, context, msg_to_edit=msg)


async def start_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if context.args:
        arg = context.args[0]
        if arg == "solicitar_sufevica":
            await _handle_solicitar_access_flow(update, context, req_type="sufevica")
            return
        elif arg == "solicitar_cliente" or arg == "solicitar":
            await _handle_solicitar_access_flow(update, context, req_type="cliente")
            return

    if not _allowed(update):
        await _deny(update, context)
        return
    context.user_data["voice_mode"] = False
    context.user_data.pop("pending_doc", None)
    context.user_data.pop("active_menu", None)
    context.user_data.pop("awaiting_emit_docs", None)
    context.user_data.pop("admin_state", None)
    context.user_data.pop("admin_new_user", None)
    context.user_data.pop("share_doc", None)
    if update.message:
        await _show_startup_menu(update, context)


def _process_parsed_ocr_invoice(fc: object) -> dict:
    from decimal import Decimal
    from . import tributario_engine
    
    # 1. Validar RIF
    rif_valido = tributario_engine.validar_rif_venezolano(fc.proveedor_rif)
    
    # 2. Conversión a Bolívares si es USD
    moneda = (getattr(fc, "moneda_original", "VES") or "VES").strip().upper()
    
    # Normalizar montos
    def _clean_val(v):
        if v is None:
            return Decimal("0.00")
        if isinstance(v, (int, float)):
            try:
                return Decimal(str(v))
            except Exception:
                return Decimal("0.00")
        s = str(v).strip()
        if not s:
            return Decimal("0.00")
        # Clean currency prefixes like Bs., Bs, bs., bsf, ves, $, etc.
        import re
        s = re.sub(r'(?i)^(bsf\.?|bs\.?|ves\.?|usd\.?|\$)\s*', '', s)
        s = s.replace(" ", "")
        if re.search(r"\d+\.\d{3},\d{2}$", s) or (
            "," in s and "." in s and s.rfind(",") > s.rfind(".")
        ):
            s = s.replace(".", "").replace(",", ".")
        elif "," in s and "." not in s:
            s = s.replace(",", ".")
        s = re.sub(r"[^\d.\-]", "", s)
        try:
            return Decimal(s)
        except Exception:
            return Decimal("0.00")
            
    subtotal_val = _clean_val(fc.subtotal)
    exento_val = _clean_val(fc.monto_exento)
    base_val = _clean_val(fc.base_imponible)
    iva_val = _clean_val(fc.monto_iva)
    total_val = _clean_val(fc.total)
    
    tasa_usada = fc.tasa_cambio
    
    if moneda == "USD":
        # Determinar tasa de cambio
        rate = Decimal("0.00")
        if fc.tasa_cambio:
            try:
                rate = Decimal(str(fc.tasa_cambio).replace(",", "."))
            except Exception:
                pass
        if rate <= 0:
            # Fallback a tasa BCV del día
            try:
                rate = Decimal(str(get_current_bcv_rate()))
            except Exception:
                rate = Decimal("39.50")  # Fallback definitivo razonable
        
        # Guardar la tasa usada en formato legible
        tasa_usada = f"{rate:.4f}"
        
        # Multiplicar los valores
        subtotal_val = (subtotal_val * rate).quantize(Decimal("0.01"))
        exento_val = (exento_val * rate).quantize(Decimal("0.01"))
        base_val = (base_val * rate).quantize(Decimal("0.01"))
        iva_val = (iva_val * rate).quantize(Decimal("0.01"))
        total_val = (total_val * rate).quantize(Decimal("0.01"))
    
    # 3. Determinar alícuota sugerida de ISLR
    # Si base_val es 0, o si el monto exento es igual al subtotal o al total facturado, la alícuota sugerida es 0%
    if base_val == 0 or exento_val == total_val or (subtotal_val > 0 and exento_val == subtotal_val):
        alicuota_sugerida = Decimal("0.00")
    else:
        concepto_sugerido = fc.tipo_documento or "Prestación de Servicios en General"
        alicuota_sugerida = tributario_engine.obtener_alicuota_islr_sugerida(concepto_sugerido)
        
    return {
        "tipo_documento": fc.tipo_documento,
        "fecha_emision": fc.fecha_emision,
        "fecha_vencimiento": fc.fecha_vencimiento,
        "numero_documento": fc.numero_documento,
        "numero_control": fc.numero_control,
        "proveedor": fc.proveedor,
        "proveedor_rif": fc.proveedor_rif,
        "proveedor_telefono": fc.proveedor_telefono,
        "direccion_fiscal_proveedor": fc.direccion_fiscal_proveedor,
        "receptor": fc.receptor,
        "receptor_rif": fc.receptor_rif,
        "subtotal": f"{subtotal_val:.2f}",
        "monto_exento": f"{exento_val:.2f}",
        "base_imponible": f"{base_val:.2f}",
        "monto_iva": f"{iva_val:.2f}",
        "total": f"{total_val:.2f}",
        "contribuyente_tipo": fc.contribuyente_tipo,
        "tasa_cambio": tasa_usada,
        "rif_valido": rif_valido,
        "islr_rate": str(alicuota_sugerida),
        "islr_concept": "Servicios en General (Jurídicos: 2%)" if alicuota_sugerida == Decimal("0.02") else "Honorarios Profesionales / Fletes (Jurídicos: 3%)" if alicuota_sugerida == Decimal("0.03") else "Publicidad, Propaganda y Comisiones (5%)" if alicuota_sugerida == Decimal("0.05") else "Compra de Mercancía / No sujeto"
    }

def _process_parsed_ocr_sale(fc: object, ctx) -> dict:
    from decimal import Decimal
    from . import tributario_engine
    
    rif_valido = tributario_engine.validar_rif_venezolano(fc.receptor_rif)
    moneda = (getattr(fc, "moneda_original", "VES") or "VES").strip().upper()
    
    def _clean_val(v):
        if v is None:
            return Decimal("0.00")
        if isinstance(v, (int, float)):
            try:
                return Decimal(str(v))
            except Exception:
                return Decimal("0.00")
        s = str(v).strip()
        if not s:
            return Decimal("0.00")
        import re
        s = re.sub(r'(?i)^(bsf\.?|bs\.?|ves\.?|usd\.?|\$)\s*', '', s)
        s = s.replace(" ", "")
        if re.search(r"\d+\.\d{3},\d{2}$", s) or (
            "," in s and "." in s and s.rfind(",") > s.rfind(".")
        ):
            s = s.replace(".", "").replace(",", ".")
        elif "," in s and "." not in s:
            s = s.replace(",", ".")
        s = re.sub(r"[^\d.\-]", "", s)
        try:
            return Decimal(s)
        except Exception:
            return Decimal("0.00")
            
    subtotal_val = _clean_val(fc.subtotal)
    exento_val = _clean_val(fc.monto_exento)
    base_val = _clean_val(fc.base_imponible)
    iva_val = _clean_val(fc.monto_iva)
    total_val = _clean_val(fc.total)
    
    tasa_usada = fc.tasa_cambio
    
    if moneda == "USD":
        rate = Decimal("0.00")
        if fc.tasa_cambio:
            try:
                rate = Decimal(str(fc.tasa_cambio).replace(",", "."))
            except Exception:
                pass
        if rate <= 0:
            try:
                rate = Decimal(str(get_current_bcv_rate()))
            except Exception:
                rate = Decimal("39.50")
        
        tasa_usada = f"{rate:.4f}"
        subtotal_val = (subtotal_val * rate).quantize(Decimal("0.01"))
        exento_val = (exento_val * rate).quantize(Decimal("0.01"))
        base_val = (base_val * rate).quantize(Decimal("0.01"))
        iva_val = (iva_val * rate).quantize(Decimal("0.01"))
        total_val = (total_val * rate).quantize(Decimal("0.01"))
    
    tipo_doc = "Factura"
    tipo_lower = str(fc.tipo_documento).lower()
    if "credito" in tipo_lower or "crédito" in tipo_lower:
        tipo_doc = "Nota de Credito"
    elif "debito" in tipo_lower or "débito" in tipo_lower:
        tipo_doc = "Nota de Debito"
        
    return {
        "clasificacion": tipo_doc,
        "fecha_emision": fc.fecha_emision,
        "numero_documento": fc.numero_documento,
        "numero_control": fc.numero_control,
        "proveedor": fc.proveedor or ctx.company_name,
        "proveedor_rif": fc.proveedor_rif or ctx.company_rif,
        "receptor": fc.receptor,
        "receptor_rif": fc.receptor_rif,
        "subtotal": f"{subtotal_val:.2f}",
        "monto_exento": f"{exento_val:.2f}",
        "base_imponible": f"{base_val:.2f}",
        "monto_iva": f"{iva_val:.2f}",
        "total": f"{total_val:.2f}",
        "tasa_cambio": tasa_usada,
        "rif_valido": rif_valido,
    }

def _process_parsed_ocr_reporte_z(z_data: dict) -> dict:
    from decimal import Decimal
    
    def _clean_val(v):
        if v is None:
            return Decimal("0.00")
        if isinstance(v, (int, float)):
            try:
                return Decimal(str(v))
            except Exception:
                return Decimal("0.00")
        s = str(v).strip()
        if not s:
            return Decimal("0.00")
        import re
        s = re.sub(r'(?i)^(bsf\.?|bs\.?|ves\.?|usd\.?|\$)\s*', '', s)
        s = s.replace(" ", "")
        if re.search(r"\d+\.\d{3},\d{2}$", s) or (
            "," in s and "." in s and s.rfind(",") > s.rfind(".")
        ):
            s = s.replace(".", "").replace(",", ".")
        elif "," in s and "." not in s:
            s = s.replace(",", ".")
        s = re.sub(r"[^\d.\-]", "", s)
        try:
            return Decimal(s)
        except Exception:
            return Decimal("0.00")
            
    subtotal_val = _clean_val(z_data.get("sub_total"))
    exento_val = _clean_val(z_data.get("monto_exento"))
    base_val = _clean_val(z_data.get("base_imponible"))
    iva_val = _clean_val(z_data.get("iva"))
    total_val = _clean_val(z_data.get("total"))
    
    if base_val > 0 and iva_val == 0:
        iva_val = (base_val * Decimal("0.16")).quantize(Decimal("0.01"))
    if total_val == 0:
        total_val = subtotal_val + iva_val
    if subtotal_val == 0:
        subtotal_val = base_val + exento_val
        
    return {
        "numero_reporte": z_data.get("numero_reporte"),
        "fecha_emision": z_data.get("fecha_emision"),
        "sub_total": f"{subtotal_val:.2f}",
        "base_imponible": f"{base_val:.2f}",
        "monto_exento": f"{exento_val:.2f}",
        "iva": f"{iva_val:.2f}",
        "total": f"{total_val:.2f}",
    }

async def _send_ocr_reporte_z_card(update: Update, context: ContextTypes.DEFAULT_TYPE, msg_to_edit=None) -> None:
    pending = context.user_data.get("pending_ocr_reporte_z")
    if not pending:
        return
        
    text = (
        f"📊 *REPORTE Z EXTRAÍDO* 📊\n\n"
        f"🔢 *Reporte Nro:* `{pending['numero_reporte'] or '—'}`\n"
        f"📅 *Fecha Emisión:* {pending['fecha_emision'] or '—'}\n\n"
        f"-----------------------------------------\n"
        f"💵 *Subtotal:* {pending['sub_total']} Bs\n"
        f"ex *Monto Exento:* {pending['monto_exento']} Bs\n"
        f"💰 *Base Imponible:* {pending['base_imponible']} Bs\n"
        f"⚡ *IVA (16%):* {pending['iva']} Bs\n"
        f"💸 *Total Reportado:* {pending['total']} Bs\n\n"
        f"👇 *¿Deseas registrar este Reporte Z de ventas diarias?*"
    )
    
    kb = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("✅ Confirmar y Guardar", callback_data="ocr_confirm_reporte_z"),
            InlineKeyboardButton("❌ Cancelar", callback_data="ocr_cancel_photo")
        ]
    ])
    
    if msg_to_edit:
        await msg_to_edit.edit_text(text, reply_markup=kb, parse_mode="Markdown")
    else:
        msg = update.effective_message
        if msg:
            await msg.reply_text(text, reply_markup=kb, parse_mode="Markdown")

async def _send_ocr_sale_card(update: Update, context: ContextTypes.DEFAULT_TYPE, msg_to_edit=None) -> None:
    pending = context.user_data.get("pending_ocr_sale")
    if not pending:
        return
        
    rif_status = "✅ Válido" if pending["rif_valido"] else "❌ INVÁLIDO (Módulo 11)"
    tipo_doc_desc = "FACTURA EMITIDA (VENTA)" if pending["clasificacion"] == "Factura" else "NOTA DE CRÉDITO EMITIDA" if pending["clasificacion"] == "Nota de Credito" else "NOTA DE DÉBITO EMITIDA"
    
    text = (
        f"📈 *{tipo_doc_desc} EXTRAÍDA* 📈\n\n"
        f"👤 *Cliente:* {pending['receptor'] or '—'}\n"
        f"🆔 *RIF Cliente:* `{pending['receptor_rif'] or '—'}` ({rif_status})\n"
        f"📅 *Fecha Emisión:* {pending['fecha_emision'] or '—'}\n"
        f"🔢 *Documento Nro:* `{pending['numero_documento'] or '—'}`\n"
        f"🎛️ *Nro Control:* `{pending['numero_control'] or '—'}`\n\n"
        f"-----------------------------------------\n"
        f"💵 *Subtotal:* {pending['subtotal'] or '0.00'} Bs\n"
        f"ex *Monto Exento:* {pending['monto_exento'] or '0.00'} Bs\n"
        f"💰 *Base Imponible:* {pending['base_imponible'] or '0.00'} Bs\n"
        f"⚡ *IVA (16%):* {pending['monto_iva'] or '0.00'} Bs\n"
        f"💸 *Total General:* {pending['total'] or '0.00'} Bs\n"
    )
    if pending['tasa_cambio']:
        text += f"💱 *Tasa Cambio:* {pending['tasa_cambio']} Bs/$\n"
        
    text += (
        f"-----------------------------------------\n"
        f"👇 *Confirma para guardar este registro de ventas:*"
    )
    
    kb = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("✅ Confirmar y Guardar", callback_data="ocr_confirm_sale"),
            InlineKeyboardButton("❌ Cancelar", callback_data="ocr_cancel_photo")
        ]
    ])
    
    if msg_to_edit:
        await msg_to_edit.edit_text(text, reply_markup=kb, parse_mode="Markdown")
    else:
        msg = update.effective_message
        if msg:
            await msg.reply_text(text, reply_markup=kb, parse_mode="Markdown")


async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    msg = update.effective_message
    if not msg:
        return
    if not _allowed(update):
        await _deny(update)
        return
    
    user = user_manager.get_user(update.effective_user.id)
    if user and user.get("role") == "tributos_only":
        await msg.reply_text("❌ Tu nivel de autorización (\"Tributos Only\") solo te permite consultar y generar reportes, no subir o procesar imágenes en el bot.")
        return

    if not msg.photo:
        return

    # Interceptar RIF para admin (escanear RIF)
    if context.user_data.get("awaiting_admin_rif_photo"):
        target_uid = context.user_data.get("admin_edit_target_uid")
        context.user_data.pop("awaiting_admin_rif_photo", None)
        context.user_data.pop("admin_edit_target_uid", None)
        
        status_msg = await msg.reply_text("⏳ *Procesando imagen del RIF con Gemini...*", parse_mode="Markdown")
        try:
            photo = msg.photo[-1]
            tg_file = await context.bot.get_file(photo.file_id)
            img_data = await tg_file.download_as_bytearray()
            from PIL import Image
            import io
            image = Image.open(io.BytesIO(img_data))
            
            from . import ocr_extract
            rif_data = ocr_extract.extract_rif_data_from_image(image)
            
            razon_social = rif_data.get("razon_social", "").strip()
            rif_val = rif_data.get("rif", "").strip().upper()
            
            if not rif_val or not razon_social:
                await status_msg.edit_text("❌ No se pudo extraer la Razón Social o el RIF de la imagen. Por favor, ingrésalos manualmente.")
                await _show_admin_user_detail(update, context, target_uid)
                return
                
            # Validate RIF
            if not tributario_engine.validar_rif_venezolano(rif_val):
                await status_msg.edit_text(f"❌ RIF extraído (`{rif_val}`) no es válido. Edítalo manualmente.")
                await _show_admin_user_detail(update, context, target_uid)
                return
                
            user_manager.update_user_field(target_uid, "company_rif", rif_val)
            user_manager.update_user_field(target_uid, "company_name", razon_social)
            
            await status_msg.edit_text(
                f"✅ *RIF Escaneado y Configurado!*\n\n"
                f"• *Razón Social:* `{razon_social}`\n"
                f"• *RIF:* `{rif_val}`",
                parse_mode="Markdown"
            )
            await _show_admin_user_detail(update, context, target_uid)
        except Exception as e:
            logger.error(f"Error al procesar RIF desde foto: {e}")
            await status_msg.edit_text(f"❌ *Error al escanear RIF:* `{e}`", parse_mode="Markdown")
            await _show_admin_user_detail(update, context, target_uid)
        return

    # Interceptar firma y sello para admin (cargar firma y sello)
    if context.user_data.get("awaiting_admin_company_signature"):
        target_uid = context.user_data.get("admin_edit_target_uid")
        context.user_data.pop("awaiting_admin_company_signature", None)
        context.user_data.pop("admin_edit_target_uid", None)
        
        ctx = CompanyContext(target_uid)
        if not ctx.is_custom:
            await msg.reply_text("❌ Solo las empresas personalizadas pueden subir firma y sello.")
            return
            
        status_msg = await msg.reply_text("📥 *Guardando imagen de firma y sello del cliente...*", parse_mode="Markdown")
        try:
            photo = msg.photo[-1]
            tg_file = await context.bot.get_file(photo.file_id)
            target_path = ctx.dir_path / "firma_sello_transparente.png"
            target_path.parent.mkdir(parents=True, exist_ok=True)
            await tg_file.download_to_drive(custom_path=str(target_path))
            await status_msg.edit_text(f"✅ *Firma y Sello del cliente actualizados con éxito!*", parse_mode="Markdown")
            await _show_admin_user_detail(update, context, target_uid)
        except Exception as e:
            logger.error(f"Error al guardar firma y sello de cliente desde foto: {e}")
            await status_msg.edit_text(f"❌ *Error al guardar la firma y sello de cliente:* `{e}`", parse_mode="Markdown")
            await _show_admin_user_detail(update, context, target_uid)
        return

    # Interceptar firma y sello
    if context.user_data.get("awaiting_company_signature"):
        context.user_data.pop("awaiting_company_signature", None)
        ctx = _get_company_context(update)
        if not ctx.is_custom:
            await msg.reply_text("❌ Solo las empresas personalizadas pueden subir firma y sello.")
            return
        
        status_msg = await msg.reply_text("📥 *Guardando imagen de firma y sello...*", parse_mode="Markdown")
        try:
            photo = msg.photo[-1]
            tg_file = await context.bot.get_file(photo.file_id)
            target_path = ctx.dir_path / "firma_sello_transparente.png"
            target_path.parent.mkdir(parents=True, exist_ok=True)
            await tg_file.download_to_drive(custom_path=str(target_path))
            await status_msg.edit_text(f"✅ *Firma y Sello actualizados con éxito!*\nGuardado en `{target_path.name}`", parse_mode="Markdown")
            await _show_company_config_menu(update, context)
        except Exception as e:
            logger.error(f"Error al guardar firma y sello desde foto: {e}")
            await status_msg.edit_text(f"❌ *Error al guardar la firma y sello:* `{e}`", parse_mode="Markdown")
        return

    pending_doc = context.user_data.get("pending_doc")
    if not pending_doc:
        active_menu = context.user_data.get("active_menu")
        if active_menu in ("cotizacion", "nota"):
            pending_doc = {
                "type": active_menu,
                "awaiting": "text_data"
            }
            context.user_data["pending_doc"] = pending_doc

    if pending_doc and pending_doc.get("awaiting") in ("search_barcode", "search_ocr"):
        photo = msg.photo[-1]
        mode = pending_doc.get("awaiting")
        
        status_msg = await msg.reply_text(
            "📥 *Escaneando código de barras con IA (Gemini)...*" if mode == "search_barcode" else "📥 *Analizando imagen del producto con IA (Gemini OCR)...*", 
            parse_mode="Markdown"
        )
        
        suffix = ".jpg"
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            tmp_path = Path(tmp.name)
            
        try:
            tg_file = await context.bot.get_file(photo.file_id)
            await tg_file.download_to_drive(tmp_path)
            
            from PIL import Image
            from . import ocr_extract
            import os
            
            img = Image.open(tmp_path)
            
            if mode == "search_barcode":
                query_val = ocr_extract.extract_barcode_from_image(img)
            else:
                query_val = ocr_extract.extract_product_query_from_image(img)
                
            if query_val != "NONE":
                ctx = _get_company_context(update)
                products = []
                if mode == "search_barcode":
                    products = excel_store.search_products_in_excel(ctx.productos_path, query_val, search_by="code")
                else:
                    # For OCR, try searching by description first, then code
                    products = excel_store.search_products_in_excel(ctx.productos_path, query_val, search_by="desc")
                    if not products:
                        products = excel_store.search_products_in_excel(ctx.productos_path, query_val, search_by="code")
                
                if products:
                    if len(products) == 1:
                        product = products[0]
                        pending_doc["selected_product"] = product
                        pending_doc["awaiting"] = "input_qty"
                        
                        try:
                            await msg.delete()
                        except Exception:
                            pass
                        try:
                            await status_msg.delete()
                        except Exception:
                            pass
                            
                        prompt_id = pending_doc.pop("prompt_message_id", None)
                        if prompt_id:
                            try:
                                await context.bot.delete_message(chat_id=msg.chat_id, message_id=prompt_id)
                            except Exception:
                                pass
                                
                        import html
                        desc_esc = html.escape(product['description'])
                        code_esc = html.escape(product['code'])
                        kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Volver al Constructor", callback_data="coti_build_main")]])
                        prompt = await msg.reply_text(
                            f"✅ <b>Producto Encontrado:</b>\n\n"
                            f"Producto: <b>{desc_esc}</b> (<code>{code_esc}</code>)\n"
                            f"Precio Unitario: <b>${product['price']:.2f}</b>\n\n"
                            f"Por favor, escribe la <b>cantidad</b> a cotizar / usar para este producto:",
                            reply_markup=kb,
                            parse_mode="HTML"
                        )
                        pending_doc["prompt_message_id"] = prompt.message_id
                    else:
                        pending_doc["awaiting"] = "select_product"
                        try:
                            await msg.delete()
                        except Exception:
                            pass
                        try:
                            await status_msg.delete()
                        except Exception:
                            pass
                            
                        prompt_id = pending_doc.pop("prompt_message_id", None)
                        if prompt_id:
                            try:
                                await context.bot.delete_message(chat_id=msg.chat_id, message_id=prompt_id)
                            except Exception:
                                pass
                                
                        pending_doc["temp_search_results"] = products[:8]
                        kb_list = []
                        for idx, p in enumerate(products[:8]):
                            btn_text = f"[{p['code']}] {p['description'][:25]} (${p['price']:.2f})"
                            kb_list.append([InlineKeyboardButton(btn_text, callback_data=f"coti_build_select_p:{idx}")])
                        kb_list.append([
                            InlineKeyboardButton("📸 Escanear de Nuevo", callback_data="coti_build_search_barcode" if mode == "search_barcode" else "coti_build_search_ocr"),
                            InlineKeyboardButton("🔙 Volver al Constructor", callback_data="coti_build_main")
                        ])
                        kb = InlineKeyboardMarkup(kb_list)
                        prompt = await msg.reply_text(
                            f"🔍 *Múltiples coincidencias encontradas para \"{html.escape(query_val)}\" ({len(products)}):*\n"
                            f"Por favor, selecciona el producto exacto de abajo:",
                            reply_markup=kb,
                            parse_mode="Markdown"
                        )
                        pending_doc["prompt_message_id"] = prompt.message_id
                else:
                    try:
                        await msg.delete()
                    except Exception:
                        pass
                    try:
                        await status_msg.delete()
                    except Exception:
                        pass
                        
                    prompt_id = pending_doc.pop("prompt_message_id", None)
                    if prompt_id:
                        try:
                            await context.bot.delete_message(chat_id=msg.chat_id, message_id=prompt_id)
                        except Exception:
                            pass
                            
                    kb = InlineKeyboardMarkup([
                        [
                            InlineKeyboardButton("📸 Intentar de Nuevo", callback_data="coti_build_search_barcode" if mode == "search_barcode" else "coti_build_search_ocr"),
                            InlineKeyboardButton("🔙 Volver al Constructor", callback_data="coti_build_main")
                        ]
                    ])
                    prompt = await msg.reply_text(
                        f"🔍 *Búsqueda:* \"{html.escape(query_val)}\"\n\n"
                        f"❌ El producto con este código o texto no se encuentra registrado en el inventario Excel.\n\n"
                        f"¿Deseas intentar con otra imagen?",
                        reply_markup=kb,
                        parse_mode="Markdown"
                    )
                    pending_doc["prompt_message_id"] = prompt.message_id
            else:
                try:
                    await msg.delete()
                except Exception:
                    pass
                try:
                    await status_msg.delete()
                except Exception:
                    pass
                    
                prompt_id = pending_doc.pop("prompt_message_id", None)
                if prompt_id:
                    try:
                        await context.bot.delete_message(chat_id=msg.chat_id, message_id=prompt_id)
                    except Exception:
                        pass
                        
                kb = InlineKeyboardMarkup([
                    [
                        InlineKeyboardButton("📸 Intentar de Nuevo", callback_data="coti_build_search_barcode" if mode == "search_barcode" else "coti_build_search_ocr"),
                        InlineKeyboardButton("🔙 Volver al Constructor", callback_data="coti_build_main")
                    ]
                ])
                prompt = await msg.reply_text(
                    "❌ No se pudo identificar el producto o código en la imagen provista.\n\n"
                    "Por favor asegúrate de que la foto esté nítida y bien iluminada. ¿Deseas intentar de nuevo?",
                    reply_markup=kb,
                    parse_mode="Markdown"
                )
                pending_doc["prompt_message_id"] = prompt.message_id
        except Exception as e:
            logger.error(f"Error procesando imagen para busqueda de producto: {e}")
            await msg.reply_text(f"⚠️ Ocurrió un error al procesar el producto por imagen: {e}")
        finally:
            try:
                if tmp_path.exists():
                    os.unlink(tmp_path)
            except Exception:
                pass
        return

    elif pending_doc:
        photo = msg.photo[-1]
        doc_type = pending_doc["type"]
        status_msg = await msg.reply_text("📥 *Analizando documento con IA (Gemini OCR)...*", parse_mode="Markdown")
        
        suffix = ".jpg"
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            tmp_path = Path(tmp.name)
            
        try:
            tg_file = await context.bot.get_file(photo.file_id)
            await tg_file.download_to_drive(tmp_path)
            
            from PIL import Image
            from . import ocr_extract
            import os
            
            img = Image.open(tmp_path)
            extracted = ocr_extract.extract_document_data_from_image(img)
            
            # Mapear cliente y productos
            client_info = {
                "name": extracted.get("client_name") or "",
                "rif": extracted.get("client_rif") or "",
                "address": extracted.get("client_address") or "",
                "phone": extracted.get("client_phone") or "",
                "salesman": "FREDDY LOPEZ",
                "saleType": "Contado",
                "note": ""
            }
            if client_info["rif"]:
                client_info["rif"] = _normalize_rif(client_info["rif"])

            items = []
            for it in extracted.get("items", []):
                code = (it.get("code") or "").strip()
                desc = (it.get("desc") or "").strip()
                try:
                    qty = float(it.get("qty") or 1.0)
                except Exception:
                    qty = 1.0
                try:
                    price = float(it.get("priceUsd") or 0.0)
                except Exception:
                    price = 0.0
                items.append({
                    "code": code,
                    "desc": desc,
                    "qty": qty,
                    "priceUsd": price,
                    "totalUsd": qty * price
                })

            if items:
                pending_doc["parsed_data"] = {
                    "docType": doc_type,
                    "currency": "usd",
                    "exchangeRate": get_current_bcv_rate(),
                    "docNumber": "",
                    "docDate": date.today().strftime("%Y-%m-%d"),
                    "client": client_info,
                    "items": items
                }
                
                try:
                    await msg.delete()
                except Exception:
                    pass
                try:
                    await status_msg.delete()
                except Exception:
                    pass
                    
                start_prompt_id = pending_doc.pop("start_prompt_message_id", None)
                if start_prompt_id:
                    try:
                        await context.bot.delete_message(chat_id=msg.chat_id, message_id=start_prompt_id)
                    except Exception:
                        pass
                
                pending_doc["awaiting"] = "edit_card"
                await _send_client_data_card(update, context, first_time=True)
            else:
                await status_msg.edit_text(
                    "⚠️ *No pude identificar productos en el documento.*\n\n"
                    "Por favor asegúrate de que la foto esté nítida y bien iluminada. ¿Deseas intentar de nuevo?",
                    parse_mode="Markdown"
                )
        except Exception as e:
            logger.error(f"Error procesando documento OCR en constructor: {e}")
            await status_msg.edit_text(f"⚠️ Ocurrió un error al procesar el documento con OCR: {e}")
        finally:
            try:
                if tmp_path.exists():
                    os.unlink(tmp_path)
            except Exception:
                pass
        return

    photo = msg.photo[-1]
    status_msg = await msg.reply_text("📥 *Procesando imagen con IA (Gemini)...*", parse_mode="Markdown")
    
    suffix = ".jpg"
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp_path = Path(tmp.name)
        
    try:
        tg_file = await context.bot.get_file(photo.file_id)
        await tg_file.download_to_drive(tmp_path)
        
        from PIL import Image
        from . import ocr_extract
        
        img = Image.open(tmp_path)
        category = ocr_extract.classify_image_type(img)
        logger.info("Imagen clasificada como: %s", category)
        
        tributos_mode = context.user_data.get("tributos_mode")
        ctx = _get_company_context(update)
        
        def _normalize_rif_simple(r: str) -> str:
            if not r:
                return ""
            import re
            return re.sub(r"[^A-Z0-9]", "", str(r).upper().strip())
            
        company_rif_clean = _normalize_rif_simple(ctx.company_rif)
        
        if category in ("factura", "nota_credito"):
            fc = ocr_extract.extract_invoice_from_image(img)
            
            # Determinar de forma inteligente si es venta o compra
            is_sale = False
            prov_rif_clean = _normalize_rif_simple(fc.proveedor_rif)
            rec_rif_clean = _normalize_rif_simple(fc.receptor_rif)
            
            if prov_rif_clean and prov_rif_clean == company_rif_clean:
                is_sale = True
            elif rec_rif_clean and rec_rif_clean == company_rif_clean:
                is_sale = False
            else:
                if tributos_mode == "venta":
                    is_sale = True
                else:
                    is_sale = False
                    
            if is_sale:
                context.user_data["pending_ocr_sale"] = _process_parsed_ocr_sale(fc, ctx)
                await _send_ocr_sale_card(update, context, status_msg)
            else:
                context.user_data["pending_ocr_invoice"] = _process_parsed_ocr_invoice(fc)
                await _send_ocr_invoice_card(update, context, status_msg)
                
        elif category == "reporte_z" or tributos_mode == "reporte_z":
            z_data = ocr_extract.extract_reporte_z_from_image(img)
            context.user_data["pending_ocr_reporte_z"] = _process_parsed_ocr_reporte_z(z_data)
            await _send_ocr_reporte_z_card(update, context, status_msg)
            
        elif category == "documento_comercial":
            saved_path = Path(tempfile.gettempdir()) / f"doc_{photo.file_id}.jpg"
            import shutil
            shutil.copy(tmp_path, saved_path)
            context.user_data["pending_unknown_image"] = str(saved_path)
            
            kb = InlineKeyboardMarkup([
                [
                    InlineKeyboardButton("📋 Elaborar Cotización", callback_data="ocr_force_cotizacion"),
                    InlineKeyboardButton("📦 Elaborar Nota de Entrega", callback_data="ocr_force_nota"),
                ],
                [
                    InlineKeyboardButton("❌ Cancelar", callback_data="ocr_cancel_photo"),
                ]
            ])
            await status_msg.edit_text(
                "📋 *He detectado una Nota de Entrega, Cotización o Tabla de Excel.*\n\n"
                "¿Qué tipo de documento deseas elaborar con esta imagen?",
                reply_markup=kb,
                parse_mode="Markdown"
            )
            
        elif category == "retencion_iva":
            ret_iva = ocr_extract.extract_from_image(img)
            context.user_data["pending_ocr_ret_iva"] = {
                "fecha_emision": ret_iva.fecha_emision,
                "numero_comprobante": ret_iva.numero_comprobante,
                "rif": ret_iva.rif,
                "fechas_facturas": ret_iva.fechas_facturas,
                "numeros_facturas": ret_iva.numeros_facturas,
                "controles_facturas": ret_iva.controles_facturas,
                "total_compra_con_iva": ret_iva.total_compra_iva,
                "base_imponible": ret_iva.base_imponible,
                "iva_retenido": ret_iva.iva_retenido,
                "raw_text": ret_iva.raw_text
            }
            await _send_ocr_ret_iva_card(update, context, status_msg)
            
        elif category == "retencion_islr":
            ret_islr = ocr_extract.extract_islr_from_image(img)
            context.user_data["pending_ocr_ret_islr"] = ret_islr
            await _send_ocr_ret_islr_card(update, context, status_msg)
            
        else:
            saved_path = Path(tempfile.gettempdir()) / f"unknown_{photo.file_id}.jpg"
            import shutil
            shutil.copy(tmp_path, saved_path)
            context.user_data["pending_unknown_image"] = str(saved_path)
            
            kb = InlineKeyboardMarkup([
                [
                    InlineKeyboardButton("📄 Factura de Compra", callback_data="ocr_force_factura"),
                    InlineKeyboardButton("📈 Factura/Nota de Venta", callback_data="ocr_force_sale"),
                ],
                [
                    InlineKeyboardButton("📊 Reporte Z", callback_data="ocr_force_reporte_z"),
                    InlineKeyboardButton("📥 Retención IVA", callback_data="ocr_force_ret_iva"),
                ],
                [
                    InlineKeyboardButton("💸 Retención ISLR", callback_data="ocr_force_ret_islr"),
                    InlineKeyboardButton("❌ Cancelar", callback_data="ocr_cancel_photo"),
                ]
            ])
            await status_msg.edit_text(
                "❓ *No pude determinar el tipo de documento automáticamente.*\n\n"
                "Por favor, selecciona qué tipo de documento es para procesarlo:",
                reply_markup=kb,
                parse_mode="Markdown"
            )
            
    except Exception as e:
        logger.exception("Error al procesar foto con OCR")
        await status_msg.edit_text(
            f"❌ *Error al procesar la imagen con OCR:*\n`{e!s}`",
            parse_mode="Markdown"
        )
    finally:
        tmp_path.unlink(missing_ok=True)


async def _send_ocr_invoice_card(update: Update, context: ContextTypes.DEFAULT_TYPE, msg_to_edit=None) -> None:
    pending = context.user_data.get("pending_ocr_invoice")
    if not pending:
        return
        
    rif_status = "✅ Válido" if pending["rif_valido"] else "❌ INVÁLIDO (Módulo 11)"
    contrib_status = pending["contribuyente_tipo"] or "No especificado"
    
    base_val = Decimal("0")
    try:
        base_val = Decimal(pending["base_imponible"])
    except Exception:
        pass
    rate_val = Decimal(pending["islr_rate"])
    islr_retenido_est = (base_val * rate_val).quantize(Decimal("0.01"))
    
    text = (
        f"📄 *FACTURA DE COMPRA EXTRAÍDA* 📄\n\n"
        f"🏢 *Proveedor:* {pending['proveedor'] or '—'}\n"
        f"🆔 *RIF Proveedor:* `{pending['proveedor_rif'] or '—'}` ({rif_status})\n"
        f"👤 *Contribuyente:* {contrib_status}\n"
        f"📅 *Fecha Emisión:* {pending['fecha_emision'] or '—'}\n"
        f"🔢 *Factura Nro:* `{pending['numero_documento'] or '—'}`\n"
        f"🎛️ *Nro Control:* `{pending['numero_control'] or '—'}`\n\n"
        f"-----------------------------------------\n"
        f"💵 *Subtotal:* {pending['subtotal'] or '0.00'} Bs\n"
        f"ex *Monto Exento:* {pending['monto_exento'] or '0.00'} Bs\n"
        f"💰 *Base Imponible:* {pending['base_imponible'] or '0.00'} Bs\n"
        f"⚡ *IVA (16%):* {pending['monto_iva'] or '0.00'} Bs\n"
        f"💸 *Total Factura:* {pending['total'] or '0.00'} Bs\n"
    )
    if pending['tasa_cambio']:
        text += f"💱 *Tasa Cambio:* {pending['tasa_cambio']} Bs/$\n"
        
    text += (
        f"-----------------------------------------\n"
        f"✍️ *Retención de ISLR sugerida:*\n"
        f" 🔸 *Concepto:* {pending['islr_concept']}\n"
        f" 🔸 *Alícuota:* {(rate_val * 100):.1f}%\n"
        f" 👉 *ISLR Retenido Estimado:* `{excel_store._format_monto_ves(islr_retenido_est)}` Bs\n\n"
        f"👇 *Confirma para guardar o ajusta la alícuota de ISLR:*"
    )
    
    kb = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("✅ Confirmar y Guardar", callback_data="ocr_confirm_invoice"),
        ],
        [
            InlineKeyboardButton("ISLR: 0%", callback_data="ocr_islr_set_0.00"),
            InlineKeyboardButton("ISLR: 2%", callback_data="ocr_islr_set_0.02"),
            InlineKeyboardButton("ISLR: 3%", callback_data="ocr_islr_set_0.03"),
            InlineKeyboardButton("ISLR: 5%", callback_data="ocr_islr_set_0.05"),
        ],
        [
            InlineKeyboardButton("❌ Cancelar", callback_data="ocr_cancel_photo"),
        ]
    ])
    
    if msg_to_edit:
        await msg_to_edit.edit_text(text, reply_markup=kb, parse_mode="Markdown")
    else:
        msg = update.effective_message
        if msg:
            await msg.reply_text(text, reply_markup=kb, parse_mode="Markdown")


async def _send_ocr_ret_iva_card(update: Update, context: ContextTypes.DEFAULT_TYPE, msg_to_edit=None) -> None:
    pending = context.user_data.get("pending_ocr_ret_iva")
    if not pending:
        return
        
    text = (
        f"📥 *COMPROBANTE RETENCIÓN IVA EXTRAÍDO* 📥\n\n"
        f"🔢 *Nro Comprobante:* `{pending['numero_comprobante'] or '—'}`\n"
        f"📅 *Fecha Emisión:* {pending['fecha_emision'] or '—'}\n"
        f"🆔 *RIF Proveedor:* `{pending['rif'] or '—'}`\n"
        f"🔢 *Facturas Afectadas:* `{pending['numeros_facturas'] or '—'}`\n"
        f"🎛️ *Nro Controles:* `{pending['controles_facturas'] or '—'}`\n\n"
        f"-----------------------------------------\n"
        f"💵 *Total Compra:* {pending['total_compra_con_iva'] or '0.00'} Bs\n"
        f"💰 *Base Imponible:* {pending['base_imponible'] or '0.00'} Bs\n"
        f"💸 *IVA Retenido:* `{pending['iva_retenido'] or '0.00'}` Bs\n\n"
        f"👇 *¿Deseas registrar este comprobante de retención recibido?*"
    )
    
    kb = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("✅ Confirmar y Guardar", callback_data="ocr_confirm_ret_iva"),
            InlineKeyboardButton("❌ Cancelar", callback_data="ocr_cancel_photo")
        ]
    ])
    
    if msg_to_edit:
        await msg_to_edit.edit_text(text, reply_markup=kb, parse_mode="Markdown")
    else:
        msg = update.effective_message
        if msg:
            await msg.reply_text(text, reply_markup=kb, parse_mode="Markdown")


async def _send_ocr_ret_islr_card(update: Update, context: ContextTypes.DEFAULT_TYPE, msg_to_edit=None) -> None:
    pending = context.user_data.get("pending_ocr_ret_islr")
    if not pending:
        return
        
    text = (
        f"💸 *COMPROBANTE RETENCIÓN ISLR EXTRAÍDO* 💸\n\n"
        f"🔢 *Nro Comprobante:* `{pending['numero_comprobante'] or '—'}`\n"
        f"📅 *Fecha Emisión:* {pending['fecha_emision'] or '—'}\n"
        f"🏢 *Proveedor/Retenido:* {pending['proveedor'] or '—'}\n"
        f"🆔 *RIF Proveedor:* `{pending['proveedor_rif'] or '—'}`\n"
        f"✍️ *Concepto Retención:* {pending['concepto_retencion'] or '—'}\n"
        f"🔢 *Factura Afectada:* `{pending['numero_documento'] or '—'}`\n"
        f"🎛️ *Nro Control:* `{pending['numero_control'] or '—'}`\n\n"
        f"-----------------------------------------\n"
        f"💵 *Total Factura:* {pending['total_factura'] or '0.00'} Bs\n"
        f"💰 *Base Imponible:* {pending['base_imponible'] or '0.00'} Bs\n"
        f"📊 *Alícuota:* {pending['porcentaje_retencion'] or '0.00'}%\n"
        f"💸 *ISLR Retenido:* `{pending['islr_retenido'] or '0.00'}` Bs\n\n"
        f"👇 *¿Deseas registrar este comprobante de retención de ISLR?*"
    )
    
    kb = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("✅ Confirmar y Guardar", callback_data="ocr_confirm_ret_islr"),
            InlineKeyboardButton("❌ Cancelar", callback_data="ocr_cancel_photo")
        ]
    ])
    
    if msg_to_edit:
        await msg_to_edit.edit_text(text, reply_markup=kb, parse_mode="Markdown")
    else:
        msg = update.effective_message
        if msg:
            await msg.reply_text(text, reply_markup=kb, parse_mode="Markdown")


async def handle_ocr_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    q = update.callback_query
    if not q:
        return
        
    if not _allowed(update):
        await q.answer("❌ Acceso no autorizado / Suscripción Expirada.", show_alert=True)
        return
    if not _check_permission(update, "tributos"):
        await q.answer("❌ No tienes privilegios para realizar operaciones de Tributos.", show_alert=True)
        return
        
    await q.answer()
    data = (q.data or "").strip()
    msg = q.message
    if not msg:
        return
        
    if data == "ocr_cancel_photo":
        context.user_data.pop("pending_ocr_invoice", None)
        context.user_data.pop("pending_ocr_sale", None)
        context.user_data.pop("pending_ocr_reporte_z", None)
        context.user_data.pop("pending_ocr_ret_iva", None)
        context.user_data.pop("pending_ocr_ret_islr", None)
        saved_path_str = context.user_data.pop("pending_unknown_image", None)
        if saved_path_str:
            try:
                Path(saved_path_str).unlink(missing_ok=True)
            except Exception:
                pass
        try:
            await q.delete_message()
        except Exception:
            pass
        await msg.reply_text("❌ Procesamiento de imagen cancelado.")
        
    elif data.startswith("ocr_islr_set_"):
        pending = context.user_data.get("pending_ocr_invoice")
        if not pending:
            await msg.reply_text("Información de factura perdida.")
            return
        rate_str = data.replace("ocr_islr_set_", "")
        pending["islr_rate"] = rate_str
        rate_dec = Decimal(rate_str)
        if rate_dec == Decimal("0.02"):
            pending["islr_concept"] = "Servicios en General (Jurídicos: 2%)"
        elif rate_dec == Decimal("0.03"):
            pending["islr_concept"] = "Honorarios Profesionales / Fletes (Jurídicos: 3%)"
        elif rate_dec == Decimal("0.05"):
            pending["islr_concept"] = "Publicidad, Propaganda y Comisiones (5%)"
        else:
            pending["islr_concept"] = "Compra de Mercancía / No sujeto"
            
        await _send_ocr_invoice_card(update, context, msg_to_edit=msg)
        
    elif data.startswith("ocr_force_"):
        saved_path_str = context.user_data.get("pending_unknown_image")
        if not saved_path_str:
            await msg.reply_text("No se encontró la imagen para re-procesar.")
            return
            
        saved_path = Path(saved_path_str)
        if not saved_path.exists():
            await msg.reply_text("El archivo de imagen temporal ya no existe.")
            return
            
        force_type = data.replace("ocr_force_", "")
        await msg.edit_text(f"⏳ *Forzando procesamiento de imagen como {force_type.upper()}...*", parse_mode="Markdown")
        
        try:
            from PIL import Image
            from . import ocr_extract
            img = Image.open(saved_path)
            
            if force_type == "factura":
                fc = ocr_extract.extract_invoice_from_image(img)
                context.user_data["pending_ocr_invoice"] = _process_parsed_ocr_invoice(fc)
                context.user_data.pop("pending_unknown_image", None)
                saved_path.unlink(missing_ok=True)
                await _send_ocr_invoice_card(update, context, msg_to_edit=msg)
                
            elif force_type == "sale":
                fc = ocr_extract.extract_invoice_from_image(img)
                context.user_data["pending_ocr_sale"] = _process_parsed_ocr_sale(fc, ctx)
                context.user_data.pop("pending_unknown_image", None)
                saved_path.unlink(missing_ok=True)
                await _send_ocr_sale_card(update, context, msg_to_edit=msg)
                
            elif force_type == "reporte_z":
                z_data = ocr_extract.extract_reporte_z_from_image(img)
                context.user_data["pending_ocr_reporte_z"] = _process_parsed_ocr_reporte_z(z_data)
                context.user_data.pop("pending_unknown_image", None)
                saved_path.unlink(missing_ok=True)
                await _send_ocr_reporte_z_card(update, context, msg_to_edit=msg)
                
            elif force_type == "ret_iva":
                ret_iva = ocr_extract.extract_from_image(img)
                context.user_data["pending_ocr_ret_iva"] = {
                    "fecha_emision": ret_iva.fecha_emision,
                    "numero_comprobante": ret_iva.numero_comprobante,
                    "rif": ret_iva.rif,
                    "fechas_facturas": ret_iva.fechas_facturas,
                    "numeros_facturas": ret_iva.numeros_facturas,
                    "controles_facturas": ret_iva.controles_facturas,
                    "total_compra_con_iva": ret_iva.total_compra_iva,
                    "base_imponible": ret_iva.base_imponible,
                    "iva_retenido": ret_iva.iva_retenido,
                    "raw_text": ret_iva.raw_text
                }
                context.user_data.pop("pending_unknown_image", None)
                saved_path.unlink(missing_ok=True)
                await _send_ocr_ret_iva_card(update, context, msg_to_edit=msg)
                
            elif force_type == "ret_islr":
                ret_islr = ocr_extract.extract_islr_from_image(img)
                context.user_data["pending_ocr_ret_islr"] = ret_islr
                context.user_data.pop("pending_unknown_image", None)
                saved_path.unlink(missing_ok=True)
                await _send_ocr_ret_islr_card(update, context, msg_to_edit=msg)
                
            elif force_type in ("cotizacion", "nota"):
                extracted = ocr_extract.extract_document_data_from_image(img)
                
                # Keep client info empty as requested by user
                client_info = {
                    "name": "",
                    "rif": "",
                    "address": "",
                    "phone": "",
                    "salesman": "FREDDY LOPEZ",
                    "saleType": "Contado",
                    "note": ""
                }
                
                items = []
                for it in extracted.get("items", []):
                    code = (it.get("code") or "").strip()
                    desc = (it.get("desc") or "").strip()
                    try:
                        qty = float(it.get("qty") or 1.0)
                    except Exception:
                        qty = 1.0
                    try:
                        price = float(it.get("priceUsd") or 0.0)
                    except Exception:
                        price = 0.0
                    items.append({
                        "code": code,
                        "desc": desc,
                        "qty": qty,
                        "priceUsd": price,
                        "totalUsd": qty * price
                    })
                
                if items:
                    context.user_data["pending_doc"] = {
                        "type": force_type,
                        "awaiting": "edit_card",
                        "parsed_data": {
                            "docType": force_type,
                            "currency": "usd",
                            "exchangeRate": get_current_bcv_rate(),
                            "docNumber": "",
                            "docDate": date.today().strftime("%Y-%m-%d"),
                            "client": client_info,
                            "items": items
                        }
                    }
                    context.user_data["active_menu"] = force_type
                    context.user_data.pop("pending_unknown_image", None)
                    saved_path.unlink(missing_ok=True)
                    try:
                        await msg.delete()
                    except Exception:
                        pass
                    await _send_client_data_card(update, context, first_time=True)
                else:
                    await msg.edit_text(
                        "⚠️ *No pude extraer productos válidos de la imagen.*\n\n"
                        "Por favor asegúrate de que la foto de la cotización o tabla de Excel sea nítida.",
                        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancelar", callback_data="ocr_cancel_photo")]])
                    )
                    
        except Exception as e:
            logger.exception("Error al forzar clasificación de foto")
            await msg.reply_text(f"❌ Error al procesar imagen forzada: {e!s}")
            
    elif data == "ocr_confirm_sale":
        pending = context.user_data.get("pending_ocr_sale")
        if not pending:
            await msg.reply_text("No hay venta pendiente de confirmación.")
            return
            
        try:
            ctx = _get_company_context(update)
            inserted = excel_store.append_venta_record(
                ctx.facturas_emitidas_path,
                clasificacion=pending["clasificacion"],
                estado="REGISTRADO",
                fecha=pending["fecha_emision"],
                numero_documento=pending["numero_documento"],
                razon_social=pending["receptor"],
                rif=pending["receptor_rif"],
                base_imponible=pending["base_imponible"],
                iva=pending["monto_iva"],
                total=pending["total"],
                texto_origen="Registrado desde OCR de foto.",
            )
            
            if not inserted:
                await msg.reply_text(f"⚠️ La {pending['clasificacion']} Nro {pending['numero_documento']} ya se encuentra registrada.")
                context.user_data.pop("pending_ocr_sale", None)
                try:
                    await q.delete_message()
                except Exception:
                    pass
                return
                
            await msg.reply_text(f"✅ {pending['clasificacion']} Nro {pending['numero_documento']} registrada con éxito en {ctx.facturas_emitidas_path.name}.")
            context.user_data.pop("pending_ocr_sale", None)
            try:
                await q.delete_message()
            except Exception:
                pass
        except Exception as e:
            logger.exception("Error al confirmar venta desde OCR")
            await msg.reply_text(f"❌ Error al registrar venta: {e!s}")

    elif data == "ocr_confirm_reporte_z":
        pending = context.user_data.get("pending_ocr_reporte_z")
        if not pending:
            await msg.reply_text("No hay reporte Z pendiente de confirmación.")
            return
            
        try:
            ctx = _get_company_context(update)
            inserted = excel_store.append_reporte_z_nuevo(
                ctx.reportes_z_path,
                numero_reporte=pending["numero_reporte"],
                fecha_emision=pending["fecha_emision"],
                sub_total=pending["sub_total"],
                base_imponible=pending["base_imponible"],
                monto_exento=pending["monto_exento"],
                iva=pending["iva"],
                total=pending["total"],
                texto_origen="Registrado desde OCR de foto.",
            )
            
            if not inserted:
                await msg.reply_text(f"⚠️ El Reporte Z Nro {pending['numero_reporte']} ya se encuentra registrado.")
                context.user_data.pop("pending_ocr_reporte_z", None)
                try:
                    await q.delete_message()
                except Exception:
                    pass
                return
                
            await msg.reply_text(f"✅ Reporte Z Nro {pending['numero_reporte']} registrado con éxito en {ctx.reportes_z_path.name}.")
            context.user_data.pop("pending_ocr_reporte_z", None)
            try:
                await q.delete_message()
            except Exception:
                pass
        except Exception as e:
            logger.exception("Error al confirmar reporte Z desde OCR")
            await msg.reply_text(f"❌ Error al registrar reporte Z: {e!s}")

    elif data == "ocr_confirm_invoice":
        pending = context.user_data.get("pending_ocr_invoice")
        if not pending:
            await msg.reply_text("No hay factura pendiente de confirmación.")
            return
            
        try:
            ctx = _get_company_context(update)
            inserted = excel_store.append_factura_compra(
                ctx.facturas_recibidas_path,
                tipo_documento=pending["tipo_documento"],
                fecha_emision=pending["fecha_emision"],
                fecha_vencimiento=pending["fecha_vencimiento"],
                numero_documento=pending["numero_documento"],
                numero_control=pending["numero_control"],
                proveedor=pending["proveedor"],
                proveedor_rif=pending["proveedor_rif"],
                proveedor_telefono=pending["proveedor_telefono"],
                direccion_fiscal_proveedor=pending["direccion_fiscal_proveedor"],
                receptor=pending["receptor"],
                receptor_rif=pending["receptor_rif"],
                subtotal=pending["subtotal"],
                monto_exento=pending["monto_exento"],
                base_imponible=pending["base_imponible"],
                monto_iva=pending["monto_iva"],
                total=pending["total"],
                texto_resumen="Registrado desde OCR de foto.",
            )
            
            if not inserted:
                await msg.reply_text(f"⚠️ La factura Nro {pending['numero_documento']} del proveedor {pending['proveedor']} ya se encuentra registrada.")
                context.user_data.pop("pending_ocr_invoice", None)
                try:
                    await q.delete_message()
                except Exception:
                    pass
                return
                
            await msg.reply_text(f"✅ Factura Nro {pending['numero_documento']} registrada con éxito en {ctx.facturas_recibidas_path.name}.")
            
            rate_val = Decimal(pending["islr_rate"])
            if rate_val > 0:
                base_val = Decimal(pending["base_imponible"] or "0")
                total_val = Decimal(pending["total"] or "0")
                islr_retenido_est = (base_val * rate_val).quantize(Decimal("0.01"))
                
                emission_date = tributario_engine._parse_row_date(pending["fecha_emision"]) or date.today()
                num_comp = excel_store.next_retencion_islr_number(ctx.retenciones_islr_dir, emission_date=emission_date)
                periodo_fiscal = _periodo_fiscal(emission_date)
                
                monthly_path = excel_store.monthly_retencion_islr_path(ctx.retenciones_islr_dir, emission_date)
                excel_store.append_retencion_islr(
                    monthly_path,
                    numero_comprobante=num_comp,
                    fecha_emision=pending["fecha_emision"],
                    periodo_fiscal=periodo_fiscal,
                    proveedor=pending["proveedor"],
                    proveedor_rif=pending["proveedor_rif"],
                    concepto_retencion=pending["islr_concept"],
                    numero_documento=pending["numero_documento"],
                    numero_control=pending["numero_control"],
                    base_imponible=base_val,
                    porcentaje_retencion=rate_val,
                    islr_retenido=islr_retenido_est,
                    total_factura=total_val,
                )
                
                ctx.generados_dir.mkdir(parents=True, exist_ok=True)
                pdf_filename = f"COMPROBANTE_RETENCION_ISLR_{num_comp}.pdf"
                pdf_path = ctx.generados_dir / pdf_filename
                
                try:
                    excel_store.export_comprobante_islr_pdf(
                        out_path=pdf_path,
                        numero_comprobante=num_comp,
                        fecha_emision=pending["fecha_emision"],
                        periodo_fiscal=periodo_fiscal,
                        proveedor=pending["proveedor"],
                        proveedor_rif=pending["proveedor_rif"],
                        concepto_retencion=pending["islr_concept"],
                        base_imponible=base_val,
                        porcentaje_retencion=rate_val,
                        islr_retenido=islr_retenido_est,
                        total_factura=total_val,
                        numero_documento=pending["numero_documento"],
                        numero_control=pending["numero_control"],
                    )
                    
                    # Registrar el documento generado en el historial
                    _register_document_in_history(
                        update=update,
                        doc_type="retencion_islr",
                        doc_number=num_comp,
                        client_name=pending["proveedor"],
                        client_rif=pending["proveedor_rif"],
                        total_amount=f"Bs. {islr_retenido_est:,.2f}",
                        temp_pdf_path=pdf_path
                    )
                    
                    pdf_filename_disp = f"COMPROBANTE_RETENCION_ISLR_{num_comp}.pdf"
                    await msg.reply_document(
                        document=str(pdf_path),
                        filename=pdf_filename_disp,
                        caption=f"📄 *Comprobante Oficial de Retención de ISLR Nro {num_comp}* generado con éxito para el proveedor {pending['proveedor']}.",
                        parse_mode="Markdown"
                    )
                finally:
                    pass
                    
        except Exception as e:
            logger.exception("Error al confirmar factura desde OCR")
            await msg.reply_text(f"❌ Error al guardar la factura/retención: {e!s}")
        finally:
            context.user_data.pop("pending_ocr_invoice", None)
            try:
                await q.delete_message()
            except Exception:
                pass
                
    elif data == "ocr_confirm_ret_iva":
        pending = context.user_data.get("pending_ocr_ret_iva")
        if not pending:
            await msg.reply_text("No hay retención de IVA pendiente.")
            return
            
        try:
            ctx = _get_company_context(update)
            inserted = excel_store.append_record(
                ctx.excel_path,
                fecha_emision=pending["fecha_emision"],
                numero_comprobante=pending["numero_comprobante"],
                rif=pending["rif"],
                fechas_facturas=pending["fechas_facturas"],
                numeros_facturas=pending["numeros_facturas"],
                controles_facturas=pending["controles_facturas"],
                total_compra_con_iva=pending["total_compra_con_iva"],
                base_imponible=pending["base_imponible"],
                iva_retenido=pending["iva_retenido"],
                ocr_snippet=f"Comprobante IVA registrado por OCR: {pending['raw_text'][:200]}"
            )
            
            if not inserted:
                await msg.reply_text(f"⚠️ La retención de IVA Nro {pending['numero_comprobante']} ya se encuentra registrada.")
            else:
                await msg.reply_text(f"✅ Retención de IVA Nro {pending['numero_comprobante']} guardada con éxito en {ctx.excel_path.name}.")
                
        except Exception as e:
            logger.exception("Error al guardar retención IVA desde OCR")
            await msg.reply_text(f"❌ Error al registrar retención IVA: {e!s}")
        finally:
            context.user_data.pop("pending_ocr_ret_iva", None)
            try:
                await q.delete_message()
            except Exception:
                pass
                
    elif data == "ocr_confirm_ret_islr":
        pending = context.user_data.get("pending_ocr_ret_islr")
        if not pending:
            await msg.reply_text("No hay retención de ISLR pendiente.")
            return
            
        try:
            ctx = _get_company_context(update)
            emission_date = tributario_engine._parse_row_date(pending["fecha_emision"]) or date.today()
            monthly_path = excel_store.monthly_retencion_islr_path(ctx.retenciones_islr_dir, emission_date)
            num_comp = pending["numero_comprobante"]
            
            # Validar duplicados de comprobantes de ISLR
            dup_info = excel_store.check_retencion_islr_exists(ctx.retenciones_islr_dir, num_comp)
            if dup_info:
                context.user_data["pending_replace_islr"] = {
                    "type": "ocr",
                    "data": pending,
                    "monthly_path": str(monthly_path),
                }
                kb = InlineKeyboardMarkup([
                    [
                        InlineKeyboardButton("✅ Sí, sustituir", callback_data=f"rep_ret_conf:islr:{num_comp}"),
                        InlineKeyboardButton("❌ Cancelar", callback_data="rep_ret_cancel")
                    ]
                ])
                await msg.reply_text(
                    f"⚠️ La retención de ISLR Nro {num_comp} ya existe.\n"
                    f"¿Deseas sustituir el comprobante de ISLR existente?",
                    reply_markup=kb
                )
                try:
                    await q.delete_message()
                except Exception:
                    pass
                return
            
            await _save_retencion_islr_ocr(update, context, pending, ctx, monthly_path)
            
        finally:
            context.user_data.pop("pending_ocr_ret_islr", None)
            try:
                await q.delete_message()
            except Exception:
                pass


async def handle_voice(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not _allowed(update):
        await _deny(update)
        return
        
    user = user_manager.get_user(update.effective_user.id)
    if user and user.get("role") == "tributos_only":
        await update.message.reply_text("❌ Tu nivel de autorización (\"Tributos Only\") solo te permite consultar y generar reportes, no usar comandos de voz en el bot.")
        return

    context.user_data.setdefault("voice_mode", False)
    if not update.message or not update.message.voice:
        return
    voice = update.message.voice
    tg_file = await context.bot.get_file(voice.file_id)
    suffix = ".ogg"
    if tg_file.file_path:
        sfx = Path(tg_file.file_path).suffix
        if sfx:
            suffix = sfx
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        out_path = Path(tmp.name)
    try:
        await tg_file.download_to_drive(out_path)
        text = await asyncio.to_thread(transcribe_audio_file, out_path)
    except Exception as e:  # noqa: BLE001
        logger.exception("Transcripción fallida")
        await update.message.reply_text(
            "No pude transcribir el audio. Si no usas OpenAI, instala ffmpeg y configura "
            "OPENAI_API_KEY para Whisper, o revisa el micrófono/formato.\n"
            f"Detalle: {e!s}"
        )
        return
    finally:
        out_path.unlink(missing_ok=True)
    if update.message:
        await update.message.reply_text(f"Transcripción: «{text}»")
    await _process_intent(update, context, text)
    if context.user_data.get("voice_mode"):
        context.user_data["voice_mode"] = False
        if update.message:
            await update.message.reply_text(
                "Comando de voz procesado. Si quieres otro, pulsa de nuevo «Activar comando de voz».",
                reply_markup=_main_keyboard(update.effective_user.id),
            )


def find_generated_document(generados_dir: Path, doc_type: str, user_input: str, ctx) -> Path | None:
    import re
    cleaned = re.sub(r"[^A-Za-z0-9]", "", user_input).strip()
    if not cleaned:
        return None
    
    # 1. Intentar buscar en historico_documentos.json
    history = []
    if ctx.historico_json_path.exists():
        try:
            import json
            with open(ctx.historico_json_path, "r", encoding="utf-8") as f:
                history = json.load(f)
        except Exception:
            pass
            
    for entry in history:
        if entry.get("doc_type") == doc_type:
            entry_num = re.sub(r"[^A-Za-z0-9]", "", str(entry.get("doc_number", ""))).strip()
            if entry_num and (entry_num == cleaned or entry_num.lstrip("0") == cleaned.lstrip("0")):
                pdf_filename = entry.get("pdf_filename")
                if pdf_filename:
                    path = generados_dir / pdf_filename
                    if path.exists():
                        return path

    # 2. Fallback: Buscar usando patrones de coincidencia glob
    patterns = []
    if doc_type == "cotizacion":
        patterns = [
            f"COTIZACIÓN_{cleaned.zfill(6)}.pdf",
            f"COTIZACIÓN_*{cleaned}*.pdf",
            f"COTIZACION_*{cleaned}*.pdf",
            f"*{cleaned}*.pdf"
        ]
    elif doc_type == "nota":
        patterns = [
            f"NOTA DE ENTREGA_{cleaned.zfill(6)}.pdf",
            f"NOTA DE ENTREGA_*{cleaned}*.pdf",
            f"*{cleaned}*.pdf"
        ]
    elif doc_type == "retencion_iva":
        patterns = [
            f"COMPROBANTE-RET-{cleaned}.pdf",
            f"COMPROBANTE-RET-{cleaned}.xlsx",
            f"COMPROBANTE-RET-*{cleaned}*.pdf",
            f"COMPROBANTE-RET-*{cleaned}*.xlsx"
        ]
    elif doc_type == "retencion_islr":
        patterns = [
            f"COMPROBANTE_RETENCION_ISLR_{cleaned}.pdf",
            f"COMPROBANTE_RETENCION_ISLR_*{cleaned}*.pdf",
            f"*{cleaned}*.pdf"
        ]

    for pattern in patterns:
        for filepath in generados_dir.glob(pattern):
            if filepath.is_file():
                return filepath
                
    return None


async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    msg = update.effective_message
    if not msg:
        return
    text = (msg.text or msg.caption or "").strip()
    if not text:
        return
    is_channel = update.channel_post is not None or update.edited_channel_post is not None

    pending_doc = context.user_data.get("pending_doc")
    has_active_state = (
        pending_doc is not None or
        context.user_data.get("awaiting_company_name") is not None or
        context.user_data.get("awaiting_company_rif") is not None or
        context.user_data.get("awaiting_company_email") is not None or
        context.user_data.get("awaiting_company_phone") is not None or
        context.user_data.get("awaiting_company_address") is not None or
        context.user_data.get("awaiting_company_signature") is not None or
        context.user_data.get("awaiting_company_next_cotizacion") is not None or
        context.user_data.get("awaiting_company_next_nota") is not None or
        context.user_data.get("awaiting_emit_docs") is not None or
        context.user_data.get("admin_state") is not None or
        context.user_data.get("share_doc") is not None or
        context.user_data.get("awaiting_reprint_num") is not None or
        context.user_data.get("awaiting_delete_ret_num") is not None
    )

    # Procesar siempre publicaciones de canal y además chat SUFEVICA detectado, si no hay un flujo activo.
    if (is_channel or _is_sufevica_chat(update)) and not has_active_state:
        logger.info("Procesando texto de canal/chat detectado (chat_id=%s).", msg.chat_id)
        await _process_intent(update, context, text)
        return

    # Mensajes privados: aplicar restricción por usuario.
    if not _allowed(update):
        await _deny(update)
        return

    context.user_data.setdefault("voice_mode", False)
    
    ctx = _get_company_context(update)
    if ctx.is_custom and ctx.company_type == "Ordinario":
        normalized_text = text.lower()
        if "generar retencion" in normalized_text or "generar retención" in normalized_text or "emitir retencion" in normalized_text or "emitir retención" in normalized_text or text == SUBMENU_GENERAR_RETENCION:
            await msg.reply_text("⚠️ Opción Bloqueada: Los Contribuyentes Ordinarios no emiten comprobantes de retención de acuerdo con las normativas del SENIAT.")
            return

    # Interceptar entradas de texto del usuario para reimpresión manual
    if context.user_data.get("awaiting_reprint_num"):
        doc_type = context.user_data.pop("awaiting_reprint_num")
        user_input = text.strip()
        
        pdf_path = find_generated_document(ctx.generados_dir, doc_type, user_input, ctx)
        if pdf_path and pdf_path.exists():
            doc_names = {
                "cotizacion": "Cotizacion",
                "nota": "Nota_de_Entrega",
                "retencion_iva": "Retencion_IVA",
                "retencion_islr": "Retencion_ISLR"
            }
            doc_name = doc_names.get(doc_type, "Documento")
            
            await msg.reply_document(
                document=str(pdf_path),
                filename=f"{doc_name}_{user_input}.pdf" if doc_type != "retencion_iva" or pdf_path.suffix == ".pdf" else f"{doc_name}_{user_input}.xlsx",
                caption=f"📄 *Reimpresión (Manual):* {doc_name} Nro {user_input}",
                parse_mode="Markdown"
            )
        else:
            doc_names_es = {
                "cotizacion": "Cotización",
                "nota": "Nota de Entrega",
                "retencion_iva": "Retención de IVA",
                "retencion_islr": "Retención de ISLR"
            }
            doc_name_es = doc_names_es.get(doc_type, "Documento")
            await msg.reply_text(
                f"❌ No encontré el archivo físico del documento de tipo *{doc_name_es}* con número `{user_input}` en el servidor.",
                parse_mode="Markdown"
            )
        return

    # Interceptar entradas de texto del usuario para eliminación de retención
    if context.user_data.get("awaiting_delete_ret_num"):
        ret_type = context.user_data.pop("awaiting_delete_ret_num")
        user_input = text.strip()
        
        # Validar existencia
        if ret_type == "iva":
            dup_info = excel_store.check_retencion_emitida_exists(ctx.retenciones_emitidas_dir, user_input)
            label = "IVA"
        else:
            dup_info = excel_store.check_retencion_islr_exists(ctx.retenciones_islr_dir, user_input)
            label = "ISLR"
            
        if dup_info:
            kb = InlineKeyboardMarkup([
                [
                    InlineKeyboardButton("🔥 Sí, eliminar", callback_data=f"delete_ret_confirm:{ret_type}:{user_input}"),
                    InlineKeyboardButton("❌ Cancelar", callback_data="delete_ret_cancel")
                ]
            ])
            await msg.reply_text(
                f"⚠️ *ADVERTENCIA DE ELIMINACIÓN* ⚠️\n\n"
                f"¿Estás seguro de que deseas eliminar permanentemente el comprobante de *{label}* Nro `{user_input}`?\n"
                f"Esta acción es irreversible y restablecerá los documentos involucrados a estatus pendiente.",
                reply_markup=kb,
                parse_mode="Markdown"
            )
        else:
            await msg.reply_text(
                f"❌ No encontré ningún comprobante de retención de *{label}* con número `{user_input}` en los registros.",
                parse_mode="Markdown"
            )
        return

    # Interceptar entradas de texto del usuario para configuración de empresa (FlashTax)
    user_id = update.effective_user.id
    if context.user_data.get("awaiting_company_name"):
        context.user_data.pop("awaiting_company_name", None)
        new_name = text.strip()
        user_manager.update_user_field(user_id, "company_name", new_name)
        await msg.reply_text(f"✅ Razón Social actualizada a: *{new_name}*", parse_mode="Markdown")
        await _show_company_config_menu(update, context)
        return

    elif context.user_data.get("awaiting_company_rif"):
        new_rif = text.strip().upper()
        # Validar RIF
        if not tributario_engine.validar_rif_venezolano(new_rif):
            await msg.reply_text(
                "❌ *RIF Inválido*\n\n"
                "El RIF ingresado no es válido de acuerdo a las especificaciones del SENIAT (ej: `J-40194130-3`).\n"
                "Por favor, verifícalo e ingresa un RIF correcto:",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancelar", callback_data="cfg_company_back")]])
            )
            return
        context.user_data.pop("awaiting_company_rif", None)
        user_manager.update_user_field(user_id, "company_rif", new_rif)
        await msg.reply_text(f"✅ RIF de la empresa actualizado a: `{new_rif}`", parse_mode="Markdown")
        await _show_company_config_menu(update, context)
        return

    elif context.user_data.get("awaiting_company_email"):
        new_email = text.strip()
        email_regex = r"^[\w\.-]+@[\w\.-]+\.\w+$"
        if not re.match(email_regex, new_email):
            await msg.reply_text(
                "❌ *Correo Electrónico Inválido*\n\n"
                "Por favor, ingresa una dirección de correo válida para el contador o presiona cancelar:",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancelar", callback_data="cfg_company_back")]])
            )
            return
        context.user_data.pop("awaiting_company_email", None)
        user_manager.update_user_field(user_id, "company_email", new_email)
        await msg.reply_text(f"✅ Correo del contador actualizado a: `{new_email}`", parse_mode="Markdown")
        await _show_company_config_menu(update, context)
        return

    elif context.user_data.get("awaiting_company_phone"):
        new_phone = text.strip()
        context.user_data.pop("awaiting_company_phone", None)
        user_manager.update_user_field(user_id, "company_phone", new_phone)
        await msg.reply_text(f"✅ Teléfono de contacto actualizado a: `{new_phone}`", parse_mode="Markdown")
        await _show_company_config_menu(update, context)
        return

    elif context.user_data.get("awaiting_company_address"):
        new_address = text.strip()
        context.user_data.pop("awaiting_company_address", None)
        user_manager.update_user_field(user_id, "company_address", new_address)
        await msg.reply_text(f"✅ Dirección fiscal actualizada a:\n_{new_address}_", parse_mode="Markdown")
        await _show_company_config_menu(update, context)
        return

    elif context.user_data.get("awaiting_company_next_cotizacion"):
        val = text.strip()
        if not val.isdigit() or int(val) <= 0:
            await msg.reply_text(
                "❌ *Número Inválido*\n\n"
                "Debe ser un número entero mayor a cero (ej: `18`).\n"
                "Por favor, verifícalo e ingrésalo correctamente o envía /start para cancelar:",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancelar", callback_data="cfg_company_back")]])
            )
            return
        context.user_data.pop("awaiting_company_next_cotizacion", None)
        num_val = int(val)
        user_manager.update_user_field(user_id, "next_cotizacion", num_val)
        await msg.reply_text(f"✅ Próxima Cotización actualizada a: `{num_val:06d}`", parse_mode="Markdown")
        await _show_company_config_menu(update, context)
        return

    elif context.user_data.get("awaiting_company_next_nota"):
        val = text.strip()
        if not val.isdigit() or int(val) <= 0:
            await msg.reply_text(
                "❌ *Número Inválido*\n\n"
                "Debe ser un número entero mayor a cero (ej: `10`).\n"
                "Por favor, verifícalo e ingrésalo correctamente o envía /start para cancelar:",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancelar", callback_data="cfg_company_back")]])
            )
            return
        context.user_data.pop("awaiting_company_next_nota", None)
        num_val = int(val)
        user_manager.update_user_field(user_id, "next_nota", num_val)
        await msg.reply_text(f"✅ Próxima Nota de Entrega actualizada a: `{num_val:06d}`", parse_mode="Markdown")
        await _show_company_config_menu(update, context)
        return
    
    # Limpiar flujos/estados si es navegación de menús
    if text in {
        COTI_BUTTON, NOTA_BUTTON, TRIBUTOS_BUTTON, ADMIN_PANEL_BUTTON, 
        SUBMENU_VOLVER, REPORT_VOLVER_TRIBUTOS, VOICE_BUTTON, VOICE_CANCEL_BUTTON,
        SUBMENU_CARGAR_FACTURA, SUBMENU_RETENCION_RECIBIDA, SUBMENU_REPORTE_Z,
        SUBMENU_FACTURA_EMITIDA, SUBMENU_GENERAR_RETENCION, SUBMENU_GENERAR_REPORTES,
        REPORT_IVA_BUTTON, REPORT_RETENCIONES_BUTTON, REPORT_FACTURAS_BUTTON, REPORT_PENDIENTES_BUTTON,
        SUBMENU_ELIMINAR_RETENCION
    }:
        context.user_data.pop("pending_doc", None)
        context.user_data.pop("awaiting_emit_docs", None)
        context.user_data.pop("admin_state", None)
        context.user_data.pop("admin_new_user", None)
        context.user_data.pop("share_doc", None)
    
    # Interceptar entradas de texto del administrador para registro de usuarios
    admin_state = context.user_data.get("admin_state")
    if admin_state:
        if admin_state == "awaiting_new_user_id":
            input_val = text.strip()
            if not input_val.isdigit():
                await msg.reply_text("❌ El ID de Telegram debe ser un número entero. Escribe un ID válido o envía /start para cancelar:")
                return
            context.user_data["admin_new_user"]["id"] = input_val
            context.user_data["admin_state"] = "awaiting_new_user_name"
            await msg.reply_text(
                f"ID de Telegram guardado: `{input_val}`.\n\n"
                "Ahora escribe el *Nombre / Descripción* de este cliente (ej: `Carlos Gómez` o `Empresa ABC`):",
                parse_mode="Markdown"
            )
            return
            
        elif admin_state == "awaiting_new_user_name":
            input_val = text.strip()
            context.user_data["admin_new_user"]["name"] = input_val
            context.user_data.pop("admin_state", None)
            
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("🏢 Cliente (FlashTax)", callback_data="admin_new_role:nueva_empresa")],
                [InlineKeyboardButton("🏛️ Tributos Only", callback_data="admin_new_role:tributos_only")],
                [InlineKeyboardButton("📋 Cotizaciones Only", callback_data="admin_new_role:cotizaciones_only")],
                [InlineKeyboardButton("⭐ Acceso Total", callback_data="admin_new_role:full_access")],
                [InlineKeyboardButton("❌ Cancelar", callback_data="admin_main")]
            ])
            await msg.reply_text(
                f"Nombre guardado: *{input_val}*\n\n"
                "🛡️ Selecciona el *Rol* para el cliente:",
                reply_markup=kb,
                parse_mode="Markdown"
            )
            return

        elif admin_state == "awaiting_admin_field_edit":
            target_uid = context.user_data.get("admin_edit_target_uid")
            field = context.user_data.get("admin_edit_field")
            val = text.strip()
            
            if field == "company_rif":
                new_rif = val.upper()
                if not tributario_engine.validar_rif_venezolano(new_rif):
                    await msg.reply_text(
                        "❌ *RIF Inválido*\n\n"
                        "El RIF ingresado no es válido de acuerdo a las especificaciones del SENIAT (ej: `J-40194130-3`).\n"
                        "Por favor, verifícalo e ingresa un RIF correcto o envía /start para cancelar:",
                        parse_mode="Markdown",
                        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancelar", callback_data=f"admin_edit:{target_uid}")]])
                    )
                    return
                val = new_rif
                
            elif field == "last_correlative":
                if not (len(val) == 14 and val.isdigit()):
                    await msg.reply_text(
                        "❌ *Correlativo Inválido*\n\n"
                        "El correlativo debe ser un número de 14 dígitos (ej: `20260600000000`).\n"
                        "Por favor, verifícalo e ingrésalo correctamente o envía /start para cancelar:",
                        parse_mode="Markdown",
                        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancelar", callback_data=f"admin_edit:{target_uid}")]])
                    )
                    return

            elif field in ("next_cotizacion", "next_nota"):
                if not (val.isdigit() and int(val) > 0):
                    await msg.reply_text(
                        "❌ *Número Inválido*\n\n"
                        "Debe ser un número entero mayor a cero (ej: `18`).\n"
                        "Por favor, verifícalo e ingrésalo correctamente o envía /start para cancelar:",
                        parse_mode="Markdown",
                        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancelar", callback_data=f"admin_edit:{target_uid}")]])
                    )
                    return
                val = int(val)

            user_manager.update_user_field(target_uid, field, val)
            context.user_data.pop("admin_state", None)
            context.user_data.pop("admin_edit_target_uid", None)
            context.user_data.pop("admin_edit_field", None)
            
            await msg.reply_text(f"✅ Campo actualizado con éxito.", parse_mode="Markdown")
            await _show_admin_user_detail(update, context, target_uid)
            return

    # Interceptar entrada de texto del botón "Generar Retención"
    if context.user_data.get("awaiting_emit_docs"):
        context.user_data.pop("awaiting_emit_docs", None)
        emit_docs = _parse_emitir_retencion_request(text)
        if emit_docs is None:
            emit_docs = [x.strip() for x in re.split(r'[|,\s]+', text) if x.strip()]
            
        if not emit_docs:
            await msg.reply_text("❌ No se detectaron números de factura válidos. Operación cancelada.")
            return
            
        await _start_emitir_retencion_flow(update, context, emit_docs)
        return

    # Botones principales e interactivos
    if text == "🚀 Menú de Inicio":
        await _show_startup_menu(update, context)
        return

    elif text == ADMIN_PANEL_BUTTON:
        user = user_manager.get_user(update.effective_user.id)
        if user and user.get("role") == "admin":
            await _show_admin_panel(update, context)
        else:
            await msg.reply_text("❌ No tienes privilegios para acceder al Panel de Administración.")
        return

    elif text == "⚙️ Configurar Empresa":
        user = user_manager.get_user(update.effective_user.id)
        if user and user.get("role") == "nueva_empresa":
            await _show_company_config_menu(update, context)
        else:
            await msg.reply_text("❌ Esta opción solo está disponible para usuarios con privilegios de Empresa (FlashTax).")
        return

    elif text == TRIBUTOS_BUTTON:
        if not _check_permission(update, "tributos"):
            await msg.reply_text("❌ No tienes privilegios para acceder al módulo de Tributos.")
            return
        context.user_data["active_menu"] = "tributos"
        await msg.reply_text(
            "🏛️ *Módulo de Tributos y Control Fiscal*\n\n"
            "Selecciona una de las opciones de abajo:",
            reply_markup=_tributos_submenu_keyboard(update.effective_user.id),
            parse_mode="Markdown"
        )
        return

    elif text == COTI_BUTTON:
        if not _check_permission(update, "cotizaciones"):
            await msg.reply_text("❌ No tienes privilegios para acceder al módulo de Cotizaciones.")
            return
        context.user_data["active_menu"] = "cotizacion"
        await _start_document_flow(update, context, "cotizacion")
        return

    elif text == NOTA_BUTTON:
        if not _check_permission(update, "cotizaciones"):
            await msg.reply_text("❌ No tienes privilegios para acceder al módulo de Notas de Entrega.")
            return
        context.user_data["active_menu"] = "nota"
        await _start_document_flow(update, context, "nota")
        return

    elif text == HISTORIAL_BUTTON:
        context.user_data.pop("active_menu", None)
        await _send_history_menu(update, context)
        return

    elif text == VOICE_BUTTON:
        context.user_data["voice_mode"] = True
        await msg.reply_text(
            "Modo voz activado. Envía ahora tu nota de voz con el requerimiento.",
            reply_markup=_main_keyboard(update.effective_user.id),
        )
        return

    elif text == VOICE_CANCEL_BUTTON:
        context.user_data["voice_mode"] = False
        await msg.reply_text(
            "Modo voz desactivado.",
            reply_markup=_main_keyboard(update.effective_user.id),
        )
        return

    elif text == SUBMENU_VOLVER:
        context.user_data["active_menu"] = "main"
        await msg.reply_text(
            "🏛️ *Menú Principal*\n\n"
            "Selecciona una opción del menú para continuar:",
            reply_markup=_main_keyboard(update.effective_user.id),
            parse_mode="Markdown"
        )
        return

    elif text == REPORT_VOLVER_TRIBUTOS:
        if not _check_permission(update, "tributos"):
            await msg.reply_text("❌ No tienes privilegios para acceder al módulo de Tributos.")
            return
        await msg.reply_text(
            "🏛️ *Módulo de Tributos y Control Fiscal*\n\n"
            "Selecciona una de las opciones de abajo:",
            reply_markup=_tributos_submenu_keyboard(update.effective_user.id),
            parse_mode="Markdown"
        )
        return

    elif text in (SUBMENU_CARGAR_FACTURA, SUBMENU_RETENCION_RECIBIDA, SUBMENU_REPORTE_Z, SUBMENU_FACTURA_EMITIDA, SUBMENU_GENERAR_RETENCION, SUBMENU_ELIMINAR_RETENCION):
        if not _check_permission(update, "tributos"):
            await msg.reply_text("❌ No tienes privilegios para acceder al módulo de Tributos.")
            return
        if not _can_modify_tributos(update):
            await msg.reply_text("❌ Tu nivel de autorización (\"Tributos Only\") solo te permite consultar y generar reportes, no registrar nueva información en el bot.")
            return

        if text == SUBMENU_CARGAR_FACTURA:
            context.user_data["tributos_mode"] = "compra"
            await msg.reply_text(
                "📥 *Cargar Facturas Recibidas (Compras)*\n\n"
                "Puedes cargar facturas de las siguientes formas:\n"
                "1️⃣ Envía la imagen o foto de la factura física.\n"
                "2️⃣ Envía el archivo PDF o XML de la factura digital.\n"
                "3️⃣ Envía una nota de voz dictando los datos.\n"
                "4️⃣ Pega el texto copiado de la factura."
            )
        elif text == SUBMENU_RETENCION_RECIBIDA:
            context.user_data["tributos_mode"] = "retencion_recibida"
            await msg.reply_text(
                "🧾 *Cargar Retenciones Recibidas (Clientes)*\n\n"
                "Puedes registrar retenciones de IVA/ISLR de las siguientes formas:\n"
                "1️⃣ Envía la imagen/foto o PDF del comprobante de retención.\n"
                "2️⃣ Escribe los datos con el formato: `registrar retencion, fecha: DD/MM/AAAA, comprobante: XXXXXX, ...`"
            )
        elif text == SUBMENU_REPORTE_Z:
            context.user_data["tributos_mode"] = "reporte_z"
            await msg.reply_text(
                "📊 *Cargar Reporte Z de Ventas Diarias*\n\n"
                "Envía la imagen del reporte Z impreso de tu máquina fiscal, o escribe sus datos de ventas directamente en texto."
            )
        elif text == SUBMENU_FACTURA_EMITIDA:
            context.user_data["tributos_mode"] = "venta"
            await msg.reply_text(
                "📈 *Cargar Factura Emitida (Ventas)*\n\n"
                "Registra tus facturas de ventas emitidas:\n"
                "1️⃣ Envía la foto o el PDF de la factura emitida.\n"
                "2️⃣ Escribe los datos correspondientes en texto."
            )
        elif text == SUBMENU_GENERAR_RETENCION:
            await msg.reply_text(
                "✍️ *Generar Comprobante de Retención de IVA*\n\n"
                "Por favor, escribe el o los números de factura (separados por coma o barra vertical `|`) "
                "a las cuales les deseas emitir el comprobante (ej: `00007553` o `00007553|00007554`):"
            )
            context.user_data["awaiting_emit_docs"] = True
        elif text == SUBMENU_ELIMINAR_RETENCION:
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("✍️ Retención de IVA", callback_data="delete_ret_sel:iva")],
                [InlineKeyboardButton("🏛️ Retención de ISLR", callback_data="delete_ret_sel:islr")],
                [InlineKeyboardButton("❌ Cancelar", callback_data="delete_ret_cancel")]
            ])
            await msg.reply_text(
                "🗑️ *Eliminar Comprobante de Retención*\n\n"
                "Selecciona el tipo de retención que deseas eliminar:",
                reply_markup=kb,
                parse_mode="Markdown"
            )
        return

    elif text == SUBMENU_GENERAR_REPORTES:
        if not _check_permission(update, "tributos"):
            await msg.reply_text("❌ No tienes privilegios para acceder al módulo de Tributos.")
            return
        await msg.reply_text(
            "📋 *Generación de Reportes Tributarios y Fiscales*\n\n"
            "Selecciona qué reporte deseas generar:",
            reply_markup=_reportes_submenu_keyboard(),
            parse_mode="Markdown"
        )
        return

    elif text == REPORT_IVA_BUTTON:
        if not _check_permission(update, "tributos"):
            await msg.reply_text("❌ No tienes privilegios para acceder al módulo de Tributos.")
            return
        today = date.today()
        fortnight = 1 if today.day <= 15 else 2
        report = tributario_engine.get_compromiso_tributario_report(today.year, today.month, fortnight)
        text_report = format_tributos_report(report)
        kb = _tributos_keyboard(today.year, today.month, fortnight, _generate_short_summary(report))
        await msg.reply_text(text_report, reply_markup=kb, parse_mode="Markdown")
        return

    elif text == REPORT_RETENCIONES_BUTTON:
        if not _check_permission(update, "tributos"):
            await msg.reply_text("❌ No tienes privilegios para acceder al módulo de Tributos.")
            return
        await msg.reply_text(
            "🧾 *Reporte de Retenciones Recibidas*\n\n"
            "Para generar el reporte detallado, por favor escribe el rango de fechas en el siguiente formato:\n"
            "`retenciones recibidas del DD/MM/AAAA al DD/MM/AAAA en pdf` (o `en excel`)"
        )
        return

    elif text == REPORT_FACTURAS_BUTTON:
        if not _check_permission(update, "tributos"):
            await msg.reply_text("❌ No tienes privilegios para acceder al módulo de Tributos.")
            return
        await msg.reply_text(
            "📥 *Reporte de Facturas Cargadas (Compras)*\n\n"
            "Por favor, escribe el rango de fechas en el siguiente formato:\n"
            "`compras del DD/MM/AAAA al DD/MM/AAAA`"
        )
        return

    elif text == REPORT_PENDIENTES_BUTTON:
        if not _check_permission(update, "tributos"):
            await msg.reply_text("❌ No tienes privilegios para acceder al módulo de Tributos.")
            return
        status_msg = await msg.reply_text("⏳ *Buscando facturas cargadas sin comprobante de retención emitido...*", parse_mode="Markdown")
        await _generate_pending_withholdings_report(update, context, status_msg)
        return

    elif text.lower().startswith("compras del "):
        if not _check_permission(update, "tributos"):
            await msg.reply_text("❌ No tienes privilegios para acceder al módulo de Tributos.")
            return
        pattern = r"compras\s+del\s+(\d{1,2}/\d{1,2}/\d{4})\s+al\s+(\d{1,2}/\d{1,2}/\d{4})"
        match = re.match(pattern, text, re.IGNORECASE)
        if not match:
            await msg.reply_text("❌ Formato inválido. Usa: `compras del DD/MM/AAAA al DD/MM/AAAA`")
            return
        date_from_str, date_to_str = match.groups()
        date_from = _parse_user_date(date_from_str)
        date_to = _parse_user_date(date_to_str)
        if not date_from or not date_to:
            await msg.reply_text("❌ Fecha inválida. Usa el formato DD/MM/AAAA.")
            return
            
        status_msg = await msg.reply_text("⏳ *Generando reporte de compras cargadas en el rango de fechas...*", parse_mode="Markdown")
        await _generate_purchases_loaded_report(update, context, date_from, date_to, status_msg)
        return
    
    share_doc = context.user_data.get("share_doc")
    if share_doc and share_doc.get("awaiting") == "client_email":
        email_input = text.strip()
        if email_input.lower() in ("cancelar", "cancel"):
            context.user_data.pop("share_doc", None)
            await msg.reply_text("❌ Envío de correo cancelado.")
            return
            
        email_regex = r"^[\w\.-]+@[\w\.-]+\.\w+$"
        if not re.match(email_regex, email_input):
            await msg.reply_text(
                "❌ Dirección de correo electrónico inválida.\n"
                "Por favor, escribe una dirección de correo válida (ej. cliente@correo.com) o escribe `cancelar`:"
            )
            return
            
        share_doc["awaiting"] = None
        status_msg = await msg.reply_text(
            f"⏳ *Enviando {share_doc['title']} por correo SMTP a `{email_input}`...*",
            parse_mode="Markdown"
        )
        
        asyncio.create_task(_send_document_email_async(update, context, email_input, status_msg))
        return
        
    pending_doc = context.user_data.get("pending_doc")
    if pending_doc:
        state = pending_doc.get("awaiting")
        
        if state == "search_code":
            try:
                await msg.delete()
            except Exception:
                pass
            
            prompt_id = pending_doc.pop("prompt_message_id", None)
            if prompt_id:
                try:
                    await context.bot.delete_message(chat_id=msg.chat_id, message_id=prompt_id)
                except Exception:
                    pass
                    
            ctx = _get_company_context(update)
            products = excel_store.search_products_in_excel(ctx.productos_path, text, search_by="code")
            
            if products:
                if len(products) == 1:
                    product = products[0]
                    pending_doc["selected_product"] = product
                    pending_doc["awaiting"] = "input_qty"
                    import html
                    desc_esc = html.escape(product['description'])
                    code_esc = html.escape(product['code'])
                    kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Volver al Constructor", callback_data="coti_build_main")]])
                    prompt = await msg.reply_text(
                        f"🔢 <b>CANTIDAD DE PRODUCTO</b>\n\n"
                        f"Has seleccionado: <b>{desc_esc}</b> (<code>{code_esc}</code>)\n"
                        f"Precio Unitario: <b>${product['price']:.2f}</b>\n\n"
                        f"Por favor, escribe la <b>cantidad</b> a cotizar / usar para este producto:",
                        reply_markup=kb,
                        parse_mode="HTML"
                    )
                    pending_doc["prompt_message_id"] = prompt.message_id
                else:
                    pending_doc["awaiting"] = "select_product"
                    pending_doc["temp_search_results"] = products[:8]
                    kb_list = []
                    for idx, p in enumerate(products[:8]):
                        btn_text = f"[{p['code']}] {p['description'][:25]} (${p['price']:.2f})"
                        kb_list.append([InlineKeyboardButton(btn_text, callback_data=f"coti_build_select_p:{idx}")])
                    kb_list.append([
                        InlineKeyboardButton("🔍 Buscar de Nuevo", callback_data="coti_build_search_code"),
                        InlineKeyboardButton("🔙 Volver al Constructor", callback_data="coti_build_main")
                    ])
                    kb = InlineKeyboardMarkup(kb_list)
                    prompt = await msg.reply_text(
                        f"🔍 *Múltiples coincidencias encontradas ({len(products)}):*\n"
                        f"Por favor, selecciona el producto exacto de la lista de abajo:",
                        reply_markup=kb,
                        parse_mode="Markdown"
                    )
                    pending_doc["prompt_message_id"] = prompt.message_id
            else:
                kb = InlineKeyboardMarkup([
                    [
                        InlineKeyboardButton("🔍 Buscar de Nuevo", callback_data="coti_build_search_code"),
                        InlineKeyboardButton("🔙 Volver al Constructor", callback_data="coti_build_main")
                    ]
                ])
                prompt = await msg.reply_text(
                    f"❌ No se encontró ningún producto con el código `{text}`.\n\n"
                    f"¿Deseas intentar de nuevo?",
                    reply_markup=kb,
                    parse_mode="Markdown"
                )
                pending_doc["prompt_message_id"] = prompt.message_id
            return

        elif state == "search_desc":
            try:
                await msg.delete()
            except Exception:
                pass
            
            prompt_id = pending_doc.pop("prompt_message_id", None)
            if prompt_id:
                try:
                    await context.bot.delete_message(chat_id=msg.chat_id, message_id=prompt_id)
                except Exception:
                    pass
                    
            ctx = _get_company_context(update)
            products = excel_store.search_products_in_excel(ctx.productos_path, text, search_by="desc")
            
            if products:
                pending_doc["awaiting"] = "select_product"
                pending_doc["temp_search_results"] = products[:8]
                kb_list = []
                for idx, p in enumerate(products[:8]):
                    btn_text = f"[{p['code']}] {p['description'][:25]} (${p['price']:.2f})"
                    kb_list.append([InlineKeyboardButton(btn_text, callback_data=f"coti_build_select_p:{idx}")])
                    
                extra_text = ""
                if len(products) > 8:
                    extra_text = f"⚠️ Se encontraron {len(products)} resultados. Se muestran los primeros 8. Sé más específico en tu búsqueda si es necesario.\n\n"
                
                kb_list.append([
                    InlineKeyboardButton("🔎 Buscar de Nuevo", callback_data="coti_build_search_desc"),
                    InlineKeyboardButton("🔙 Volver al Constructor", callback_data="coti_build_main")
                ])
                kb = InlineKeyboardMarkup(kb_list)
                prompt = await msg.reply_text(
                    f"🔎 *Productos que coinciden con su búsqueda:*\n\n"
                    f"{extra_text}"
                    f"Por favor, selecciona el producto exacto de la lista de abajo:",
                    reply_markup=kb,
                    parse_mode="Markdown"
                )
                pending_doc["prompt_message_id"] = prompt.message_id
            else:
                kb = InlineKeyboardMarkup([
                    [
                        InlineKeyboardButton("🔎 Buscar de Nuevo", callback_data="coti_build_search_desc"),
                        InlineKeyboardButton("🔙 Volver al Constructor", callback_data="coti_build_main")
                    ]
                ])
                prompt = await msg.reply_text(
                    f"❌ No se encontró ningún producto que contenga `{text}` en su descripción.\n\n"
                    f"¿Deseas intentar de nuevo?",
                    reply_markup=kb,
                    parse_mode="Markdown"
                )
                pending_doc["prompt_message_id"] = prompt.message_id
            return
            
        elif state == "input_qty":
            try:
                qty_val = float(text.replace(",", "."))
                if qty_val <= 0:
                    raise ValueError()
            except ValueError:
                await msg.reply_text("⚠️ Por favor ingresa una cantidad válida mayor a cero (ej. 10 o 5.5).")
                return
                
            try:
                await msg.delete()
            except Exception:
                pass
                
            prompt_id = pending_doc.pop("prompt_message_id", None)
            if prompt_id:
                try:
                    await context.bot.delete_message(chat_id=msg.chat_id, message_id=prompt_id)
                except Exception:
                    pass
            
            product = pending_doc.pop("selected_product", None)
            doc_data = pending_doc["parsed_data"]
            if product:
                found = False
                p_code = (product["code"] or "").strip().upper()
                is_generic = p_code in ("", "S/C", "S.C.", "S/D", "S.D.", "SIN CODIGO", "SIN CÓDIGO")
                for it in doc_data["items"]:
                    if it.get("code") == product["code"] and (not is_generic or it.get("desc") == product["description"]):
                        it["qty"] = float(it["qty"]) + qty_val
                        it["totalUsd"] = it["qty"] * float(it["priceUsd"])
                        found = True
                        break
                if not found:
                    doc_data["items"].append({
                        "code": product["code"],
                        "desc": product["description"],
                        "qty": qty_val,
                        "priceUsd": product["price"],
                        "totalUsd": qty_val * product["price"]
                    })
            
            pending_doc["awaiting"] = "builder_main"
            await _send_interactive_builder_card(update, context, first_time=False)
            return

        elif state == "edit_item_qty":
            idx = pending_doc.get("edit_item_idx")
            doc_data = pending_doc["parsed_data"]
            items = doc_data["items"]
            
            try:
                qty_val = float(text.replace(",", "."))
                if qty_val <= 0:
                    raise ValueError()
            except ValueError:
                await msg.reply_text("⚠️ Por favor ingresa una cantidad válida mayor a cero (ej. 10 o 5.5).")
                return
                
            try:
                await msg.delete()
            except Exception:
                pass
                
            prompt_id = pending_doc.pop("prompt_message_id", None)
            if prompt_id:
                try:
                    await context.bot.delete_message(chat_id=msg.chat_id, message_id=prompt_id)
                except Exception:
                    pass
            
            if idx is not None and 0 <= idx < len(items):
                items[idx]["qty"] = qty_val
                items[idx]["totalUsd"] = qty_val * float(items[idx]["priceUsd"])
                
            pending_doc["awaiting"] = "builder_main"
            await _send_builder_items_editor(update, context)
            return

        elif state == "edit_item_price":
            idx = pending_doc.get("edit_item_idx")
            doc_data = pending_doc["parsed_data"]
            items = doc_data["items"]
            
            try:
                price_val = float(text.replace(",", "."))
                if price_val <= 0:
                    raise ValueError()
            except ValueError:
                await msg.reply_text("⚠️ Por favor ingresa un precio válido mayor a cero (ej. 15.50).")
                return
                
            try:
                await msg.delete()
            except Exception:
                pass
                
            prompt_id = pending_doc.pop("prompt_message_id", None)
            if prompt_id:
                try:
                    await context.bot.delete_message(chat_id=msg.chat_id, message_id=prompt_id)
                except Exception:
                    pass
            
            if idx is not None and 0 <= idx < len(items):
                items[idx]["priceUsd"] = price_val
                items[idx]["totalUsd"] = float(items[idx]["qty"]) * price_val
                
            pending_doc["awaiting"] = "builder_main"
            await _send_builder_items_editor(update, context)
            return

        if state == "text_data":
            doc_type = pending_doc["type"]
            doc_data = _parse_document_text_explicit(text, doc_type)

            # Fallback a Gemini si falla el parser local de expresiones regulares
            if doc_data is None and config.GEMINI_API_KEY:
                try:
                    from . import ocr_extract
                    extracted = ocr_extract.parse_document_text_with_gemini(text)
                    if extracted and extracted.get("items"):
                        client_info = {
                            "name": extracted.get("client_name") or "",
                            "rif": extracted.get("client_rif") or "",
                            "address": extracted.get("client_address") or "",
                            "phone": extracted.get("client_phone") or "",
                            "salesman": "FREDDY LOPEZ",
                            "saleType": "Contado",
                            "note": ""
                        }
                        if client_info["rif"]:
                            client_info["rif"] = _normalize_rif(client_info["rif"])

                        doc_data = {
                            "docType": doc_type,
                            "currency": "usd",
                            "exchangeRate": get_current_bcv_rate(),
                            "docNumber": "",
                            "docDate": date.today().strftime("%Y-%m-%d"),
                            "client": client_info,
                            "items": extracted.get("items")
                        }
                except Exception as e:
                    logger.error(f"Error en fallback Gemini para texto: {e}")

            if doc_data is not None:
                # Formatear items crudos sin cruzar con inventario
                for it in doc_data["items"]:
                    it["qty"] = float(it.get("qty") or 1.0)
                    it["priceUsd"] = float(it.get("priceUsd") or 0.0)
                    it["totalUsd"] = it["qty"] * it["priceUsd"]
                    it["code"] = (it.get("code") or "").strip()
                    it["desc"] = (it.get("desc") or "").strip()

                # Eliminar el mensaje de prompt inicial
                start_prompt_id = pending_doc.pop("start_prompt_message_id", None)
                if start_prompt_id:
                    try:
                        await context.bot.delete_message(chat_id=msg.chat_id, message_id=start_prompt_id)
                    except Exception:
                        pass

                # Eliminar el mensaje del usuario con el texto pegado
                try:
                    await msg.delete()
                except Exception:
                    pass

                pending_doc["parsed_data"] = doc_data
                pending_doc["awaiting"] = "edit_card"
                await _send_client_data_card(update, context, first_time=True)
                return
            else:
                await msg.reply_text(
                    "⚠️ No pude identificar productos en el texto pegado.\n\n"
                    "Por favor, asegúrate de incluir cantidades y precios (ej: `10 x PARP-220 Perno a 2.50`), o escríbelo en columnas limpias.\n\n"
                    "Vuelve a intentarlo pegando el texto aquí, o envía `/start` para cancelar.",
                    parse_mode="Markdown"
                )
                return
                
        elif state in ("input_name", "input_rif", "input_address", "input_phone", "input_salesman", "input_note", "input_rate"):
            doc_data = pending_doc["parsed_data"]
            client = doc_data["client"]
            
            val = text.strip()
            
            if state == "input_rate":
                try:
                    rate_val = float(val.replace(",", "."))
                    if rate_val <= 0:
                        raise ValueError()
                    doc_data["exchangeRate"] = rate_val
                except ValueError:
                    await msg.reply_text("⚠️ Por favor ingresa una tasa de cambio válida mayor a cero (ej. 40.50 o 40,50).")
                    return
            else:
                # Si el usuario responde omitir, vacio, ninguno, etc., dejarlo vacío (se omitirá el dato en el PDF/HTML)
                if val.lower() in ("omitir", "ninguno", "vacio", "vacío", "omitido", "cancelar"):
                    val = ""
                    
                if state == "input_name":
                    client["name"] = val
                elif state == "input_rif":
                    client["rif"] = _normalize_rif(val) if val else ""
                elif state == "input_address":
                    client["address"] = val
                elif state == "input_phone":
                    client["phone"] = val
                elif state == "input_salesman":
                    client["salesman"] = val
                elif state == "input_note":
                    client["note"] = val
                
            # Eliminar el mensaje de prompt anterior
            prompt_id = pending_doc.pop("prompt_message_id", None)
            if prompt_id:
                try:
                    await context.bot.delete_message(chat_id=msg.chat_id, message_id=prompt_id)
                except Exception:
                    pass
            
            # Eliminar el mensaje de texto enviado por el usuario para mantener limpio el chat
            try:
                await msg.delete()
            except Exception:
                pass
                
            pending_doc["awaiting"] = "edit_card"
            await _send_client_data_card(update, context, first_time=False)
            return

    pending_email = context.user_data.get("pending_email_report")
    if pending_email and pending_email.get("awaiting") == "email_address":
        email_input = text.strip()
        if email_input.lower() in ("cancelar", "cancel"):
            context.user_data.pop("pending_email_report", None)
            await msg.reply_text("❌ Envío de reportes por correo cancelado.")
            return
            
        if email_input.lower() == "contador":
            if config.DEFAULT_ACCOUNTANT_EMAIL:
                email_input = config.DEFAULT_ACCOUNTANT_EMAIL
            else:
                await msg.reply_text(
                    "⚠️ No has configurado `DEFAULT_ACCOUNTANT_EMAIL` en tu archivo `.env`.\n"
                    "Por favor, escribe una dirección de correo válida para continuar:"
                )
                return
                
        email_regex = r"^[\w\.-]+@[\w\.-]+\.\w+$"
        if not re.match(email_regex, email_input):
            await msg.reply_text(
                "❌ Dirección de correo electrónico inválida.\n"
                "Por favor, escribe una dirección de correo válida (ej. nombre@servidor.com):"
            )
            return
            
        y = pending_email["year"]
        m = pending_email["month"]
        f = pending_email["fortnight"]
        context.user_data.pop("pending_email_report", None)
        
        status_msg = await msg.reply_text(
            f"⏳ *Generando reportes y conectando al servidor SMTP...*\n"
            f"Despachando correo a: `{email_input}`",
            parse_mode="Markdown"
        )
        
        asyncio.create_task(_send_reportes_async_workflow(update, context, y, m, f, email_input, status_msg))
        return

    pending = context.user_data.get("pending_emit_ret")
    if pending and pending.get("awaiting") == "manual_number":
        manual_num = re.sub(r"\D", "", text)
        if not (len(manual_num) == 14 and manual_num.isdigit()):
            await msg.reply_text(
                "Número inválido. Debe tener 14 dígitos (YYYYMM########)."
            )
            return
        pending["manual_num"] = manual_num
        pending["awaiting"] = None
        kb = InlineKeyboardMarkup(
            [[
                InlineKeyboardButton("Fecha de hoy", callback_data="emit_date_today"),
                InlineKeyboardButton("Fecha manual", callback_data="emit_date_manual"),
            ]]
        )
        await msg.reply_text("Número manual guardado. Ahora selecciona fecha de emisión.", reply_markup=kb)
        return
    if pending and pending.get("awaiting") == "manual_date":
        d = _parse_user_date(text)
        if d is None:
            await msg.reply_text("Fecha inválida. Usa DD/MM/AAAA.")
            return
        pending["emission_date"] = d.strftime("%d/%m/%Y")
        pending["awaiting"] = None
        kb = InlineKeyboardMarkup(
            [[
                InlineKeyboardButton("PDF", callback_data="emit_fmt_pdf"),
                InlineKeyboardButton("Excel", callback_data="emit_fmt_excel"),
            ]]
        )
        await msg.reply_text("Fecha guardada. Selecciona formato del comprobante.", reply_markup=kb)
        return

    if context.user_data.get("voice_mode"):
        await msg.reply_text(
            "Modo voz activo: envía una nota de voz o pulsa «Cancelar voz».",
            reply_markup=_main_keyboard(update.effective_user.id),
        )
        return
    await _process_intent(update, context, text)


async def _send_reportes_async_workflow(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    year: int,
    month: int,
    fortnight: int,
    recipient_email: str,
    status_msg,
) -> None:
    from .tributario_engine import get_fortnight_range, get_compromiso_tributario_report
    from .email_sender import send_report_email
    import tempfile
    
    ctx = _get_company_context(update)
    start_date, end_date = get_fortnight_range(year, month, fortnight)
    
    # 1. Generar datos y reportes quincenales
    records = excel_store.retenciones_by_document_date(
        ctx.excel_path,
        date_from=start_date,
        date_to=end_date,
    )
    
    # Archivos temporales para adjuntar
    temp_files: list[Path] = []
    attachments: list[Path] = []
    
    try:

        month_names = {
            1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
            7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
        }
        period_str = f"Q{fortnight} - {month_names.get(month, str(month))} / {year}"

        # 1. Excel de Retenciones Recibidas
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            xls_path = Path(tmp.name)
            temp_files.append(xls_path)
        ret_headers = ["#", "Fecha Emisión", "Nro Comprobante", "Cliente RIF", "Factura Afectada", "Control Factura", "Base (Bs)", "IVA Retenido (Bs)"]
        ret_rows = []
        for idx, rec in enumerate(records):
            ret_rows.append([
                idx + 1,
                rec.fecha_emision.strftime("%d/%m/%Y"),
                rec.numero_comprobante,
                rec.rif,
                rec.numeros_facturas,
                rec.controles_facturas,
                float(rec.base_imponible or 0),
                float(rec.iva_retenido)
            ])
        excel_store.generate_premium_report_excel(
            xls_path,
            title="Resumen Quincenal de Retenciones de IVA Recibidas",
            period_str=period_str,
            headers=ret_headers,
            rows=ret_rows,
            numeric_cols=[6, 7],
            sum_cols=[6, 7],
            emisor_nombre=ctx.company_name,
            emisor_rif=ctx.company_rif,
            emisor_telefono=ctx.company_phone or "No definido",
            emisor_direccion=ctx.company_address or "No definida"
        )
        prof_xls = xls_path.with_name(f"RETENCIONES-RECIBIDAS-Q{fortnight}-{month}-{year}.xlsx")
        xls_path.rename(prof_xls)
        attachments.append(prof_xls)
        temp_files.append(prof_xls)
        
        # 2. Excel de Facturas Recibidas / Compras con Retención Emitida
        purchases_rows = excel_store.load_purchases_by_date_range(
            ctx.retenciones_emitidas_dir,
            date_from=start_date,
            date_to=end_date
        )
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            purchases_xls_path = Path(tmp.name)
            temp_files.append(purchases_xls_path)
        excel_store.generate_premium_report_excel(
            purchases_xls_path,
            title="Resumen Quincenal de Facturas Recibidas (Compras)",
            period_str=period_str,
            headers=["#", "Fecha", "Correlativo", "Proveedor", "RIF", "Nro Factura", "Nro Control", "Base (Bs)", "IVA (Bs)", "Monto (Bs)", "Retención (Bs)"],
            rows=purchases_rows,
            numeric_cols=[7, 8, 9, 10],
            sum_cols=[7, 8, 9, 10],
            emisor_nombre=ctx.company_name,
            emisor_rif=ctx.company_rif,
            emisor_telefono=ctx.company_phone or "No definido",
            emisor_direccion=ctx.company_address or "No definida"
        )
        prof_purchases_xls = purchases_xls_path.with_name(f"FACTURAS-RECIBIDAS-Q{fortnight}-{month}-{year}.xlsx")
        purchases_xls_path.rename(prof_purchases_xls)
        attachments.append(prof_purchases_xls)
        temp_files.append(prof_purchases_xls)

        # 3. Excel de Facturas Emitidas / Ventas
        sales_rows = excel_store.load_sales_by_date_range(
            ctx.facturas_emitidas_path,
            date_from=start_date,
            date_to=end_date
        )
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            sales_xls_path = Path(tmp.name)
            temp_files.append(sales_xls_path)
        excel_store.generate_premium_report_excel(
            sales_xls_path,
            title="Resumen Quincenal de Facturas Emitidas (Ventas)",
            period_str=period_str,
            headers=["#", "Fecha", "Nro Factura", "Cliente", "RIF", "Base (Bs)", "IVA (Bs)", "Monto (Bs)"],
            rows=sales_rows,
            numeric_cols=[5, 6, 7],
            sum_cols=[5, 6, 7],
            emisor_nombre=ctx.company_name,
            emisor_rif=ctx.company_rif,
            emisor_telefono=ctx.company_phone or "No definido",
            emisor_direccion=ctx.company_address or "No definida"
        )
        prof_sales_xls = sales_xls_path.with_name(f"FACTURAS-EMITIDAS-Q{fortnight}-{month}-{year}.xlsx")
        sales_xls_path.rename(prof_sales_xls)
        attachments.append(prof_sales_xls)
        temp_files.append(prof_sales_xls)

        # 4. Excel de Reportes Z / Ventas Diarias
        z_rows = excel_store.load_reportes_z_by_date_range(
            ctx.reportes_z_path,
            date_from=start_date,
            date_to=end_date
        )
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            z_xls_path = Path(tmp.name)
            temp_files.append(z_xls_path)
        excel_store.generate_premium_report_excel(
            z_xls_path,
            title="Resumen Quincenal de Cierres Z (Ventas Diarias)",
            period_str=period_str,
            headers=["#", "Fecha", "Nro Reporte Z", "Subtotal (Bs)", "Exento (Bs)", "Base (Bs)", "IVA (Bs)", "Total (Bs)"],
            rows=z_rows,
            numeric_cols=[3, 4, 5, 6, 7],
            sum_cols=[3, 4, 5, 6, 7],
            emisor_nombre=ctx.company_name,
            emisor_rif=ctx.company_rif,
            emisor_telefono=ctx.company_phone or "No definido",
            emisor_direccion=ctx.company_address or "No definida"
        )
        prof_z_xls = z_xls_path.with_name(f"REPORTES-Z-Q{fortnight}-{month}-{year}.xlsx")
        z_xls_path.rename(prof_z_xls)
        attachments.append(prof_z_xls)
        temp_files.append(prof_z_xls)
            
        # 2. Generar cuerpo del correo
        report = get_compromiso_tributario_report(
            year, month, fortnight,
            facturas_emitidas_path=ctx.facturas_emitidas_path,
            reportes_z_path=ctx.reportes_z_path,
            retenciones_emitidas_dir=ctx.retenciones_emitidas_dir,
            excel_path=ctx.excel_path,
            retenciones_islr_dir=ctx.retenciones_islr_dir
        )
        report_text = format_tributos_report(report)
        
        subject = f"Reporte Tributario Quincenal {ctx.company_name} - Q{fortnight} {month}/{year}"
        
        body = (
            f"Estimado destinatario,\n\n"
            f"Se adjuntan los reportes financieros y las planillas fiscales de la empresa "
            f"{ctx.company_name} para la quincena evaluada:\n\n"
            f"--------------------------------------------------\n"
            f"{report_text}\n"
            f"--------------------------------------------------\n\n"
            f"Atentamente,\n"
            f"Bot Financiero Automatizado ({ctx.company_name})"
        )
        
        # 3. Despachar el correo electrónico en segundo plano de forma no bloqueante
        await asyncio.to_thread(
            send_report_email,
            recipient_email,
            subject,
            body,
            attachments,
        )
        
        # Actualizar estado en Telegram
        await status_msg.edit_text(
            f"✅ *¡Reportes enviados exitosamente!*\n\n"
            f"Se despacharon todas las planillas quincenales (PDF y Excels) de la *Quincena {fortnight}* a la dirección: `{recipient_email}`",
            parse_mode="Markdown"
        )
    except Exception as e:
        logger.exception("Error en el envío asíncrono de correo")
        await status_msg.edit_text(
            f"❌ *Error al enviar los reportes por correo SMTP*\n\n"
            f"• *Destinatario:* `{recipient_email}`\n"
            f"• *Detalle del error:* `{e!s}`\n\n"
            f"Por favor, revisa tus credenciales SMTP en el archivo `.env` o el estado de la conexión a internet.",
            parse_mode="Markdown"
        )
    finally:
        # Eliminar archivos temporales creados para evitar fugas de espacio en disco
        for path in temp_files:
            try:
                path.unlink(missing_ok=True)
            except Exception:
                pass


async def _send_reportes_telegram_async(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    year: int,
    month: int,
    fortnight: int,
    status_msg,
) -> None:
    from .tributario_engine import get_fortnight_range
    import tempfile
    
    ctx = _get_company_context(update)
    logger.info("Iniciando _send_reportes_telegram_async para %d/%d Q%d", year, month, fortnight)
    temp_files: list[Path] = []
    attachments: list[tuple[Path, str]] = []
    
    try:
        start_date, end_date = get_fortnight_range(year, month, fortnight)
        
        # 1. Generar datos y reportes quincenales
        records = excel_store.retenciones_by_document_date(
            ctx.excel_path,
            date_from=start_date,
            date_to=end_date,
        )
        
        month_names = {
            1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
            7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
        }
        period_str = f"Q{fortnight} - {month_names.get(month, str(month))} / {year}"

        # 1. Excel de Retenciones Recibidas
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            xls_path = Path(tmp.name)
            temp_files.append(xls_path)
        ret_headers = ["#", "Fecha Emisión", "Nro Comprobante", "Cliente RIF", "Factura Afectada", "Control Factura", "Base (Bs)", "IVA Retenido (Bs)"]
        ret_rows = []
        for idx, rec in enumerate(records):
            ret_rows.append([
                idx + 1,
                rec.fecha_emision.strftime("%d/%m/%Y"),
                rec.numero_comprobante,
                rec.rif,
                rec.numeros_facturas,
                rec.controles_facturas,
                float(rec.base_imponible or 0),
                float(rec.iva_retenido)
            ])
        excel_store.generate_premium_report_excel(
            xls_path,
            title="Resumen Quincenal de Retenciones de IVA Recibidas",
            period_str=period_str,
            headers=ret_headers,
            rows=ret_rows,
            numeric_cols=[6, 7],
            sum_cols=[6, 7],
            emisor_nombre=ctx.company_name,
            emisor_rif=ctx.company_rif,
            emisor_telefono=ctx.company_phone or "No definido",
            emisor_direccion=ctx.company_address or "No definida"
        )
        attachments.append((xls_path, f"RETENCIONES-RECIBIDAS-Q{fortnight}-{month}-{year}.xlsx"))
        
        # 2. Excel de Facturas Recibidas / Compras con Retención Emitida
        purchases_rows = excel_store.load_purchases_by_date_range(
            ctx.retenciones_emitidas_dir,
            date_from=start_date,
            date_to=end_date
        )
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            purchases_xls_path = Path(tmp.name)
            temp_files.append(purchases_xls_path)
        excel_store.generate_premium_report_excel(
            purchases_xls_path,
            title="Resumen Quincenal de Facturas Recibidas (Compras)",
            period_str=period_str,
            headers=["#", "Fecha", "Correlativo", "Proveedor", "RIF", "Nro Factura", "Nro Control", "Base (Bs)", "IVA (Bs)", "Monto (Bs)", "Retención (Bs)"],
            rows=purchases_rows,
            numeric_cols=[7, 8, 9, 10],
            sum_cols=[7, 8, 9, 10],
            emisor_nombre=ctx.company_name,
            emisor_rif=ctx.company_rif,
            emisor_telefono=ctx.company_phone or "No definido",
            emisor_direccion=ctx.company_address or "No definida"
        )
        attachments.append((purchases_xls_path, f"FACTURAS-RECIBIDAS-Q{fortnight}-{month}-{year}.xlsx"))

        # 3. Excel de Facturas Emitidas / Ventas
        sales_rows = excel_store.load_sales_by_date_range(
            ctx.facturas_emitidas_path,
            date_from=start_date,
            date_to=end_date
        )
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            sales_xls_path = Path(tmp.name)
            temp_files.append(sales_xls_path)
        excel_store.generate_premium_report_excel(
            sales_xls_path,
            title="Resumen Quincenal de Facturas Emitidas (Ventas)",
            period_str=period_str,
            headers=["#", "Fecha", "Nro Factura", "Cliente", "RIF", "Base (Bs)", "IVA (Bs)", "Monto (Bs)"],
            rows=sales_rows,
            numeric_cols=[5, 6, 7],
            sum_cols=[5, 6, 7],
            emisor_nombre=ctx.company_name,
            emisor_rif=ctx.company_rif,
            emisor_telefono=ctx.company_phone or "No definido",
            emisor_direccion=ctx.company_address or "No definida"
        )
        attachments.append((sales_xls_path, f"FACTURAS-EMITIDAS-Q{fortnight}-{month}-{year}.xlsx"))

        # 4. Excel de Reportes Z / Ventas Diarias
        z_rows = excel_store.load_reportes_z_by_date_range(
            ctx.reportes_z_path,
            date_from=start_date,
            date_to=end_date
        )
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            z_xls_path = Path(tmp.name)
            temp_files.append(z_xls_path)
        excel_store.generate_premium_report_excel(
            z_xls_path,
            title="Resumen Quincenal de Cierres Z (Ventas Diarias)",
            period_str=period_str,
            headers=["#", "Fecha", "Nro Reporte Z", "Subtotal (Bs)", "Exento (Bs)", "Base (Bs)", "IVA (Bs)", "Total (Bs)"],
            rows=z_rows,
            numeric_cols=[3, 4, 5, 6, 7],
            sum_cols=[3, 4, 5, 6, 7],
            emisor_nombre=ctx.company_name,
            emisor_rif=ctx.company_rif,
            emisor_telefono=ctx.company_phone or "No definido",
            emisor_direccion=ctx.company_address or "No definida"
        )
        attachments.append((z_xls_path, f"REPORTES-Z-Q{fortnight}-{month}-{year}.xlsx"))
            
        # 5. Enviar cada archivo por Telegram
        for path, filename in attachments:
            logger.info("Enviando documento %s a Telegram...", filename)
            with open(path, "rb") as f:
                await context.bot.send_document(
                    chat_id=update.effective_message.chat_id,
                    document=f,
                    filename=filename,
                    caption=f"📊 Reporte: {filename}"
                )
                
        # Actualizar estado en Telegram
        await status_msg.edit_text(
            f"✅ *¡Reportes quincenales generados y enviados!*\n\n"
            f"Se enviaron los 4 archivos Excel de la *Quincena {fortnight}* del mes *{month}/{year}* directamente a este chat.",
            parse_mode="Markdown"
        )
        logger.info("Reportes generados y enviados con éxito.")
    except Exception as e:
        logger.exception("Error al generar y enviar reportes por Telegram")
        try:
            await status_msg.edit_text(
                f"❌ Error al generar los reportes:\n\n{e!s}"
            )
        except Exception as edit_err:
            logger.error("No se pudo editar el mensaje de error: %s", edit_err)
    finally:
        # Eliminar archivos temporales creados para evitar fugas de espacio en disco
        for path in temp_files:
            try:
                path.unlink(missing_ok=True)
            except Exception:
                pass


async def _send_seniat_txt_telegram_async(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    year: int,
    month: int,
    fortnight: int,
    status_msg,
) -> None:
    import tempfile
    
    ctx = _get_company_context(update)
    logger.info("Iniciando _send_seniat_txt_telegram_async para %d/%d Q%d", year, month, fortnight)
    temp_files: list[Path] = []
    
    try:
        # 1. Generar contenido TXT usando tributario_engine
        emitidas_txt, recibidas_txt = tributario_engine.generate_seniat_txt_data(
            year, month, fortnight,
            retenciones_emitidas_dir=ctx.retenciones_emitidas_dir,
            facturas_recibidas_path=ctx.facturas_recibidas_path,
            excel_path=ctx.excel_path,
            emitter_rif=ctx.company_rif
        )
        
        # 2. Crear archivos temporales
        # Retenciones Emitidas
        emit_filename = f"RETENCIONES_IVA_EMITIDAS_{year}_{month:02d}_Q{fortnight}.txt"
        emit_path = Path(tempfile.gettempdir()) / emit_filename
        emit_path.write_text(emitidas_txt, encoding="utf-8")
        temp_files.append(emit_path)
        
        # Retenciones Recibidas
        recib_filename = f"RETENCIONES_IVA_RECIBIDAS_{year}_{month:02d}_Q{fortnight}.txt"
        recib_path = Path(tempfile.gettempdir()) / recib_filename
        recib_path.write_text(recibidas_txt, encoding="utf-8")
        temp_files.append(recib_path)
        
        # 3. Enviar archivos a Telegram
        sent_count = 0
        if emitidas_txt.strip():
            logger.info("Enviando TXT de emitidas a Telegram...")
            with open(emit_path, "rb") as f:
                await context.bot.send_document(
                    chat_id=update.effective_message.chat_id,
                    document=f,
                    filename=emit_filename,
                    caption=f"📝 TXT Retenciones Emitidas (Compras) - Q{fortnight} {month:02d}/{year}"
                )
            sent_count += 1
            
        if recibidas_txt.strip():
            logger.info("Enviando TXT de recibidas a Telegram...")
            with open(recib_path, "rb") as f:
                await context.bot.send_document(
                    chat_id=update.effective_message.chat_id,
                    document=f,
                    filename=recib_filename,
                    caption=f"📝 TXT Retenciones Recibidas (Ventas) - Q{fortnight} {month:02d}/{year}"
                )
            sent_count += 1
            
        if sent_count > 0:
            await status_msg.edit_text(
                f"✅ *¡Archivos TXT generados con éxito!*\n\n"
                f"Se han enviado los archivos correspondientes a la *Quincena {fortnight}* del mes *{month}/{year}* a este chat para su declaración en el portal del SENIAT.",
                parse_mode="Markdown"
            )
        else:
            await status_msg.edit_text(
                f"⚠️ No se encontraron retenciones registradas en la *Quincena {fortnight}* del mes *{month}/{year}*.",
                parse_mode="Markdown"
            )
            
    except Exception as e:
        logger.exception("Error al generar y enviar archivos TXT SENIAT por Telegram")
        try:
            await status_msg.edit_text(
                f"❌ Error al generar los archivos TXT:\n\n{e!s}"
            )
        except Exception as edit_err:
            logger.error("No se pudo editar el mensaje de error: %s", edit_err)
    finally:
        for path in temp_files:
            try:
                path.unlink(missing_ok=True)
            except Exception:
                pass


def _generate_short_summary(report: dict[str, object]) -> str:
    month_names = {
        1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr", 5: "May", 6: "Jun",
        7: "Jul", 8: "Ago", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic"
    }
    y = report["year"]
    m = report["month"]
    f = report["fortnight"]
    m_name = month_names.get(m, str(m))
    
    iva = report["iva_neto_pagar_efectivo"]
    ret = report["retenciones_emitidas"]
    anticipo = report["anticipo_islr"]
    total = report["total_compromisos_a_pagar"]
    
    return (
        f"*Resumen Tributario SUFEVICA - Q{f} {m_name} {y}*\n"
        f"• IVA Neto: {excel_store._format_monto_ves(iva)} Bs\n"
        f"• Retenciones IVA: {excel_store._format_monto_ves(ret)} Bs\n"
        f"• Anticipo ISLR: {excel_store._format_monto_ves(anticipo)} Bs\n"
        f"👉 *Total a Pagar:* {excel_store._format_monto_ves(total)} Bs\n"
        f"_Generado por el Bot Financiero SUFEVICA._"
    )


def format_tributos_report(report: dict[str, object]) -> str:
    month_names = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
        7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }
    y = report["year"]
    m = report["month"]
    f = report["fortnight"]
    m_name = month_names.get(m, str(m))
    
    start_str = report["start_date"].strftime("%d/%m/%Y")
    end_str = report["end_date"].strftime("%d/%m/%Y")
    
    v_base = report["ventas_base"]
    v_iva = report["ventas_iva"]
    v_cnt = report["ventas_count"]
    
    c_base = report["compras_base"]
    c_iva = report["compras_iva"]
    c_cnt = report["compras_count"]
    
    ret_rec = report["retenciones_recibidas"]
    ret_rec_cnt = report["retenciones_recibidas_count"]
    
    ret_emi = report["retenciones_emitidas"]
    ret_emi_cnt = report["retenciones_emitidas_count"]
    
    ret_islr = report.get("retenciones_islr_compras", Decimal("0"))
    ret_islr_cnt = report.get("retenciones_islr_compras_count", 0)
    
    iva_neto = report["iva_neto_pagar"]
    pago_iva = report["iva_neto_pagar_efectivo"]
    anticipo = report["anticipo_islr"]
    total = report["total_compromisos_a_pagar"]
    
    due_date_str = report["due_date"].strftime("%d/%m/%Y")
    
    text = (
        f"🏛️ *CONTROL TRIBUTARIO - SUFEVICA* 🏛️\n"
        f"📅 *Período:* {f}ra Quincena de {m_name} {y}\n"
        f"🕒 *Rango:* {start_str} al {end_str}\n"
        f"⚠️ *Límite de Pago (RIF termina en 3):* `{due_date_str}`\n\n"
        f"--------------------------------------------------\n\n"
        f"📊 *Impuesto al Valor Agregado (IVA)*:\n"
        f" 🔸 *Ventas (Débito Fiscal):* {excel_store._format_monto_ves(v_iva)} Bs ({v_cnt} doc)\n"
        f" 🔸 *Compras (Crédito Fiscal):* {excel_store._format_monto_ves(c_iva)} Bs ({c_cnt} doc)\n"
        f" 🔸 *Retenciones Recibidas:* {excel_store._format_monto_ves(ret_rec)} Bs ({ret_rec_cnt} doc)\n"
    )
    
    if iva_neto >= 0:
        text += f" 👉 *IVA Neto a Pagar:* `{excel_store._format_monto_ves(iva_neto)}` Bs\n\n"
    else:
        text += f" 👉 *Excedente a Favor:* `{excel_store._format_monto_ves(abs(iva_neto))}` Bs (Crédito Fiscal)\n\n"
        
    text += (
        f"📤 *Retenciones de IVA a Enterar (Proveedores)*:\n"
        f" 🔸 *Retenido en Compras:* `{excel_store._format_monto_ves(ret_emi)}` Bs ({ret_emi_cnt} doc)\n"
        f"    _(Este monto se paga en su totalidad al SENIAT)_\n\n"
        f"💸 *Retenciones de ISLR a Enterar (Proveedores)*:\n"
        f" 🔸 *Retenido en Compras:* `{excel_store._format_monto_ves(ret_islr)}` Bs ({ret_islr_cnt} doc)\n"
        f"    _(Este monto se paga al SENIAT por retenciones efectuadas)_\n\n"
        f"💸 *Anticipo de ISLR (1% sobre Ventas)*:\n"
        f" 🔸 *Base imponible:* {excel_store._format_monto_ves(v_base)} Bs\n"
        f" 👉 *Anticipo a pagar:* `{excel_store._format_monto_ves(anticipo)}` Bs\n\n"
        f"--------------------------------------------------\n\n"
        f"💰 *TOTAL COMPROMISOS ESTIMADOS:* `{excel_store._format_monto_ves(total)}` Bs\n"
        f"_(Monto a pagar en banco/SENIAT para este corte)_"
    )
    return text


def _tributos_keyboard(year: int, month: int, fortnight: int, report_text: str = "") -> InlineKeyboardMarkup:
    prev_y, prev_m, prev_f = year, month, fortnight
    if fortnight == 2:
        prev_f = 1
    else:
        prev_f = 2
        if month == 1:
            prev_m = 12
            prev_y -= 1
        else:
            prev_m -= 1
            
    next_y, next_m, next_f = year, month, fortnight
    if fortnight == 1:
        next_f = 2
    else:
        next_f = 1
        if month == 12:
            next_m = 1
            next_y += 1
        else:
            next_m += 1
            
    keyboard = [
        [
            InlineKeyboardButton("📊 Detalle IVA", callback_data=f"tributos_detiva_{year}_{month}_{fortnight}"),
            InlineKeyboardButton("💸 Detalle ISLR", callback_data=f"tributos_detislr_{year}_{month}_{fortnight}"),
        ],
    ]
    
    if report_text:
        import urllib.parse
        # Codificar texto para compartir
        encoded_text = urllib.parse.quote(report_text)
        wa_url = f"https://api.whatsapp.com/send?text={encoded_text}"
        
        keyboard.append([
            InlineKeyboardButton("🟢 Compartir por WhatsApp", url=wa_url),
        ])
        
    keyboard.extend([
        [
            InlineKeyboardButton("📥 Descargar Reportes (Excel)", callback_data=f"tributos_download_{year}_{month}_{fortnight}")
        ],
        [
            InlineKeyboardButton("📝 Generar TXT SENIAT", callback_data=f"tributos_seniattxt_{year}_{month}_{fortnight}")
        ],
        [
            InlineKeyboardButton("📤 Enviar Reportes por Correo (SMTP)", callback_data=f"tributos_sendemail_{year}_{month}_{fortnight}")
        ],
        [
            InlineKeyboardButton("◀️ Quincena Ant.", callback_data=f"tributos_period_{prev_y}_{prev_m}_{prev_f}"),
            InlineKeyboardButton("Quincena Sig. ▶️", callback_data=f"tributos_period_{next_y}_{next_m}_{next_f}"),
        ],
        [
            InlineKeyboardButton("📅 Seleccionar Otro Mes", callback_data="tributos_selmonth"),
        ],
    ])
    return InlineKeyboardMarkup(keyboard)


def _months_selection_keyboard() -> InlineKeyboardMarkup:
    keyboard = []
    month_names = {
        1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr", 5: "May", 6: "Jun",
        7: "Jul", 8: "Ago", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic"
    }
    today = date.today()
    current_year = today.year
    current_month = today.month
    
    row = []
    for i in range(6):
        m = current_month - i
        y = current_year
        if m <= 0:
            m += 12
            y -= 1
        btn_text = f"{month_names[m]} {y}"
        row.append(InlineKeyboardButton(f"Q1 {btn_text}", callback_data=f"tributos_period_{y}_{m}_1"))
        row.append(InlineKeyboardButton(f"Q2 {btn_text}", callback_data=f"tributos_period_{y}_{m}_2"))
        if len(row) == 2:
            keyboard.append(row)
            row = []
    if row:
        keyboard.append(row)
    keyboard.append([InlineKeyboardButton("❌ Volver al menú principal", callback_data=f"tributos_period_{today.year}_{today.month}_1")])
    return InlineKeyboardMarkup(keyboard)


def _format_iva_details(year: int, month: int, fortnight: int, user_id: int | str | None = None) -> str:
    ctx = CompanyContext(user_id)
    start_date, end_date = tributario_engine.get_fortnight_range(year, month, fortnight)
    from .tributario_engine import _parse_row_date
    
    compras_details = []
    path_c = ctx.facturas_recibidas_path
    if path_c.exists():
        try:
            wb = load_workbook(path_c, read_only=True, data_only=True)
            ws = wb.active
            headers = excel_store._headers_index(ws)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                f_emi = _parse_row_date(excel_store._cell(row, headers, "Fecha_emision", None))
                if f_emi and start_date <= f_emi <= end_date:
                    prov = str(excel_store._cell(row, headers, "Proveedor", "-"))[:15]
                    rif = str(excel_store._cell(row, headers, "Proveedor_RIF", "-"))
                    iva_val = excel_store._parse_monto_cell(excel_store._cell(row, headers, "Monto_IVA", None)) or Decimal("0")
                    compras_details.append(f"• {f_emi.strftime('%d/%m')}: {prov} ({rif}) | IVA: {excel_store._format_monto_ves(iva_val)} Bs")
            wb.close()
        except Exception as e:
            compras_details.append(f"Error cargando compras: {e}")
            
    ventas_details = []
    for path_v in [ctx.facturas_emitidas_path, ctx.reportes_z_path]:
        if not path_v.exists():
            continue
        try:
            wb = load_workbook(path_v, read_only=True, data_only=True)
            ws = wb.active
            headers = excel_store._headers_index(ws)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                fecha_cell = excel_store._cell(row, headers, "Fecha_emision", None) or excel_store._cell(row, headers, "Fecha", None)
                f_doc = _parse_row_date(fecha_cell)
                if f_doc and start_date <= f_doc <= end_date:
                    doc = str(excel_store._cell(row, headers, "Numero_reporte", None) or excel_store._cell(row, headers, "Numero_documento", "-"))
                    cli = str(excel_store._cell(row, headers, "Razon_social", None) or "VENTAS DIARIAS")[:15]
                    iva_val = excel_store._parse_monto_cell(excel_store._cell(row, headers, "IVA", None)) or Decimal("0")
                    file_label = "Fact" if "EMITIDAS" in path_v.name else "RepZ"
                    ventas_details.append(f"• {f_doc.strftime('%d/%m')} [{file_label} {doc}]: {cli} | IVA: {excel_store._format_monto_ves(iva_val)} Bs")
            wb.close()
        except Exception as e:
            ventas_details.append(f"Error cargando ventas: {e}")
            
    ret_details = []
    path_r = ctx.excel_path
    if path_r.exists():
        try:
            wb = load_workbook(path_r, read_only=True, data_only=True)
            ws = wb.active
            headers = excel_store._headers_index(ws)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                f_emi = _parse_row_date(excel_store._cell(row, headers, "Fecha_emision", None))
                if f_emi and start_date <= f_emi <= end_date:
                    comp = str(excel_store._cell(row, headers, "Numero_comprobante", "-"))[:10]
                    cli = str(excel_store._cell(row, headers, "RIF", "-"))
                    ret_val = excel_store._parse_monto_cell(excel_store._cell(row, headers, "IVA_retenido", None)) or Decimal("0")
                    ret_details.append(f"• {f_emi.strftime('%d/%m')} [Comp {comp}]: {cli} | Ret: {excel_store._format_monto_ves(ret_val)} Bs")
            wb.close()
        except Exception as e:
            ret_details.append(f"Error cargando retenciones: {e}")
            
    text = f"📊 *DETALLE IVA QUINCENAL* ({start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')})\n\n"
    
    text += "*🛒 compras / crédito fiscal:*\n"
    if compras_details:
        text += "\n".join(compras_details) + "\n\n"
    else:
        text += "No hay compras registradas en este período.\n\n"
        
    text += "*📈 ventas / débito fiscal:*\n"
    if ventas_details:
        text += "\n".join(ventas_details) + "\n\n"
    else:
        text += "No hay ventas registradas en este período.\n\n"
        
    text += "*📥 retenciones recibidas:*\n"
    if ret_details:
        text += "\n".join(ret_details) + "\n"
    else:
        text += "No hay retenciones de clientes en este período.\n"
        
    return text


def _format_islr_details(year: int, month: int, fortnight: int, user_id: int | str | None = None) -> str:
    ctx = CompanyContext(user_id)
    start_date, end_date = tributario_engine.get_fortnight_range(year, month, fortnight)
    from .tributario_engine import _parse_row_date
    
    islr_details = []
    path_dir = ctx.retenciones_islr_dir
    # Buscar archivos de Excel en el directorio de retenciones de ISLR
    if path_dir.exists():
        try:
            for path_r in path_dir.glob("*.xlsx"):
                wb = load_workbook(path_r, read_only=True, data_only=True)
                ws = wb.active
                headers = excel_store._headers_index(ws)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue
                    f_emi = _parse_row_date(excel_store._cell(row, headers, "Fecha_emision", None) or excel_store._cell(row, headers, "Fecha", None))
                    if f_emi and start_date <= f_emi <= end_date:
                        comp = str(excel_store._cell(row, headers, "Numero_comprobante", None) or excel_store._cell(row, headers, "Comprobante", "-"))[:12]
                        prov = str(excel_store._cell(row, headers, "Proveedor", None) or excel_store._cell(row, headers, "Razon_social", "-"))[:15]
                        rif = str(excel_store._cell(row, headers, "Proveedor_RIF", None) or excel_store._cell(row, headers, "RIF", "-"))
                        monto_ret = excel_store._parse_monto_cell(excel_store._cell(row, headers, "Monto_ISLR_retenido", None) or excel_store._cell(row, headers, "ISLR_retenido", None)) or Decimal("0")
                        islr_details.append(f"• {f_emi.strftime('%d/%m')} [Comp {comp}]: {prov} ({rif}) | Ret: {excel_store._format_monto_ves(monto_ret)} Bs")
                wb.close()
        except Exception as e:
            islr_details.append(f"Error cargando ISLR: {e}")
            
    text = f"📊 *DETALLE ISLR QUINCENAL* ({start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')})\n\n"
    text += "*💸 retenciones de ISLR emitidas:*\n"
    if islr_details:
        text += "\n".join(islr_details) + "\n"
    else:
        text += "No hay retenciones de ISLR registradas en este período.\n"
        
    return text


async def _show_company_config_menu(update: Update, context: ContextTypes.DEFAULT_TYPE, msg_to_edit=None) -> None:
    ctx = _get_company_context(update)
    if not ctx.is_custom:
        await update.effective_message.reply_text("❌ Esta opción solo está disponible para usuarios con privilegios de Empresa (FlashTax).")
        return

    firma_personalizada = (ctx.dir_path / "firma_sello_transparente.png").exists()
    firma_status = "✅ Personalizada" if firma_personalizada else "❌ Por defecto (SUFEVICA)"

    user = user_manager.get_user(ctx.user_id) if ctx.user_id else None
    next_coti = user.get("next_cotizacion", 1) if user else 1
    next_nt = user.get("next_nota", 1) if user else 1

    text = (
        f"🏢 *Configuración de la Empresa (FlashTax)*\n\n"
        f"• *Razón Social:* {ctx.company_name}\n"
        f"• *RIF:* `{ctx.company_rif}`\n"
        f"• *Tipo de Contribuyente:* {ctx.company_type}\n"
        f"• *Correo del Contador:* `{ctx.company_email or 'no definido'}`\n"
        f"• *Teléfono:* `{ctx.company_phone or 'no definido'}`\n"
        f"• *Dirección:* `{ctx.company_address or 'no definida'}`\n"
        f"• *Firma y Sello:* {firma_status}\n"
        f"• *Próxima Cotización:* `{next_coti:06d}`\n"
        f"• *Próxima Nota de Entrega:* `{next_nt:06d}`\n\n"
        f"Seleccione el campo que desea modificar usando los botones de abajo:"
    )

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("📝 Modificar Razón Social", callback_data="cfg_company_name")],
        [InlineKeyboardButton("🆔 Modificar RIF", callback_data="cfg_company_rif")],
        [InlineKeyboardButton(f"🏛️ Contribuyente: {ctx.company_type}", callback_data="cfg_company_type")],
        [InlineKeyboardButton("📧 Modificar Correo", callback_data="cfg_company_email")],
        [InlineKeyboardButton("📞 Modificar Teléfono", callback_data="cfg_company_phone")],
        [InlineKeyboardButton("📍 Modificar Dirección Fiscal", callback_data="cfg_company_address")],
        [InlineKeyboardButton("✍️ Subir Firma y Sello", callback_data="cfg_company_signature")],
        [InlineKeyboardButton("🔢 Próxima Cotización", callback_data="cfg_next_cotizacion"), InlineKeyboardButton("🔢 Próxima Nota", callback_data="cfg_next_nota")],
        [InlineKeyboardButton("🔙 Volver al Inicio", callback_data="cfg_company_close")]
    ])

    if msg_to_edit:
        await msg_to_edit.edit_text(text, reply_markup=kb, parse_mode="Markdown")
    else:
        await update.effective_message.reply_text(text, reply_markup=kb, parse_mode="Markdown")


async def handle_cfg_company_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    q = update.callback_query
    if not q:
        return
    await q.answer()
    data = q.data
    msg = q.message
    if not msg:
        return
    user_id = update.effective_user.id
    
    user = user_manager.get_user(user_id)
    if not user or user.get("role") != "nueva_empresa":
        await msg.reply_text("❌ No autorizado.")
        return

    context.user_data.pop("awaiting_company_name", None)
    context.user_data.pop("awaiting_company_rif", None)
    context.user_data.pop("awaiting_company_email", None)
    context.user_data.pop("awaiting_company_phone", None)
    context.user_data.pop("awaiting_company_address", None)
    context.user_data.pop("awaiting_company_signature", None)
    context.user_data.pop("awaiting_company_next_cotizacion", None)
    context.user_data.pop("awaiting_company_next_nota", None)

    if data == "cfg_company_close":
        await _show_startup_menu(update, context, msg_to_edit=msg)
        return

    elif data == "cfg_company_type":
        current_type = user.get("company_type", "Especial")
        new_type = "Ordinario" if current_type == "Especial" else "Especial"
        user_manager.update_user_field(user_id, "company_type", new_type)
        await _show_company_config_menu(update, context, msg_to_edit=msg)
        return

    elif data == "cfg_company_name":
        context.user_data["awaiting_company_name"] = True
        await msg.edit_text(
            "📝 *Modificar Razón Social*\n\nEscribe el nuevo nombre legal de tu empresa:",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancelar", callback_data="cfg_company_back")]])
        )

    elif data == "cfg_company_rif":
        context.user_data["awaiting_company_rif"] = True
        await msg.edit_text(
            "🆔 *Modificar RIF*\n\nEscribe el RIF de tu empresa (debe ser válido según las normativas del SENIAT, ej: `J-40194130-3`):",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancelar", callback_data="cfg_company_back")]])
        )

    elif data == "cfg_company_email":
        context.user_data["awaiting_company_email"] = True
        await msg.edit_text(
            "📧 *Modificar Correo del Contador*\n\nEscribe la dirección de correo a la que deseas enviar reportes automáticos:",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancelar", callback_data="cfg_company_back")]])
        )

    elif data == "cfg_company_phone":
        context.user_data["awaiting_company_phone"] = True
        await msg.edit_text(
            "📞 *Modificar Teléfono de Contacto*\n\nEscribe el teléfono que aparecerá impreso en tus documentos (cotizaciones/notas):",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancelar", callback_data="cfg_company_back")]])
        )

    elif data == "cfg_company_address":
        context.user_data["awaiting_company_address"] = True
        await msg.edit_text(
            "📍 *Modificar Dirección Fiscal*\n\nEscribe la dirección fiscal de tu empresa:",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancelar", callback_data="cfg_company_back")]])
        )

    elif data == "cfg_company_signature":
        context.user_data["awaiting_company_signature"] = True
        await msg.edit_text(
            "✍️ *Subir Firma y Sello*\n\n"
            "Por favor, envía por chat la *imagen* (foto o archivo JPG/PNG) que contiene tu firma y sello.\n\n"
            "Idealmente, utiliza una imagen con fondo transparente para que se vea profesional sobre tus PDF.",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancelar", callback_data="cfg_company_back")]])
        )

    elif data == "cfg_next_cotizacion":
        context.user_data["awaiting_company_next_cotizacion"] = True
        await msg.edit_text(
            "🔢 *Modificar Siguiente Cotización*\n\nEscribe el próximo número correlativo que deseas asignar a tus cotizaciones (ej: `18`):",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancelar", callback_data="cfg_company_back")]])
        )

    elif data == "cfg_next_nota":
        context.user_data["awaiting_company_next_nota"] = True
        await msg.edit_text(
            "🔢 *Modificar Siguiente Nota de Entrega*\n\nEscribe el próximo número correlativo que deseas asignar a tus notas de entrega (ej: `10`):",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancelar", callback_data="cfg_company_back")]])
        )

    elif data == "cfg_company_back":
        await _show_company_config_menu(update, context, msg_to_edit=msg)





async def handle_share_email_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    q = update.callback_query
    if not q:
        return
        
    if not _allowed(update):
        await q.answer("❌ Acceso no autorizado / Suscripción Expirada.", show_alert=True)
        return
    if not _check_permission(update, "cotizaciones"):
        await q.answer("❌ No tienes privilegios para realizar operaciones de Cotizaciones.", show_alert=True)
        return
        
    await q.answer()
    msg = q.message
    if not msg:
        return
        
    # Check if we have SMTP configured
    if not config.SMTP_SERVER or not config.SMTP_USER or not config.SMTP_PASSWORD:
        await msg.reply_text(
            "❌ *Configuración de correo incompleta*\n\n"
            "Para poder enviar cotizaciones por correo, debes configurar las variables SMTP (`SMTP_SERVER`, `SMTP_PORT`, `SMTP_USER`, `SMTP_PASSWORD`) en tu archivo `.env`.",
            parse_mode="Markdown"
        )
        return
        
    share_doc = context.user_data.get("share_doc")
    if not share_doc:
        await msg.reply_text("❌ No encontré información del documento para compartir. Inicia el proceso nuevamente.")
        return
        
    share_doc["awaiting"] = "client_email"
    
    kb = InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancelar", callback_data="share_cancel")]])
    await msg.reply_text(
        f"📧 *Enviar {share_doc['title']} Nro {share_doc['doc_number']} por Correo*\n\n"
        f"Por favor, escribe la dirección de correo electrónico del cliente a donde deseas enviar el PDF:",
        reply_markup=kb,
        parse_mode="Markdown"
    )

async def handle_share_cancel_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    q = update.callback_query
    if not q:
        return
        
    if not _allowed(update):
        await q.answer("❌ Acceso no autorizado / Suscripción Expirada.", show_alert=True)
        return
    if not _check_permission(update, "cotizaciones"):
        await q.answer("❌ No tienes privilegios para realizar operaciones de Cotizaciones.", show_alert=True)
        return
        
    await q.answer()
    msg = q.message
    if not msg:
        return
    context.user_data.pop("share_doc", None)
    try:
        await q.delete_message()
    except Exception:
        pass
    await msg.reply_text("❌ Envío de correo cancelado.")


async def _send_document_email_async(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    recipient_email: str,
    status_msg,
) -> None:
    from .email_sender import send_report_email
    
    share_doc = context.user_data.get("share_doc")
    if not share_doc:
        await status_msg.edit_text("❌ Error: Información del documento perdida.")
        return
        
    pdf_path = Path(share_doc["pdf_path"])
    title = share_doc["title"]
    doc_num = share_doc["doc_number"]
    client_name = share_doc["client_name"]
    total_str = share_doc["total_amount"]
    
    subject = f"{title} Nro {doc_num} - SUFEVICA"
    body = (
        f"Estimado/a {client_name},\n\n"
        f"Adjunto a este correo encontrará su {title} Nro {doc_num} por un monto de {total_str}.\n\n"
        f"Agradecemos su preferencia.\n\n"
        f"Atentamente,\n"
        f"FREDDY LOPEZ (SUFEVICA)"
    )
    
    try:
        await asyncio.to_thread(
            send_report_email,
            recipient_email,
            subject,
            body,
            [pdf_path]
        )
        await status_msg.edit_text(
            f"✅ *¡Correo enviado con éxito!*\n\n"
            f"La {title.lower()} Nro *{doc_num}* ha sido enviada exitosamente a la dirección: `{recipient_email}`",
            parse_mode="Markdown"
        )
    except Exception as e:
        logger.exception("Error al enviar documento por correo")
        await status_msg.edit_text(
            f"❌ *Error al enviar el correo*\n\n"
            f"• *Destinatario:* `{recipient_email}`\n"
            f"• *Detalle:* `{e!s}`\n\n"
            f"Por favor verifica la conexión o tus credenciales SMTP en el `.env`.",
            parse_mode="Markdown"
        )
    finally:
        context.user_data.pop("share_doc", None)


def _get_and_increment_correlativo(doc_type: str, user_id: int | str | None = None) -> str:
    correlativos_path = Path(__file__).resolve().parent / "modulo_cotizaciones" / "correlativos.json"
    key = "next_cotizacion" if doc_type == "cotizacion" else "next_nota"
    num = None
    
    # 1. Intentar obtener el correlativo del perfil de usuario (usuarios.json)
    if user_id is not None:
        user = user_manager.get_user(user_id)
        if user:
            num = user.get(key)
            if num is not None:
                try:
                    num = int(num)
                except (ValueError, TypeError):
                    num = None
                    
    # 2. Si no está en el perfil, buscar en el archivo global correlativos.json
    if num is None:
        global_data = {"cotizacion": 1, "nota": 1}
        if correlativos_path.exists():
            try:
                import json
                global_data = json.loads(correlativos_path.read_text(encoding="utf-8"))
            except Exception:
                pass
        num = global_data.get(doc_type, 1)
        try:
            num = int(num)
        except (ValueError, TypeError):
            num = 1
            
    # 3. Incrementar el correlativo actual
    next_num = num + 1
    
    # 4. Guardar en el perfil de usuario (si existe)
    if user_id is not None:
        user_manager.update_user_field(user_id, key, next_num)
        
    # 5. Guardar también en el archivo global correlativos.json para retrocompatibilidad
    try:
        import json
        global_data = {"cotizacion": 1, "nota": 1}
        if correlativos_path.exists():
            try:
                global_data = json.loads(correlativos_path.read_text(encoding="utf-8"))
            except Exception:
                pass
        global_data[doc_type] = next_num
        correlativos_path.write_text(json.dumps(global_data, indent=4), encoding="utf-8")
    except Exception:
        pass
        
    return f"{num:06d}"


async def _start_document_flow(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    doc_type: str,
) -> None:
    msg = update.effective_message
    if not msg:
        return
    
    emoji = "📋" if doc_type == "cotizacion" else "📦"
    title_up = "COTIZACIÓN" if doc_type == "cotizacion" else "NOTA DE ENTREGA"
    
    context.user_data["active_menu"] = doc_type
    context.user_data["pending_doc"] = {
        "type": doc_type,
        "awaiting": "text_data"
    }
    
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("🛠️ Constructor Interactivo", callback_data="coti_build_start")],
        [InlineKeyboardButton("❌ Cancelar", callback_data="coti_build_cancel")]
    ])
    
    prompt = await msg.reply_text(
        f"{emoji} *NUEVA {title_up}* {emoji}\n\n"
        f"Por favor, *pega o escribe el texto* con los datos del cliente y los productos, o *envía una foto/imagen de la orden/documento* (cotización, factura o nota de entrega) para extraerlos automáticamente usando OCR con IA.\n\n"
        f"También puedes presionar el botón de abajo para iniciar el *Constructor Interactivo* y seleccionar productos directamente desde el inventario Excel.",
        reply_markup=kb,
        parse_mode="Markdown"
    )
    context.user_data["pending_doc"]["start_prompt_message_id"] = prompt.message_id


async def _process_document_text_immediate(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    doc_type: str,
    text_content: str,
) -> None:
    msg = update.effective_message
    if not msg:
        return
        
    doc_data = _parse_document_text_explicit(text_content, doc_type)
    
    # Fallback a Gemini si falla el parser local de expresiones regulares
    if doc_data is None and config.GEMINI_API_KEY:
        try:
            from . import ocr_extract
            extracted = ocr_extract.parse_document_text_with_gemini(text_content)
            if extracted and extracted.get("items"):
                client_info = {
                    "name": extracted.get("client_name") or "",
                    "rif": extracted.get("client_rif") or "",
                    "address": extracted.get("client_address") or "",
                    "phone": extracted.get("client_phone") or "",
                    "salesman": "FREDDY LOPEZ",
                    "saleType": "Contado",
                    "note": ""
                }
                if client_info["rif"]:
                    client_info["rif"] = _normalize_rif(client_info["rif"])

                doc_data = {
                    "docType": doc_type,
                    "currency": "usd",
                    "exchangeRate": get_current_bcv_rate(),
                    "docNumber": "",
                    "docDate": date.today().strftime("%Y-%m-%d"),
                    "client": client_info,
                    "items": extracted.get("items")
                }
        except Exception as e:
            logger.error(f"Error en fallback Gemini para texto inmediato: {e}")

    if doc_data is not None:
        # Formatear items crudos sin cruzar con inventario
        for it in doc_data["items"]:
            it["qty"] = float(it.get("qty") or 1.0)
            it["priceUsd"] = float(it.get("priceUsd") or 0.0)
            it["totalUsd"] = it["qty"] * it["priceUsd"]
            it["code"] = (it.get("code") or "").strip()
            it["desc"] = (it.get("desc") or "").strip()
        context.user_data["pending_doc"] = {
            "type": doc_type,
            "awaiting": "edit_card",
            "parsed_data": doc_data
        }
        try:
            await msg.delete()
        except Exception:
            pass
        await _send_client_data_card(update, context, first_time=True)
    else:
        context.user_data["pending_doc"] = {
            "type": doc_type,
            "awaiting": "text_data"
        }
        await msg.reply_text(
            "⚠️ No pude identificar productos en el texto que incluiste después del comando.\n\n"
            "Por favor, vuelve a intentarlo *pegando o escribiendo solo el texto* con los datos aquí (asegúrate de incluir cantidades y precios, ej: `10 x PARP-220 Perno a 2.50`):\n\n"
            "O envía `/start` para cancelar.",
            parse_mode="Markdown"
        )


async def cotizacion_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not _allowed(update):
        await _deny(update)
        return
    msg = update.effective_message
    if not msg:
        return
        
    if not _check_permission(update, "cotizaciones"):
        await msg.reply_text("❌ No tienes privilegios para acceder al módulo de Cotizaciones.")
        return
        
    text_content = ""
    if msg.text:
        parts = msg.text.split(None, 1)
        if len(parts) > 1:
            text_content = parts[1].strip()
            
    if text_content:
        await _process_document_text_immediate(update, context, "cotizacion", text_content)
    else:
        await _start_document_flow(update, context, "cotizacion")


async def nota_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not _allowed(update):
        await _deny(update)
        return
    msg = update.effective_message
    if not msg:
        return
        
    if not _check_permission(update, "cotizaciones"):
        await msg.reply_text("❌ No tienes privilegios para acceder al módulo de Cotizaciones.")
        return
        
    text_content = ""
    if msg.text:
        parts = msg.text.split(None, 1)
        if len(parts) > 1:
            text_content = parts[1].strip()
            
    if text_content:
        await _process_document_text_immediate(update, context, "nota", text_content)
    else:
        await _start_document_flow(update, context, "nota")


async def _send_client_data_card(update: Update, context: ContextTypes.DEFAULT_TYPE, first_time: bool = False) -> None:
    msg = update.effective_message
    if not msg:
        return
    pending_doc = context.user_data.get("pending_doc")
    if not pending_doc or "parsed_data" not in pending_doc:
        return
    doc_data = pending_doc["parsed_data"]
    client = doc_data["client"]
    doc_type = doc_data["docType"]
    title_up = "COTIZACIÓN" if doc_type == "cotizacion" else "NOTA DE ENTREGA"
    emoji = "📋" if doc_type == "cotizacion" else "📦"
    
    rate = doc_data.get("exchangeRate", get_current_bcv_rate())
    text = (
        f"📝 *DATOS DEL CLIENTE DETECTADOS ({title_up})* {emoji}\n\n"
        f"👤 *Cliente:* {client.get('name') or '_[Omitido / Vacío]_'}\n"
        f"🆔 *RIF/CI:* {client.get('rif') or '_[Omitido / Vacío]_'}\n"
        f"📍 *Dirección:* {client.get('address') or '_[Omitido / Vacío]_'}\n"
        f"📞 *Teléfono:* {client.get('phone') or '_[Omitido / Vacío]_'}\n"
        f"👔 *Vendedor:* {client.get('salesman') or '_[Omitido / Vacío]_'}\n"
        f"💳 *Condición de Pago:* {client.get('saleType') or 'Contado'}\n"
        f"📝 *Nota:* {client.get('note') or '_[Nota por defecto]_'}\n"
        f"💵 *Tasa BCV:* Bs. {rate:,.2f}\n\n"
        f"¿Deseas modificar o añadir alguno de estos datos antes de continuar?\n"
        f"_(Cualquiera de estos datos puede ser omitido para dejarlo en blanco)_"
    )
    done_lbl = "⏩ VOLVER AL CONSTRUCTOR" if pending_doc.get("builder_mode") else "⏩ CONTINUAR A LA MONEDA"
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("👤 Cliente", callback_data="coti_edit_name"), InlineKeyboardButton("🆔 RIF/CI", callback_data="coti_edit_rif")],
        [InlineKeyboardButton("📍 Dirección", callback_data="coti_edit_address"), InlineKeyboardButton("📞 Teléfono", callback_data="coti_edit_phone")],
        [InlineKeyboardButton("👔 Vendedor", callback_data="coti_edit_salesman"), InlineKeyboardButton("💳 Condición: " + client.get("saleType", "Contado"), callback_data="coti_edit_saletype")],
        [InlineKeyboardButton("📝 Nota", callback_data="coti_edit_note"), InlineKeyboardButton("💵 Tasa BCV", callback_data="coti_edit_rate")],
        [InlineKeyboardButton(done_lbl, callback_data="coti_edit_done")]
    ])
    menu_message_id = pending_doc.get("menu_message_id") if pending_doc else None
    chat_id = update.effective_chat.id if update.effective_chat else None
    
    if first_time or not menu_message_id or not chat_id:
        sent_msg = await context.bot.send_message(
            chat_id=chat_id,
            text=text,
            reply_markup=kb,
            parse_mode="Markdown"
        )
        if pending_doc:
            pending_doc["menu_message_id"] = sent_msg.message_id
    else:
        try:
            await context.bot.edit_message_text(
                chat_id=chat_id,
                message_id=menu_message_id,
                text=text,
                reply_markup=kb,
                parse_mode="Markdown"
            )
        except Exception:
            sent_msg = await context.bot.send_message(
                chat_id=chat_id,
                text=text,
                reply_markup=kb,
                parse_mode="Markdown"
            )
            if pending_doc:
                pending_doc["menu_message_id"] = sent_msg.message_id


async def handle_cotizaciones_edit_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    q = update.callback_query
    if not q:
        return
        
    if not _allowed(update):
        await q.answer("❌ Acceso no autorizado / Suscripción Expirada.", show_alert=True)
        return
    if not _check_permission(update, "cotizaciones"):
        await q.answer("❌ No tienes privilegios para realizar operaciones de Cotizaciones.", show_alert=True)
        return
        
    await q.answer()
    data = (q.data or "").strip()
    msg = q.message
    if not msg:
        return
    pending_doc = context.user_data.get("pending_doc")
    if not pending_doc or "parsed_data" not in pending_doc:
        await msg.reply_text("❌ No hay ningún documento pendiente de procesar.")
        return
    doc_data = pending_doc["parsed_data"]
    client = doc_data["client"]
    
    prompt = None
    if data == "coti_edit_name":
        pending_doc["awaiting"] = "input_name"
        prompt = await msg.reply_text("👤 Envía el *Nombre del Cliente* (o escribe `omitir` para dejarlo en blanco):", parse_mode="Markdown")
    elif data == "coti_edit_rif":
        pending_doc["awaiting"] = "input_rif"
        prompt = await msg.reply_text("🆔 Envía el *RIF o C.I. del Cliente* (o escribe `omitir` para dejarlo en blanco):", parse_mode="Markdown")
    elif data == "coti_edit_address":
        pending_doc["awaiting"] = "input_address"
        prompt = await msg.reply_text("📍 Envía la *Dirección Fiscal del Cliente* (o escribe `omitir` para dejarlo en blanco):", parse_mode="Markdown")
    elif data == "coti_edit_phone":
        pending_doc["awaiting"] = "input_phone"
        prompt = await msg.reply_text("📞 Envía el *Teléfono del Cliente* (o escribe `omitir` para dejarlo en blanco):", parse_mode="Markdown")
    elif data == "coti_edit_salesman":
        pending_doc["awaiting"] = "input_salesman"
        prompt = await msg.reply_text("👔 Envía el *Nombre del Vendedor* (o escribe `omitir` para dejarlo en blanco):", parse_mode="Markdown")
    elif data == "coti_edit_note":
        pending_doc["awaiting"] = "input_note"
        prompt = await msg.reply_text("📝 Envía el texto de la *Nota* personalizada para el pie de página (o escribe `omitir` para usar la nota por defecto):", parse_mode="Markdown")
    elif data == "coti_edit_rate":
        pending_doc["awaiting"] = "input_rate"
        prompt = await msg.reply_text("💵 Envía el valor de la *Tasa de Cambio BCV* (ej. `40.50`):", parse_mode="Markdown")
        
    if prompt:
        pending_doc["prompt_message_id"] = prompt.message_id
    elif data == "coti_edit_saletype":
        current = client.get("saleType", "Contado")
        client["saleType"] = "Crédito" if current == "Contado" else "Contado"
        await _send_client_data_card(update, context)
    elif data == "coti_edit_done":
        if pending_doc.get("builder_mode"):
            pending_doc["awaiting"] = "builder_main"
            try:
                await q.delete_message()
            except Exception:
                pass
            await _send_interactive_builder_card(update, context, first_time=True)
        else:
            pending_doc["awaiting"] = "currency"
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("Dólares Americanos ($)", callback_data="coti_curr_usd"), InlineKeyboardButton("Bolívares (Bs.)", callback_data="coti_curr_ves")]
            ])
            try:
                await q.delete_message()
            except Exception:
                pass
            await msg.reply_text("💵 ¿En qué moneda deseas que se exprese el documento por defecto al abrirse?", reply_markup=kb, parse_mode="Markdown")


async def _send_interactive_builder_card(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    first_time: bool = False,
) -> None:
    msg = update.effective_message
    if not msg:
        return
    pending_doc = context.user_data.get("pending_doc")
    if not pending_doc or "parsed_data" not in pending_doc:
        return
        
    doc_data = pending_doc["parsed_data"]
    client = doc_data["client"]
    items = doc_data["items"]
    doc_type = doc_data["docType"]
    currency = doc_data.get("currency", "usd")
    rate = doc_data.get("exchangeRate", get_current_bcv_rate())
    
    title_up = "COTIZACIÓN" if doc_type == "cotizacion" else "NOTA DE ENTREGA"
    emoji = "📋" if doc_type == "cotizacion" else "📦"
    
    # Calcular totales
    total_usd = sum(float(it.get("qty", 1.0)) * float(it.get("priceUsd", 0.0)) for it in items)
    conv_rate = float(rate) if currency == "ves" else 1.0
    total_conv = total_usd * conv_rate
    symbol = "$" if currency == "usd" else "Bs."
    formatted_total = f"{total_conv:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    
    # Construir listado de ítems
    items_text = ""
    for idx, it in enumerate(items, 1):
        qty = float(it.get("qty", 1.0))
        price = float(it.get("priceUsd", 0.0)) * conv_rate
        subt = qty * price
        formatted_subt = f"{subt:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        formatted_price = f"{price:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        
        desc = it.get("desc", "")
        if len(desc) > 30:
            desc = desc[:27] + "..."
            
        items_text += f"*{idx}.* `{it.get('code')}` - {desc}\n    _{qty} x {symbol} {formatted_price} = {symbol} {formatted_subt}_\n"
        
    if not items_text:
        items_text = "_[No hay productos agregados todavía]_\n"
        
    client_name = client.get("name") or "_[No especificado]_"
    client_rif = client.get("rif") or "_[No especificado]_"
    
    text = (
        f"🛠️ *CONSTRUCTOR INTERACTIVO DE {title_up}* {emoji}\n\n"
        f"👤 *Cliente:* {client_name}\n"
        f"🆔 *RIF/CI:* {client_rif}\n"
        f"💵 *Moneda:* {currency.upper()} | *Tasa BCV:* Bs. {rate:,.2f}\n\n"
        f"🛒 *PRODUCTOS:* \n{items_text}\n"
        f"💰 *TOTAL APROXIMADO:* *{symbol} {formatted_total}*\n\n"
        f"Usa los botones de abajo para buscar y agregar productos desde el Excel, configurar los datos del cliente, o generar el PDF final."
    )
    
    kb = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("🔍 Buscar por Código", callback_data="coti_build_search_code"),
            InlineKeyboardButton("🔎 Buscar por Descripción", callback_data="coti_build_search_desc")
        ],
        [
            InlineKeyboardButton("📷 Escanear Código Barra", callback_data="coti_build_search_barcode"),
            InlineKeyboardButton("📸 Escanear Producto (OCR)", callback_data="coti_build_search_ocr")
        ],
        [
            InlineKeyboardButton("👤 Configurar Cliente", callback_data="coti_build_edit_client"),
            InlineKeyboardButton("💵 Cambiar Moneda", callback_data="coti_build_toggle_curr")
        ],
        [
            InlineKeyboardButton("🛒 Ver / Editar Ítems", callback_data="coti_build_view_items")
        ],
        [
            InlineKeyboardButton("✅ GENERAR DOCUMENTO", callback_data="coti_build_generate"),
            InlineKeyboardButton("❌ CANCELAR", callback_data="coti_build_cancel")
        ]
    ])
    
    menu_message_id = pending_doc.get("menu_message_id")
    chat_id = update.effective_chat.id if update.effective_chat else None
    
    if first_time or not menu_message_id or not chat_id:
        # Borrar el prompt anterior de entrada de texto si existe
        start_prompt_id = pending_doc.pop("start_prompt_message_id", None)
        if start_prompt_id and chat_id:
            try:
                await context.bot.delete_message(chat_id=chat_id, message_id=start_prompt_id)
            except Exception:
                pass
        
        sent_msg = await context.bot.send_message(
            chat_id=chat_id,
            text=text,
            reply_markup=kb,
            parse_mode="Markdown"
        )
        pending_doc["menu_message_id"] = sent_msg.message_id
    else:
        try:
            await context.bot.edit_message_text(
                chat_id=chat_id,
                message_id=menu_message_id,
                text=text,
                reply_markup=kb,
                parse_mode="Markdown"
            )
        except Exception:
            sent_msg = await context.bot.send_message(
                chat_id=chat_id,
                text=text,
                reply_markup=kb,
                parse_mode="Markdown"
            )
            pending_doc["menu_message_id"] = sent_msg.message_id


async def _send_builder_items_editor(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    msg = update.effective_message
    if not msg:
        return
    pending_doc = context.user_data.get("pending_doc")
    if not pending_doc or "parsed_data" not in pending_doc:
        return
    
    doc_data = pending_doc["parsed_data"]
    items = doc_data["items"]
    currency = doc_data.get("currency", "usd")
    rate = doc_data.get("exchangeRate", get_current_bcv_rate())
    symbol = "$" if currency == "usd" else "Bs."
    conv_rate = float(rate) if currency == "ves" else 1.0
    
    text = (
        "🛒 *VISTA PREVIA DEL DOCUMENTO*\n\n"
        "Aquí puedes ver la lista de productos agregados, sus cantidades y precios. "
        "Usa los botones de cada fila para modificar o eliminar un ítem:\n\n"
    )
    
    total_val = 0.0
    kb_list = []
    for idx, it in enumerate(items):
        qty = float(it.get("qty", 1.0))
        price_unit = float(it.get("priceUsd", 0.0)) * conv_rate
        subt = qty * price_unit
        total_val += subt
        
        formatted_price = f"{price_unit:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        formatted_subt = f"{subt:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        
        desc_short = it.get('desc', '')
        if len(desc_short) > 28:
            desc_short = desc_short[:25] + "..."
            
        text += (
            f"*{idx + 1}. {it.get('code')}* - {desc_short}\n"
            f"   • Cant: *{qty}* | P.U.: *{symbol} {formatted_price}* | Sub: *{symbol} {formatted_subt}*\n\n"
        )
        
        btn_cant = InlineKeyboardButton(f"{idx+1}. ✏️ Cant", callback_data=f"coti_build_edit_qty:{idx}")
        btn_precio = InlineKeyboardButton(f"{idx+1}. ✏️ Precio", callback_data=f"coti_build_edit_price:{idx}")
        btn_delete = InlineKeyboardButton(f"{idx+1}. ❌ Eliminar", callback_data=f"coti_build_delete_item:{idx}")
        kb_list.append([btn_cant, btn_precio, btn_delete])
        
    if not items:
        text += "_[No hay productos agregados todavía]_\n"
        
    formatted_total = f"{total_val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    text += f"----------------------------------\n💰 *TOTAL:* *{symbol} {formatted_total}*\n"
    
    kb_list.append([
        InlineKeyboardButton("🧹 Vaciar Todo", callback_data="coti_build_clear_items"),
        InlineKeyboardButton("🔙 Volver al Constructor", callback_data="coti_build_main")
    ])
    kb = InlineKeyboardMarkup(kb_list)
    
    chat_id = update.effective_chat.id if update.effective_chat else None
    menu_message_id = pending_doc.get("menu_message_id")
    
    if menu_message_id and chat_id:
        try:
            await context.bot.edit_message_text(
                chat_id=chat_id,
                message_id=menu_message_id,
                text=text,
                reply_markup=kb,
                parse_mode="Markdown"
            )
            return
        except Exception:
            pass
            
    sent_msg = await msg.reply_text(
        text,
        reply_markup=kb,
        parse_mode="Markdown"
    )
    pending_doc["menu_message_id"] = sent_msg.message_id


async def handle_builder_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    q = update.callback_query
    if not q:
        return
        
    if not _allowed(update):
        await q.answer("❌ Acceso no autorizado / Suscripción Expirada.", show_alert=True)
        return
    if not _check_permission(update, "cotizaciones"):
        await q.answer("❌ No tienes privilegios para realizar operaciones de Cotizaciones.", show_alert=True)
        return
        
    await q.answer()
    data = (q.data or "").strip()
    msg = q.message
    if not msg:
        return
        
    pending_doc = context.user_data.get("pending_doc")
    
    # Para coti_build_start, inicializar estructura de constructor interactivo
    if data == "coti_build_start":
        doc_type = "cotizacion"
        if pending_doc:
            doc_type = pending_doc.get("type", "cotizacion")
        
        pending_doc = {
            "type": doc_type,
            "builder_mode": True,
            "awaiting": "builder_main",
            "parsed_data": {
                "docType": doc_type,
                "docNumber": "",
                "client": {
                    "name": "",
                    "rif": "",
                    "address": "",
                    "phone": "",
                    "salesman": "",
                    "saleType": "Contado",
                    "note": "",
                },
                "items": [],
                "currency": "usd",
                "exchangeRate": get_current_bcv_rate(),
            }
        }
        context.user_data["pending_doc"] = pending_doc
        
        try:
            await q.delete_message()
        except Exception:
            pass
            
        await _send_interactive_builder_card(update, context, first_time=True)
        return
        
    if not pending_doc or "parsed_data" not in pending_doc:
        await msg.reply_text("❌ No hay ningún documento en construcción activa. Envía /cotizacion o /nota para iniciar.")
        return
        
    doc_data = pending_doc["parsed_data"]
    
    if data == "coti_build_main":
        pending_doc["awaiting"] = "builder_main"
        prompt_id = pending_doc.pop("prompt_message_id", None)
        if prompt_id:
            try:
                await context.bot.delete_message(chat_id=msg.chat_id, message_id=prompt_id)
            except Exception:
                pass
        await _send_interactive_builder_card(update, context, first_time=False)
        
    elif data == "coti_build_search_code":
        pending_doc["awaiting"] = "search_code"
        try:
            await q.delete_message()
        except Exception:
            pass
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Volver al Constructor", callback_data="coti_build_main")]])
        prompt = await msg.reply_text(
            "🔍 *BUSCAR POR CÓDIGO*\n\nPor favor, escribe el *código de producto* a buscar (ej. `TUB-001` o `PER`):",
            reply_markup=kb,
            parse_mode="Markdown"
        )
        pending_doc["prompt_message_id"] = prompt.message_id
        
    elif data == "coti_build_search_barcode":
        pending_doc["awaiting"] = "search_barcode"
        try:
            await q.delete_message()
        except Exception:
            pass
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Volver al Constructor", callback_data="coti_build_main")]])
        prompt = await msg.reply_text(
            "📷 *ESCANEAR CÓDIGO DE BARRAS*\n\nPor favor, *toma una foto* al código de barras o código QR de tu producto (desde tu móvil o PC) y *envíamela* por aquí. Yo la decodificaré para buscar el producto:",
            reply_markup=kb,
            parse_mode="Markdown"
        )
        pending_doc["prompt_message_id"] = prompt.message_id

    elif data == "coti_build_search_ocr":
        pending_doc["awaiting"] = "search_ocr"
        try:
            await q.delete_message()
        except Exception:
            pass
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Volver al Constructor", callback_data="coti_build_main")]])
        prompt = await msg.reply_text(
            "📸 *ESCANEAR PRODUCTO CON OCR*\n\nPor favor, *toma una foto* a la etiqueta, empaque o catálogo del producto y *envíamela* por aquí. Yo leeré el texto con Inteligencia Artificial para buscarlo en tu inventario:",
            reply_markup=kb,
            parse_mode="Markdown"
        )
        pending_doc["prompt_message_id"] = prompt.message_id

    elif data == "coti_build_search_ocr":
        pending_doc["awaiting"] = "search_ocr"
        try:
            await q.delete_message()
        except Exception:
            pass
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Volver al Constructor", callback_data="coti_build_main")]])
        prompt = await msg.reply_text(
            "📸 *ESCANEAR PRODUCTO CON OCR*\n\nPor favor, *toma una foto* a la etiqueta, empaque o catálogo del producto y *envíamela* por aquí. Yo leeré el texto con Inteligencia Artificial para buscarlo en tu inventario:",
            reply_markup=kb,
            parse_mode="Markdown"
        )
        pending_doc["prompt_message_id"] = prompt.message_id
        
    elif data == "coti_build_search_desc":
        pending_doc["awaiting"] = "search_desc"
        try:
            await q.delete_message()
        except Exception:
            pass
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Volver al Constructor", callback_data="coti_build_main")]])
        prompt = await msg.reply_text(
            "🔎 *BUSCAR POR DESCRIPCIÓN*\n\nPor favor, escribe la *descripción o palabra clave* del producto a buscar (ej. `tubo` o `codo`):",
            reply_markup=kb,
            parse_mode="Markdown"
        )
        pending_doc["prompt_message_id"] = prompt.message_id
        
    elif data == "coti_build_toggle_curr":
        curr = doc_data.get("currency", "usd")
        doc_data["currency"] = "ves" if curr == "usd" else "usd"
        await _send_interactive_builder_card(update, context, first_time=False)
        
    elif data == "coti_build_edit_client":
        pending_doc["awaiting"] = "edit_card"
        try:
            await q.delete_message()
        except Exception:
            pass
        await _send_client_data_card(update, context, first_time=True)
        
    elif data == "coti_build_view_items":
        if pending_doc and pending_doc.get("awaiting") in ("edit_item_qty", "edit_item_price"):
            pending_doc["awaiting"] = "builder_main"
        await _send_builder_items_editor(update, context)
        
    elif data.startswith("coti_build_edit_qty:"):
        idx = int(data.split(":")[1])
        items = doc_data["items"]
        if 0 <= idx < len(items):
            pending_doc["awaiting"] = "edit_item_qty"
            pending_doc["edit_item_idx"] = idx
            item = items[idx]
            try:
                await q.delete_message()
            except Exception:
                pass
            kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancelar", callback_data="coti_build_view_items")]])
            prompt = await msg.reply_text(
                f"📝 *EDITAR CANTIDAD*\n\n"
                f"Producto: *{item.get('desc')}* (`{item.get('code')}`)\n"
                f"Cantidad actual: *{item.get('qty')}*\n\n"
                f"Por favor, escribe la *nueva cantidad* (mayor a cero):",
                reply_markup=kb,
                parse_mode="Markdown"
            )
            pending_doc["prompt_message_id"] = prompt.message_id
            
    elif data.startswith("coti_build_edit_price:"):
        idx = int(data.split(":")[1])
        items = doc_data["items"]
        if 0 <= idx < len(items):
            pending_doc["awaiting"] = "edit_item_price"
            pending_doc["edit_item_idx"] = idx
            item = items[idx]
            try:
                await q.delete_message()
            except Exception:
                pass
            kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancelar", callback_data="coti_build_view_items")]])
            prompt = await msg.reply_text(
                f"📝 *EDITAR PRECIO UNITARIO*\n\n"
                f"Producto: *{item.get('desc')}* (`{item.get('code')}`)\n"
                f"Precio actual: *${float(item.get('priceUsd', 0.0)):,.2f}*\n\n"
                f"Por favor, escribe el *nuevo precio unitario* en USD (mayor a cero):",
                reply_markup=kb,
                parse_mode="Markdown"
            )
            pending_doc["prompt_message_id"] = prompt.message_id
        
    elif data == "coti_build_clear_items":
        doc_data["items"] = []
        await _send_builder_items_editor(update, context)
        
    elif data.startswith("coti_build_delete_item:"):
        idx = int(data.split(":")[1])
        if 0 <= idx < len(doc_data["items"]):
            doc_data["items"].pop(idx)
        await _send_builder_items_editor(update, context)
        
    elif data.startswith("coti_build_select_p:"):
        p_val = data.split(":", 1)[1]
        product = None
        if p_val.isdigit() and "temp_search_results" in pending_doc:
            idx = int(p_val)
            if 0 <= idx < len(pending_doc["temp_search_results"]):
                product = pending_doc["temp_search_results"][idx]
                
        if not product:
            # Fallback en caso de que no esté en la caché temporal (ej. reinicio)
            ctx = _get_company_context(update)
            products = excel_store.search_products_in_excel(ctx.productos_path, p_val, search_by="code")
            if products:
                product = products[0]
                for p in products:
                    if p["code"].lower() == p_val.lower():
                        product = p
                        break
                        
        if product:
            pending_doc["selected_product"] = product
            pending_doc["awaiting"] = "input_qty"
            
            try:
                await q.delete_message()
            except Exception:
                pass
                
            import html
            desc_esc = html.escape(product['description'])
            code_esc = html.escape(product['code'])
            kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Volver al Constructor", callback_data="coti_build_main")]])
            prompt = await msg.reply_text(
                f"🔢 <b>CANTIDAD DE PRODUCTO</b>\n\n"
                f"Has seleccionado: <b>{desc_esc}</b> (<code>{code_esc}</code>)\n"
                f"Precio Unitario: <b>${product['price']:.2f}</b>\n\n"
                f"Por favor, escribe la <b>cantidad</b> a cotizar / usar para este producto:",
                reply_markup=kb,
                parse_mode="HTML"
            )
            pending_doc["prompt_message_id"] = prompt.message_id
        else:
            await msg.reply_text("❌ Error: No se encontró la información del producto. Inténtalo nuevamente.")
            
    elif data == "coti_build_generate":
        if not doc_data.get("items"):
            await msg.reply_text("⚠️ No se puede generar un documento sin productos. Agrega al menos uno primero.")
            return
            
        doc_type = doc_data["docType"]
        user_id = update.effective_user.id if update.effective_user else None
        correlativo = _get_and_increment_correlativo(doc_type, user_id)
        doc_data["docNumber"] = correlativo
        
        try:
            await q.delete_message()
        except Exception:
            pass
            
        await _generate_document_from_parsed_data(update, context, doc_data)
        context.user_data.pop("pending_doc", None)
        context.user_data.pop("active_menu", None)
        
    elif data == "coti_build_cancel":
        doc_type = pending_doc.get("type", "documento")
        title_up = "COTIZACIÓN" if doc_type == "cotizacion" else "NOTA DE ENTREGA"
        
        try:
            await q.delete_message()
        except Exception:
            pass
            
        context.user_data.pop("pending_doc", None)
        context.user_data.pop("active_menu", None)
        await msg.reply_text(f"❌ Flujo de creación de *{title_up}* cancelado.", parse_mode="Markdown")


async def handle_cotizaciones_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
) -> None:
    q = update.callback_query
    if not q:
        return
        
    if not _allowed(update):
        await q.answer("❌ Acceso no autorizado / Suscripción Expirada.", show_alert=True)
        return
    if not _check_permission(update, "cotizaciones"):
        await q.answer("❌ No tienes privilegios para realizar operaciones de Cotizaciones.", show_alert=True)
        return
        
    await q.answer()
    data = (q.data or "").strip()
    msg = q.message
    if not msg:
        return
        
    pending_doc = context.user_data.get("pending_doc")
    if not pending_doc or "parsed_data" not in pending_doc:
        await msg.reply_text("❌ No hay ningún documento pendiente de procesar. Inicia el flujo nuevamente.")
        return
        
    currency = "ves" if "ves" in data else "usd"
    doc_data = pending_doc["parsed_data"]
    doc_data["currency"] = currency
    
    # Asignar correlativo automático secuencial
    doc_type = doc_data["docType"]
    user_id = update.effective_user.id if update.effective_user else None
    correlativo = _get_and_increment_correlativo(doc_type, user_id)
    doc_data["docNumber"] = correlativo
    
    # Eliminar mensaje de espera
    try:
        await q.delete_message()
    except Exception:
        pass
        
    await _generate_document_from_parsed_data(update, context, doc_data)
    context.user_data.pop("pending_doc", None)


async def tributos_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not _allowed(update):
        await _deny(update)
        return
        
    if not _check_permission(update, "tributos"):
        await update.effective_message.reply_text("❌ No tienes privilegios para acceder al módulo de Tributos.")
        return
    today = date.today()
    fortnight = 1 if today.day <= 15 else 2
    report = tributario_engine.get_compromiso_tributario_report(today.year, today.month, fortnight)
    text = format_tributos_report(report)
    kb = _tributos_keyboard(today.year, today.month, fortnight, _generate_short_summary(report))
    
    msg = update.effective_message
    if msg:
        await msg.reply_text(text, reply_markup=kb, parse_mode="Markdown")


async def handle_tributos_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
) -> None:
    q = update.callback_query
    if not q:
        return
        
    if not _allowed(update):
        await q.answer("❌ Acceso no autorizado / Suscripción Expirada.", show_alert=True)
        return
    if not _check_permission(update, "tributos"):
        await q.answer("❌ No tienes privilegios para realizar operaciones de Tributos.", show_alert=True)
        return
        
    await q.answer()
    data = (q.data or "").strip()
    msg = q.message
    if not msg:
        return
        
    if data.startswith("delete_ret_sel:"):
        ret_type = data.split(":", 1)[1]
        context.user_data["awaiting_delete_ret_num"] = ret_type
        
        sample = "20260600000381" if ret_type == "iva" else "20260600000002"
        lbl = "IVA" if ret_type == "iva" else "ISLR"
        
        await msg.edit_text(
            f"✏️ *Eliminar Retención de {lbl}*\n\n"
            f"Por favor, escribe y envía el número de comprobante que deseas eliminar (ej: `{sample}`):",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancelar", callback_data="delete_ret_cancel")]])
        )
        return
        
    elif data == "delete_ret_cancel":
        context.user_data.pop("awaiting_delete_ret_num", None)
        await msg.edit_text(
            "❌ Operación cancelada. El comprobante no ha sido modificado."
        )
        return
        
    elif data.startswith("delete_ret_confirm:"):
        parts = data.split(":")
        if len(parts) < 3:
            return
        ret_type = parts[1]
        num_comp = parts[2]
        
        ctx = _get_company_context(update)
        
        if ret_type == "iva":
            dup_info = excel_store.check_retencion_emitida_exists(ctx.retenciones_emitidas_dir, num_comp)
            if dup_info:
                path, row_idx, _ = dup_info
                excel_store.delete_retencion_emitida_row(path, row_idx)
                logger.info(f"Admin deleted IVA retention {num_comp} from {path} row {row_idx}")
                await msg.edit_text(
                    f"✅ *Comprobante de IVA Nro {num_comp} eliminado con éxito.*\n\n"
                    f"Cualquier factura asociada ha sido devuelta al listado de pendientes.",
                    parse_mode="Markdown"
                )
            else:
                await msg.edit_text(f"❌ El comprobante de IVA {num_comp} ya no existe en los registros.")
                
        elif ret_type == "islr":
            dup_info = excel_store.check_retencion_islr_exists(ctx.retenciones_islr_dir, num_comp)
            if dup_info:
                path, row_idx, _ = dup_info
                excel_store.delete_retencion_islr_row(path, row_idx)
                logger.info(f"Admin deleted ISLR retention {num_comp} from {path} row {row_idx}")
                await msg.edit_text(
                    f"✅ *Comprobante de ISLR Nro {num_comp} eliminado con éxito.*\n\n"
                    f"El registro ha sido removido de forma permanente.",
                    parse_mode="Markdown"
                )
            else:
                await msg.edit_text(f"❌ El comprobante de ISLR {num_comp} ya no existe en los registros.")
        return

    if data.startswith("tributos_period_"):
        parts = data.split("_")
        y = int(parts[2])
        m = int(parts[3])
        f = int(parts[4])
        ctx = _get_company_context(update)
        report = tributario_engine.get_compromiso_tributario_report(
            y, m, f,
            facturas_emitidas_path=ctx.facturas_emitidas_path,
            reportes_z_path=ctx.reportes_z_path,
            retenciones_emitidas_dir=ctx.retenciones_emitidas_dir,
            excel_path=ctx.excel_path,
            retenciones_islr_dir=ctx.retenciones_islr_dir
        )
        text = format_tributos_report(report)
        kb = _tributos_keyboard(y, m, f, _generate_short_summary(report))
        await msg.edit_text(text, reply_markup=kb, parse_mode="Markdown")
        
    elif data.startswith("tributos_detiva_"):
        parts = data.split("_")
        y = int(parts[2])
        m = int(parts[3])
        f = int(parts[4])
        text = _format_iva_details(y, m, f, update.effective_user.id)
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Volver al Reporte", callback_data=f"tributos_period_{y}_{m}_{f}")]])
        await msg.edit_text(text, reply_markup=kb, parse_mode="Markdown")
        
    elif data.startswith("tributos_detislr_"):
        parts = data.split("_")
        y = int(parts[2])
        m = int(parts[3])
        f = int(parts[4])
        text = _format_islr_details(y, m, f, update.effective_user.id)
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Volver al Reporte", callback_data=f"tributos_period_{y}_{m}_{f}")]])
        await msg.edit_text(text, reply_markup=kb, parse_mode="Markdown")
        
    elif data == "tributos_selmonth":
        text = "📅 *Selecciona la quincena y el mes que deseas consultar:*"
        kb = _months_selection_keyboard()
        await msg.edit_text(text, reply_markup=kb, parse_mode="Markdown")
        
    elif data.startswith("tributos_download_"):
        parts = data.split("_")
        y = int(parts[2])
        m = int(parts[3])
        f = int(parts[4])
        
        status_msg = await msg.reply_text(
            f"⏳ *Generando los 4 reportes quincenales en formato Excel...*",
            parse_mode="Markdown"
        )
        asyncio.create_task(_send_reportes_telegram_async(update, context, y, m, f, status_msg))
        
    elif data.startswith("tributos_seniattxt_"):
        parts = data.split("_")
        y = int(parts[2])
        m = int(parts[3])
        f = int(parts[4])
        
        status_msg = await msg.reply_text(
            f"⏳ *Generando archivos TXT según la normativa del SENIAT...*",
            parse_mode="Markdown"
        )
        asyncio.create_task(_send_seniat_txt_telegram_async(update, context, y, m, f, status_msg))
        
    elif data.startswith("tributos_sendemail_"):
        parts = data.split("_")
        y = int(parts[2])
        m = int(parts[3])
        f = int(parts[4])
        
        # Verificar si SMTP está configurado
        if not config.SMTP_SERVER or not config.SMTP_USER or not config.SMTP_PASSWORD:
            await msg.reply_text(
                "❌ *Función de Correo SMTP no configurada*\n\n"
                "Para poder enviar reportes directamente en segundo plano con archivos adjuntos, "
                "debes definir las siguientes variables en tu archivo `.env`:\n\n"
                "• `SMTP_SERVER` (ej: smtp.gmail.com)\n"
                "• `SMTP_PORT` (ej: 587)\n"
                "• `SMTP_USER` (tu dirección de correo)\n"
                "• `SMTP_PASSWORD` (tu contraseña de aplicación)\n\n"
                "_Mientras tanto, puedes presionar el botón '📧 Correo Rápido' de arriba para enviar desde tu propio cliente._",
                parse_mode="Markdown"
            )
            return
            
        context.user_data["pending_email_report"] = {
            "year": y,
            "month": m,
            "fortnight": f,
            "awaiting": "email_address"
        }
        
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancelar Envío", callback_data="tributos_cancelemail")]])
        default_lbl = f"`{config.DEFAULT_ACCOUNTANT_EMAIL}`" if config.DEFAULT_ACCOUNTANT_EMAIL else "_no configurado_"
        await msg.reply_text(
            "📧 *Envío de Reportes por Correo (SMTP)*\n\n"
            "Por favor, escribe la dirección de correo electrónico del destinatario a donde deseas enviar los reportes:\n\n"
            f"👉 *O escribe la palabra *`contador`* para enviar al destinatario predeterminado:* {default_lbl}",
            reply_markup=kb,
            parse_mode="Markdown"
        )
        
    elif data == "tributos_cancelemail":
        context.user_data.pop("pending_email_report", None)
        await msg.reply_text("❌ Envío de reportes por correo cancelado.")


async def _show_admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE, msg_to_edit=None) -> None:
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("👥 Listar Clientes", callback_data="admin_list")],
        [InlineKeyboardButton("➕ Registrar Cliente", callback_data="admin_new_start")],
        [InlineKeyboardButton("🔙 Volver al Inicio", callback_data="work_panel:start")]
    ])
    bot_username = context.bot.username
    request_link_sufevica = f"https://t.me/{bot_username}?start=solicitar_sufevica"
    request_link_cliente = f"https://t.me/{bot_username}?start=solicitar_cliente"
    
    text = (
        "⚙️ *Panel de Administración de Clientes y Usuarios*\n\n"
        "Selecciona una opción para gestionar suscripciones, roles y cuotas de uso:\n\n"
        "🔗 *Enlaces de Solicitud de Acceso:*\n"
        f"• *Usuarios SUFEVICA:* `{request_link_sufevica}`\n"
        f"• *Clientes (FlashTax):* `{request_link_cliente}`"
    )
    
    if msg_to_edit:
        await msg_to_edit.edit_text(text, reply_markup=kb, parse_mode="Markdown")
    else:
        await update.effective_message.reply_text(text, reply_markup=kb, parse_mode="Markdown")


async def _show_admin_user_detail(update: Update, context: ContextTypes.DEFAULT_TYPE, uid: str, msg_to_edit=None) -> None:
    u_info = user_manager.get_user(uid)
    if not u_info:
        err_text = "Usuario no encontrado."
        err_kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Volver al Listado", callback_data="admin_list")]])
        if msg_to_edit:
            await msg_to_edit.edit_text(err_text, reply_markup=err_kb)
        else:
            await update.effective_message.reply_text(err_text, reply_markup=err_kb)
        return

    # Detail Text Formatting
    status_emoji = "🟢" if u_info.get("status") == "active" else "🔴"
    status_lbl = "APROBADO" if u_info.get("status") == "active" else "BLOQUEADO"
    
    # Persistent registration date
    reg_date = u_info.get("registration_date")
    if not reg_date:
        reg_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        user_manager.update_user_field(uid, "registration_date", reg_date)

    role = u_info.get("role")
    if role == "admin":
        role_lbl = "Administrador"
    elif role == "nueva_empresa":
        role_lbl = "Propietario (Analistas: Ninguno)"
    elif role == "tributos_only":
        role_lbl = "Tributos Only"
    elif role == "cotizaciones_only":
        role_lbl = "Cotizaciones Only"
    elif role == "tributos_and_cotizaciones":
        role_lbl = "Tributos + Cotizaciones"
    else:
        role_lbl = "Acceso Total"

    limit_ops = u_info.get("limit_ops", -1)
    if limit_ops == -1:
        plan_lbl = "Plan Premium (3 meses)"
    else:
        plan_lbl = f"Plan Estándar ({limit_ops} ops)"

    exp_date = u_info.get("expiration_date", "never")
    exp_lbl = exp_date if exp_date != "never" else "Sin Límite"

    company_name = u_info.get("company_name", "(sin configurar)")
    company_rif = u_info.get("company_rif", "(sin configurar)")
    company_address = u_info.get("company_address", "(sin configurar)")
    company_phone = u_info.get("company_phone", "(sin configurar)")
    
    # Resolve actual last correlativo dynamically
    prefix = datetime.now().strftime("%Y%m")
    monthly_path = _reten_emit_monthly_path(date.today(), uid)
    max_seq_excel = excel_store.max_seq_retencion_emitida(monthly_path.parent, emission_date=date.today())
    
    max_seq_config = 0
    config_correlative = u_info.get("last_correlative", "20260600000000")
    if config_correlative and str(config_correlative).startswith(prefix):
        seq_part = str(config_correlative)[len(prefix):]
        if seq_part.isdigit():
            max_seq_config = int(seq_part)
            
    max_seq = max(max_seq_excel, max_seq_config)
    last_correlative = f"{prefix}{max_seq:08d}"
    
    next_coti = u_info.get("next_cotizacion", 1)
    next_nt = u_info.get("next_nota", 1)

    text = (
        f"Nombre: {u_info.get('name')} ({u_info.get('company_email') or 'sin correo'})\n"
        f"ID: `{uid}`\n"
        f"Estado: {status_emoji} *{status_lbl}*\n"
        f"Registrado: {reg_date}\n"
        f"👥 *Rol:* {role_lbl}\n"
        f"✅ *Membresía:* {plan_lbl}\n"
        f"📅 *Vence:* {exp_lbl}\n\n"
        f"🏢 *Agente:* {company_name}\n"
        f"🔹 *RIF:* `{company_rif}`\n"
        f"📍 *Dirección:* {company_address}\n"
        f"📞 *Teléfono:* `{company_phone}`\n"
        f"🔢 *Último correlativo:* `{last_correlative}`\n"
        f"📋 *Próxima Cotización:* `{next_coti:06d}`\n"
        f"📦 *Próxima Nota:* `{next_nt:06d}`"
    )

    status_btn_text = "🚫 BLOQUEAR CLIENTE" if u_info.get("status") == "active" else "🟢 DESBLOQUEAR CLIENTE"

    kb_list = [
        [InlineKeyboardButton(status_btn_text, callback_data=f"admin_edit_status:{uid}")],
        [InlineKeyboardButton("💳 GESTIONAR SUSCRIPCIÓN", callback_data=f"admin_sub_menu:{uid}")],
        [InlineKeyboardButton("✍️ EDITAR RAZÓN SOCIAL", callback_data=f"admin_edit_field:{uid}:company_name")],
        [InlineKeyboardButton("✍️ EDITAR RIF", callback_data=f"admin_edit_field:{uid}:company_rif")],
        [InlineKeyboardButton("✍️ EDITAR DIRECCIÓN", callback_data=f"admin_edit_field:{uid}:company_address")],
        [InlineKeyboardButton("✍️ EDITAR TELÉFONO", callback_data=f"admin_edit_field:{uid}:company_phone")],
        [InlineKeyboardButton("📸 ESCANEAR RIF (FOTO)", callback_data=f"admin_scan_rif:{uid}")],
        [InlineKeyboardButton("🖊️ CARGAR FIRMA Y SELLO", callback_data=f"admin_upload_sig:{uid}")],
        [InlineKeyboardButton("🔢 EDITAR RETENCIÓN", callback_data=f"admin_edit_field:{uid}:last_correlative")],
        [
            InlineKeyboardButton("📋 EDITAR COTIZACIÓN", callback_data=f"admin_edit_field:{uid}:next_cotizacion"),
            InlineKeyboardButton("📦 EDITAR NOTA", callback_data=f"admin_edit_field:{uid}:next_nota")
        ],
        [InlineKeyboardButton("🔍 REENVIAR COMPROBANTE", callback_data=f"admin_resend_auth:{uid}")],
        [InlineKeyboardButton("🔙 VOLVER A CLIENTES", callback_data="admin_list")]
    ]

    kb = InlineKeyboardMarkup(kb_list)
    if msg_to_edit:
        await msg_to_edit.edit_text(text, reply_markup=kb, parse_mode="Markdown")
    else:
        await update.effective_message.reply_text(text, reply_markup=kb, parse_mode="Markdown")


async def handle_admin_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    q = update.callback_query
    if not q:
        return
    await q.answer()
    data = (q.data or "").strip()
    msg = q.message
    if not msg:
        return
        
    user_id = update.effective_user.id
    user = user_manager.get_user(user_id)
    if not user or user.get("role") != "admin":
        await q.answer("❌ No autorizado.", show_alert=True)
        return
        
    if data == "admin_main":
        await _show_admin_panel(update, context, msg_to_edit=msg)
        return
        
    elif data.startswith("admin_req_"):
        parts = data.split(":")
        action = parts[0]
        target_uid = parts[1]
        
        if action == "admin_req_start":
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("📅 Prueba (5 días)", callback_data=f"admin_req_plan:{target_uid}:5")],
                [InlineKeyboardButton("📅 Plan Estándar (30 días)", callback_data=f"admin_req_plan:{target_uid}:30")],
                [InlineKeyboardButton("📅 Plan Premium (3 meses)", callback_data=f"admin_req_plan:{target_uid}:90")],
                [InlineKeyboardButton("❌ Rechazar Solicitud", callback_data=f"admin_req_reject:{target_uid}")]
            ])
            await msg.edit_text(
                f"👤 *Autorizar Usuario (ID: {target_uid})*\n\n"
                f"Paso 1: Selecciona la duración del plan:",
                reply_markup=kb,
                parse_mode="Markdown"
            )
            return

        elif action == "admin_req_plan":
            days = parts[2]
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("🏛️ Tributos Only", callback_data=f"admin_req_role:{target_uid}:{days}:tributos_only")],
                [InlineKeyboardButton("📋 Cotizaciones Only", callback_data=f"admin_req_role:{target_uid}:{days}:cotizaciones_only")],
                [InlineKeyboardButton("🏢 Nueva Empresa (FlashTax)", callback_data=f"admin_req_role:{target_uid}:{days}:nueva_empresa")],
                [InlineKeyboardButton("⭐ Acceso Total", callback_data=f"admin_req_role:{target_uid}:{days}:full_access")],
                [InlineKeyboardButton("🔙 Atrás", callback_data=f"admin_req_start:{target_uid}")]
            ])
            plan_name = "Prueba (5 días)" if days == "5" else ("Plan Estándar (30 días)" if days == "30" else "Plan Premium (3 meses)")
            await msg.edit_text(
                f"👤 *Autorizar Usuario (ID: {target_uid})*\n\n"
                f"• *Plan seleccionado:* {plan_name}\n\n"
                f"Paso 2: Selecciona los privilegios (rol):",
                reply_markup=kb,
                parse_mode="Markdown"
            )
            return

        elif action == "admin_req_role":
            import html
            days = int(parts[2])
            role = parts[3]
            
            try:
                chat = await context.bot.get_chat(int(target_uid))
                name = f"{chat.first_name} {chat.last_name or ''}".strip()
                username_str = f"@{chat.username}" if chat.username else "sin username"
            except Exception:
                name = "Usuario Solicitante"
                username_str = "desconocido"

            # Calcular fecha de vencimiento
            exp_date = (date.today() + timedelta(days=days)).strftime("%Y-%m-%d")
            
            # Registrar usuario en la base de datos
            user_manager.register_user(
                user_id=target_uid,
                name=name,
                role=role,
                expiration_date=exp_date,
                limit_ops=-1
            )

            role_lbl = "Tributos Only" if role == "tributos_only" else ("Cotizaciones Only" if role == "cotizaciones_only" else "Tributos + Cotizaciones" if role == "tributos_and_cotizaciones" else "Nueva Empresa (FlashTax)" if role == "nueva_empresa" else "Acceso Total")
            plan_lbl = "Prueba (5 días)" if days == 5 else ("Plan Estándar (30 días)" if days == 30 else "Plan Premium (3 meses)")

            name_escaped = html.escape(name)
            username_escaped = html.escape(username_str)
            await msg.edit_text(
                f"✅ <b>¡Usuario Autorizado con Éxito!</b>\n\n"
                f"• <b>Nombre:</b> {name_escaped} ({username_escaped})\n"
                f"• <b>ID:</b> <code>{target_uid}</code>\n"
                f"• <b>Plan:</b> {plan_lbl}\n"
                f"• <b>Rol:</b> {role_lbl}\n"
                f"• <b>Vencimiento:</b> <code>{exp_date}</code>",
                parse_mode="HTML"
            )

            # Notificar al usuario
            try:
                await context.bot.send_message(
                    chat_id=int(target_uid),
                    text=f"🎉 *¡Tu solicitud de acceso ha sido aprobada!*\n\n"
                         f"El administrador te ha concedido acceso con los siguientes detalles:\n\n"
                         f"• *Plan:* {plan_lbl}\n"
                         f"• *Vencimiento:* `{exp_date}`\n"
                         f"• *Privilegios:* {role_lbl}\n\n"
                         f"Presiona /start para activar el menú y comenzar a usar el bot.",
                    parse_mode="Markdown",
                    reply_markup=_main_keyboard(target_uid)
                )
            except Exception as e:
                logger.error(f"No se pudo notificar al usuario {target_uid}: {e}")
            return

        elif action == "admin_req_sufevica_start":
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("⬜ 🏛️ Tributos (Ver/Reporte)", callback_data=f"admin_req_sufevica_toggle:{target_uid}:1:0")],
                [InlineKeyboardButton("⬜ 📋 Cotizaciones", callback_data=f"admin_req_sufevica_toggle:{target_uid}:0:1")],
                [InlineKeyboardButton("🟢 Continuar", callback_data=f"admin_req_sufevica_confirm:{target_uid}:0:0")],
                [InlineKeyboardButton("❌ Rechazar Solicitud", callback_data=f"admin_req_reject:{target_uid}")]
            ])
            await msg.edit_text(
                f"🛡️ *Autorizar Usuario SUFEVICA (ID: {target_uid})*\n\n"
                f"El tiempo de uso es indefinido.\n\n"
                f"Selecciona las opciones del rol para el usuario y presiona Continuar:",
                reply_markup=kb,
                parse_mode="Markdown"
            )
            return

        elif action == "admin_req_sufevica_toggle":
            target_uid = parts[1]
            tributos_val = int(parts[2])
            cotizaciones_val = int(parts[3])
            
            trib_emoji = "✅" if tributos_val else "⬜"
            coti_emoji = "✅" if cotizaciones_val else "⬜"
            
            next_trib_val = 0 if tributos_val else 1
            next_coti_val = 0 if cotizaciones_val else 1
            
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton(f"{trib_emoji} 🏛️ Tributos (Ver/Reporte)", callback_data=f"admin_req_sufevica_toggle:{target_uid}:{next_trib_val}:{cotizaciones_val}")],
                [InlineKeyboardButton(f"{coti_emoji} 📋 Cotizaciones", callback_data=f"admin_req_sufevica_toggle:{target_uid}:{tributos_val}:{next_coti_val}")],
                [InlineKeyboardButton("🟢 Continuar", callback_data=f"admin_req_sufevica_confirm:{target_uid}:{tributos_val}:{cotizaciones_val}")],
                [InlineKeyboardButton("❌ Rechazar Solicitud", callback_data=f"admin_req_reject:{target_uid}")]
            ])
            
            await msg.edit_text(
                f"🛡️ *Autorizar Usuario SUFEVICA (ID: {target_uid})*\n\n"
                f"El tiempo de uso es indefinido.\n\n"
                f"Selecciona las opciones del rol para el usuario y presiona Continuar:",
                reply_markup=kb,
                parse_mode="Markdown"
            )
            return

        elif action == "admin_req_sufevica_confirm":
            import html
            target_uid = parts[1]
            tributos_val = int(parts[2])
            cotizaciones_val = int(parts[3])
            
            if not tributos_val and not cotizaciones_val:
                await q.answer("⚠️ Selecciona al menos una opción para continuar.", show_alert=True)
                return
                
            if tributos_val and cotizaciones_val:
                role = "tributos_and_cotizaciones"
            elif tributos_val:
                role = "tributos_only"
            else:
                role = "cotizaciones_only"
                
            try:
                chat = await context.bot.get_chat(int(target_uid))
                name = f"{chat.first_name} {chat.last_name or ''}".strip()
                username_str = f"@{chat.username}" if chat.username else "sin username"
            except Exception:
                name = "Usuario Solicitante"
                username_str = "desconocido"

            user_manager.register_user(
                user_id=target_uid,
                name=name,
                role=role,
                expiration_date="never",
                limit_ops=-1
            )

            role_lbl = "Tributos Only" if role == "tributos_only" else ("Cotizaciones Only" if role == "cotizaciones_only" else "Tributos + Cotizaciones")
            name_escaped = html.escape(name)
            username_escaped = html.escape(username_str)
            await msg.edit_text(
                f"✅ <b>¡Usuario SUFEVICA Autorizado con Éxito!</b>\n\n"
                f"• <b>Nombre:</b> {name_escaped} ({username_escaped})\n"
                f"• <b>ID:</b> <code>{target_uid}</code>\n"
                f"• <b>Plan / Duración:</b> Indefinido (Sin límite)\n"
                f"• <b>Rol:</b> {role_lbl}",
                parse_mode="HTML"
            )

            try:
                await context.bot.send_message(
                    chat_id=int(target_uid),
                    text=f"🎉 *¡Tu solicitud de acceso como SUFEVICA ha sido aprobada!*\n\n"
                         f"El administrador te ha concedido acceso con los siguientes detalles:\n\n"
                         f"• *Privilegios:* {role_lbl}\n"
                         f"• *Vencimiento:* Indefinido (Sin límite)\n\n"
                         f"Presiona /start para activar el menú y comenzar a usar el bot.",
                    parse_mode="Markdown",
                    reply_markup=_main_keyboard(target_uid)
                )
            except Exception as e:
                logger.error(f"No se pudo notificar al usuario {target_uid}: {e}")
            return

        elif action == "admin_req_cliente_start":
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("📅 Prueba (5 días)", callback_data=f"admin_req_cliente_plan:{target_uid}:5")],
                [InlineKeyboardButton("📅 Plan Estándar (30 días)", callback_data=f"admin_req_cliente_plan:{target_uid}:30")],
                [InlineKeyboardButton("📅 Plan Premium (3 meses)", callback_data=f"admin_req_cliente_plan:{target_uid}:90")],
                [InlineKeyboardButton("❌ Rechazar Solicitud", callback_data=f"admin_req_reject:{target_uid}")]
            ])
            await msg.edit_text(
                f"🏢 *Autorizar Cliente (ID: {target_uid})*\n\n"
                f"El rol asignado será automáticamente 'Nueva Empresa' (Propietario).\n\n"
                f"Selecciona la duración del plan del cliente:",
                reply_markup=kb,
                parse_mode="Markdown"
            )
            return

        elif action == "admin_req_cliente_plan":
            import html
            days = int(parts[2])
            try:
                chat = await context.bot.get_chat(int(target_uid))
                name = f"{chat.first_name} {chat.last_name or ''}".strip()
                username_str = f"@{chat.username}" if chat.username else "sin username"
            except Exception:
                name = "Usuario Solicitante"
                username_str = "desconocido"

            exp_date = (date.today() + timedelta(days=days)).strftime("%Y-%m-%d")
            
            user_manager.register_user(
                user_id=target_uid,
                name=name,
                role="nueva_empresa",
                expiration_date=exp_date,
                limit_ops=-1
            )

            plan_lbl = "Prueba (5 días)" if days == 5 else ("Plan Estándar (30 días)" if days == 30 else "Plan Premium (3 meses)")
            name_escaped = html.escape(name)
            username_escaped = html.escape(username_str)
            await msg.edit_text(
                f"✅ <b>¡Cliente Autorizado con Éxito!</b>\n\n"
                f"• <b>Nombre:</b> {name_escaped} ({username_escaped})\n"
                f"• <b>ID:</b> <code>{target_uid}</code>\n"
                f"• <b>Plan:</b> {plan_lbl}\n"
                f"• <b>Rol:</b> Propietario (FlashTax)\n"
                f"• <b>Vencimiento:</b> <code>{exp_date}</code>",
                parse_mode="HTML"
            )

            try:
                await context.bot.send_message(
                    chat_id=int(target_uid),
                    text=f"🎉 *¡Tu solicitud de acceso como Cliente ha sido aprobada!*\n\n"
                         f"El administrador te ha concedido acceso con los siguientes detalles:\n\n"
                         f"• *Plan:* {plan_lbl}\n"
                         f"• *Vencimiento:* `{exp_date}`\n"
                         f"• *Privilegios:* Propietario (FlashTax)\n\n"
                         f"Presiona /start para activar el menú y comenzar a usar el bot.",
                    parse_mode="Markdown",
                    reply_markup=_main_keyboard(target_uid)
                )
            except Exception as e:
                logger.error(f"No se pudo notificar al usuario {target_uid}: {e}")
            return

        elif action == "admin_req_reject":
            await msg.edit_text(
                f"❌ *Solicitud Rechazada*\n\nSe ha denegado la solicitud de acceso para el ID `{target_uid}`.",
                parse_mode="Markdown"
            )
            try:
                await context.bot.send_message(
                    chat_id=int(target_uid),
                    text="❌ *Solicitud Denegada*\n\nTu solicitud para acceder al bot ha sido rechazada por el administrador.",
                    parse_mode="Markdown"
                )
            except Exception as e:
                logger.error(f"No se pudo enviar notificación de rechazo al usuario {target_uid}: {e}")
            return

    elif data == "admin_close":
        try:
            await msg.delete()
        except Exception:
            await msg.edit_text("⚙️ Panel de administración cerrado.")
        return
        
    elif data == "admin_list":
        users_data = user_manager.load_users().get("users", {})
        if not users_data:
            await msg.edit_text(
                "No hay usuarios registrados.",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Volver", callback_data="admin_main")]])
            )
            return
            
        text = "👥 *Usuarios Registrados:*\n\nSelecciona un usuario para ver detalles y editarlo:"
        kb_list = []
        for uid, u_info in users_data.items():
            role_emoji = "🛡️" if u_info.get("role") == "admin" else "👤"
            status_emoji = "✅" if u_info.get("status") == "active" else "🚫"
            btn_text = f"{status_emoji} {role_emoji} {u_info.get('name')} ({uid})"
            kb_list.append([InlineKeyboardButton(btn_text, callback_data=f"admin_edit:{uid}")])
            
        kb_list.append([InlineKeyboardButton("🔙 Volver", callback_data="admin_main")])
        await msg.edit_text(text, reply_markup=InlineKeyboardMarkup(kb_list), parse_mode="Markdown")
        return
        
    elif data.startswith("admin_edit:"):
        uid = data.split(":")[1]
        await _show_admin_user_detail(update, context, uid, msg_to_edit=msg)
        return

    elif data.startswith("admin_sub_menu:"):
        uid = data.split(":")[1]
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("🔄 Renovar plan / Duración", callback_data=f"admin_edit_dur:{uid}")],
            [InlineKeyboardButton("🛡️ Cambiar Rol", callback_data=f"admin_edit_role:{uid}")],
            [InlineKeyboardButton("📊 Modificar Cuota", callback_data=f"admin_edit_quota:{uid}")],
            [InlineKeyboardButton("🔙 Volver al Detalle", callback_data=f"admin_edit:{uid}")]
        ])
        await msg.edit_text("💳 *Gestión de Suscripción de Usuario*\n\nSeleccione el campo a modificar:", reply_markup=kb, parse_mode="Markdown")
        return

    elif data.startswith("admin_edit_field:"):
        parts = data.split(":")
        uid = parts[1]
        field = parts[2]
        field_lbl = (
            "Razón Social" if field == "company_name" 
            else "RIF" if field == "company_rif" 
            else "Dirección Fiscal" if field == "company_address" 
            else "Teléfono" if field == "company_phone" 
            else "Siguiente Cotización" if field == "next_cotizacion"
            else "Siguiente Nota de Entrega" if field == "next_nota"
            else "Último Correlativo"
        )
        
        context.user_data["admin_edit_target_uid"] = uid
        context.user_data["admin_edit_field"] = field
        context.user_data["admin_state"] = "awaiting_admin_field_edit"
        
        await msg.edit_text(
            f"📝 *Editar {field_lbl}*\n\n"
            f"Escribe el nuevo valor para {field_lbl}:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancelar", callback_data=f"admin_edit:{uid}")]])
        )
        return

    elif data.startswith("admin_scan_rif:"):
        uid = data.split(":")[1]
        context.user_data["admin_edit_target_uid"] = uid
        context.user_data["admin_state"] = "awaiting_admin_rif_photo"
        await msg.edit_text(
            "📸 *Escanear RIF de Cliente*\n\n"
            "Por favor, envía la *foto o imagen del RIF* del cliente para extraer la Razón Social y RIF usando Gemini:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancelar", callback_data=f"admin_edit:{uid}")]])
        )
        return

    elif data.startswith("admin_upload_sig:"):
        uid = data.split(":")[1]
        context.user_data["admin_edit_target_uid"] = uid
        context.user_data["awaiting_admin_company_signature"] = True
        await msg.edit_text(
            "🖊️ *Cargar Firma y Sello de Cliente*\n\n"
            "Por favor, envía la *imagen o archivo PDF/PNG* de la firma y sello del cliente:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancelar", callback_data=f"admin_edit:{uid}")]])
        )
        return

    elif data.startswith("admin_resend_auth:"):
        uid = data.split(":")[1]
        u_info = user_manager.get_user(uid)
        if u_info:
            role = u_info.get("role")
            role_lbl = "Tributos Only" if role == "tributos_only" else ("Cotizaciones Only" if role == "cotizaciones_only" else "Tributos + Cotizaciones" if role == "tributos_and_cotizaciones" else "Nueva Empresa (FlashTax)" if role == "nueva_empresa" else "Acceso Total")
            limit_ops = u_info.get("limit_ops", -1)
            plan_lbl = "Plan Estándar (100 ops)" if limit_ops == 100 else "Plan Premium" if limit_ops == -1 else "Prueba"
            exp_date = u_info.get("expiration_date", "never")
            try:
                await context.bot.send_message(
                    chat_id=int(uid),
                    text=f"🎉 *¡Tu suscripción y acceso al bot están activos!*\n\n"
                         f"Detalles de tu cuenta:\n\n"
                         f"• *Plan:* {plan_lbl}\n"
                         f"• *Vencimiento:* `{exp_date}`\n"
                         f"• *Privilegios:* {role_lbl}\n\n"
                         f"Presiona /start para activar el menú y comenzar a usar el bot.",
                    reply_markup=_main_keyboard(uid),
                    parse_mode="Markdown"
                )
                await q.answer("✅ Comprobante reenviado con éxito al usuario.", show_alert=True)
            except Exception as e:
                logger.error(f"Error reenviando comprobante al usuario {uid}: {e}")
                await q.answer(f"❌ Error al enviar mensaje: {e}", show_alert=True)
        return
        
    elif data.startswith("admin_edit_status:"):
        uid = data.split(":")[1]
        u_info = user_manager.get_user(uid)
        if u_info:
            new_status = "suspended" if u_info.get("status") == "active" else "active"
            user_manager.update_user_status(uid, new_status)
            await q.answer(f"Estado actualizado a {new_status}")
            q.data = f"admin_edit:{uid}"
            await handle_admin_callback(update, context)
        return
        
    elif data.startswith("admin_edit_role:"):
        uid = data.split(":")[1]
        u_info = user_manager.get_user(uid)
        if not u_info:
            await q.answer("❌ Usuario no encontrado.", show_alert=True)
            return
            
        current_role = u_info.get("role")
        tributos_val = 1 if current_role in ("tributos_only", "tributos_and_cotizaciones", "full_access", "nueva_empresa") else 0
        cotizaciones_val = 1 if current_role in ("cotizaciones_only", "tributos_and_cotizaciones", "full_access", "nueva_empresa") else 0
        
        trib_emoji = "✅" if tributos_val else "⬜"
        coti_emoji = "✅" if cotizaciones_val else "⬜"
        
        next_trib_val = 0 if tributos_val else 1
        next_coti_val = 0 if cotizaciones_val else 1
        
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton(f"{trib_emoji} 🏛️ Tributos (Ver/Reporte)", callback_data=f"admin_edit_role_toggle:{uid}:{next_trib_val}:{cotizaciones_val}")],
            [InlineKeyboardButton(f"{coti_emoji} 📋 Cotizaciones", callback_data=f"admin_edit_role_toggle:{uid}:{tributos_val}:{next_coti_val}")],
            [InlineKeyboardButton("🟢 Continuar", callback_data=f"admin_edit_role_confirm:{uid}:{tributos_val}:{cotizaciones_val}")],
            [InlineKeyboardButton("🔙 Atrás", callback_data=f"admin_edit:{uid}")]
        ])
        await msg.edit_text(
            f"🔄 *Modificar Rol del Usuario {uid}*\n\n"
            f"Selecciona las opciones del rol para el usuario y presiona Continuar:",
            reply_markup=kb,
            parse_mode="Markdown"
        )
        return

    elif data.startswith("admin_edit_role_toggle:"):
        parts = data.split(":")
        uid = parts[1]
        tributos_val = int(parts[2])
        cotizaciones_val = int(parts[3])
        
        trib_emoji = "✅" if tributos_val else "⬜"
        coti_emoji = "✅" if cotizaciones_val else "⬜"
        
        next_trib_val = 0 if tributos_val else 1
        next_coti_val = 0 if cotizaciones_val else 1
        
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton(f"{trib_emoji} 🏛️ Tributos (Ver/Reporte)", callback_data=f"admin_edit_role_toggle:{uid}:{next_trib_val}:{cotizaciones_val}")],
            [InlineKeyboardButton(f"{coti_emoji} 📋 Cotizaciones", callback_data=f"admin_edit_role_toggle:{uid}:{tributos_val}:{next_coti_val}")],
            [InlineKeyboardButton("🟢 Continuar", callback_data=f"admin_edit_role_confirm:{uid}:{tributos_val}:{cotizaciones_val}")],
            [InlineKeyboardButton("🔙 Atrás", callback_data=f"admin_edit:{uid}")]
        ])
        await msg.edit_text(
            f"🔄 *Modificar Rol del Usuario {uid}*\n\n"
            f"Selecciona las opciones del rol para el usuario y presiona Continuar:",
            reply_markup=kb,
            parse_mode="Markdown"
        )
        return

    elif data.startswith("admin_edit_role_confirm:"):
        parts = data.split(":")
        uid = parts[1]
        tributos_val = int(parts[2])
        cotizaciones_val = int(parts[3])
        
        if not tributos_val and not cotizaciones_val:
            await q.answer("⚠️ Selecciona al menos una opción para continuar.", show_alert=True)
            return
            
        if tributos_val and cotizaciones_val:
            new_role = "tributos_and_cotizaciones"
        elif tributos_val:
            new_role = "tributos_only"
        else:
            new_role = "cotizaciones_only"
            
        user_manager.update_user_field(uid, "role", new_role)
        await q.answer(f"Rol cambiado a {new_role}")
        
        # Redirigir de vuelta al detalle
        await _show_admin_user_detail(update, context, uid, msg_to_edit=msg)
        return
        
    elif data.startswith("admin_edit_dur:"):
        uid = data.split(":")[1]
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("📅 +30 días", callback_data=f"admin_edit_dur_set:{uid}:30")],
            [InlineKeyboardButton("📅 +90 días", callback_data=f"admin_edit_dur_set:{uid}:90")],
            [InlineKeyboardButton("📅 +365 días", callback_data=f"admin_edit_dur_set:{uid}:365")],
            [InlineKeyboardButton("📅 Sin Límite (never)", callback_data=f"admin_edit_dur_set:{uid}:never")],
            [InlineKeyboardButton("🔙 Atrás", callback_data=f"admin_edit:{uid}")]
        ])
        await msg.edit_text("📅 *Selecciona vigencia de la suscripción:*", reply_markup=kb, parse_mode="Markdown")
        return
        
    elif data.startswith("admin_edit_dur_set:"):
        parts = data.split(":")
        uid = parts[1]
        days_str = parts[2]
        if days_str == "never":
            exp_date = "never"
        else:
            days = int(days_str)
            exp_date = (date.today() + timedelta(days=days)).strftime("%Y-%m-%d")
            
        user_manager.update_user_field(uid, "expiration_date", exp_date)
        user_manager.update_user_field(uid, "status", "active")
        await q.answer(f"Vencimiento establecido en {exp_date}")
        q.data = f"admin_edit:{uid}"
        await handle_admin_callback(update, context)
        return
        
    elif data.startswith("admin_edit_quota:"):
        uid = data.split(":")[1]
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("📊 50 operaciones", callback_data=f"admin_edit_quota_set:{uid}:50")],
            [InlineKeyboardButton("📊 100 operaciones", callback_data=f"admin_edit_quota_set:{uid}:100")],
            [InlineKeyboardButton("📊 Ilimitado", callback_data=f"admin_edit_quota_set:{uid}:-1")],
            [InlineKeyboardButton("🔄 Resetear consumo a 0", callback_data=f"admin_edit_quota_set:{uid}:reset")],
            [InlineKeyboardButton("🔙 Atrás", callback_data=f"admin_edit:{uid}")]
        ])
        await msg.edit_text("📊 *Selecciona límite mensual de operaciones:*", reply_markup=kb, parse_mode="Markdown")
        return
        
    elif data.startswith("admin_edit_quota_set:"):
        parts = data.split(":")
        uid = parts[1]
        val = parts[2]
        if val == "reset":
            user_manager.update_user_field(uid, "consumed_ops", 0)
            await q.answer("Consumo reseteado a 0")
        else:
            limit = int(val)
            user_manager.update_user_field(uid, "limit_ops", limit)
            await q.answer(f"Límite establecido a {limit}")
            
        q.data = f"admin_edit:{uid}"
        await handle_admin_callback(update, context)
        return
        
    elif data == "admin_new_start":
        context.user_data["admin_state"] = "awaiting_new_user_id"
        context.user_data["admin_new_user"] = {}
        await msg.edit_text(
            "➕ *Registrar Nuevo Cliente*\n\n"
            "Escribe el *ID de Telegram* del nuevo cliente (ej. `123456789`):",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancelar", callback_data="admin_main")]]),
            parse_mode="Markdown"
        )
        return
        
    elif data.startswith("admin_new_role:"):
        role = data.split(":")[1]
        context.user_data["admin_new_user"]["role"] = role
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("📅 30 días", callback_data="admin_new_dur:30")],
            [InlineKeyboardButton("📅 90 días", callback_data="admin_new_dur:90")],
            [InlineKeyboardButton("📅 365 días", callback_data="admin_new_dur:365")],
            [InlineKeyboardButton("📅 Sin Límite (never)", callback_data="admin_new_dur:never")],
            [InlineKeyboardButton("❌ Cancelar", callback_data="admin_main")]
        ])
        await msg.edit_text("📅 *Selecciona vigencia para el nuevo usuario:*", reply_markup=kb, parse_mode="Markdown")
        return
        
    elif data.startswith("admin_new_dur:"):
        days_str = data.split(":")[1]
        if days_str == "never":
            exp_date = "never"
        else:
            days = int(days_str)
            exp_date = (date.today() + timedelta(days=days)).strftime("%Y-%m-%d")
            
        context.user_data["admin_new_user"]["expiration_date"] = exp_date
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("📊 50 operaciones", callback_data="admin_new_quota:50")],
            [InlineKeyboardButton("📊 100 operaciones", callback_data="admin_new_quota:100")],
            [InlineKeyboardButton("📊 Ilimitado", callback_data="admin_new_quota:-1")],
            [InlineKeyboardButton("❌ Cancelar", callback_data="admin_main")]
        ])
        await msg.edit_text("📊 *Selecciona cuota mensual para el nuevo usuario:*", reply_markup=kb, parse_mode="Markdown")
        return
        
    elif data.startswith("admin_new_quota:"):
        limit = int(data.split(":")[1])
        new_u = context.user_data.get("admin_new_user")
        if not new_u or "id" not in new_u:
            await msg.edit_text("Error en el flujo de registro. Inténtalo de nuevo.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Volver", callback_data="admin_main")]]))
            return
            
        user_manager.register_user(
            user_id=new_u["id"],
            name=new_u["name"],
            role=new_u["role"],
            expiration_date=new_u["expiration_date"],
            limit_ops=limit
        )
        
        context.user_data.pop("admin_state", None)
        context.user_data.pop("admin_new_user", None)
        
        role_lbl = "Cliente (FlashTax)" if new_u['role'] == "nueva_empresa" else ("Tributos Only" if new_u['role'] == "tributos_only" else "Cotizaciones Only" if new_u['role'] == "cotizaciones_only" else "Tributos + Cotizaciones" if new_u['role'] == "tributos_and_cotizaciones" else "Acceso Total")
        await msg.edit_text(
            f"✅ *Cliente `{new_u['name']}` registrado con éxito!*\n\n"
            f"• ID: `{new_u['id']}`\n"
            f"• Rol: `{role_lbl}`\n"
            f"• Expiración: `{new_u['expiration_date']}`\n"
            f"• Límite ops: `{limit if limit != -1 else 'Ilimitado'}`",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Ir al Panel Principal", callback_data="admin_main")]]),
            parse_mode="Markdown"
        )
        return


async def admin_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not _allowed(update):
        await _deny(update)
        return
    user = user_manager.get_user(update.effective_user.id)
    if not user or user.get("role") != "admin":
        await update.effective_message.reply_text("❌ No tienes privilegios de administrador.")
        return
    await _show_admin_panel(update, context)


async def _generate_purchases_loaded_report(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    date_from: date,
    date_to: date,
    status_msg,
) -> None:
    msg = update.effective_message
    if not msg:
        return
        
    ctx = _get_company_context(update)
    path = ctx.facturas_recibidas_path
    if not path.exists():
        await status_msg.edit_text("⚠️ No hay facturas de compra registradas en el sistema.")
        return
        
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = excel_store._headers_index(ws)
        
        rows: list[list[object]] = []
        count = 0
        total_subtotal = Decimal("0")
        total_iva = Decimal("0")
        total_monto = Decimal("0")
        
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
            if not row:
                continue
            f_cell = excel_store._cell(row, headers, "Fecha_emision", None)
            f_doc = excel_store._parse_fecha_cell(f_cell)
            if f_doc is None or f_doc < date_from or f_doc > date_to:
                continue
                
            num_doc = str(excel_store._cell(row, headers, "Numero_documento", "-"))
            prov = str(excel_store._cell(row, headers, "Proveedor", "-"))
            rif = str(excel_store._cell(row, headers, "Proveedor_RIF", "-"))
            
            sub = excel_store._parse_monto_cell(excel_store._cell(row, headers, "Subtotal", None)) or Decimal("0")
            exento = excel_store._parse_monto_cell(excel_store._cell(row, headers, "Monto_exento", None)) or Decimal("0")
            base = excel_store._parse_monto_cell(excel_store._cell(row, headers, "Base_imponible", None)) or Decimal("0")
            iva = excel_store._parse_monto_cell(excel_store._cell(row, headers, "Monto_iva", None)) or Decimal("0")
            tot = excel_store._parse_monto_cell(excel_store._cell(row, headers, "Total", None)) or Decimal("0")
            
            count += 1
            total_subtotal += sub
            total_iva += iva
            total_monto += tot
            
            rows.append([
                count,
                f_doc.strftime("%d/%m/%Y"),
                num_doc,
                prov,
                rif,
                float(sub),
                float(exento),
                float(base),
                float(iva),
                float(tot)
            ])
            
        wb.close()
        
        if count == 0:
            await status_msg.edit_text(f"No se encontraron facturas cargadas entre el {date_from.strftime('%d/%m/%Y')} y el {date_to.strftime('%d/%m/%Y')}.")
            return
            
        period_str = f"Del {date_from.strftime('%d/%m/%Y')} al {date_to.strftime('%d/%m/%Y')}"
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            report_xls_path = Path(tmp.name)
            
        excel_store.generate_premium_report_excel(
            report_xls_path,
            title="Reporte de Facturas Recibidas Cargadas",
            period_str=period_str,
            headers=["#", "Fecha Emisión", "Nro Documento", "Proveedor", "RIF", "Subtotal", "Exento", "Base Imponible", "IVA", "Total"],
            rows=rows,
            numeric_cols=[5, 6, 7, 8, 9],
            sum_cols=[5, 6, 7, 8, 9],
            emisor_nombre=ctx.company_name,
            emisor_rif=ctx.company_rif,
            emisor_telefono=ctx.company_phone or "No definido",
            emisor_direccion=ctx.company_address or "No definida"
        )
        
        renamed_xls = report_xls_path.with_name(f"COMPRAS-CARGADAS-{date_from.strftime('%d%m%Y')}-AL-{date_to.strftime('%d%m%Y')}.xlsx")
        report_xls_path.rename(renamed_xls)
        
        summary_text = (
            f"📥 *Reporte de Facturas Cargadas Generado*\n\n"
            f"📅 *Período:* {period_str}\n"
            f"📄 *Total documentos:* `{count}`\n"
            f"💰 *Total Subtotal:* `{excel_store._format_monto_ves(total_subtotal)}` Bs\n"
            f"💸 *Total IVA:* `{excel_store._format_monto_ves(total_iva)}` Bs\n"
            f"🛍️ *Total Facturado:* `{excel_store._format_monto_ves(total_monto)}` Bs\n"
        )
        
        await status_msg.delete()
        await msg.reply_document(
            document=str(renamed_xls),
            filename=renamed_xls.name,
            caption=summary_text,
            parse_mode="Markdown"
        )
        renamed_xls.unlink(missing_ok=True)
        
    except Exception as e:
        logger.exception("Error al generar reporte de compras cargadas")
        await status_msg.edit_text(f"❌ Ocurrió un error al generar el reporte: {e!s}")


async def _generate_pending_withholdings_report(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    status_msg,
) -> None:
    msg = update.effective_message
    if not msg:
        return
        
    ctx = _get_company_context(update)
    path = ctx.facturas_recibidas_path
    if not path.exists():
        await status_msg.edit_text("⚠️ No hay facturas de compra registradas en el sistema.")
        return
        
    try:
        def _norm_doc(v: str) -> str:
            return re.sub(r"\s+", "", str(v or "")).strip().upper()
            
        emitted_docs = set()
        base_dir = ctx.retenciones_emitidas_dir
        if base_dir.is_dir():
            for filepath in base_dir.glob("RETEN-EMIT-*.xlsx"):
                try:
                    temp_wb = load_workbook(filepath, read_only=True, data_only=True)
                    temp_ws = temp_wb.active
                    temp_headers = excel_store._headers_index(temp_ws)
                    for r in temp_ws.iter_rows(min_row=2, values_only=True):
                        if not r:
                            continue
                        doc_cell = str(excel_store._cell(r, temp_headers, "Documentos", "") or "").strip()
                        if doc_cell:
                            for d in re.split(r'[|,\s]+', doc_cell):
                                d_norm = _norm_doc(d)
                                if d_norm:
                                    emitted_docs.add(d_norm)
                    temp_wb.close()
                except Exception as e:
                    logger.error(f"Error cargando comprobantes de {filepath}: {e}")
                    
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = excel_store._headers_index(ws)
        
        rows: list[list[object]] = []
        count = 0
        total_subtotal = Decimal("0")
        total_iva = Decimal("0")
        total_monto = Decimal("0")
        
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
            if not row:
                continue
            
            tipo_doc = str(excel_store._cell(row, headers, "Tipo_documento", "Factura")).strip()
            num_doc = str(excel_store._cell(row, headers, "Numero_documento", "") or "").strip()
            num_doc_norm = _norm_doc(num_doc)
            
            if not num_doc_norm or num_doc_norm in emitted_docs:
                continue
                
            f_cell = excel_store._cell(row, headers, "Fecha_emision", None)
            f_doc = excel_store._parse_fecha_cell(f_cell)
            fecha_str = f_doc.strftime("%d/%m/%Y") if f_doc else "-"
            
            prov = str(excel_store._cell(row, headers, "Proveedor", "-"))
            rif = str(excel_store._cell(row, headers, "Proveedor_RIF", "-"))
            
            sub = excel_store._parse_monto_cell(excel_store._cell(row, headers, "Subtotal", None)) or Decimal("0")
            exento = excel_store._parse_monto_cell(excel_store._cell(row, headers, "Monto_exento", None)) or Decimal("0")
            base = excel_store._parse_monto_cell(excel_store._cell(row, headers, "Base_imponible", None)) or Decimal("0")
            iva = excel_store._parse_monto_cell(excel_store._cell(row, headers, "Monto_iva", None)) or Decimal("0")
            tot = excel_store._parse_monto_cell(excel_store._cell(row, headers, "Total", None)) or Decimal("0")
            
            count += 1
            total_subtotal += sub
            total_iva += iva
            total_monto += tot
            
            rows.append([
                count,
                fecha_str,
                num_doc,
                prov,
                rif,
                float(sub),
                float(exento),
                float(base),
                float(iva),
                float(tot)
            ])
            
        wb.close()
        
        if count == 0:
            await status_msg.edit_text("🎉 *¡Excelente! No hay facturas pendientes de retención.* todas las facturas cargadas ya tienen su comprobante emitido.")
            return
            
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            report_xls_path = Path(tmp.name)
            
        excel_store.generate_premium_report_excel(
            report_xls_path,
            title="Facturas de Compra Pendientes de Retención de IVA",
            period_str=f"Al día de hoy: {date.today().strftime('%d/%m/%Y')}",
            headers=["#", "Fecha Emisión", "Nro Documento", "Proveedor", "RIF", "Subtotal", "Exento", "Base Imponible", "IVA", "Total"],
            rows=rows,
            numeric_cols=[5, 6, 7, 8, 9],
            sum_cols=[5, 6, 7, 8, 9],
            emisor_nombre=ctx.company_name,
            emisor_rif=ctx.company_rif,
            emisor_telefono=ctx.company_phone or "No definido",
            emisor_direccion=ctx.company_address or "No definida"
        )
        
        renamed_xls = report_xls_path.with_name(f"FACTURAS-SIN-RETENCION-{date.today().strftime('%d%m%Y')}.xlsx")
        report_xls_path.rename(renamed_xls)
        
        summary_text = (
            f"⚠️ *Facturas sin Comprobante de Retención Emitido*\n\n"
            f"📄 *Total pendientes:* `{count}`\n"
            f"💰 *Total Subtotal:* `{excel_store._format_monto_ves(total_subtotal)}` Bs\n"
            f"💸 *Total IVA Pendiente:* `{excel_store._format_monto_ves(total_iva)}` Bs\n"
            f"🛍️ *Total Facturado Pendiente:* `{excel_store._format_monto_ves(total_monto)}` Bs\n\n"
            f"Se adjunta el reporte detallado con los datos de las facturas."
        )
        
        await status_msg.delete()
        await msg.reply_document(
            document=str(renamed_xls),
            filename=renamed_xls.name,
            caption=summary_text,
            parse_mode="Markdown"
        )
        renamed_xls.unlink(missing_ok=True)
        
    except Exception as e:
        logger.exception("Error al generar reporte de facturas sin retención")
        await status_msg.edit_text(f"❌ Ocurrió un error al generar el reporte: {e!s}")


async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Manejador global de errores para capturar y registrar excepciones no controladas."""
    logger.error("Excepción capturada mientras se procesaba una actualización:", exc_info=context.error)
    
    # Intentar responder al usuario si es posible
    if isinstance(update, Update) and update.effective_message:
        try:
            import traceback
            tb_list = traceback.format_exception(None, context.error, context.error.__traceback__)
            tb_string = "".join(tb_list)
            
            u = update.effective_user
            if u and u.id == config.ALLOWED_USER_ID:
                # Cortar si excede el límite de mensaje de Telegram
                err_msg = f"⚠️ *Error capturado:*\n```python\n{tb_string[:3800]}\n```"
                await update.effective_message.reply_text(err_msg, parse_mode="Markdown")
            else:
                await update.effective_message.reply_text(
                    "⚠️ Ocurrió un error inesperado al procesar tu solicitud.\n"
                    "El detalle técnico ha sido registrado en los logs del servidor."
                )
        except Exception:
            pass


def build_application() -> Application:
    # Configurar un cliente HTTP con tiempos de espera mayores (30 segundos) para conexiones lentas o inestables
    from telegram.request import HTTPXRequest
    req = HTTPXRequest(connect_timeout=30.0, read_timeout=30.0)
    
    app = Application.builder().token(config.BOT_TOKEN).request(req).post_init(post_init).build()
    app.add_handler(CommandHandler("start", start_cmd))
    app.add_handler(CommandHandler("mi_id", mi_id_cmd))
    app.add_handler(CommandHandler("admin", admin_cmd))
    app.add_handler(CommandHandler("tributos", tributos_cmd))
    app.add_handler(CommandHandler("cotizacion", cotizacion_cmd))
    app.add_handler(CommandHandler("nota", nota_cmd))
    app.add_handler(CommandHandler("descargar_excel", descargar_excel_cmd))
    app.add_handler(CallbackQueryHandler(handle_admin_callback, pattern=r"^admin_"))
    app.add_handler(CallbackQueryHandler(handle_emit_retention_callback, pattern=r"^emit_"))
    app.add_handler(CallbackQueryHandler(handle_tributos_callback, pattern=r"^tributos_"))
    app.add_handler(CallbackQueryHandler(handle_cotizaciones_edit_callback, pattern=r"^coti_edit_"))
    app.add_handler(CallbackQueryHandler(handle_cotizaciones_callback, pattern=r"^coti_curr_"))
    app.add_handler(CallbackQueryHandler(handle_builder_callback, pattern=r"^coti_build_"))
    app.add_handler(CallbackQueryHandler(handle_share_email_callback, pattern=r"^share_email$"))
    app.add_handler(CallbackQueryHandler(handle_share_cancel_callback, pattern=r"^share_cancel$"))
    app.add_handler(CallbackQueryHandler(handle_ocr_callback, pattern=r"^ocr_"))
    app.add_handler(CallbackQueryHandler(handle_user_request_access_callback, pattern=r"^user_request_access"))
    app.add_handler(CallbackQueryHandler(handle_cfg_company_callback, pattern=r"^cfg_company_"))
    app.add_handler(CallbackQueryHandler(handle_work_panel_callback, pattern=r"^work_panel:"))
    app.add_handler(CallbackQueryHandler(handle_history_callback, pattern=r"^history_"))
    app.add_handler(CallbackQueryHandler(handle_replace_retention_callback, pattern=r"^rep_ret_"))

    app.add_handler(MessageHandler(filters.VOICE, handle_voice))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    # Acepta texto normal y también caption de fotos/documentos.
    app.add_handler(MessageHandler((filters.TEXT | filters.CAPTION) & ~filters.COMMAND, handle_text))
    
    # Registrar el manejador global de errores
    app.add_error_handler(error_handler)
    return app


def main() -> None:
    # 1. Inicializar el sistema de logs (tanto consola como archivo rotativo bot.log)
    _setup_logging()
    logger.info("Iniciando el bot financiero de Telegram...")
    
    # 2. Bucle de ejecución resiliente con autoreconexión y detección de conflictos
    reconnect_delay = 15
    while True:
        try:
            app = build_application()
            if config.RENDER_EXTERNAL_URL:
                logger.info("Bot en marcha (Webhook en Render). Usuario permitido: %s", config.ALLOWED_USER_ID)
                webhook_url = f"{config.RENDER_EXTERNAL_URL}/{config.BOT_TOKEN}"
                logger.info("Configurando webhook en: %s en el puerto %d", webhook_url, config.PORT)
                app.run_webhook(
                    listen="0.0.0.0",
                    port=config.PORT,
                    url_path=config.BOT_TOKEN,
                    webhook_url=webhook_url,
                    allowed_updates=Update.ALL_TYPES,
                    close_loop=False
                )
            else:
                if not config.FORCE_LOCAL_POLLING:
                    logger.critical(
                        "🛑 EJECUCIÓN LOCAL ABORTADA PARA EVITAR CONFLICTOS CON LA NUBE:\n"
                        "Se detectó un intento de iniciar el bot localmente en modo polling.\n"
                        "Hacer esto eliminaría la configuración del Webhook en Render y dejaría al bot "
                        "de producción inactivo.\n"
                        "Si realmente deseas probar este bot localmente y suspender temporalmente el de producción, "
                        "agrega 'FORCE_LOCAL_POLLING=true' en tu archivo .env local antes de iniciar."
                    )
                    import sys
                    sys.exit(1)

                logger.info("Bot en marcha (polling). Usuario permitido: %s", config.ALLOWED_USER_ID)
                # Ejecutar el bot. Este método bloquea hasta que recibe señal de parada (Ctrl+C, SIGINT, SIGTERM)
                app.run_polling(allowed_updates=Update.ALL_TYPES, close_loop=False)
            logger.info("El bot se ha detenido de manera limpia.")
            break
            
        except Conflict as e:
            logger.critical(
                "¡CONFLICTO DE TOKEN DETECTADO! Hay otra instancia de este bot ejecutándose con el mismo token en otro lugar.\n"
                "Por favor, detén todas las otras ventanas de terminal o procesos de python.exe que estén usando este token.\n"
                "Detalle del error: %s\n"
                "Reintentando conexión en 60 segundos...", e
            )
            time.sleep(60)
            
        except (NetworkError, TimedOut) as e:
            logger.warning(
                "Falla temporal de red detectada: %s.\n"
                "Reintentando establecer conexión en %d segundos...", e, reconnect_delay
            )
            time.sleep(reconnect_delay)
            
        except TelegramError as e:
            logger.error(
                "Error en la API de Telegram: %s.\n"
                "Reintentando en 30 segundos...", e
            )
            time.sleep(30)
            
        except Exception as e:
            logger.exception(
                "Ocurrió un error inesperado e inusual en el bucle principal: %s.\n"
                "Reintentando arranque automático en 30 segundos...", e
            )
            time.sleep(30)
