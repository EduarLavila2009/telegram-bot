"""Lectura y escritura del Excel consolidado."""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook, load_workbook

HEADERS = [
    "Fecha_emision",
    "Numero_comprobante",
    "RIF",
    "Fechas_facturas",
    "Numeros_facturas",
    "Controles_facturas",
    "Total_compra_con_iva",
    "Base_imponible",
    "IVA_retenido",
    "Fecha_registro",
    "Texto_OCR_resumen",
]

FACTURA_COMPRA_HEADERS = [
    "Tipo_documento",
    "Fecha_emision",
    "Fecha_vencimiento",
    "Numero_documento",
    "Numero_control",
    "Proveedor",
    "Proveedor_RIF",
    "Proveedor_Telefono",
    "Direccion_fiscal_proveedor",
    "Receptor",
    "Receptor_RIF",
    "Subtotal",
    "Monto_exento",
    "Base_imponible",
    "Monto_IVA",
    "Total",
    "Fecha_registro",
    "Texto_resumen",
]

RETENCION_EMITIDA_HEADERS = [
    "Numero_comprobante",
    "Fecha_emision",
    "Periodo_fiscal",
    "Proveedor",
    "Proveedor_RIF",
    "DIRECCION FISCAL PROV.",
    "Documentos",
    "Controles",
    "Base_imponible_total",
    "IVA_total",
    "Porcentaje_retencion",
    "IVA_retenido_total",
    "Formato_salida",
    "Fecha_registro",
]


@dataclass
class RetencionRecord:
    fecha_emision: date
    numero_comprobante: str
    rif: str
    fechas_facturas: str
    numeros_facturas: str
    controles_facturas: str
    total_compra_con_iva: Decimal | None
    base_imponible: Decimal | None
    iva_retenido: Decimal
    fecha_registro: str
    texto_ocr: str


@dataclass
class FacturaCompraRow:
    tipo_documento: str
    fecha_emision: str
    fecha_vencimiento: str
    numero_documento: str
    numero_control: str
    proveedor: str
    proveedor_rif: str
    proveedor_telefono: str
    direccion_fiscal_proveedor: str
    receptor: str
    receptor_rif: str
    subtotal: Decimal | None
    monto_exento: Decimal | None
    base_imponible: Decimal | None
    monto_iva: Decimal | None
    total: Decimal | None


def ensure_factura_compra_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas_compra"
    ws.append(FACTURA_COMPRA_HEADERS)
    wb.save(path)


def _excel_numeric_cell(s: str) -> float | str:
    if not str(s).strip():
        return ""
    try:
        return float(Decimal(str(s)))
    except InvalidOperation:
        return str(s)


def _ascii_fold_key(label: str) -> str:
    base = unicodedata.normalize("NFKD", label or "")
    ascii_only = "".join(ch for ch in base if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", "", ascii_only.lower())


def _canonical_factura_header(cell_header: str) -> str | None:
    """Mapea un encabezado de hoja al nombre esperado si el usuario renombró columnas."""
    k = _ascii_fold_key(cell_header)
    if not k:
        return None
    synonyms: dict[str, str] = {
        # Canónicos igual que FACTURA_COMPRA_HEADERS
        "tipodedocumento": "Tipo_documento",
        "tipo": "Tipo_documento",
        "fechaemision": "Fecha_emision",
        "fechadevencimiento": "Fecha_vencimiento",
        "vencimiento": "Fecha_vencimiento",
        "numerodedocumento": "Numero_documento",
        "numerodocumento": "Numero_documento",
        "nrodocumento": "Numero_documento",
        "ndocumento": "Numero_documento",
        "numerodecontrol": "Numero_control",
        "nrocontrol": "Numero_control",
        "control": "Numero_control",
        "proveedor": "Proveedor",
        "nombreproveedor": "Proveedor",
        "proveedorrif": "Proveedor_RIF",
        "rifproveedor": "Proveedor_RIF",
        "telefono": "Proveedor_Telefono",
        "telefonoproveedor": "Proveedor_Telefono",
        "proveedortelefono": "Proveedor_Telefono",
        "tlf": "Proveedor_Telefono",
        "direccionfiscalprov": "Direccion_fiscal_proveedor",
        "direccionfiscalproveedor": "Direccion_fiscal_proveedor",
        "direccionproveedor": "Direccion_fiscal_proveedor",
        "direccionprov": "Direccion_fiscal_proveedor",
        "dirfiscalprov": "Direccion_fiscal_proveedor",
        "receptor": "Receptor",
        "cliente": "Receptor",
        "receptorrif": "Receptor_RIF",
        "rifreceptor": "Receptor_RIF",
        "subtotal": "Subtotal",
        "subtotalbs": "Subtotal",
        "montoexento": "Monto_exento",
        "exento": "Monto_exento",
        "baseimponible": "Base_imponible",
        "basegravable": "Base_imponible",
        "montoiva": "Monto_IVA",
        "iva": "Monto_IVA",
        "total": "Total",
        "totalapagar": "Total",
        "totalbs": "Total",
        "fechaderegistro": "Fecha_registro",
        "registro": "Fecha_registro",
        "textoresumen": "Texto_resumen",
        "resumen": "Texto_resumen",
        "observaciones": "Texto_resumen",
    }
    return synonyms.get(k)


def append_factura_compra(
    path: Path,
    *,
    tipo_documento: str,
    fecha_emision: str,
    fecha_vencimiento: str,
    numero_documento: str,
    numero_control: str,
    proveedor: str,
    proveedor_rif: str,
    proveedor_telefono: str,
    direccion_fiscal_proveedor: str,
    receptor: str,
    receptor_rif: str,
    subtotal: str,
    monto_exento: str,
    base_imponible: str,
    monto_iva: str,
    total: str,
    texto_resumen: str = "",
) -> bool:
    ensure_factura_compra_workbook(path)
    wb = load_workbook(path)
    ws = wb.active
    headers = _headers_index(ws)
    
    # Verificar duplicidad por Proveedor_RIF + Numero_documento
    num_doc_clean = str(numero_documento).strip()
    rif_clean = str(proveedor_rif).strip().upper()
    if num_doc_clean and num_doc_clean != "-" and num_doc_clean.lower() != "s/n" and rif_clean:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            existing_num = _cell(row, headers, "Numero_documento", None)
            existing_rif = _cell(row, headers, "Proveedor_RIF", None)
            if (existing_num is not None and str(existing_num).strip() == num_doc_clean and
                existing_rif is not None and str(existing_rif).strip().upper() == rif_clean):
                wb.close()
                return False

    resolved: set[str] = set()
    for cell_title in headers:
        cn = str(cell_title or "").strip()
        if not cn:
            continue
        field = (
            cn
            if cn in FACTURA_COMPRA_HEADERS
            else (_canonical_factura_header(cn) or "")
        )
        if field:
            resolved.add(field)
    missing = [h for h in FACTURA_COMPRA_HEADERS if h not in resolved]
    if missing:
        wb.close()
        raise ValueError(
            "El Excel de facturas compra no tiene todas las columnas necesarias "
            f"(faltan: {', '.join(missing)}). Renombra columnas al formato estándar "
            "o elimina el archivo para regenerarlo."
        )
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or ()
    fecha_registro = datetime.now().isoformat(timespec="seconds")
    resumen_cell = (texto_resumen or "")[:500]
    payload_by_header: dict[str, object] = {
        "Tipo_documento": tipo_documento,
        "Fecha_emision": fecha_emision,
        "Fecha_vencimiento": fecha_vencimiento,
        "Numero_documento": numero_documento,
        "Numero_control": numero_control,
        "Proveedor": proveedor,
        "Proveedor_RIF": proveedor_rif,
        "Proveedor_Telefono": proveedor_telefono,
        "Direccion_fiscal_proveedor": direccion_fiscal_proveedor,
        "Receptor": receptor,
        "Receptor_RIF": receptor_rif,
        "Subtotal": _excel_numeric_cell(subtotal),
        "Monto_exento": _excel_numeric_cell(monto_exento),
        "Base_imponible": _excel_numeric_cell(base_imponible),
        "Monto_IVA": _excel_numeric_cell(monto_iva),
        "Total": _excel_numeric_cell(total),
        "Fecha_registro": fecha_registro,
        "Texto_resumen": resumen_cell,
    }
    row_out: list[object] = []
    for raw_name in header_row:
        name = str(raw_name or "").strip()
        if not name:
            row_out.append("")
            continue
        key = (
            name
            if name in payload_by_header
            else (_canonical_factura_header(name) or "")
        )
        if key and key in payload_by_header:
            row_out.append(payload_by_header[key])
        else:
            row_out.append("")
    ws.append(row_out)
    wb.save(path)
    wb.close()
    return True


def ensure_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado"
    ws.append(HEADERS)
    wb.save(path)


def _headers_index(ws) -> dict[str, int]:
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or ()
    out: dict[str, int] = {}
    for idx, name in enumerate(header_row):
        k = str(name or "").strip()
        if k:
            out[k] = idx
    return out


def _cell(row: tuple, headers: dict[str, int], key: str, default: str = ""):
    idx = headers.get(key)
    if idx is None or idx >= len(row):
        return default
    v = row[idx]
    return default if v is None else v


def _norm_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", _normalize_text(value))


def _is_duplicate_record(
    ws,
    headers: dict[str, int],
    *,
    numero_comprobante: str,
    rif: str,
    fecha_emision: str,
    iva_retenido: str,
) -> bool:
    comp_new = _norm_key(numero_comprobante)
    rif_new = _norm_key(rif)
    fecha_new = _norm_key(fecha_emision)
    iva_new = _norm_key(iva_retenido)
    for row in ws.iter_rows(min_row=2, values_only=True):
        comp_old = _norm_key(str(_cell(row, headers, "Numero_comprobante", "")))
        rif_old = _norm_key(str(_cell(row, headers, "RIF", "")))
        fecha_old = _norm_key(str(_cell(row, headers, "Fecha_emision", "")))
        iva_old = _norm_key(str(_cell(row, headers, "IVA_retenido", "")))
        if comp_new and comp_old and comp_new == comp_old and rif_new == rif_old:
            return True
        if not comp_new and rif_new and fecha_new and iva_new:
            if rif_new == rif_old and fecha_new == fecha_old and iva_new == iva_old:
                return True
    return False


def append_record(
    path: Path,
    *,
    fecha_emision: str,
    numero_comprobante: str,
    rif: str,
    fechas_facturas: str,
    numeros_facturas: str,
    controles_facturas: str,
    total_compra_con_iva: str,
    base_imponible: str,
    iva_retenido: str,
    ocr_snippet: str = "",
) -> bool:
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb.active
    headers = _headers_index(ws)
    for required in HEADERS:
        if required not in headers:
            wb.close()
            raise ValueError(
                "El Excel existente no tiene la estructura nueva. "
                "Crea uno nuevo o adapta encabezados al formato actual."
            )
    if _is_duplicate_record(
        ws,
        headers,
        numero_comprobante=numero_comprobante,
        rif=rif,
        fecha_emision=fecha_emision,
        iva_retenido=iva_retenido,
    ):
        wb.close()
        return False
    ws.append(
        [
            fecha_emision,
            numero_comprobante,
            rif,
            fechas_facturas,
            numeros_facturas,
            controles_facturas,
            total_compra_con_iva,
            base_imponible,
            iva_retenido,
            datetime.now().isoformat(timespec="seconds"),
            (ocr_snippet or "")[:500],
        ]
    )
    wb.save(path)
    wb.close()
    return True


def _parse_monto_cell(value: str | int | float | None) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    s = str(value).strip()
    if not s:
        return None
    # Clean currency prefixes like Bs., Bs, bs., bsf, ves, $, etc.
    s = re.sub(r'(?i)^(bsf\.?|bs\.?|ves\.?|usd\.?|\$)\s*', '', s)
    s = s.replace(" ", "")
    if re.search(r"\d+\.\d{3},\d{2}$", s) or (
        "," in s and "." in s and s.rfind(",") > s.rfind(".")
    ):
        s = s.replace(".", "").replace(",", ".")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")
    s = re.sub(r"[^\d.\-]", "", s)
    try:
        return Decimal(s)
    except InvalidOperation:
        return None



def parse_amount_ves_string(value: str | None) -> Decimal | None:
    """Convierte montos tipo 74.009,12 o 84823,40 (Bs) a Decimal."""
    return _parse_monto_cell(value)


def _format_monto_ves(value: Decimal | int | float | str | None) -> str:
    """Miles con punto y decimales con coma (ej. 200.000,00)."""
    if value is None:
        return "0,00"
    try:
        d = Decimal(str(value))
    except InvalidOperation:
        return "0,00"
    d = d.quantize(Decimal("0.01"))
    sign = "-" if d < 0 else ""
    d = abs(d)
    entero = int(d)
    centavos = int((d - Decimal(entero)) * 100)
    s_entero = str(entero)
    partes: list[str] = []
    while s_entero:
        partes.append(s_entero[-3:])
        s_entero = s_entero[:-3]
    int_fmt = ".".join(reversed(partes)) if partes else "0"
    return f"{sign}{int_fmt},{centavos:02d}"


def _parse_fecha_cell(value: str | int | float | None) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def summary_for_date(path: Path, target: date) -> tuple[int, Decimal]:
    """Devuelve (cantidad de filas del día, suma de montos del día)."""
    if not path.exists():
        return 0, Decimal("0")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = _headers_index(ws)
    rows = 0
    total = Decimal("0")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        fecha_cell = _cell(row, headers, "Fecha_emision", None)
        monto_cell = _cell(row, headers, "IVA_retenido", None)
        d = _parse_fecha_cell(fecha_cell)
        if d != target:
            continue
        m = _parse_monto_cell(monto_cell)
        if m is None:
            continue
        rows += 1
        total += m
    wb.close()
    return rows, total


def _normalize_text(value: str | None) -> str:
    if not value:
        return ""
    base = unicodedata.normalize("NFKD", value)
    return "".join(ch for ch in base if not unicodedata.combining(ch)).lower()


def _is_retencion_text(value: str | None) -> bool:
    t = _normalize_text(value)
    if not t:
        return False
    return "retencion" in t or "retenciones" in t


_RIF_LIKE = re.compile(
    r"\b([VEJPGvejpg])[\s.\-]*(\d{7,10})[\s.\-]*(\d)\b",
)


def _row_qualifies_for_retencion_report(
    ocr_cell: str,
    numero_comprobante: str,
    rif_cell: str,
) -> bool:
    """Incluye filas consolidadas aunque el resumen no repita la palabra 'retencion'."""
    if _is_retencion_text(ocr_cell):
        return True
    t = _normalize_text(ocr_cell)
    if "comprobante" in t and ("iva" in t or "reten" in t):
        return True
    if "manual" in t or "texto/voz" in t:
        return True
    comp_digits = re.sub(r"\D", "", str(numero_comprobante or ""))
    if len(comp_digits) >= 10:
        return True
    if _RIF_LIKE.search(str(rif_cell or "")):
        return True
    return False


def retenciones_by_document_date(
    path: Path,
    *,
    date_from: date,
    date_to: date,
) -> list[RetencionRecord]:
    """Filtra retenciones por FECHA del documento, no por fecha de registro."""
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = _headers_index(ws)
    out: list[RetencionRecord] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        fecha_cell = _cell(row, headers, "Fecha_emision", None)
        num_comp_cell = _cell(row, headers, "Numero_comprobante", "")
        rif_cell = _cell(row, headers, "RIF", "")
        fechas_fact_cell = _cell(row, headers, "Fechas_facturas", "")
        nums_fact_cell = _cell(row, headers, "Numeros_facturas", "")
        ctrls_fact_cell = _cell(row, headers, "Controles_facturas", "")
        total_compra_cell = _cell(row, headers, "Total_compra_con_iva", None)
        base_cell = _cell(row, headers, "Base_imponible", None)
        iva_ret_cell = _cell(row, headers, "IVA_retenido", None)
        fecha_registro_cell = _cell(row, headers, "Fecha_registro", "")
        ocr_cell = _cell(row, headers, "Texto_OCR_resumen", "")

        fecha_doc = _parse_fecha_cell(fecha_cell)
        if fecha_doc is None or fecha_doc < date_from or fecha_doc > date_to:
            continue
        if not _row_qualifies_for_retencion_report(
            str(ocr_cell),
            str(num_comp_cell or ""),
            str(rif_cell or ""),
        ):
            continue
        iva_retenido = _parse_monto_cell(iva_ret_cell)
        if iva_retenido is None:
            continue
        out.append(
            RetencionRecord(
                fecha_emision=fecha_doc,
                numero_comprobante=str(num_comp_cell or "").strip(),
                rif=str(rif_cell or "").strip(),
                fechas_facturas=str(fechas_fact_cell or "").strip(),
                numeros_facturas=str(nums_fact_cell or "").strip(),
                controles_facturas=str(ctrls_fact_cell or "").strip(),
                total_compra_con_iva=_parse_monto_cell(total_compra_cell),
                base_imponible=_parse_monto_cell(base_cell),
                iva_retenido=iva_retenido,
                fecha_registro=str(fecha_registro_cell or "").strip(),
                texto_ocr=str(ocr_cell or "").strip(),
            )
        )
    wb.close()
    out.sort(key=lambda x: (x.fecha_emision, x.rif, x.iva_retenido))
    return out


def export_retenciones_excel(records: list[RetencionRecord], out_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Retenciones"
    ws.append(
        [
            "Fecha_emision",
            "Numero_comprobante",
            "RIF",
            "Fechas_facturas",
            "Numeros_facturas",
            "Controles_facturas",
            "Total_compra_con_iva",
            "Base_imponible",
            "IVA_retenido",
            "Fecha_registro",
            "Texto_OCR_resumen",
        ]
    )
    for rec in records:
        ws.append(
            [
                rec.fecha_emision.strftime("%d/%m/%Y"),
                rec.numero_comprobante,
                rec.rif,
                rec.fechas_facturas,
                rec.numeros_facturas,
                rec.controles_facturas,
                str(rec.total_compra_con_iva or ""),
                str(rec.base_imponible or ""),
                str(rec.iva_retenido),
                rec.fecha_registro,
                rec.texto_ocr[:500],
            ]
        )
    wb.save(out_path)
    return out_path


def export_retenciones_pdf(
    records: list[RetencionRecord],
    out_path: Path,
    *,
    date_from: date,
    date_to: date,
) -> Path:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    c = canvas.Canvas(str(out_path), pagesize=letter)
    width, height = letter
    y = height - 40
    c.setFont("Helvetica-Bold", 12)
    c.drawString(40, y, "Reporte de retenciones recibidas")
    y -= 16
    c.setFont("Helvetica", 10)
    c.drawString(
        40,
        y,
        f"Rango por fecha de documento: {date_from.strftime('%d/%m/%Y')} - {date_to.strftime('%d/%m/%Y')}",
    )
    y -= 24

    headers_pdf = "Fecha Emision | Nro Comp | RIF | IVA Retenido"
    c.setFont("Helvetica-Bold", 9)
    c.drawString(40, y, headers_pdf)
    y -= 14
    c.setFont("Helvetica", 9)

    total = Decimal("0")
    for rec in records:
        total += rec.iva_retenido
        line = (
            f"{rec.fecha_emision.strftime('%d/%m/%Y')} | "
            f"{(rec.numero_comprobante or '-')[:12]} | "
            f"{(rec.rif or '-')[:14]} | {rec.iva_retenido}"
        )
        c.drawString(40, y, line[:105])
        y -= 12
        if y < 60:
            c.showPage()
            y = height - 40
            c.setFont("Helvetica", 9)

    y -= 8
    if y < 60:
        c.showPage()
        y = height - 40
    c.setFont("Helvetica-Bold", 10)
    c.drawString(40, y, f"Total registros: {len(records)}")
    y -= 14
    c.drawString(40, y, f"Monto total: {total}")
    c.save()
    return out_path


def load_facturas_by_document_numbers(
    path: Path,
    doc_numbers: list[str],
) -> list[FacturaCompraRow]:
    if not path.exists():
        return []
    def _norm_doc(v: str) -> str:
        return re.sub(r"\s+", "", str(v or "")).strip().upper()

    wanted = {_norm_doc(x) for x in doc_numbers if _norm_doc(x)}
    if not wanted:
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = _headers_index(ws)
    out: list[FacturaCompraRow] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        nro = _norm_doc(_cell(row, headers, "Numero_documento", "") or "")
        if not nro or nro not in wanted:
            continue
        out.append(
            FacturaCompraRow(
                tipo_documento=str(_cell(row, headers, "Tipo_documento", "") or "").strip(),
                fecha_emision=str(_cell(row, headers, "Fecha_emision", "") or "").strip(),
                fecha_vencimiento=str(_cell(row, headers, "Fecha_vencimiento", "") or "").strip(),
                numero_documento=nro,
                numero_control=str(_cell(row, headers, "Numero_control", "") or "").strip(),
                proveedor=str(_cell(row, headers, "Proveedor", "") or "").strip(),
                proveedor_rif=str(_cell(row, headers, "Proveedor_RIF", "") or "").strip(),
                proveedor_telefono=str(_cell(row, headers, "Proveedor_Telefono", "") or "").strip(),
                direccion_fiscal_proveedor=str(
                    _cell(row, headers, "Direccion_fiscal_proveedor", "") or ""
                ).strip(),
                receptor=str(_cell(row, headers, "Receptor", "") or "").strip(),
                receptor_rif=str(_cell(row, headers, "Receptor_RIF", "") or "").strip(),
                subtotal=_parse_monto_cell(_cell(row, headers, "Subtotal", None)),
                monto_exento=_parse_monto_cell(_cell(row, headers, "Monto_exento", None)),
                base_imponible=_parse_monto_cell(_cell(row, headers, "Base_imponible", None)),
                monto_iva=_parse_monto_cell(_cell(row, headers, "Monto_IVA", None)),
                total=_parse_monto_cell(_cell(row, headers, "Total", None)),
            )
        )
    wb.close()
    return out


def monthly_retencion_emitida_path(base_dir: Path, emission_date: date) -> Path:
    return base_dir / f"RETEN-EMIT-{emission_date.strftime('%Y-%m')}.xlsx"


def _normalize_numero_comprobante_emitido(raw: str | int | float | None) -> str:
    """Normaliza el correlativo leído desde Excel (texto, entero o notación científica)."""
    if raw is None:
        return ""
    if isinstance(raw, bool):
        return ""
    if isinstance(raw, int):
        return str(raw)
    if isinstance(raw, float):
        if raw.is_integer():
            return str(int(raw))
        return format(raw, "f").split(".")[0]
    s = str(raw).strip()
    if not s:
        return ""
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".", 1)[0]
    if re.fullmatch(r"\d+(\.\d+)?[eE][+-]?\d+", s):
        try:
            f = float(s)
            if f.is_integer():
                return str(int(f))
        except ValueError:
            pass
    digits = re.sub(r"\D", "", s)
    return digits


def _seq_from_numero_comprobante(raw: str | int | float | None, *, prefix: str) -> int | None:
    numero = _normalize_numero_comprobante_emitido(raw)
    if not numero.startswith(prefix):
        return None
    seq_part = numero[len(prefix) :]
    if len(seq_part) != 8 or not seq_part.isdigit():
        return None
    return int(seq_part)


def _max_seq_retencion_emitida_in_workbook(path: Path, *, prefix: str) -> int:
    if not path.exists():
        return 0
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        headers = _headers_index(ws)
        max_seq = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw = _cell(row, headers, "Numero_comprobante", "")
            seq = _seq_from_numero_comprobante(raw, prefix=prefix)
            if seq is not None:
                max_seq = max(max_seq, seq)
        return max_seq
    finally:
        wb.close()


def max_seq_retencion_emitida(base_dir: Path, *, emission_date: date) -> int:
    """Último secuencial (8 dígitos) registrado para YYYYMM en todos los libros del directorio."""
    prefix = emission_date.strftime("%Y%m")
    max_seq = 0
    if base_dir.is_dir():
        for path in sorted(base_dir.glob("RETEN-EMIT-*.xlsx")):
            max_seq = max(max_seq, _max_seq_retencion_emitida_in_workbook(path, prefix=prefix))
    monthly = monthly_retencion_emitida_path(base_dir, emission_date)
    max_seq = max(max_seq, _max_seq_retencion_emitida_in_workbook(monthly, prefix=prefix))
    return max_seq


def ensure_ventas_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Documentos"
    ws.append([
        "Clasificacion", "Estado", "Fecha", "Numero_documento", "Razon_social", 
        "RIF", "Base_imponible", "IVA", "Total", "Monto_retenido", "Dato_fiscal", 
        "Emisor", "Sujeto_retenido", "Texto_origen", "Fecha_registro", "Motivo_clasificacion"
    ])
    wb.save(path)


def append_venta_record(
    path: Path,
    *,
    clasificacion: str,
    estado: str,
    fecha: str,
    numero_documento: str,
    razon_social: str,
    rif: str,
    base_imponible: str,
    iva: str,
    total: str,
    monto_retenido: str = "",
    dato_fiscal: str = "",
    emisor: str = "SUMINISTROS FERRETEROS VITTORIA (SUFEVICA), C.A.",
    sujeto_retenido: str = "",
    texto_origen: str = "",
    motivo_clasificacion: str = "",
) -> bool:
    ensure_ventas_workbook(path)
    wb = load_workbook(path)
    ws = wb.active
    headers = _headers_index(ws)
    
    # Verificar duplicados por Numero_documento
    num_doc_clean = str(numero_documento).strip()
    if num_doc_clean and num_doc_clean != "-" and num_doc_clean.lower() != "s/n":
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            existing_num = _cell(row, headers, "Numero_documento", None)
            if existing_num is not None and str(existing_num).strip() == num_doc_clean:
                wb.close()
                return False
                
    ws.append([
        clasificacion,
        estado,
        fecha,
        numero_documento,
        razon_social,
        rif,
        _excel_numeric_cell(base_imponible),
        _excel_numeric_cell(iva),
        _excel_numeric_cell(total),
        _excel_numeric_cell(monto_retenido),
        dato_fiscal,
        emisor,
        sujeto_retenido,
        (texto_origen or "")[:500],
        datetime.now().isoformat(timespec="seconds"),
        motivo_clasificacion,
    ])
    wb.save(path)
    wb.close()
    return True


def ensure_reporte_z_nuevo_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Reportes Z"
    ws.append([
        "Numero_reporte", "Fecha_emision", "Sub_total", "Base_imponible", 
        "Monto_exento", "IVA", "Total", "Texto_origen", "Fecha_registro"
    ])
    wb.save(path)


def append_reporte_z_nuevo(
    path: Path,
    *,
    numero_reporte: str,
    fecha_emision: str,
    sub_total: str,
    base_imponible: str,
    monto_exento: str,
    iva: str,
    total: str,
    texto_origen: str = "",
) -> bool:
    ensure_reporte_z_nuevo_workbook(path)
    wb = load_workbook(path)
    ws = wb.active
    headers = _headers_index(ws)
    
    # Verificar duplicados por Numero_reporte
    num_rep_clean = str(numero_reporte).strip()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        existing_num = _cell(row, headers, "Numero_reporte", None)
        if existing_num is not None and str(existing_num).strip() == num_rep_clean:
            wb.close()
            return False
            
    ws.append([
        numero_reporte,
        fecha_emision,
        _excel_numeric_cell(sub_total),
        _excel_numeric_cell(base_imponible),
        _excel_numeric_cell(monto_exento),
        _excel_numeric_cell(iva),
        _excel_numeric_cell(total),
        (texto_origen or "")[:500],
        datetime.now().isoformat(timespec="seconds"),
    ])
    wb.save(path)
    wb.close()
    return True


def ensure_retencion_emitida_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Retenciones_emitidas"
    ws.append(RETENCION_EMITIDA_HEADERS)
    wb.save(path)


def next_retencion_emitida_number(path: Path, *, emission_date: date) -> str:
    """
    Siguiente correlativo: YYYYMM + secuencial de 8 dígitos.

    Lee todos los RETEN-EMIT-*.xlsx del directorio base (carpeta de retenciones
    emitidas) y toma el mayor secuencial ya registrado para ese mes.
    """
    ensure_retencion_emitida_workbook(path)
    prefix = emission_date.strftime("%Y%m")
    max_seq = max_seq_retencion_emitida(path.parent, emission_date=emission_date)
    return f"{prefix}{max_seq + 1:08d}"


def append_retencion_emitida(
    path: Path,
    *,
    numero_comprobante: str,
    fecha_emision: str,
    periodo_fiscal: str,
    proveedor: str,
    proveedor_rif: str,
    direccion_fiscal_prov: str,
    documentos: str,
    controles: str,
    base_imponible_total: Decimal,
    iva_total: Decimal,
    porcentaje_retencion: Decimal,
    iva_retenido_total: Decimal,
    formato_salida: str,
) -> None:
    ensure_retencion_emitida_workbook(path)
    wb = load_workbook(path)
    ws = wb.active
    ws.append(
        [
            numero_comprobante,
            fecha_emision,
            periodo_fiscal,
            proveedor,
            proveedor_rif,
            direccion_fiscal_prov,
            documentos,
            controles,
            float(base_imponible_total),
            float(iva_total),
            float(porcentaje_retencion),
            float(iva_retenido_total),
            formato_salida,
            datetime.now().isoformat(timespec="seconds"),
        ]
    )
    wb.save(path)
    wb.close()


RETENCION_ISLR_HEADERS = [
    "Numero_comprobante",
    "Fecha_emision",
    "Periodo_fiscal",
    "Proveedor",
    "Proveedor_RIF",
    "Concepto_retencion",
    "Numero_documento",
    "Numero_control",
    "Base_imponible",
    "Porcentaje_retencion",
    "ISLR_retenido",
    "Total_factura",
    "Fecha_registro"
]

def ensure_retencion_islr_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Retenciones_ISLR"
    ws.append(RETENCION_ISLR_HEADERS)
    wb.save(path)

def append_retencion_islr(
    path: Path,
    *,
    numero_comprobante: str,
    fecha_emision: str,
    periodo_fiscal: str,
    proveedor: str,
    proveedor_rif: str,
    concepto_retencion: str,
    numero_documento: str,
    numero_control: str,
    base_imponible: Decimal,
    porcentaje_retencion: Decimal,
    islr_retenido: Decimal,
    total_factura: Decimal,
) -> None:
    ensure_retencion_islr_workbook(path)
    wb = load_workbook(path)
    ws = wb.active
    ws.append(
        [
            numero_comprobante,
            fecha_emision,
            periodo_fiscal,
            proveedor,
            proveedor_rif,
            concepto_retencion,
            numero_documento,
            numero_control,
            float(base_imponible),
            float(porcentaje_retencion),
            float(islr_retenido),
            float(total_factura),
            datetime.now().isoformat(timespec="seconds"),
        ]
    )
    wb.save(path)
    wb.close()


def export_comprobante_islr_pdf(
    *,
    out_path: Path,
    numero_comprobante: str,
    fecha_emision: str,
    periodo_fiscal: str,
    proveedor: str,
    proveedor_rif: str,
    concepto_retencion: str,
    base_imponible: Decimal,
    porcentaje_retencion: Decimal,
    islr_retenido: Decimal,
    total_factura: Decimal,
    numero_documento: str,
    numero_control: str,
) -> Path:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    c = canvas.Canvas(str(out_path), pagesize=letter)
    width, height = letter
    left = 40
    right = width - 40
    y = height - 50

    # Header
    c.setFont("Helvetica-Bold", 12)
    c.drawString(left, y, "SUMINISTROS FERRETEROS VITTORIA (SUFEVICA), C.A.")
    c.setFont("Helvetica", 9)
    c.drawRightString(right, y, f"Comprobante Nro: {numero_comprobante}")
    y -= 13
    c.setFont("Helvetica-Bold", 10)
    c.drawString(left, y, "RIF: J-40194130-3")
    c.drawRightString(right, y, f"Fecha Emisión: {fecha_emision}")
    y -= 25

    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(width / 2, y, "COMPROBANTE DE RETENCIÓN DE ISLR")
    y -= 10
    c.setFont("Helvetica-Oblique", 7.5)
    c.drawCentredString(width / 2, y, "Decreto N° 1.808 - Reglamento Parcial de la Ley de Impuesto sobre la Renta sobre Retenciones")
    y -= 25

    # Beneficiary Info Box
    box_y = y
    box_h = 75
    c.rect(left, box_y - box_h, right - left, box_h)
    
    c.setFont("Helvetica-Bold", 9)
    c.drawString(left + 10, box_y - 15, "DATOS DEL BENEFICIARIO (PROVEEDOR RETENIDO)")
    c.setFont("Helvetica", 9)
    c.drawString(left + 10, box_y - 32, f"Razón Social: {proveedor}")
    c.drawString(left + 10, box_y - 47, f"RIF: {proveedor_rif}")
    c.drawString(left + 10, box_y - 62, f"Período Fiscal: {periodo_fiscal}")
    y -= box_h + 25

    # Table headers
    headers = ["Factura Nro", "Control Nro", "Concepto de Retención", "Total Factura", "Base Imponible", "% Ret.", "ISLR Retenido"]
    col_widths = [75, 75, 140, 75, 75, 40, 52] # Total width = 532
    
    table_top = y
    c.setFont("Helvetica-Bold", 8.5)
    
    # Draw header row
    c.rect(left, table_top - 20, right - left, 20, fill=0)
    cx = left
    for h, w in zip(headers, col_widths):
        c.drawCentredString(cx + w/2, table_top - 14, h)
        cx += w
        
    # Draw data row
    table_data_y = table_top - 20
    c.setFont("Helvetica", 8.5)
    c.rect(left, table_data_y - 25, right - left, 25)
    
    # Values
    pct = float(porcentaje_retencion)
    if pct < 1.0:
        pct = pct * 100.0
    values = [
        str(numero_documento),
        str(numero_control),
        str(concepto_retencion),
        _format_monto_ves(total_factura),
        _format_monto_ves(base_imponible),
        f"{pct:.1f}%",
        _format_monto_ves(islr_retenido)
    ]
    
    cx = left
    for idx, (val, w) in enumerate(zip(values, col_widths)):
        if idx in (3, 4, 6): # right align amounts
            c.drawRightString(cx + w - 5, table_data_y - 16, val)
        else:
            c.drawCentredString(cx + w/2, table_data_y - 16, val)
        cx += w
        
    y = table_data_y - 25 - 60
    
    # Signatures
    c.setFont("Helvetica-Bold", 8)
    c.line(left + 40, y, left + 180, y)
    c.drawCentredString(left + 110, y - 12, "FIRMA Y SELLO AGENTE")
    c.drawCentredString(left + 110, y - 22, "SUFEVICA, C.A.")
    
    c.line(right - 180, y, right - 40, y)
    c.drawCentredString(right - 110, y - 12, "FIRMA BENEFICIARIO")
    c.drawCentredString(right - 110, y - 22, "RECIBIDO CONFORME")

    c.save()
    return out_path


def monthly_retencion_islr_path(base_dir: Path, emission_date: date) -> Path:
    return base_dir / f"RETEN-ISLR-{emission_date.strftime('%Y-%m')}.xlsx"


def next_retencion_islr_number(base_dir: Path, *, emission_date: date) -> str:
    """Genera el correlativo de ISLR: ISLR-YYYYMM + 6 secuencial."""
    prefix = f"ISLR-{emission_date.strftime('%Y%m')}"
    base_dir.mkdir(parents=True, exist_ok=True)
    max_seq = 0
    # Buscar el mayor correlativo existente en todos los archivos de ISLR
    for path in sorted(base_dir.glob("RETEN-ISLR-*.xlsx")):
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            headers = _headers_index(ws)
            for row in ws.iter_rows(min_row=2, values_only=True):
                raw = _cell(row, headers, "Numero_comprobante", "")
                if raw and str(raw).startswith(prefix):
                    seq_part = str(raw)[len(prefix):]
                    if seq_part.isdigit():
                        max_seq = max(max_seq, int(seq_part))
        finally:
            wb.close()
    return f"{prefix}{max_seq + 1:06d}"



def export_comprobante_emitido_excel(
    *,
    out_path: Path,
    numero_comprobante: str,
    fecha_emision: str,
    periodo_fiscal: str,
    proveedor: str,
    proveedor_rif: str,
    proveedor_telefono: str,
    direccion_fiscal_prov: str,
    items: list[FacturaCompraRow],
    porcentaje_retencion: Decimal,
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comprobante"
    ws.append(["SUMINISTROS FERRETEROS VITTORIA (SUFEVICA), C.A."])
    ws.append(["RIF J-40194130-3"])
    ws.append(["COMPROBANTE DE RETENCION DEL IVA"])
    ws.append([])
    ws.append(["Numero", numero_comprobante])
    ws.append(["Fecha emision", fecha_emision])
    ws.append(["Periodo fiscal", periodo_fiscal])
    ws.append(["Proveedor", proveedor])
    ws.append(["RIF proveedor", proveedor_rif])
    ws.append(["Telefono proveedor", proveedor_telefono])
    ws.append(["Direccion fiscal proveedor", direccion_fiscal_prov])
    ws.append(["Porcentaje retencion", f"{(porcentaje_retencion * 100):.0f}%"])
    ws.append([])
    ws.append(
        [
            "Numero_documento",
            "Numero_control",
            "Fecha_emision",
            "Compras_sin_derecho_credito_IVA",
            "Base_imponible",
            "Monto_IVA",
            "IVA_retenido",
        ]
    )
    sum_base = Decimal("0")
    sum_exento = Decimal("0")
    sum_iva = Decimal("0")
    sum_ret = Decimal("0")
    for it in items:
        exento = it.monto_exento or Decimal("0")
        base = it.base_imponible or Decimal("0")
        iva = it.monto_iva or Decimal("0")
        ret = (iva * porcentaje_retencion).quantize(Decimal("0.01"))
        ws.append(
            [
                it.numero_documento,
                it.numero_control,
                it.fecha_emision,
                float(exento),
                float(base),
                float(iva),
                float(ret),
            ]
        )
        sum_exento += exento
        sum_base += base
        sum_iva += iva
        sum_ret += ret
    ws.append([])
    ws.append(["Totales", "", "", float(sum_exento), float(sum_base), float(sum_iva), float(sum_ret)])
    wb.save(out_path)
    wb.close()
    return out_path


def export_comprobante_emitido_pdf(
    *,
    out_path: Path,
    numero_comprobante: str,
    fecha_emision: str,
    periodo_fiscal: str,
    proveedor: str,
    proveedor_rif: str,
    proveedor_telefono: str,
    direccion_fiscal_prov: str,
    items: list[FacturaCompraRow],
    porcentaje_retencion: Decimal,
    firma_sello_path: Path | str | None = None,
) -> Path:
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas

    c = canvas.Canvas(str(out_path), pagesize=landscape(letter))
    width, height = landscape(letter)
    left = 20
    right = width - 20
    y = height - 24

    # Encabezado principal
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(width / 2, y, "SUMINISTROS FERRETEROS VITTORIA (SUFEVICA), C.A.")
    y -= 13
    c.setFont("Helvetica-Bold", 10)
    c.drawCentredString(width / 2, y, "RIF.- J-40194130-3")
    y -= 14
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(width / 2, y, "COMPROBANTE DE RETENCION DEL IMPUESTO AL VALOR AGREGADO")
    y -= 12
    # Texto legal en 2 líneas (como el modelo).
    c.setFont("Helvetica", 7)
    c.drawCentredString(
        width / 2,
        y,
        "(LEY IVA. ART. 11 “La Administración Tributaria podrá designar como responsables del pago del impuesto,",
    )
    y -= 9
    c.drawCentredString(
        width / 2,
        y,
        "en calidad de agentes de retención,a quienes por sus funciones públicas o por razón de sus actividades privadas "
        "intervengan en operaciones gravadas con el impuesto establecido en este Decreto con Rango, Valor y Fuerza de Ley”",
    )
    y -= 18

    def _format_rif(raw: str) -> str:
        v = str(raw or "").strip().upper().replace(" ", "")
        if not v:
            return "-"
        if "-" not in v and len(v) > 1 and v[0].isalpha():
            v = f"{v[0]}-{v[1:]}"
        return v

    # Bloques superiores
    box_h = 40
    c.rect(left, y - box_h, 290, box_h)
    c.setFont("Helvetica-Bold", 7)
    c.drawString(left + 4, y - 11, "DIRECCION AGENTE DE RETENCION")
    c.setFont("Helvetica", 7)
    c.drawString(left + 4, y - 20, "AV. JUAN DE URPIN, CC RES. VITTORIA III, EDIF. F, NIVEL PB, LOCAL NRO. 2")
    c.drawString(left + 4, y - 29, "EL ESPEJO, BARCELONA-EDO ANZOATEGUI. TELEFONOS 0281-2768765")
    c.drawString(left + 4, y - 37, "RIF.- J-40194130-3")

    c.rect(left + 300, y - box_h, 120, box_h)
    c.setFont("Helvetica-Bold", 7)
    c.drawString(left + 304, y - 11, "ANO-MES")
    c.setFont("Helvetica-Bold", 10)
    c.drawString(left + 328, y - 28, periodo_fiscal)

    c.setFont("Helvetica-Bold", 8)
    c.drawString(right - 235, y - 12, f"Comprobante Nro: {numero_comprobante}")
    c.drawString(right - 235, y - 30, f"Fecha de Emision: {fecha_emision}")
    y -= box_h + 12

    # Bloque proveedor como el formato modelo: Nombre / RIF + Telefono / Direccion fiscal
    c.setFont("Helvetica-Bold", 8)
    c.drawString(left, y, "NOMBRE O RAZON SOCIAL DEL PROVEEDOR:")
    c.setFont("Helvetica-Bold", 8)
    prov_nombre = (proveedor or "-")[:70]
    c.drawString(left + 240, y, prov_nombre)
    c.line(left + 240, y - 2, right - 10, y - 2)
    y -= 10

    c.setFont("Helvetica-Bold", 8)
    c.drawString(left, y, "RIF:")
    rif_display = _format_rif(proveedor_rif)
    if rif_display == "-":
        for it in items:
            if str(it.proveedor_rif or "").strip():
                rif_display = _format_rif(it.proveedor_rif)
                break
    c.setFont("Helvetica-Bold", 8)
    c.drawString(left + 28, y, rif_display[:20])
    c.line(left + 28, y - 2, left + 160, y - 2)

    c.setFont("Helvetica-Bold", 8)
    c.drawString(left + 190, y, "TELEFONO:")
    c.setFont("Helvetica-Bold", 8)
    c.drawString(left + 260, y, (proveedor_telefono or "-")[:30])
    c.line(left + 260, y - 2, right - 10, y - 2)
    y -= 10

    c.setFont("Helvetica-Bold", 8)
    c.drawString(left, y, "DIRECCIÓN FISCAL:")
    c.setFont("Helvetica-Bold", 8)
    addr = (direccion_fiscal_prov or "-").strip() or "-"
    c.drawString(left + 110, y, addr[:130])
    c.line(left + 110, y - 2, right - 10, y - 2)
    y -= 16

    # Tabla principal horizontal estilo formato SENIAT/SUFEVICA
    compras_sin_credito_title = "Compras sin Derecho a Crédito I.V.A."
    cols = [
        ("Oper", 24),
        ("Fecha Factura", 58),
        ("Nro Factura", 58),
        ("Nro Control", 58),
        ("N/ Deb", 34),
        ("N/ Cred", 34),
        ("Tipo", 26),
        ("Factura Afectada", 56),
        ("Total c/IVA", 64),
        (compras_sin_credito_title, 64),
        ("Base Imponible", 64),
        ("% Alicuota", 42),
        ("Impuesto IVA", 64),
        ("IVA Retenido", 66),
    ]
    table_w = sum(w for _, w in cols)
    x = left
    table_top = y
    header_h = 24
    row_h = 14
    # Reservar espacio para el pie de página y firmas (como el formato modelo).
    footer_reserved = 80
    max_rows_fit = max(1, int((table_top - footer_reserved - header_h) / row_h))
    data_rows = max(len(items), min(12, max_rows_fit))
    table_h = header_h + (data_rows * row_h)
    c.rect(x, table_top - table_h, table_w, table_h)

    # Vertical lines + headers
    cx = x
    c.setFont("Helvetica-Bold", 6.5)
    for title, w in cols:
        c.line(cx, table_top, cx, table_top - table_h)
        if title == compras_sin_credito_title:
            c.drawCentredString(cx + (w / 2), table_top - 8, "Compras sin")
            c.drawCentredString(cx + (w / 2), table_top - 14, "Derecho a Crédito")
            c.drawCentredString(cx + (w / 2), table_top - 20, "I.V.A.")
        else:
            c.drawCentredString(cx + (w / 2), table_top - 14, title)
        cx += w
    c.line(cx, table_top, cx, table_top - table_h)
    c.line(x, table_top - header_h, x + table_w, table_top - header_h)

    c.setFont("Helvetica", 7)
    sum_base = Decimal("0")
    sum_exento = Decimal("0")
    sum_iva = Decimal("0")
    sum_ret = Decimal("0")
    sum_total = Decimal("0")
    start_y = table_top - header_h - 10
    for idx in range(data_rows):
        if idx < len(items):
            it = items[idx]
        else:
            it = None
        row_y = start_y - (idx * row_h)
        c.line(x, row_y - 5, x + table_w, row_y - 5)
        if not it:
            continue
        base = it.base_imponible or Decimal("0")
        exento = it.monto_exento or Decimal("0")
        iva = it.monto_iva or Decimal("0")
        total = it.total or (base + iva)
        ret = (iva * porcentaje_retencion).quantize(Decimal("0.01"))
        values = [
            f"{idx + 1:02d}",
            (it.fecha_emision or "-")[:10],
            it.numero_documento[:12],
            (it.numero_control or "-")[:12],
            "",
            "",
            "C",
            "",
            _format_monto_ves(total),
            _format_monto_ves(exento),
            _format_monto_ves(base),
            "16",
            _format_monto_ves(iva),
            _format_monto_ves(ret),
        ]
        cx = x
        for (title, w), val in zip(cols, values):
            if title in {"Total c/IVA", "Base Imponible", "Impuesto IVA", "IVA Retenido"}:
                c.drawRightString(cx + w - 2, row_y, val)
            else:
                c.drawCentredString(cx + (w / 2), row_y, val)
            cx += w
        sum_base += base
        sum_exento += exento
        sum_iva += iva
        sum_ret += ret
        sum_total += total

    # Totales
    total_y = table_top - table_h - 14
    c.setFont("Helvetica-Bold", 9)
    c.drawString(x + 8, total_y, "TOTALES")
    c.drawRightString(x + sum(w for _, w in cols[:9]) - 4, total_y, _format_monto_ves(sum_total))
    c.drawRightString(x + sum(w for _, w in cols[:10]) - 4, total_y, _format_monto_ves(sum_exento))
    c.drawRightString(x + sum(w for _, w in cols[:11]) - 4, total_y, _format_monto_ves(sum_base))
    c.drawRightString(x + sum(w for _, w in cols[:13]) - 4, total_y, _format_monto_ves(sum_iva))
    c.drawRightString(x + table_w - 4, total_y, _format_monto_ves(sum_ret))

    # Firmas / notas (separado de la tabla, como el modelo)
    y = total_y - 18
    box_h = 64
    left_box_w = 360
    right_box_w = 360
    gap = 30

    c.setFont("Helvetica-Bold", 8)
    c.rect(left, y - box_h, left_box_w, box_h)
    c.drawCentredString(left + (left_box_w / 2), y - 12, "FIRMA Y SELLO DEL AGENTE DE RETENCION")
    sello_path = Path(firma_sello_path).expanduser() if firma_sello_path else None
    if sello_path and sello_path.exists():
        try:
            # La imagen se dibuja en tamaño fijo de 4cm x 2cm.
            img_w = 4 * cm
            img_h = 2 * cm
            c.drawImage(
                str(sello_path),
                left + ((left_box_w - img_w) / 2),
                (y - box_h) + 2,
                width=img_w,
                height=img_h,
                preserveAspectRatio=False,
                mask="auto",
                anchor="sw",
            )
        except Exception:
            # Si la imagen no se puede leer, el PDF se emite igual.
            pass

    right_x = left + left_box_w + gap
    c.rect(right_x, y - box_h, right_box_w, box_h)
    c.drawCentredString(right_x + (right_box_w / 2), y - 12, "SUJETO RETENIDO")
    c.setFont("Helvetica", 7)
    c.drawString(right_x + 10, y - 38, "FECHA DE ENTREGA:      /      /")

    y = y - box_h - 14
    c.setFont("Helvetica-Bold", 7)
    c.line(left, y, right, y)
    y -= 11
    c.drawCentredString(
        width / 2,
        y,
        "ESTE COMPROBANTE SE EMITE SEGÚN LO ESTABLECIDO EN EL ARTÍCULO 16 DE PROVIDENCIA ADMINISTRATIVA "
        "N° SNAT/2015/0049 DE FECHA 10/08/2015",
    )
    c.save()
    return out_path


def _to_float_or_none(val: object) -> float | None:
    if val is None:
        return None
    try:
        if isinstance(val, (int, float)):
            return float(val)
        d = _parse_monto_cell(str(val))
        return float(d) if d is not None else None
    except Exception:
        return None


def h_lower_contains_date_or_num(header: str, val: object) -> bool:
    h = str(header).lower()
    if any(x in h for x in ("fecha", "nro", "numero", "control", "rif", "correlativo", "tipo", "documento", "comprobante")):
        return True
    s = str(val or "").strip()
    if re.fullmatch(r"\d{2}/\d{2}/\d{4}", s) or re.fullmatch(r"\d+", s):
        return True
    return False


def generate_premium_report_excel(
    out_path: Path,
    title: str,
    period_str: str,
    headers: list[str],
    rows: list[list[object]],
    numeric_cols: list[int],
    sum_cols: list[int],
) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.views.sheetView[0].showGridLines = True

    # Font definitions
    font_family = "Segoe UI"
    font_title = Font(name=font_family, size=14, bold=True, color="1A1A1A")
    font_subtitle = Font(name=font_family, size=10, italic=True, color="555555")
    font_meta_key = Font(name=font_family, size=9, bold=True, color="333333")
    font_meta_val = Font(name=font_family, size=9, color="333333")
    font_header = Font(name=font_family, size=9, bold=True, color="FFFFFF")
    font_data = Font(name=font_family, size=9, color="333333")
    font_totals = Font(name=font_family, size=9, bold=True, color="1A1A1A")

    # Border definitions
    thin_border_side = Side(style="thin", color="CCCCCC")
    border_all_thin = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    double_bottom = Side(style="double", color="333333")
    thin_top = Side(style="thin", color="333333")
    border_totals = Border(top=thin_top, bottom=double_bottom)

    # Rellenos
    fill_header = PatternFill(start_color="3B4F66", end_color="3B4F66", fill_type="solid")
    fill_meta = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")

    # Title
    ws.cell(row=2, column=2, value=title).font = font_title
    ws.cell(row=3, column=2, value=f"Período: {period_str}").font = font_subtitle

    # Metadata Box (Rows 5 to 10)
    meta_info = [
        ("Agente de retención:", "Suministros Ferreteros Vittoria, C.A.(SUFEVICA)"),
        ("RIF:", "J-40194130-3"),
        ("Periodo:", period_str),
        ("Teléfono:", "+582812768765"),
        ("Dirección:", "AV JUAN DE URPIN CC RESIDENCIAS VITTORIA III, EDIF.F NIVEL PB LOCAL NRO. 2 SECTOR EL ESPEJO BARCELONA ANZOATEGUI ZONA POSTAL 6001"),
        ("Facturas listadas:", len(rows)),
    ]

    max_col_idx = len(headers) + 1
    end_col = max(7, max_col_idx)
    
    for r_offset, (k, v) in enumerate(meta_info):
        r = 5 + r_offset
        cell_k = ws.cell(row=r, column=2, value=k)
        cell_k.font = font_meta_key
        cell_k.alignment = Alignment(horizontal="right", vertical="center")
        
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=end_col)
        cell_v = ws.cell(row=r, column=3, value=v)
        cell_v.font = font_meta_val
        cell_v.alignment = Alignment(horizontal="left", vertical="center")

        for c in range(2, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill_meta
            
            left_side = Side(style="thin", color="BBBBBB") if c == 2 else None
            right_side = Side(style="thin", color="BBBBBB") if c == end_col else None
            top_side = Side(style="thin", color="BBBBBB") if r == 5 else None
            bottom_side = Side(style="thin", color="BBBBBB") if r == 10 else None
            cell.border = Border(left=left_side, right=right_side, top=top_side, bottom=bottom_side)

    # Table Header Row (Row 12)
    start_row = 12
    for c_idx, h in enumerate(headers):
        col = 2 + c_idx
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_all_thin

    ws.row_dimensions[start_row].height = 24

    # Data Rows (Row 13 onwards)
    current_row = start_row + 1
    for r_idx, r_data in enumerate(rows):
        is_even = (r_idx % 2 == 1)
        for c_idx, val in enumerate(r_data):
            col = 2 + c_idx
            cell = ws.cell(row=current_row, column=col)
            
            if c_idx in numeric_cols:
                n_val = _to_float_or_none(val)
                if n_val is not None:
                    cell.value = n_val
                    cell.number_format = '#,##0.00'
                else:
                    cell.value = val
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.value = val
                if c_idx == 0 or h_lower_contains_date_or_num(headers[c_idx], val):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            cell.font = font_data
            cell.border = border_all_thin
            if is_even:
                cell.fill = fill_zebra
                
        ws.row_dimensions[current_row].height = 18
        current_row += 1

    # Totals Row
    first_sum_col = min(sum_cols) if sum_cols else 1000
    totales_col = first_sum_col + 2 - 1
    if totales_col >= 2:
        cell_t = ws.cell(row=current_row, column=totales_col, value="TOTALES")
        cell_t.font = font_totals
        cell_t.alignment = Alignment(horizontal="right", vertical="center")
        cell_t.border = border_totals

    for c_idx in range(len(headers)):
        col = 2 + c_idx
        cell = ws.cell(row=current_row, column=col)
        if c_idx in sum_cols:
            col_letter = get_column_letter(col)
            if current_row - 1 >= 13:
                formula = f"=SUM({col_letter}13:{col_letter}{current_row - 1})"
                cell.value = formula
            else:
                cell.value = 0.0
            cell.number_format = '#,##0.00'
            cell.font = font_totals
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = border_totals
        else:
            cell.border = border_totals

    ws.row_dimensions[current_row].height = 20

    # Auto-adjust column widths
    for col in range(2, 2 + len(headers)):
        col_letter = get_column_letter(col)
        max_len = 0
        for r in range(start_row, current_row + 1):
            cell_val = ws.cell(row=r, column=col).value
            if cell_val is not None:
                s_val = str(cell_val)
                if s_val.startswith("="):
                    s_val = "123.456,78"
                max_len = max(max_len, len(s_val))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    ws.column_dimensions["A"].width = 3

    wb.save(out_path)
    wb.close()
    return out_path


def load_purchases_by_date_range(
    base_dir: Path,
    *,
    date_from: date,
    date_to: date,
) -> list[list[object]]:
    if not base_dir.is_dir():
        return []
    records: list[tuple[date, str, str, str, str, str, float, float, float, float]] = []
    
    for path in base_dir.glob("RETEN-EMIT-*.xlsx"):
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            headers = _headers_index(ws)
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                f_cell = _cell(row, headers, "Fecha_emision", None)
                f_doc = _parse_fecha_cell(f_cell)
                if f_doc is None or f_doc < date_from or f_doc > date_to:
                    continue
                
                num_comp = str(_cell(row, headers, "Numero_comprobante", "") or "").strip()
                prov = str(_cell(row, headers, "Proveedor", "") or "").strip()
                rif = str(_cell(row, headers, "Proveedor_RIF", "") or "").strip()
                doc = str(_cell(row, headers, "Documentos", "") or "").strip()
                ctrl = str(_cell(row, headers, "Controles", "") or "").strip()
                
                base = _parse_monto_cell(_cell(row, headers, "Base_imponible_total", None)) or Decimal("0")
                iva = _parse_monto_cell(_cell(row, headers, "IVA_total", None)) or Decimal("0")
                ret = _parse_monto_cell(_cell(row, headers, "IVA_retenido_total", None)) or Decimal("0")
                monto = base + iva
                
                records.append((
                    f_doc,
                    num_comp,
                    prov,
                    rif,
                    doc,
                    ctrl,
                    float(base),
                    float(iva),
                    float(monto),
                    float(ret)
                ))
            wb.close()
        except Exception as e:
            logger.exception("Error loading purchases from %s: %s", path, e)
            
    records.sort(key=lambda x: (x[0], x[1], x[3]))
    
    out: list[list[object]] = []
    for idx, r in enumerate(records):
        out.append([
            idx + 1,
            r[0].strftime("%d/%m/%Y"),
            r[1],
            r[2],
            r[3],
            r[4],
            r[5],
            r[6],
            r[7],
            r[8],
            r[9]
        ])
    return out


def load_sales_by_date_range(
    path: Path,
    *,
    date_from: date,
    date_to: date,
) -> list[list[object]]:
    if not path.exists():
        return []
    records: list[tuple[date, str, str, str, float, float, float]] = []
    
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = _headers_index(ws)
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            f_cell = _cell(row, headers, "Fecha", None) or _cell(row, headers, "Fecha_emision", None)
            f_doc = _parse_fecha_cell(f_cell)
            if f_doc is None or f_doc < date_from or f_doc > date_to:
                continue
            
            num_doc = str(_cell(row, headers, "Numero_documento", "") or "").strip()
            cliente = str(_cell(row, headers, "Razon_social", "") or "").strip()
            rif = str(_cell(row, headers, "RIF", "") or "").strip()
            
            base = _parse_monto_cell(_cell(row, headers, "Base_imponible", None)) or Decimal("0")
            iva = _parse_monto_cell(_cell(row, headers, "IVA", None)) or Decimal("0")
            total = _parse_monto_cell(_cell(row, headers, "Total", None)) or (base + iva)
            
            records.append((
                f_doc,
                num_doc,
                cliente,
                rif,
                float(base),
                float(iva),
                float(total)
            ))
        wb.close()
    except Exception as e:
        logger.exception("Error loading sales from %s: %s", path, e)
        
    records.sort(key=lambda x: (x[0], x[1]))
    
    out: list[list[object]] = []
    for idx, r in enumerate(records):
        out.append([
            idx + 1,
            r[0].strftime("%d/%m/%Y"),
            r[1],
            r[2],
            r[3],
            r[4],
            r[5],
            r[6]
        ])
    return out


def load_reportes_z_by_date_range(
    path: Path,
    *,
    date_from: date,
    date_to: date,
) -> list[list[object]]:
    if not path.exists():
        return []
    records: list[tuple[date, str, float, float, float, float, float]] = []
    
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = _headers_index(ws)
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            f_cell = _cell(row, headers, "Fecha_emision", None) or _cell(row, headers, "Fecha", None)
            f_doc = _parse_fecha_cell(f_cell)
            if f_doc is None or f_doc < date_from or f_doc > date_to:
                continue
            
            num_rep = str(_cell(row, headers, "Numero_reporte", "") or "").strip()
            
            sub = _parse_monto_cell(_cell(row, headers, "Sub_total", None)) or Decimal("0")
            exento = _parse_monto_cell(_cell(row, headers, "Monto_exento", None)) or Decimal("0")
            base = _parse_monto_cell(_cell(row, headers, "Base_imponible", None)) or Decimal("0")
            iva = _parse_monto_cell(_cell(row, headers, "IVA", None)) or Decimal("0")
            total = _parse_monto_cell(_cell(row, headers, "Total", None)) or Decimal("0")
            
            records.append((
                f_doc,
                num_rep,
                float(sub),
                float(exento),
                float(base),
                float(iva),
                float(total)
            ))
        wb.close()
    except Exception as e:
        logger.exception("Error loading Reportes Z from %s: %s", path, e)
        
    records.sort(key=lambda x: (x[0], x[1]))
    
    out: list[list[object]] = []
    for idx, r in enumerate(records):
        out.append([
            idx + 1,
            r[0].strftime("%d/%m/%Y"),
            r[1],
            r[2],
            r[3],
            r[4],
            r[5],
            r[6]
        ])
    return out


def search_products_in_excel(path: Path, query: str, search_by: str = "desc") -> list[dict]:
    """
    Busca productos en el archivo Excel de inventario.
    
    Argumentos:
        path: Ruta al archivo Excel
        query: Término de búsqueda (código o descripción)
        search_by: "code" o "desc"
        
    Retorna:
        Lista de diccionarios: [{"code": str, "description": str, "price": float}]
    """
    if not path.exists():
        return []
        
    wb = load_workbook(path, read_only=True, data_only=True)
    results = []
    try:
        # Intentar buscar la hoja más adecuada: active o que contenga "prod" o "inv"
        ws = wb.active
        for name in wb.sheetnames:
            if "prod" in name.lower() or "inv" in name.lower():
                ws = wb[name]
                break
                
        # Obtener la primera fila para buscar cabeceras
        first_row_iter = ws.iter_rows(max_row=1, values_only=True)
        try:
            first_row = next(first_row_iter)
        except StopIteration:
            return []
            
        if not first_row:
            return []
            
        headers = [str(cell).strip().lower() for cell in first_row if cell is not None]
        
        # Identificar columnas
        code_idx = -1
        desc_idx = -1
        price_idx = -1
        barcode_idx = -1
        
        for idx, h in enumerate(headers):
            if any(x in h for x in ("barr", "upc", "ean", "scan")):
                barcode_idx = idx
            elif any(x in h for x in ("cod", "ref")):
                code_idx = idx
            elif any(x in h for x in ("desc", "nomb", "prod", "det")):
                desc_idx = idx
            elif any(x in h for x in ("prec", "cost", "usd", "unit", "val")):
                price_idx = idx
                
        # Índices por defecto si no se detectan cabeceras
        if code_idx == -1: code_idx = 0
        if desc_idx == -1: desc_idx = 1
        if price_idx == -1: price_idx = 2
        
        query_norm = query.strip().lower()
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= max(code_idx, desc_idx, price_idx, barcode_idx):
                continue
                
            code_val = str(row[code_idx] or "").strip()
            desc_val = str(row[desc_idx] or "").strip()
            barcode_val = str(row[barcode_idx] or "").strip() if barcode_idx != -1 else ""
            price_val = row[price_idx]
            
            # Convertir precio a float
            try:
                if isinstance(price_val, str):
                    clean_price = price_val.replace("Bs.", "").replace("$", "").replace("USD", "").strip()
                    if "," in clean_price and "." in clean_price:
                        clean_price = clean_price.replace(".", "").replace(",", ".")
                    elif "," in clean_price:
                        clean_price = clean_price.replace(",", ".")
                    price_float = float(clean_price)
                elif isinstance(price_val, (int, float)):
                    price_float = float(price_val)
                else:
                    price_float = 0.0
            except Exception:
                price_float = 0.0
                
            if not code_val and not desc_val:
                continue
                
            match = False
            if search_by == "code":
                # Coincidencia si el código o el código de barras contiene el query
                match = (query_norm in code_val.lower()) or (barcode_val and query_norm in barcode_val.lower())
            else:
                # Coincidencia si la descripción contiene el query
                match = query_norm in desc_val.lower()
                
            if match:
                results.append({
                    "code": code_val,
                    "description": desc_val,
                    "price": price_float
                })
    finally:
        wb.close()
        
    return results

