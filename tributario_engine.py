"""Motor tributario para el cálculo de IVA por pagar y compromisos quincenales (SENIAT)."""

from __future__ import annotations

import logging
import re
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from . import config
from . import excel_store

logger = logging.getLogger(__name__)

# Alícuota estándar de anticipo de ISLR en Venezuela para la mayoría de contribuyentes especiales (1%)
ALICUOTA_ANTICIPO_ISLR = Decimal("0.01")


def get_fortnight_range(year: int, month: int, fortnight: int) -> tuple[date, date]:
    """
    Retorna el rango de fechas (desde, hasta) para una quincena específica de un mes/año.
    
    fortnight = 1: días 01 al 15
    fortnight = 2: días 16 al último día del mes
    """
    if fortnight == 1:
        start = date(year, month, 1)
        end = date(year, month, 15)
        return start, end
    elif fortnight == 2:
        start = date(year, month, 16)
        # Calcular el último día del mes
        if month == 12:
            end = date(year, month, 31)
        else:
            # Siguiente mes día 1 menos 1 día
            next_month = date(year, month + 1, 1)
            import datetime as dt
            end = next_month - dt.timedelta(days=1)
        return start, end
    else:
        raise ValueError("La quincena debe ser 1 o 2.")


def _parse_row_date(val: object) -> date | None:
    """Intenta parsear un valor de celda de fecha con máxima flexibilidad."""
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def get_sales_totals(start_date: date, end_date: date) -> tuple[Decimal, Decimal, int]:
    """
    Calcula totales de ventas (Base Imponible, IVA, número de registros)
    a partir de FACTURAS-EMITIDAS.xlsx y REPORTES-Z.xlsx en el rango dado.
    """
    total_base = Decimal("0")
    total_iva = Decimal("0")
    total_count = 0

    files_to_read = [config.FACTURAS_EMITIDAS_PATH, config.REPORTES_Z_PATH]

    for path in files_to_read:
        if not path.exists():
            continue
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            headers = excel_store._headers_index(ws)
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                # Se asume columnas estándar de facturas emitidas y fallbacks para reporte Z nuevo
                fecha_cell = excel_store._cell(row, headers, "Fecha_emision", None) or excel_store._cell(row, headers, "Fecha", None)
                fecha_doc = _parse_row_date(fecha_cell)
                
                if fecha_doc is None or fecha_doc < start_date or fecha_doc > end_date:
                    continue
                
                base_cell = excel_store._cell(row, headers, "Base_imponible", None)
                iva_cell = excel_store._cell(row, headers, "IVA", None)
                
                base = excel_store._parse_monto_cell(base_cell) or Decimal("0")
                iva = excel_store._parse_monto_cell(iva_cell) or Decimal("0")
                
                total_base += base
                total_iva += iva
                total_count += 1
            wb.close()
        except Exception as e:
            logger.exception("Error leyendo ventas desde %s", path)

    return total_base, total_iva, total_count


def get_purchases_totals(start_date: date, end_date: date) -> tuple[Decimal, Decimal, int]:
    """
    Calcula totales de compras (Base Imponible, IVA, número de registros)
    a partir de los Excels mensuales de retenciones emitidas (RETEN-EMIT-*.xlsx)
    en el rango dado, de acuerdo a la instrucción de obtener el crédito fiscal
    directamente de este libro en lugar del Excel de facturas recibidas.
    """
    total_base = Decimal("0")
    total_iva = Decimal("0")
    total_count = 0
    
    base_dir = config.RETENCIONES_EMITIDAS_DIR
    if not base_dir.is_dir():
        return total_base, total_iva, total_count
        
    for path in base_dir.glob("RETEN-EMIT-*.xlsx"):
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            headers = excel_store._headers_index(ws)
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                fecha_cell = excel_store._cell(row, headers, "Fecha_emision", None)
                fecha_doc = _parse_row_date(fecha_cell)
                
                if fecha_doc is None or fecha_doc < start_date or fecha_doc > end_date:
                    continue
                
                base_cell = excel_store._cell(row, headers, "Base_imponible_total", None)
                iva_cell = excel_store._cell(row, headers, "IVA_total", None)
                
                base = excel_store._parse_monto_cell(base_cell) or Decimal("0")
                iva = excel_store._parse_monto_cell(iva_cell) or Decimal("0")
                
                total_base += base
                total_iva += iva
                total_count += 1
            wb.close()
        except Exception as e:
            logger.exception("Error leyendo compras (crédito fiscal) desde %s", path)
            
    return total_base, total_iva, total_count


def get_withholdings_received_totals(start_date: date, end_date: date) -> tuple[Decimal, int]:
    """
    Calcula el total de retenciones de IVA recibidas de clientes en el rango dado.
    Se extraen de RETEN-REC.xlsx (config.EXCEL_PATH).
    """
    total_ret = Decimal("0")
    total_count = 0
    
    path = config.EXCEL_PATH
    if not path.exists():
        return total_ret, total_count
        
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = excel_store._headers_index(ws)
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            fecha_cell = excel_store._cell(row, headers, "Fecha_emision", None)
            fecha_doc = _parse_row_date(fecha_cell)
            
            if fecha_doc is None or fecha_doc < start_date or fecha_doc > end_date:
                continue
            
            ret_cell = excel_store._cell(row, headers, "IVA_retenido", None)
            ret = excel_store._parse_monto_cell(ret_cell) or Decimal("0")
            
            total_ret += ret
            total_count += 1
        wb.close()
    except Exception as e:
        logger.exception("Error leyendo retenciones recibidas desde %s", path)
        
    return total_ret, total_count


def get_withholdings_issued_totals(start_date: date, end_date: date) -> tuple[Decimal, int]:
    """
    Calcula el total de retenciones de IVA emitidas a proveedores en el rango dado.
    Busca los Excels mensuales en config.RETENCIONES_EMITIDAS_DIR.
    """
    total_ret = Decimal("0")
    total_count = 0
    
    base_dir = config.RETENCIONES_EMITIDAS_DIR
    if not base_dir.is_dir():
        return total_ret, total_count
        
    # Buscar todos los libros de retenciones emitidas del directorio
    for path in base_dir.glob("RETEN-EMIT-*.xlsx"):
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            headers = excel_store._headers_index(ws)
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                fecha_cell = excel_store._cell(row, headers, "Fecha_emision", None)
                fecha_doc = _parse_row_date(fecha_cell)
                
                if fecha_doc is None or fecha_doc < start_date or fecha_doc > end_date:
                    continue
                
                ret_cell = excel_store._cell(row, headers, "IVA_retenido_total", None)
                ret = excel_store._parse_monto_cell(ret_cell) or Decimal("0")
                
                total_ret += ret
                total_count += 1
            wb.close()
        except Exception as e:
            logger.exception("Error leyendo retenciones emitidas desde %s", path)
            
    return total_ret, total_count


def get_islr_withholdings_totals(start_date: date, end_date: date) -> tuple[Decimal, int]:
    """
    Calcula totales de ISLR retenido en compras a proveedores (Base Imponible, ISLR retenido, número de registros)
    a partir de los Excels mensuales en RETENCIONES_ISLR_DIR.
    """
    total_islr = Decimal("0")
    total_count = 0
    base_dir = config.RETENCIONES_ISLR_DIR
    if not base_dir.is_dir():
        return total_islr, total_count

    for path in base_dir.glob("RETEN-ISLR-*.xlsx"):
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            headers = excel_store._headers_index(ws)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                fecha_cell = excel_store._cell(row, headers, "Fecha_emision", None)
                fecha_doc = _parse_row_date(fecha_cell)
                if fecha_doc is None or fecha_doc < start_date or fecha_doc > end_date:
                    continue
                ret_cell = excel_store._cell(row, headers, "ISLR_retenido", None)
                ret_val = excel_store._parse_monto_cell(ret_cell) or Decimal("0")
                total_islr += ret_val
                total_count += 1
            wb.close()
        except Exception as e:
            logger.exception("Error leyendo retenciones ISLR desde %s", path)

    return total_islr, total_count


def get_seniat_due_date(year: int, month: int, fortnight: int) -> date:
    """
    Calcula la fecha de declaración y pago del SENIAT para RIF terminado en 3.
    Utiliza el Calendario Fiscal Oficial del SENIAT para el año 2026 (Gaceta Oficial 43.273).
    """
    # Calendario RIF terminado en 3 para el año 2026
    # Tabla 1: "Entre el día 01 y 15" (1ra Quincena, se declara el mismo mes)
    CALENDARIO_Q1_2026 = {
        1: 30,  # Enero
        2: 18,  # Febrero
        3: 23,  # Marzo
        4: 30,  # Abril
        5: 22,  # Mayo
        6: 18,  # Junio
        7: 23,  # Julio
        8: 18,  # Agosto
        9: 21,  # Septiembre
        10: 23, # Octubre
        11: 23, # Noviembre
        12: 28, # Diciembre
    }

    # Tabla 2: "Entre el día 16 y último" (2da Quincena, se declara el mes de declaración)
    # Mapeado por el mes de declaración (mes siguiente al de las operaciones)
    CALENDARIO_Q2_2026 = {
        1: 16,  # Enero
        2: 12,  # Febrero
        3: 16,  # Marzo
        4: 16,  # Abril
        5: 14,  # Mayo
        6: 15,  # Junio
        7: 15,  # Julio
        8: 14,  # Agosto
        9: 15,  # Septiembre
        10: 15, # Octubre
        11: 16, # Noviembre
        12: 15, # Diciembre
    }

    if fortnight == 1:
        # Se declara en el mismo mes de las operaciones
        day = CALENDARIO_Q1_2026.get(month, 22)
        return date(year, month, day)
    else:
        # Se declara en el mes siguiente
        if month == 12:
            decl_year = year + 1
            decl_month = 1
        else:
            decl_year = year
            decl_month = month + 1
        
        day = CALENDARIO_Q2_2026.get(decl_month, 5)
        return date(decl_year, decl_month, day)


def get_compromiso_tributario_report(year: int, month: int, fortnight: int) -> dict[str, object]:
    """
    Genera un informe completo estructurado de todos los compromisos tributarios
    de la quincena indicada.
    """
    start_date, end_date = get_fortnight_range(year, month, fortnight)
    
    # 1. Ventas
    v_base, v_iva, v_count = get_sales_totals(start_date, end_date)
    
    # 2. Compras
    c_base, c_iva, c_count = get_purchases_totals(start_date, end_date)
    
    # 3. Retenciones Recibidas
    ret_rec, ret_rec_count = get_withholdings_received_totals(start_date, end_date)
    
    # 4. Retenciones Emitidas
    ret_emi, ret_emi_count = get_withholdings_issued_totals(start_date, end_date)
    
    # 4b. Retenciones de ISLR a proveedores en compras
    ret_islr_compras, ret_islr_compras_count = get_islr_withholdings_totals(start_date, end_date)
    
    # 5. Cálculos netos de IVA por pagar
    # IVA a pagar = IVA Débito (Ventas) - IVA Crédito (Compras) - Retenciones Recibidas (Clientes)
    iva_neto = v_iva - c_iva - ret_rec
    
    # 6. Anticipo de ISLR
    anticipo_islr = v_base * ALICUOTA_ANTICIPO_ISLR
    
    # 7. Total general de compromisos de la quincena
    # IVA Neto + Retenciones Emitidas a Enterar + Anticipo de ISLR + Retenciones ISLR
    # Nota: Si el IVA Neto es negativo (excedente), para el pago al fisco se toma como 0.00
    pago_iva = max(Decimal("0"), iva_neto)
    total_pagos = pago_iva + ret_emi + anticipo_islr + ret_islr_compras
    
    # 8. Fecha límite del SENIAT
    due_date = get_seniat_due_date(year, month, fortnight)
    
    return {
        "year": year,
        "month": month,
        "fortnight": fortnight,
        "start_date": start_date,
        "end_date": end_date,
        "due_date": due_date,
        "ventas_base": v_base,
        "ventas_iva": v_iva,
        "ventas_count": v_count,
        "compras_base": c_base,
        "compras_iva": c_iva,
        "compras_count": c_count,
        "retenciones_recibidas": ret_rec,
        "retenciones_recibidas_count": ret_rec_count,
        "retenciones_emitidas": ret_emi,
        "retenciones_emitidas_count": ret_emi_count,
        "retenciones_islr_compras": ret_islr_compras,
        "retenciones_islr_compras_count": ret_islr_compras_count,
        "iva_neto_pagar": iva_neto,
        "iva_neto_pagar_efectivo": pago_iva,
        "anticipo_islr": anticipo_islr,
        "anticipo_islr_alicuota": ALICUOTA_ANTICIPO_ISLR,
        "total_compromisos_a_pagar": total_pagos,
    }


def clean_rif(rif: str) -> str:
    """Elimina guiones, puntos y espacios de un RIF para la normativa del SENIAT."""
    if not rif:
        return ""
    return re.sub(r"[^A-Za-z0-9]", "", str(rif)).strip().upper()


def map_document_type(doc_type: str) -> str:
    """Mapea el tipo de documento de texto a códigos SENIAT (01=Factura, 02=N/D, 03=N/C)."""
    s = str(doc_type).strip().lower()
    if "debito" in s:
        return "02"
    elif "credito" in s:
        return "03"
    else:
        return "01"


def format_seniat_txt_line(
    agent_rif: str,
    periodo: str,
    invoice_date: str,
    operation_type: str,
    doc_type: str,
    rif: str,
    doc_num: str,
    control_num: str,
    total_amount: Decimal,
    base_imponible: Decimal,
    iva_retenido: Decimal,
    doc_affected: str,
    comprobante_num: str,
    exento: Decimal,
    alicuota: Decimal,
    expediente: str = "0"
) -> str:
    """Genera una línea tabulada formateada según las normas estrictas del portal SENIAT."""
    clean_agent_rif = clean_rif(agent_rif)
    clean_rif_sujeto = clean_rif(rif)
    
    clean_doc_num = re.sub(r"[^a-zA-Z0-9\-]", "", str(doc_num)).strip()
    clean_control_num = re.sub(r"[^a-zA-Z0-9\-]", "", str(control_num)).strip()
    clean_doc_affected = re.sub(r"[^a-zA-Z0-9\-]", "", str(doc_affected)).strip() if doc_affected and str(doc_affected) != "0" else "0"
    
    total_str = f"{total_amount:.2f}"
    base_str = f"{base_imponible:.2f}"
    iva_ret_str = f"{iva_retenido:.2f}"
    exento_str = f"{exento:.2f}"
    alicuota_str = f"{alicuota:.2f}"
    
    return (
        f"{clean_agent_rif}\t"
        f"{periodo}\t"
        f"{invoice_date}\t"
        f"{operation_type}\t"
        f"{doc_type}\t"
        f"{clean_rif_sujeto}\t"
        f"{clean_doc_num}\t"
        f"{clean_control_num}\t"
        f"{total_str}\t"
        f"{base_str}\t"
        f"{iva_ret_str}\t"
        f"{clean_doc_affected}\t"
        f"{comprobante_num}\t"
        f"{exento_str}\t"
        f"{alicuota_str}\t"
        f"{expediente}\n"
    )


def generate_seniat_txt_data(year: int, month: int, fortnight: int) -> tuple[str, str]:
    """
    Genera el contenido para los archivos TXT de declaración de IVA (Emitidas y Recibidas)
    para una quincena específica.
    """
    start_date, end_date = get_fortnight_range(year, month, fortnight)
    periodo_fiscal = f"{year}{month:02d}"
    
    emitidas_lines = []
    recibidas_lines = []
    
    # ----------------------------------------------------
    # 1. RETENCIONES EMITIDAS (Compras / Proveedores)
    # ----------------------------------------------------
    base_dir = config.RETENCIONES_EMITIDAS_DIR
    if base_dir.is_dir():
        for path in sorted(base_dir.glob("RETEN-EMIT-*.xlsx")):
            try:
                wb = load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                headers = excel_store._headers_index(ws)
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue
                    fecha_cell = excel_store._cell(row, headers, "Fecha_emision", None)
                    fecha_doc = _parse_row_date(fecha_cell)
                    
                    if fecha_doc is None or fecha_doc < start_date or fecha_doc > end_date:
                        continue
                        
                    num_comprobante = str(excel_store._cell(row, headers, "Numero_comprobante", "")).strip()
                    proveedor_rif = str(excel_store._cell(row, headers, "Proveedor_RIF", "")).strip()
                    documentos = str(excel_store._cell(row, headers, "Documentos", "")).strip()
                    controles = str(excel_store._cell(row, headers, "Controles", "")).strip()
                    
                    base_imponible_total = excel_store._parse_monto_cell(excel_store._cell(row, headers, "Base_imponible_total", None)) or Decimal("0")
                    iva_total = excel_store._parse_monto_cell(excel_store._cell(row, headers, "IVA_total", None)) or Decimal("0")
                    iva_retenido_total = excel_store._parse_monto_cell(excel_store._cell(row, headers, "IVA_retenido_total", None)) or Decimal("0")
                    porcentaje_retencion = excel_store._parse_monto_cell(excel_store._cell(row, headers, "Porcentaje_retencion", None)) or Decimal("0.75")
                    
                    # Normalizar porcentaje (ej: si está guardado como 75 en vez de 0.75)
                    if porcentaje_retencion > Decimal("1.0"):
                        porcentaje_retencion = porcentaje_retencion / Decimal("100.0")
                        
                    docs = [d.strip() for d in re.split(r"[|,]", documentos) if d.strip()]
                    ctrls = [c.strip() for c in re.split(r"[|,]", controles) if c.strip()]
                    
                    # Intentar cargar detalles desde el libro de facturas recibidas (compras)
                    items = excel_store.load_facturas_by_document_numbers(config.FACTURAS_RECIBIDAS_PATH, docs)
                    items_by_doc = {str(item.numero_documento).strip().upper(): item for item in items}
                    
                    N = len(docs) or 1
                    for idx, doc_num in enumerate(docs):
                        item = items_by_doc.get(str(doc_num).strip().upper())
                        control = ctrls[idx] if idx < len(ctrls) else ("-" if not item else item.numero_control)
                        
                        if item:
                            invoice_date_parsed = _parse_row_date(item.fecha_emision)
                            invoice_date_str = invoice_date_parsed.strftime("%d/%m/%Y") if invoice_date_parsed else str(fecha_cell)
                            doc_type = map_document_type(item.tipo_documento)
                            
                            base = item.base_imponible or Decimal("0")
                            iva = item.monto_iva or Decimal("0")
                            exento = item.monto_exento or Decimal("0")
                            total = item.total or (base + iva + exento)
                            iva_ret = (iva * porcentaje_retencion).quantize(Decimal("0.01"))
                        else:
                            # Fallback: Distribuir equitativamente
                            invoice_date_str = str(fecha_cell)
                            doc_type = "01"
                            
                            base = (base_imponible_total / N).quantize(Decimal("0.01"))
                            iva = (iva_total / N).quantize(Decimal("0.01"))
                            iva_ret = (iva_retenido_total / N).quantize(Decimal("0.01"))
                            exento = Decimal("0.00")
                            total = base + iva
                            
                        alicuota = Decimal("16.00")
                        if base > 0:
                            alicuota = ((iva / base) * 100).quantize(Decimal("0.01"))
                            
                        line = format_seniat_txt_line(
                            agent_rif=config.EMITTER_RIF,
                            periodo=periodo_fiscal,
                            invoice_date=invoice_date_str,
                            operation_type="C",
                            doc_type=doc_type,
                            rif=proveedor_rif,
                            doc_num=doc_num,
                            control_num=control,
                            total_amount=total,
                            base_imponible=base,
                            iva_retenido=iva_ret,
                            doc_affected="0",
                            comprobante_num=num_comprobante,
                            exento=exento,
                            alicuota=alicuota,
                        )
                        emitidas_lines.append(line)
                wb.close()
            except Exception as e:
                logger.exception("Error leyendo retenciones emitidas para TXT en %s", path)
                
    # ----------------------------------------------------
    # 2. RETENCIONES RECIBIDAS (Ventas / Clientes)
    # ----------------------------------------------------
    path_r = config.EXCEL_PATH
    if path_r.exists():
        try:
            wb = load_workbook(path_r, read_only=True, data_only=True)
            ws = wb.active
            headers = excel_store._headers_index(ws)
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                fecha_cell = excel_store._cell(row, headers, "Fecha_emision", None)
                fecha_doc = _parse_row_date(fecha_cell)
                
                if fecha_doc is None or fecha_doc < start_date or fecha_doc > end_date:
                    continue
                    
                numero_comprobante = str(excel_store._cell(row, headers, "Numero_comprobante", "")).strip()
                rif = str(excel_store._cell(row, headers, "RIF", "")).strip()
                fechas_facturas = str(excel_store._cell(row, headers, "Fechas_facturas", "")).strip()
                numeros_facturas = str(excel_store._cell(row, headers, "Numeros_facturas", "")).strip()
                controles_facturas = str(excel_store._cell(row, headers, "Controles_facturas", "")).strip()
                
                total_compra_con_iva = excel_store._parse_monto_cell(excel_store._cell(row, headers, "Total_compra_con_iva", None)) or Decimal("0")
                base_imponible = excel_store._parse_monto_cell(excel_store._cell(row, headers, "Base_imponible", None)) or Decimal("0")
                iva_retenido = excel_store._parse_monto_cell(excel_store._cell(row, headers, "IVA_retenido", None)) or Decimal("0")
                
                dates = [d.strip() for d in re.split(r"[|,]", fechas_facturas) if d.strip()]
                numbers = [n.strip() for n in re.split(r"[|,]", numeros_facturas) if n.strip()]
                controls = [c.strip() for c in re.split(r"[|,]", controles_facturas) if c.strip()]
                
                N = max(len(numbers), 1)
                for idx in range(N):
                    doc_num = numbers[idx] if idx < len(numbers) else ("-" if N == 1 else f"FAC-REF-{idx}")
                    control = controls[idx] if idx < len(controls) else "-"
                    
                    inv_date_raw = dates[idx] if idx < len(dates) else None
                    inv_date_parsed = _parse_row_date(inv_date_raw)
                    inv_date_str = inv_date_parsed.strftime("%d/%m/%Y") if inv_date_parsed else str(fecha_cell)
                    
                    base = (base_imponible / N).quantize(Decimal("0.01"))
                    iva_ret = (iva_retenido / N).quantize(Decimal("0.01"))
                    total = (total_compra_con_iva / N).quantize(Decimal("0.01"))
                    
                    # Estimación de IVA y Exento para ventas
                    iva = (base * Decimal("0.16")).quantize(Decimal("0.01"))
                    exento = max(Decimal("0.00"), total - base - iva)
                    
                    alicuota = Decimal("16.00")
                    if base > 0:
                        alicuota = ((iva / base) * 100).quantize(Decimal("0.01"))
                        
                    line = format_seniat_txt_line(
                        agent_rif=config.EMITTER_RIF,
                        periodo=periodo_fiscal,
                        invoice_date=inv_date_str,
                        operation_type="V",
                        doc_type="01",
                        rif=rif,
                        doc_num=doc_num,
                        control_num=control,
                        total_amount=total,
                        base_imponible=base,
                        iva_retenido=iva_ret,
                        doc_affected="0",
                        comprobante_num=numero_comprobante,
                        exento=exento,
                        alicuota=alicuota,
                    )
                    recibidas_lines.append(line)
        except Exception as e:
            logger.exception("Error leyendo retenciones recibidas para TXT en %s", path_r)
            
    return "".join(emitidas_lines), "".join(recibidas_lines)


def validar_rif_venezolano(rif: str) -> bool:
    """
    Valida matemáticamente un RIF venezolano usando el algoritmo del Módulo 11 (dígito verificador).
    """
    if not rif:
        return False
    s = re.sub(r"[^A-Za-z0-9]", "", str(rif)).strip().upper()
    if len(s) != 10:
        return False
    letra = s[0]
    base_letras = {'V': 4, 'E': 8, 'J': 12, 'C': 12, 'P': 16, 'G': 20}
    if letra not in base_letras:
        return False
    
    factores = [3, 2, 7, 6, 5, 4, 3, 2]
    try:
        numeros = [int(x) for x in s[1:-1]]
        digito_ingresado = int(s[-1])
    except ValueError:
        return False
        
    suma = base_letras[letra]
    for n, f in zip(numeros, factores):
        suma += n * f
        
    residuo = suma % 11
    digito_calc = 0 if residuo == 0 else (11 - residuo)
    if digito_calc >= 10:
        digito_calc = 0
        
    return digito_calc == digito_ingresado


def obtener_alicuota_islr_sugerida(concepto: str) -> Decimal:
    """
    Retorna el porcentaje sugerido de retención de ISLR según la categoría del concepto de servicio (Decreto 1808).
    Retorna un valor Decimal de tasa (ej: Decimal("0.02") para 2%).
    """
    if not concepto:
        return Decimal("0.00")
    c = excel_store._normalize_text(str(concepto))
    if any(k in c for k in ("honorario", "profesional", "medico", "ingenier", "abogad", "contad")):
        return Decimal("0.03")  # Honorarios profesionales (persona jurídica residente: 3%)
    elif any(k in c for k in ("publicidad", "propaganda", "valla", "anuncio")):
        return Decimal("0.05")  # Publicidad y propaganda (5%)
    elif any(k in c for k in ("arrendamiento", "alquiler", "arriend")):
        return Decimal("0.05")  # Arrendamiento de inmuebles (jurídico: 5%)
    elif any(k in c for k in ("transporte", "flete", "carga", "acarreo")):
        return Decimal("0.03")  # Fletes / transporte (jurídico: 3%)
    elif any(k in c for k in ("comision", "corretaje")):
        return Decimal("0.05")  # Comisiones mercantiles (5%)
    elif any(k in c for k in ("servicio", "mantenimiento", "reparacion", "mano de obra", "instalacion", "contratista", "obra")):
        return Decimal("0.02")  # Prestación de servicios / contratistas en general (jurídico: 2%)
    else:
        return Decimal("0.00")  # Compra de mercancía u otros no sujetos a retención (0%)

